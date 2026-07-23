"""
Flask backend application for IUT Gestion Emploi Du Temps (EDT)
Manages timetables for IUT GIM Toulon
"""

from flask import Flask, request, jsonify, send_file, send_from_directory, g, session, redirect
from flask_cors import CORS
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import os
import re
import json
import math
import shutil
import hmac
import time
import threading
import logging
from logging.handlers import RotatingFileHandler
from datetime import datetime, timedelta
from pathlib import Path
import io
import zipfile
import tempfile


def _load_local_env():
    """Charge un fichier .env local (lignes KEY=VALUE) s'il existe, pour le
    développement. N'écrase jamais une variable déjà définie dans l'environnement :
    en production, définissez les variables côté hébergeur (le .env n'y est pas)."""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
    if not os.path.isfile(env_path):
        return
    try:
        with open(env_path, encoding='utf-8') as fh:
            for line in fh:
                line = line.strip()
                if not line or line.startswith('#') or '=' not in line:
                    continue
                key, _, value = line.partition('=')
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = value
    except OSError:
        pass


_load_local_env()

# Initialize Flask app
app = Flask(__name__, static_folder='static', static_url_path='')
app.secret_key = os.environ.get('EDT_SECRET_KEY', 'edt-gim-toulon-secret-2026')
# Durcissement du cookie de session :
# - HttpOnly : inaccessible au JavaScript (anti-vol via XSS)
# - SameSite=Lax : limite l'envoi du cookie en cross-site (anti-CSRF)
# - Secure : cookie envoyé uniquement en HTTPS (activer en prod via EDT_SECURE_COOKIES=1)
# - durée de vie : session expirée après 12 h
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=os.environ.get('EDT_SECURE_COOKIES', '0') == '1',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
)
# CORS restreint : la SPA est servie en same-origin par Flask, donc aucune origine
# externe n'est autorisée par défaut. Pour exposer l'API à un autre domaine,
# lister les origines dans EDT_ALLOWED_ORIGINS (séparées par des virgules).
_cors_origins = [o.strip() for o in os.environ.get('EDT_ALLOWED_ORIGINS', '').split(',') if o.strip()]
if _cors_origins:
    CORS(app, supports_credentials=True, origins=_cors_origins)

# ======================= AUTHENTIFICATION =======================
# Comptes : Admin (tous droits) + connexion enseignant par nom (lecture seule).
# Le mot de passe admin peut être surchargé via la variable d'environnement
# EDT_ADMIN_PASSWORD (recommandé en production pour ne pas l'avoir en clair).
AUTH_USERS = {
    'Admin': {
        'password': os.environ.get('EDT_ADMIN_PASSWORD', 'GestionEDT22#'),
        'role': 'admin',
    },
}
# Routes API accessibles sans être connecté
_PUBLIC_API = {'/api/login', '/api/logout', '/api/me',
               '/api/password-init', '/api/password-init/status'}

# Database configuration
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.environ.get('EDT_DB_PATH', os.path.join(_BASE_DIR, 'edt.db'))
SCHEMA_PATH = os.path.join(_BASE_DIR, 'schema.sql')
# Dossier racine des bases, rangées par sous-dossier d'année : databases/<année>/edt_<année>.db
_DB_DIR = os.environ.get('EDT_DB_DIR') or os.path.join(_BASE_DIR, 'databases')
os.makedirs(_DB_DIR, exist_ok=True)
# Dossier de stockage des fichiers de contraintes déposés par les enseignants
_CONSTRAINTS_DIR = os.environ.get('EDT_CONSTRAINTS_DIR') or os.path.join(_BASE_DIR, 'uploads', 'contraintes')

# ===== JOURNAL D'AUDIT (connexions + modifications) =====
# Écrit dans logs/audit.log (rotation 1 Mo x 5). Lisible par l'admin via /api/audit-log.
_LOG_DIR = os.path.join(_BASE_DIR, 'logs')
os.makedirs(_LOG_DIR, exist_ok=True)
AUDIT_LOG_PATH = os.path.join(_LOG_DIR, 'audit.log')
audit_log = logging.getLogger('edt.audit')
if not audit_log.handlers:
    audit_log.setLevel(logging.INFO)
    _audit_handler = RotatingFileHandler(AUDIT_LOG_PATH, maxBytes=1_000_000,
                                         backupCount=5, encoding='utf-8')
    _audit_handler.setFormatter(logging.Formatter('%(asctime)s | %(message)s'))
    audit_log.addHandler(_audit_handler)
    audit_log.propagate = False

def _audit(event, **fields):
    """Ajoute une ligne au journal d'audit (jamais bloquant)."""
    try:
        parts = [event] + [f'{k}={v}' for k, v in fields.items()]
        audit_log.info(' | '.join(str(p) for p in parts))
    except Exception:
        pass

# ===== GESTION DES ANNÉES UNIVERSITAIRES =====
SETTINGS_PATH = os.path.join(_BASE_DIR, 'edt_settings.json')
_settings_cache = None

def get_settings():
    global _settings_cache
    if _settings_cache is None:
        if os.path.exists(SETTINGS_PATH):
            with open(SETTINGS_PATH, 'r') as f:
                _settings_cache = json.load(f)
        else:
            _settings_cache = {}
    return _settings_cache

def save_settings(s):
    global _settings_cache
    _settings_cache = s
    with open(SETTINGS_PATH, 'w') as f:
        json.dump(s, f, indent=2)

def auto_detect_year():
    today = datetime.now()
    y = today.year
    return f"{y}-{y+1}" if today.month >= 8 else f"{y-1}-{y}"

def _is_year(name):
    """Vrai si name a le format AAAA-AAAA (ex : 2025-2026)."""
    return (len(name) == 9 and name[4] == '-'
            and name[:4].isdigit() and name[5:].isdigit())

def db_path_for_year(year):
    return os.path.join(_DB_DIR, year, f'edt_{year}.db')

def get_current_year():
    """Année universitaire active pour la requête courante.
    Priorité : override de requête (?year=, ex. calendrier cohorte pluriannuel)
    > année choisie dans la session (vue perso) > défaut global.
    Hors contexte requête (init, backups, scripts) : défaut global."""
    try:
        ov = getattr(g, '_year_override', None)
        if ov:
            return ov
    except RuntimeError:
        pass
    try:
        y = session.get('year')
        if y and os.path.isfile(db_path_for_year(y)):
            return y
    except RuntimeError:
        pass  # hors contexte requête Flask
    return get_settings().get('current_year', '')

def list_years():
    years = []
    if os.path.isdir(_DB_DIR):
        for name in os.listdir(_DB_DIR):
            if _is_year(name) and os.path.isfile(db_path_for_year(name)):
                years.append(name)
    return sorted(years)

# ===== CRÉATION DES ANNÉES UNIVERSITAIRES D'UNE PROMOTION =====
# Le contexte de travail est l'année universitaire (base edt_<année>.db, toutes
# cohortes S1→S6). Créer une promotion crée/complète ses 3 années universitaires
# en copiant l'année existante la plus récente qui la précède.
def _copy_db_with_migrations(src, dest):
    """Copie une base SQLite puis met à jour son schéma."""
    os.makedirs(os.path.dirname(dest), exist_ok=True)
    shutil.copy2(src, dest)
    db = sqlite3.connect(dest)
    db.row_factory = sqlite3.Row
    _apply_migrations(db)
    db.close()

def _ensure_academic_year_db(year):
    """Crée la base d'une année universitaire si absente. Source de copie :
    l'année existante la plus proche AVANT (sinon la plus proche APRÈS), sinon
    base vierge. Retourne le chemin."""
    path = db_path_for_year(year)
    if os.path.exists(path):
        return path
    try:
        start = int(str(year).split('-')[0])
    except (ValueError, IndexError):
        start = None
    src = None
    if start is not None:
        years = sorted(list_years(), key=lambda y: int(y.split('-')[0]))
        prior = [y for y in years if int(y.split('-')[0]) < start]
        later = [y for y in years if int(y.split('-')[0]) > start]
        if prior:
            src = db_path_for_year(prior[-1])   # année précédente la plus proche
        elif later:
            src = db_path_for_year(later[0])    # sinon année suivante la plus proche
    if src and os.path.exists(src):
        _copy_db_with_migrations(src, path)
    else:
        _init_fresh_db(path)
    return path

def _ensure_years_for_promo(start_year):
    """Crée/complète les 3 années universitaires couvertes par une promotion
    (start→+1, +1→+2, +2→+3), chacune copiée de la précédente. Retourne la liste créée."""
    created = []
    for i in range(3):
        year = f"{start_year + i}-{start_year + i + 1}"
        if not os.path.exists(db_path_for_year(year)):
            _ensure_academic_year_db(year)
            created.append(year)
    return created

def ensure_years_for_all_promotions():
    """Crée les années universitaires manquantes de TOUTES les promotions
    existantes (backfill idempotent). Traite les promos de la plus ancienne à la
    plus récente pour que les copies se propagent correctement."""
    try:
        rows = get_promotions_db().execute('SELECT start_year FROM promotions').fetchall()
    except sqlite3.Error:
        return []
    created = []
    for start in sorted(r['start_year'] for r in rows):
        created += _ensure_years_for_promo(start)
    return created

def _migrate_db_layout():
    """Déplace les bases edt_<année>.db de la racine vers databases/<année>/.
    Migration unique et idempotente, exécutée au démarrage."""
    for fname in os.listdir(_BASE_DIR):
        if not (fname.startswith('edt_') and fname.endswith('.db') and len(fname) == 16):
            continue
        year = fname[4:-3]
        if not _is_year(year):
            continue
        target = db_path_for_year(year)
        if os.path.exists(target):
            continue
        os.makedirs(os.path.dirname(target), exist_ok=True)
        # Déplace aussi les fichiers WAL/SHM associés s'ils existent
        for suffix in ('', '-wal', '-shm'):
            src = os.path.join(_BASE_DIR, fname + suffix)
            if os.path.exists(src):
                try:
                    shutil.move(src, target + suffix)
                except OSError:
                    pass

# ===== SAUVEGARDES (BACKUPS) DES BASES =====
# Backups rangés dans databases/<année>/backups/edt_<année>_<horodatage>.db
_BACKUP_KEEP = 30   # nombre de sauvegardes conservées par année

def _backup_dir_for_year(year):
    return os.path.join(_DB_DIR, year, 'backups')

def _db_mtime(year):
    """Dernière modification de la base (inclut WAL/SHM), 0 si absente."""
    mt = 0.0
    for suffix in ('', '-wal', '-shm'):
        p = db_path_for_year(year) + suffix
        if os.path.exists(p):
            mt = max(mt, os.path.getmtime(p))
    return mt

def _latest_backup_mtime(year):
    bdir = _backup_dir_for_year(year)
    if not os.path.isdir(bdir):
        return 0.0
    times = [os.path.getmtime(os.path.join(bdir, f))
             for f in os.listdir(bdir) if f.endswith('.db')]
    return max(times) if times else 0.0

def _create_backup(year, tag=None):
    """Copie cohérente de la base de l'année via l'API backup SQLite."""
    src_path = db_path_for_year(year)
    if not os.path.isfile(src_path):
        return None
    bdir = _backup_dir_for_year(year)
    os.makedirs(bdir, exist_ok=True)
    ts = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    name = f"edt_{year}_{ts}" + (f"_{tag}" if tag else "") + ".db"
    dest = os.path.join(bdir, name)
    src = sqlite3.connect(src_path)
    dst = sqlite3.connect(dest)
    try:
        with dst:
            src.backup(dst)
    finally:
        src.close()
        dst.close()
    _prune_backups(year)
    return dest

def _prune_backups(year, keep=_BACKUP_KEEP):
    """Ne conserve que les `keep` sauvegardes les plus récentes."""
    bdir = _backup_dir_for_year(year)
    if not os.path.isdir(bdir):
        return
    files = sorted((f for f in os.listdir(bdir) if f.endswith('.db')),
                   key=lambda f: os.path.getmtime(os.path.join(bdir, f)),
                   reverse=True)
    for f in files[keep:]:
        try:
            os.remove(os.path.join(bdir, f))
        except OSError:
            pass

def backup_if_modified(year):
    """Crée une sauvegarde uniquement si la base a changé depuis la dernière.
    Retourne le chemin créé, ou None si rien à sauvegarder."""
    if _db_mtime(year) <= _latest_backup_mtime(year):
        return None
    return _create_backup(year)

def backup_all_if_modified():
    """Sauvegarde chaque année modifiée. Retourne la liste des fichiers créés."""
    created = []
    for year in list_years():
        path = backup_if_modified(year)
        if path:
            created.append(path)
            _audit('BACKUP', year=year, file=os.path.basename(path))
    return created

# Déclenchement « sans cron » : compte gratuit PythonAnywhere = pas de tâche
# planifiée. À la 1re requête passé l'heure de backup (4h par défaut) chaque
# jour, on lance la sauvegarde conditionnelle. Le script scripts/backup_db.py
# reste utilisable pour une vraie planification (compte payant).
_BACKUP_HOUR = int(os.environ.get('EDT_BACKUP_HOUR', '4'))
_backup_check_lock = threading.Lock()
_last_backup_check_date = None

def _maybe_daily_backup():
    """Lance au plus une vérification de sauvegarde par jour (après l'heure cible)."""
    global _last_backup_check_date
    now = datetime.now()
    boundary = now.date() if now.hour >= _BACKUP_HOUR else (now.date() - timedelta(days=1))
    if _last_backup_check_date == boundary:
        return
    with _backup_check_lock:
        if _last_backup_check_date == boundary:
            return
        _last_backup_check_date = boundary   # une seule tentative par jour, même en cas d'erreur
        try:
            backup_all_if_modified()
        except Exception as e:
            _audit('BACKUP_ERROR', error=str(e))

# ===== SAUVEGARDE COMPLÈTE DU SITE (export/import ZIP, onglet Paramètres) =====
# Contenu de l'archive : bases de toutes les années + promotions + programmes,
# fichiers de notes et de contraintes, réglages globaux. Sert de plan de
# reprise (site compromis, changement de serveur…).
_FULLBACKUP_VERSION = 1

def _sqlite_snapshot_bytes(src_path):
    """Copie cohérente d'une base SQLite (API backup), renvoyée en bytes."""
    fd, tmp = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    try:
        src = sqlite3.connect(src_path)
        dst = sqlite3.connect(tmp)
        try:
            with dst:
                src.backup(dst)
        finally:
            src.close()
            dst.close()
        with open(tmp, 'rb') as fh:
            return fh.read()
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass

def _iter_dir_files(root, prefix):
    """(chemin absolu, nom dans l'archive) pour chaque fichier sous root."""
    if not os.path.isdir(root):
        return
    for dirpath, _dirs, files in os.walk(root):
        for f in files:
            p = os.path.join(dirpath, f)
            rel = os.path.relpath(p, root).replace(os.sep, '/')
            yield p, prefix + '/' + rel

def _fullbackup_safe_arcname(name):
    """Vrai si l'entrée du zip est attendue et sans traversée de chemin."""
    if name.endswith('/'):
        return False    # entrées répertoire ignorées
    if '\\' in name or name.startswith('/') or '..' in name.split('/'):
        return False
    if name in ('manifest.json', 'edt_settings.json', 'data/coefficients.json'):
        return True
    if name.startswith('databases/'):
        parts = name.split('/')
        if len(parts) == 2 and parts[1] in ('promotions.db', 'programmes.db'):
            return True
        return len(parts) == 3 and _is_year(parts[1]) and parts[2] == f'edt_{parts[1]}.db'
    return name.startswith('uploads/notes/') or name.startswith('uploads/contraintes/')

def _remove_wal_shm(db_path):
    """Supprime les fichiers WAL/SHM d'une base avant son remplacement
    (un WAL périmé corromprait la base restaurée)."""
    for suffix in ('-wal', '-shm'):
        try:
            if os.path.exists(db_path + suffix):
                os.remove(db_path + suffix)
        except OSError:
            pass

def _stash_dir(d):
    """Met de côté un dossier (renommé en .avant-import) avant restauration."""
    if not os.path.isdir(d):
        return
    stash = d + '.avant-import'
    shutil.rmtree(stash, ignore_errors=True)
    try:
        os.rename(d, stash)
    except OSError:
        shutil.rmtree(d, ignore_errors=True)

@app.route('/api/backup/full', methods=['GET'])
def full_backup_export():
    err = _require_admin()
    if err:
        return err
    years = list_years()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
        for y in years:
            z.writestr(f'databases/{y}/edt_{y}.db',
                       _sqlite_snapshot_bytes(db_path_for_year(y)))
        for name, path in (('promotions.db', _PROMOTIONS_DB),
                           ('programmes.db', _PROGRAMMES_DB)):
            if os.path.isfile(path):
                z.writestr(f'databases/{name}', _sqlite_snapshot_bytes(path))
        for p, arc in _iter_dir_files(_NOTES_DIR, 'uploads/notes'):
            z.write(p, arc)
        for p, arc in _iter_dir_files(_CONSTRAINTS_DIR, 'uploads/contraintes'):
            z.write(p, arc)
        if os.path.isfile(SETTINGS_PATH):
            z.write(SETTINGS_PATH, 'edt_settings.json')
        if os.path.isfile(_REF_COEFF_FILE):
            z.write(_REF_COEFF_FILE, 'data/coefficients.json')
        z.writestr('manifest.json', json.dumps({
            'app': 'GestionEDT',
            'version': _FULLBACKUP_VERSION,
            'created_at': datetime.now().isoformat(timespec='seconds'),
            'years': years,
        }, indent=2))
    buf.seek(0)
    _audit('FULL_BACKUP_EXPORT', years=','.join(years))
    fname = 'EDT_sauvegarde_' + datetime.now().strftime('%Y-%m-%d_%H%M%S') + '.zip'
    return send_file(buf, mimetype='application/zip',
                     as_attachment=True, download_name=fname)

@app.route('/api/backup/full', methods=['POST'])
def full_backup_import():
    err = _require_admin()
    if err:
        return err
    f = request.files.get('file')
    if not f:
        return error_response('Fichier .zip requis (champ « file »)', 400)
    try:
        z = zipfile.ZipFile(io.BytesIO(f.read()))
    except zipfile.BadZipFile:
        return error_response('Fichier invalide : archive .zip attendue', 400)
    names = [n for n in z.namelist() if not n.endswith('/')]
    if 'manifest.json' not in names:
        return error_response('Archive non reconnue : manifest.json manquant', 400)
    try:
        manifest = json.loads(z.read('manifest.json').decode('utf-8'))
    except (ValueError, UnicodeDecodeError):
        return error_response('manifest.json illisible', 400)
    if manifest.get('app') != 'GestionEDT':
        return error_response('Archive non reconnue (application différente)', 400)
    bad = [n for n in names if not _fullbackup_safe_arcname(n)]
    if bad:
        return error_response('Entrées non autorisées dans l\'archive : '
                              + ', '.join(bad[:5]), 400)

    # Sauvegarde de sécurité de l'existant avant écrasement
    for y in list_years():
        try:
            _create_backup(y, tag='avant-import')
        except Exception:
            pass
    for path in (_PROMOTIONS_DB, _PROGRAMMES_DB):
        if os.path.isfile(path):
            try:
                shutil.copy2(path, path + '.avant-import')
            except OSError:
                pass
    # Notes / contraintes : l'archive fait foi → l'existant est mis de côté
    _stash_dir(_NOTES_DIR)
    _stash_dir(_CONSTRAINTS_DIR)

    imported_years = []
    for n in names:
        if n == 'manifest.json':
            continue
        if n == 'edt_settings.json':
            target = SETTINGS_PATH
        elif n == 'data/coefficients.json':
            target = _REF_COEFF_FILE
        elif n == 'databases/promotions.db':
            target = _PROMOTIONS_DB
        elif n == 'databases/programmes.db':
            target = _PROGRAMMES_DB
        elif n.startswith('databases/'):
            year = n.split('/')[1]
            target = db_path_for_year(year)
            imported_years.append(year)
        elif n.startswith('uploads/notes/'):
            target = os.path.join(_NOTES_DIR, *n.split('/')[2:])
        else:   # uploads/contraintes/
            target = os.path.join(_CONSTRAINTS_DIR, *n.split('/')[2:])
        os.makedirs(os.path.dirname(target), exist_ok=True)
        if target.endswith('.db'):
            _remove_wal_shm(target)
        with open(target, 'wb') as out:
            out.write(z.read(n))

    # Caches et migrations de schéma sur les bases restaurées
    global _settings_cache
    _settings_cache = None
    for y in imported_years:
        try:
            db = sqlite3.connect(db_path_for_year(y))
            db.row_factory = sqlite3.Row
            _apply_migrations(db)
            db.commit()
            db.close()
        except Exception as e:
            _audit('FULL_BACKUP_IMPORT_MIGRATE_ERROR', year=y, error=str(e))
    try:
        _init_promotions_db()
    except Exception:
        pass
    try:
        _init_programmes_db()
    except Exception:
        pass

    untouched = [y for y in list_years() if y not in imported_years]
    _audit('FULL_BACKUP_IMPORT', user=session.get('user'),
           years=','.join(imported_years), untouched=','.join(untouched))
    return jsonify({
        'ok': True,
        'years_imported': imported_years,
        'years_untouched': untouched,
        'files': len(names) - 1,
        'backup_created_at': manifest.get('created_at'),
    })

def school_week_key(w, start_week=36, max_week=52):
    """Clé de tri pour l'ordre scolaire : start_week est la semaine 0.
    max_week = dernière semaine de la 1re année civile (52 ou 53 selon le calendrier ISO)."""
    offset = max_week - start_week + 1
    return w - start_week if w >= start_week else w + offset

def weeks_in_iso_year(iso_year):
    """Nombre de semaines ISO d'une année civile (52 ou 53).
    Le 28 décembre appartient toujours à la dernière semaine ISO de l'année."""
    return datetime(iso_year, 12, 28).isocalendar()[1]

def get_academic_max_week(year=None):
    """Dernière semaine de la 1re année civile de l'année universitaire.
    Vaut 53 pour les années ISO à 53 semaines (ex: 2026 → année 2026-2027), sinon 52."""
    year = year or get_current_year()
    try:
        return weeks_in_iso_year(int(str(year).split('-')[0]))
    except (ValueError, IndexError):
        return 52

# Helper function to get database connection
def _open_connection(path=None):
    """Ouvre une connexion SQLite brute (helper interne).
    Contexte = année universitaire courante (base legacy en dernier recours)."""
    year = get_current_year()
    p = path or (db_path_for_year(year) if year else DATABASE)
    db = sqlite3.connect(p, timeout=10)
    db.row_factory = sqlite3.Row
    db.execute('PRAGMA foreign_keys = ON')
    return db

def get_db():
    """Connexion à la DB de l'année universitaire courante.
    En contexte requête Flask : réutilise la connexion via flask.g
    et la ferme automatiquement via teardown_appcontext.
    Hors contexte (init, scripts) : retourne une connexion directe
    que l'appelant doit fermer manuellement."""
    try:
        if '_db' not in g:
            g._db = _open_connection()
        return g._db
    except RuntimeError:
        # Hors contexte requête Flask (init_db au démarrage, etc.)
        return _open_connection()

@app.teardown_appcontext
def close_db(exception):
    db = g.pop('_db', None)
    if db is not None:
        db.close()
    pdb = g.pop('_pdb', None)
    if pdb is not None:
        pdb.close()
    grdb = g.pop('_grdb', None)
    if grdb is not None:
        grdb.close()

# ===== Base des PROMOTIONS (cohorte sur 3 ans, indépendante de l'année) =====
# Stockée à part car une promotion suit ses étudiants/résultats d'année en année.
_PROMOTIONS_DB = os.environ.get('EDT_PROMOTIONS_DB') or os.path.join(_DB_DIR, 'promotions.db')
_STUDENT_STATUSES = ['Actif', 'RED', 'Abandon']
_PROMO_SEMESTERS = ['S1', 'S2', 'S3', 'S4', 'S5', 'S6']
# Profil d'entrée de l'étudiant (valeurs autorisées ; '' = non renseigné)
_STUDENT_BAC = ['SI2D', 'GEN', 'PRO', 'STL', 'Autre']
_STUDENT_CURSUS = ['EI', 'RE', 'PB', 'PP']            # École d'ingé / Reprise d'étude / PostBac / PostPrépa
_STUDENT_RECRUT = ['PS', 'EC', 'ADIUT']               # ParcourSup / eCandidat / ADIUT (étrangers)
_STUDENT_PROFILE = {'bac': _STUDENT_BAC, 'cursus': _STUDENT_CURSUS, 'recrutement': _STUDENT_RECRUT}

def _open_promotions_db():
    db = sqlite3.connect(_PROMOTIONS_DB, timeout=10)
    db.row_factory = sqlite3.Row
    db.execute('PRAGMA foreign_keys = ON')
    return db

def get_promotions_db():
    """Connexion à la base des promotions (partagée entre toutes les années)."""
    try:
        if '_pdb' not in g:
            g._pdb = _open_promotions_db()
        return g._pdb
    except RuntimeError:
        return _open_promotions_db()

def _apply_promotions_migrations(db):
    db.execute('PRAGMA journal_mode=WAL')
    db.execute('''
        CREATE TABLE IF NOT EXISTS promotions (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT NOT NULL UNIQUE,   -- '25-28_ALT'
            formation  TEXT NOT NULL,          -- 'ALT' / 'FTP'
            start_year INTEGER NOT NULL,        -- 2025
            end_year   INTEGER NOT NULL,        -- 2028
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT
        )
    ''')
    db.execute('''
        CREATE TABLE IF NOT EXISTS promotion_students (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            promotion_id INTEGER NOT NULL,
            numero       TEXT,
            nom          TEXT,
            prenom       TEXT,
            naissance    TEXT,
            statut       TEXT DEFAULT 'Actif',
            created_at   TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (promotion_id) REFERENCES promotions(id) ON DELETE CASCADE
        )
    ''')
    # Semestre d'abandon (S1..S6), renseigné quand statut = 'Abandon'
    try:
        db.execute("ALTER TABLE promotion_students ADD COLUMN abandon_semestre TEXT")
    except sqlite3.OperationalError:
        pass
    # Programme (Programme National) affecté à la promotion. Référence vers
    # programmes.db (fichier séparé) → pas de FK SQL, validation applicative.
    try:
        db.execute("ALTER TABLE promotions ADD COLUMN programme_id INTEGER")
    except sqlite3.OperationalError:
        pass
    # Notes par étudiant : (promotion, semestre, étudiant, matière) -> note.
    # Les matières (codes) viennent de la référence des coefficients ; suivi pluriannuel.
    db.execute('''
        CREATE TABLE IF NOT EXISTS student_marks (
            promotion_id INTEGER NOT NULL,
            semester     TEXT NOT NULL,
            student_id   INTEGER NOT NULL,
            matiere_code TEXT NOT NULL,
            note         REAL,
            PRIMARY KEY (promotion_id, semester, student_id, matiere_code),
            FOREIGN KEY (promotion_id) REFERENCES promotions(id) ON DELETE CASCADE,
            FOREIGN KEY (student_id) REFERENCES promotion_students(id) ON DELETE CASCADE
        )
    ''')
    # Mention affichée à la place de la note (ex 'ABI' = absence, comptée 0)
    try:
        db.execute("ALTER TABLE student_marks ADD COLUMN mention TEXT")
    except sqlite3.OperationalError:
        pass
    # Coefficients (matières + compétences) propres à chaque promotion : copie JSON,
    # initialisée depuis la référence globale puis éditable indépendamment.
    db.execute('''
        CREATE TABLE IF NOT EXISTS promotion_coefficients (
            promotion_id INTEGER PRIMARY KEY,
            data         TEXT NOT NULL,
            updated_at   TEXT,
            FOREIGN KEY (promotion_id) REFERENCES promotions(id) ON DELETE CASCADE
        )
    ''')
    # Sous-cohorte de chaque étudiant : 'FTP' (initiale) ou 'ALT' (alternance).
    # Une promotion = une cohorte sur 3 ans contenant les 2 sous-cohortes.
    try:
        db.execute("ALTER TABLE promotion_students ADD COLUMN formation TEXT")
    except sqlite3.OperationalError:
        pass
    # Décisions de jury : ADMJ (admis par décision de jury) sur une UE ajournée.
    # year = niveau (1..3), ue_num = numéro canonique de l'UE (1..5, cf _jury_ue_numbers).
    db.execute('''
        CREATE TABLE IF NOT EXISTS jury_decisions (
            promotion_id INTEGER NOT NULL,
            year         INTEGER NOT NULL,
            student_id   INTEGER NOT NULL,
            ue_num       INTEGER NOT NULL,
            decision     TEXT NOT NULL,
            PRIMARY KEY (promotion_id, year, student_id, ue_num),
            FOREIGN KEY (promotion_id) REFERENCES promotions(id) ON DELETE CASCADE,
            FOREIGN KEY (student_id) REFERENCES promotion_students(id) ON DELETE CASCADE
        )
    ''')
    # Effectif par année : l'année d'ENTRÉE d'un étudiant dans la cohorte (1..3).
    # 1 = intake initial ; 2/3 = redoublant entrant directement dans cette année.
    # L'effectif d'une année est calculé (report auto depuis le jury), voir _year_rosters.
    try:
        db.execute("ALTER TABLE promotion_students ADD COLUMN entry_year INTEGER DEFAULT 1")
    except sqlite3.OperationalError:
        pass
    # Profil d'entrée : BAC obtenu, cursus antérieur, méthode de recrutement.
    for _col in ('bac', 'cursus', 'recrutement'):
        try:
            db.execute(f"ALTER TABLE promotion_students ADD COLUMN {_col} TEXT")
        except sqlite3.OperationalError:
            pass
    # Ajustements manuels de l'effectif d'une année (par-dessus le calcul auto) :
    # action='remove' (retiré de l'année) ou 'add' (réintégré / ajouté à l'année).
    db.execute('''
        CREATE TABLE IF NOT EXISTS promotion_year_override (
            promotion_id INTEGER NOT NULL,
            year         INTEGER NOT NULL,
            student_id   INTEGER NOT NULL,
            action       TEXT NOT NULL,
            PRIMARY KEY (promotion_id, year, student_id),
            FOREIGN KEY (promotion_id) REFERENCES promotions(id) ON DELETE CASCADE,
            FOREIGN KEY (student_id) REFERENCES promotion_students(id) ON DELETE CASCADE
        )
    ''')
    # Sort d'un redoublant (RED) décidé par le jury : réinscrit dans la promo cible
    # (FTP/ALT) ou Abandon. target_student_id = étudiant créé dans la promo cible.
    db.execute('''
        CREATE TABLE IF NOT EXISTS promotion_red_transfer (
            promotion_id      INTEGER NOT NULL,
            student_id        INTEGER NOT NULL,
            decision          TEXT NOT NULL,
            target_student_id INTEGER,
            PRIMARY KEY (promotion_id, student_id),
            FOREIGN KEY (promotion_id) REFERENCES promotions(id) ON DELETE CASCADE,
            FOREIGN KEY (student_id) REFERENCES promotion_students(id) ON DELETE CASCADE
        )
    ''')
    # Sous-cohorte (FTP/ALT) par année : un étudiant peut basculer d'une année sur l'autre.
    # Absence de ligne = formation de base (promotion_students.formation). Report auto :
    # la formation d'une année vaut le dernier changement d'une année <= (voir _year_formation_map).
    db.execute('''
        CREATE TABLE IF NOT EXISTS promotion_year_formation (
            promotion_id INTEGER NOT NULL,
            year         INTEGER NOT NULL,
            student_id   INTEGER NOT NULL,
            formation    TEXT NOT NULL,
            PRIMARY KEY (promotion_id, year, student_id),
            FOREIGN KEY (promotion_id) REFERENCES promotions(id) ON DELETE CASCADE,
            FOREIGN KEY (student_id) REFERENCES promotion_students(id) ON DELETE CASCADE
        )
    ''')
    # Statuts simplifiés (Actif / RED / Abandon) : migre les anciennes valeurs.
    db.execute("UPDATE promotion_students SET statut='RED' WHERE statut='Redoublant'")
    db.execute("UPDATE promotion_students SET statut='Abandon' WHERE statut='Sortie'")
    # Tuteurs (universitaire + entreprise) par étudiant et par année d'étude.
    # ALT : la ligne la plus récente ≤ année affichée s'applique (héritage sauf
    # changement) ; FTP (stages 2e/3e année) : uniquement la ligne de l'année exacte.
    db.execute('''
        CREATE TABLE IF NOT EXISTS student_tutors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            promotion_id INTEGER NOT NULL,
            student_id INTEGER NOT NULL,
            year INTEGER NOT NULL CHECK (year IN (1, 2, 3)),
            tuteur_univ TEXT,
            tuteur_entreprise TEXT,
            entreprise TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (student_id, year),
            FOREIGN KEY (promotion_id) REFERENCES promotions(id) ON DELETE CASCADE,
            FOREIGN KEY (student_id) REFERENCES promotion_students(id) ON DELETE CASCADE
        )
    ''')
    # Saisie des notes PAR SOUS-MATIÈRE (onglet Saisie Notes, enseignants).
    # La note matière (student_marks) = moyenne pondérée des sous-notes,
    # recalculée automatiquement quand toutes les sous-notes comptées sont là.
    db.execute('''
        CREATE TABLE IF NOT EXISTS submatiere_marks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            promotion_id INTEGER NOT NULL,
            semester TEXT NOT NULL,
            student_id INTEGER NOT NULL,
            code TEXT NOT NULL,             -- code de la sous-matière (ex R1.05a)
            note REAL,
            mention TEXT,                   -- 'ABI' (compte 0 dans la moyenne)
            UNIQUE (promotion_id, semester, student_id, code),
            FOREIGN KEY (promotion_id) REFERENCES promotions(id) ON DELETE CASCADE,
            FOREIGN KEY (student_id) REFERENCES promotion_students(id) ON DELETE CASCADE
        )
    ''')
    # Pondération (définie par le référent de la matière) + statut provisoire/
    # définitif de chaque colonne de saisie, par face FTP/ALT. Pondération
    # absente = 1 ; 0 = sous-matière exclue (pas de note à saisir).
    db.execute('''
        CREATE TABLE IF NOT EXISTS submatiere_meta (
            promotion_id INTEGER NOT NULL,
            semester TEXT NOT NULL,
            formation TEXT NOT NULL CHECK (formation IN ('FTP', 'ALT')),
            code TEXT NOT NULL,
            weight REAL,
            definitive INTEGER DEFAULT 0,
            UNIQUE (promotion_id, semester, formation, code),
            FOREIGN KEY (promotion_id) REFERENCES promotions(id) ON DELETE CASCADE
        )
    ''')
    db.commit()
    _merge_formation_promotions(db)

def _merge_formation_promotions(db):
    """Migration : fusionne les anciennes promotions « <cohorte>_FTP » et « <cohorte>_ALT »
    en UNE cohorte « <cohorte> » dont les étudiants portent leur sous-cohorte (formation).
    Les notes (student_marks) suivent. Idempotent : ne fait rien si plus aucun nom suffixé."""
    rows = db.execute("SELECT id, name, formation, start_year, programme_id FROM promotions").fetchall()
    suffixed = [r for r in rows if str(r['name']).rsplit('_', 1)[-1] in ('FTP', 'ALT')]
    if not suffixed:
        return
    groups = {}
    for r in suffixed:
        base = str(r['name']).rsplit('_', 1)[0]
        groups.setdefault(base, []).append(r)
    for base, members in groups.items():
        # Survivant : une cohorte « base » déjà présente, sinon le 1er membre (renommé).
        survivor = next((r for r in rows if str(r['name']) == base), None)
        if survivor is None:
            survivor = members[0]
            db.execute("UPDATE promotions SET name=?, formation='MUT' WHERE id=?", (base, survivor['id']))
        else:
            db.execute("UPDATE promotions SET formation='MUT' WHERE id=?", (survivor['id'],))
        prog = survivor['programme_id']
        for m in members:
            if m['id'] == survivor['id']:
                # Étudiants déjà rattachés au survivant : taguer avec sa propre formation
                db.execute("UPDATE promotion_students SET formation=? WHERE promotion_id=? AND formation IS NULL",
                           (m['formation'], m['id']))
                continue
            db.execute("UPDATE promotion_students SET promotion_id=?, formation=? WHERE promotion_id=?",
                       (survivor['id'], m['formation'], m['id']))
            db.execute("UPDATE student_marks SET promotion_id=? WHERE promotion_id=?",
                       (survivor['id'], m['id']))
            if not prog and m['programme_id']:
                prog = m['programme_id']
            db.execute("DELETE FROM promotions WHERE id=?", (m['id'],))
        if prog:
            db.execute("UPDATE promotions SET programme_id=? WHERE id=?", (prog, survivor['id']))
    db.commit()

def _init_promotions_db():
    os.makedirs(os.path.dirname(_PROMOTIONS_DB), exist_ok=True)
    db = _open_promotions_db()
    try:
        _apply_promotions_migrations(db)
    finally:
        db.close()

# ===== Base des PROGRAMMES (Programme National réutilisable, ex 'PN_2026') =====
# Contient plusieurs programmes nommés ; une promotion en référence un (programme_id).
# Un programme regroupe matières + compétences (UE) + coefficients + heures, par
# semestre, dans la même structure JSON que les coefficients (+ champ 'hours' par
# composant). Base séparée car un programme est réutilisable d'année en année.
_PROGRAMMES_DB = os.environ.get('EDT_PROGRAMMES_DB') or os.path.join(_DB_DIR, 'programmes.db')

def _open_programmes_db():
    db = sqlite3.connect(_PROGRAMMES_DB, timeout=10)
    db.row_factory = sqlite3.Row
    db.execute('PRAGMA foreign_keys = ON')
    return db

def get_programmes_db():
    """Connexion à la base des programmes (partagée entre années et promotions)."""
    try:
        if '_grdb' not in g:
            g._grdb = _open_programmes_db()
        return g._grdb
    except RuntimeError:
        return _open_programmes_db()

def _apply_programmes_migrations(db):
    db.execute('PRAGMA journal_mode=WAL')
    db.execute('''
        CREATE TABLE IF NOT EXISTS programmes (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT NOT NULL UNIQUE,   -- 'PN_2026'
            label      TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT
        )
    ''')
    db.execute('''
        CREATE TABLE IF NOT EXISTS programme_data (
            programme_id INTEGER PRIMARY KEY,
            data         TEXT NOT NULL,        -- JSON {S1..S6: {components, competences}}
            updated_at   TEXT,
            FOREIGN KEY (programme_id) REFERENCES programmes(id) ON DELETE CASCADE
        )
    ''')
    db.commit()

def _init_programmes_db():
    os.makedirs(os.path.dirname(_PROGRAMMES_DB), exist_ok=True)
    db = _open_programmes_db()
    try:
        _apply_programmes_migrations(db)
        # Seed : si aucun programme, créer 'PN_2026' depuis la référence des coefficients.
        if not db.execute('SELECT 1 FROM programmes LIMIT 1').fetchone():
            data = _ensure_volumes(_ref_coeffs_defaults() or {})
            cur = db.execute('''INSERT INTO programmes(name, label, updated_at)
                                VALUES('PN_2026', 'Programme National 2026', CURRENT_TIMESTAMP)''')
            db.execute('''INSERT INTO programme_data(programme_id, data, updated_at)
                          VALUES(?,?,CURRENT_TIMESTAMP)''',
                       (cur.lastrowid, json.dumps(data, ensure_ascii=False)))
            db.commit()
    finally:
        db.close()

def _apply_migrations(db):
    """Applique toutes les migrations de schéma sur une connexion DB ouverte"""
    db.execute('PRAGMA journal_mode=WAL')
    db.execute('''
        CREATE TABLE IF NOT EXISTS semester_special_weeks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            semester_id INTEGER NOT NULL,
            week_number INTEGER NOT NULL,
            week_type TEXT NOT NULL CHECK(week_type IN ('vacation_ftp', 'company_alt')),
            FOREIGN KEY (semester_id) REFERENCES semesters(id) ON DELETE CASCADE,
            UNIQUE(semester_id, week_number, week_type)
        )
    ''')
    db.execute('''
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
    ''')
    db.execute("INSERT OR IGNORE INTO app_settings (key, value) VALUES ('academic_start_week', '36')")
    db.execute("INSERT OR IGNORE INTO app_settings (key, value) VALUES ('academic_end_week', '26')")
    # Semaines de début/fin par parité de semestre (impair: S1/S3/S5 ; pair: S2/S4/S6)
    db.execute("INSERT OR IGNORE INTO app_settings (key, value) VALUES ('semester_odd_start_week', '36')")
    db.execute("INSERT OR IGNORE INTO app_settings (key, value) VALUES ('semester_odd_end_week', '4')")
    db.execute("INSERT OR IGNORE INTO app_settings (key, value) VALUES ('semester_even_start_week', '5')")
    db.execute("INSERT OR IGNORE INTO app_settings (key, value) VALUES ('semester_even_end_week', '26')")
    for col in ['start_week', 'end_week']:
        try:
            db.execute(f'ALTER TABLE courses ADD COLUMN {col} INTEGER')
        except sqlite3.OperationalError:
            pass
    # Dates par défaut du semestre : si 1, les semaines sont calculées dynamiquement
    try:
        db.execute('ALTER TABLE courses ADD COLUMN default_weeks INTEGER DEFAULT 0')
    except sqlite3.OperationalError:
        pass
    # Description / contenu pédagogique de la matière (éditable par admin ou intervenant)
    try:
        db.execute('ALTER TABLE courses ADD COLUMN content TEXT')
    except sqlite3.OperationalError:
        pass
    # Contenu officiel du Programme national (PN), éditable par l'admin uniquement
    try:
        db.execute('ALTER TABLE courses ADD COLUMN content_pn TEXT')
    except sqlite3.OperationalError:
        pass
    # Contraintes d'emploi du temps saisies par les enseignants (texte + fichier joint)
    db.execute('''
        CREATE TABLE IF NOT EXISTS teacher_constraints (
            teacher_id INTEGER PRIMARY KEY,
            content TEXT,
            file_path TEXT,
            file_original TEXT,
            updated_at TEXT,
            FOREIGN KEY (teacher_id) REFERENCES teachers(id) ON DELETE CASCADE
        )
    ''')
    try:
        db.execute('ALTER TABLE course_sessions ADD COLUMN sessions_per_week_max INTEGER DEFAULT 1')
    except sqlite3.OperationalError:
        pass
    db.execute('''
        CREATE TABLE IF NOT EXISTS course_ordering (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            course_id_pred INTEGER NOT NULL,
            course_id_succ INTEGER NOT NULL,
            min_gap_weeks  INTEGER DEFAULT 0,
            created_at     TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (course_id_pred) REFERENCES courses(id) ON DELETE CASCADE,
            FOREIGN KEY (course_id_succ) REFERENCES courses(id) ON DELETE CASCADE,
            UNIQUE(course_id_pred, course_id_succ)
        )
    ''')
    # Nouveau calendrier spécial (remplace semester_special_weeks dans l'optimiseur)
    db.execute('''
        CREATE TABLE IF NOT EXISTS special_calendar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            week_number INTEGER NOT NULL,
            week_type TEXT NOT NULL,
            UNIQUE(week_number, week_type)
        )
    ''')
    # Commentaire libre par semaine (affiché en répartition + export Excel)
    # `formations` = formations concernées (ex : 'FTP,ALT', 'FTP', 'ALT')
    # `years` = années (year_group) concernées (ex : '1,2,3', '1', '2,3')
    db.execute('''
        CREATE TABLE IF NOT EXISTS week_comments (
            week_number INTEGER PRIMARY KEY,
            comment TEXT NOT NULL,
            formations TEXT NOT NULL DEFAULT 'FTP,ALT',
            years TEXT NOT NULL DEFAULT '1,2,3'
        )
    ''')
    _wc_cols = [r[1] for r in db.execute("PRAGMA table_info(week_comments)").fetchall()]
    if 'formations' not in _wc_cols:
        db.execute("ALTER TABLE week_comments ADD COLUMN formations TEXT NOT NULL DEFAULT 'FTP,ALT'")
    if 'years' not in _wc_cols:
        db.execute("ALTER TABLE week_comments ADD COLUMN years TEXT NOT NULL DEFAULT '1,2,3'")
    # Code texte par catégorie de calendrier (affiché en ligne verticale en répartition)
    db.execute('''
        CREATE TABLE IF NOT EXISTS special_category_codes (
            week_type TEXT PRIMARY KEY,
            code TEXT NOT NULL
        )
    ''')
    # Statut enseignant : Titulaire / Vacataire
    if 'status' not in [r[1] for r in db.execute("PRAGMA table_info(teachers)").fetchall()]:
        db.execute("ALTER TABLE teachers ADD COLUMN status TEXT DEFAULT 'Titulaire'")
    # Mot de passe enseignant (haché) : débloque l'accès lecture seule aux Promotions.
    # L'enseignant le crée lui-même au login (bouton d'initialisation) une fois
    # autorisé par l'admin (password_allowed=1). NB : la table teachers est par année
    # scolaire — l'autorisation se gère dans l'année active de l'admin ; les années
    # créées par copie l'héritent.
    if 'password_hash' not in [r[1] for r in db.execute("PRAGMA table_info(teachers)").fetchall()]:
        db.execute("ALTER TABLE teachers ADD COLUMN password_hash TEXT")
    if 'password_allowed' not in [r[1] for r in db.execute("PRAGMA table_info(teachers)").fetchall()]:
        db.execute("ALTER TABLE teachers ADD COLUMN password_allowed INTEGER DEFAULT 0")
        # Compat : un mot de passe déjà défini vaut autorisation
        db.execute("UPDATE teachers SET password_allowed = 1 WHERE password_hash IS NOT NULL")
    # Enseignant référent par MATIÈRE (groupe de sous-matières, clé = code sans
    # lettre finale, cf. _mat_base_key) et par face (FTP/ALT) — choix MANUEL de
    # l'admin, prioritaire sur la règle automatique (titulaire unique intervenant).
    db.execute('DROP TABLE IF EXISTS course_referents')  # 1er jet abandonné (référent par sous-matière)
    db.execute('''
        CREATE TABLE IF NOT EXISTS matiere_referents (
            matiere_key TEXT NOT NULL,
            formation TEXT NOT NULL CHECK (formation IN ('FTP', 'ALT')),
            teacher_id INTEGER,
            PRIMARY KEY (matiere_key, formation),
            FOREIGN KEY (teacher_id) REFERENCES teachers(id) ON DELETE SET NULL
        )
    ''')
    # Auto-migration depuis semester_special_weeks (données legacy)
    db.execute('''
        INSERT OR IGNORE INTO special_calendar (week_number, week_type)
        SELECT DISTINCT week_number, 'vacation_ftp'
        FROM semester_special_weeks WHERE week_type = 'vacation_ftp'
    ''')
    db.execute('''
        INSERT OR IGNORE INTO special_calendar (week_number, week_type)
        SELECT DISTINCT ssw.week_number, 'company_alt_y' || s.year_group
        FROM semester_special_weeks ssw
        JOIN semesters s ON ssw.semester_id = s.id
        WHERE ssw.week_type = 'company_alt' AND s.year_group IN (1, 2, 3)
    ''')
    # Vacances FTP désormais spécifiques à l'année de progression (1/2/3).
    # Réplique l'ancien 'vacation_ftp' (partagé) vers y1/y2/y3, puis le supprime.
    if db.execute("SELECT 1 FROM special_calendar WHERE week_type='vacation_ftp' LIMIT 1").fetchone():
        for _yg in (1, 2, 3):
            db.execute('''INSERT OR IGNORE INTO special_calendar (week_number, week_type)
                          SELECT week_number, ? FROM special_calendar WHERE week_type='vacation_ftp' ''',
                       (f'vacation_ftp_y{_yg}',))
        db.execute("DELETE FROM special_calendar WHERE week_type='vacation_ftp'")
    cursor = db.execute('SELECT COUNT(*) as cnt FROM semesters')
    if cursor.fetchone()['cnt'] == 0:
        for code, yg, name in [
            ('S1', 1, 'Semestre 1'), ('S2', 1, 'Semestre 2'),
            ('S3', 2, 'Semestre 3'), ('S4', 2, 'Semestre 4'),
            ('S5', 3, 'Semestre 5'), ('S6', 3, 'Semestre 6'),
        ]:
            db.execute('INSERT OR IGNORE INTO semesters (code, year_group, name) VALUES (?, ?, ?)',
                       (code, yg, name))
    # Normaliser les variantes TP (TP12, TP 12, TP8, TP 8, etc.) → "TP"
    # D'abord fusionner les heures quand un "TP" existe déjà pour le même cours/formation
    db.execute('''
        UPDATE course_sessions SET
            total_hours = total_hours + COALESCE((
                SELECT SUM(cs2.total_hours) FROM course_sessions cs2
                WHERE cs2.course_id = course_sessions.course_id
                  AND cs2.formation_type = course_sessions.formation_type
                  AND cs2.teaching_type != 'TP'
                  AND UPPER(REPLACE(cs2.teaching_type, ' ', '')) LIKE 'TP%'
            ), 0),
            nb_sessions = nb_sessions + COALESCE((
                SELECT SUM(cs2.nb_sessions) FROM course_sessions cs2
                WHERE cs2.course_id = course_sessions.course_id
                  AND cs2.formation_type = course_sessions.formation_type
                  AND cs2.teaching_type != 'TP'
                  AND UPPER(REPLACE(cs2.teaching_type, ' ', '')) LIKE 'TP%'
            ), 0)
        WHERE teaching_type = 'TP'
          AND EXISTS (
                SELECT 1 FROM course_sessions cs2
                WHERE cs2.course_id = course_sessions.course_id
                  AND cs2.formation_type = course_sessions.formation_type
                  AND cs2.teaching_type != 'TP'
                  AND UPPER(REPLACE(cs2.teaching_type, ' ', '')) LIKE 'TP%'
          )
    ''')
    # Supprimer les variantes TP déjà fusionnées
    db.execute('''
        DELETE FROM course_sessions
        WHERE teaching_type != 'TP'
          AND UPPER(REPLACE(teaching_type, ' ', '')) LIKE 'TP%'
          AND EXISTS (
                SELECT 1 FROM course_sessions cs2
                WHERE cs2.course_id = course_sessions.course_id
                  AND cs2.formation_type = course_sessions.formation_type
                  AND cs2.teaching_type = 'TP'
          )
    ''')
    # Renommer les variantes TP restantes (pas de "TP" existant) → "TP"
    db.execute('''
        UPDATE course_sessions SET teaching_type = 'TP'
        WHERE teaching_type != 'TP'
          AND UPPER(REPLACE(teaching_type, ' ', '')) LIKE 'TP%'
    ''')

    # Déduplication course_sessions (nettoyer les doublons existants, garder le plus ancien)
    db.execute('''
        DELETE FROM course_sessions WHERE id NOT IN (
            SELECT MIN(id) FROM course_sessions
            GROUP BY course_id, formation_type, teaching_type
        )
    ''')
    # Index unique pour empêcher les futurs doublons
    try:
        db.execute('''
            CREATE UNIQUE INDEX IF NOT EXISTS idx_cs_unique
            ON course_sessions(course_id, formation_type, teaching_type)
        ''')
    except sqlite3.OperationalError:
        pass
    # Colonne mutualized sur courses (CM/TD communs FTP+ALT)
    try:
        db.execute('ALTER TABLE courses ADD COLUMN mutualized INTEGER DEFAULT 0')
    except sqlite3.OperationalError:
        pass
    # Type de TP de la matière : TP12 (groupes de 12) ou TP8 (groupes de 8) —
    # détermine le nombre de groupes utilisé pour le calcul du service (Bilan global)
    try:
        db.execute("ALTER TABLE courses ADD COLUMN tp_type TEXT DEFAULT 'TP12'")
    except sqlite3.OperationalError:
        pass
    # Nombre de groupes de TD/TP par (année, formation) pour le calcul du service
    db.execute('''
        CREATE TABLE IF NOT EXISTS promotion_groups (
            year_group     INTEGER NOT NULL,
            formation_type INTEGER NOT NULL,
            cm_groups      INTEGER NOT NULL DEFAULT 1,
            td_groups      INTEGER NOT NULL DEFAULT 1,
            tp_groups      INTEGER NOT NULL DEFAULT 1,
            tp8_groups     INTEGER NOT NULL DEFAULT 1,
            pt_groups      INTEGER NOT NULL DEFAULT 1,
            PRIMARY KEY (year_group, formation_type)
        )
    ''')
    # Colonne tp8_groups (nb de groupes des TP à 8) sur les bases existantes :
    # initialisée sur tp_groups (comportement inchangé tant que non ajustée)
    if 'tp8_groups' not in [r[1] for r in db.execute("PRAGMA table_info(promotion_groups)").fetchall()]:
        db.execute("ALTER TABLE promotion_groups ADD COLUMN tp8_groups INTEGER NOT NULL DEFAULT 1")
        db.execute("UPDATE promotion_groups SET tp8_groups = tp_groups")
    # Valeurs par défaut : FTP 2 TD / 3 TP, ALT et MUT 1 TD / 1 TP, CM/PT = 1
    _PG_DEFAULTS = [
        # year_group, formation_type, cm, td, tp, tp8, pt
        (1, 0, 1, 2, 3, 3, 1), (1, 1, 1, 1, 1, 1, 1), (1, 2, 1, 1, 1, 1, 1),
        (2, 0, 1, 2, 3, 3, 1), (2, 1, 1, 1, 1, 1, 1), (2, 2, 1, 1, 1, 1, 1),
        (3, 0, 1, 2, 3, 3, 1), (3, 1, 1, 1, 1, 1, 1), (3, 2, 1, 1, 1, 1, 1),
    ]
    for yg, ft, cm, td, tp, tp8, pt in _PG_DEFAULTS:
        db.execute('''
            INSERT OR IGNORE INTO promotion_groups
                (year_group, formation_type, cm_groups, td_groups, tp_groups, tp8_groups, pt_groups)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (yg, ft, cm, td, tp, tp8, pt))

    # Groupes par SEMESTRE (remplace l'ancienne granularité par année de
    # promotion_groups) : nb de groupes TD / TP12 / TP8 / PT pour chaque
    # (semestre, face). formation_type : 0=FTP, 1=ALT, 2=Promo entière
    # (utilisé quand le semestre est mutualisé). CM = toujours 1 groupe.
    db.execute('''
        CREATE TABLE IF NOT EXISTS semester_groups (
            semester_code  TEXT NOT NULL,
            formation_type INTEGER NOT NULL,
            td_groups      INTEGER NOT NULL DEFAULT 1,
            tp_groups      INTEGER NOT NULL DEFAULT 1,
            tp8_groups     INTEGER NOT NULL DEFAULT 1,
            pt_groups      INTEGER NOT NULL DEFAULT 1,
            PRIMARY KEY (semester_code, formation_type)
        )
    ''')
    # Reprise des valeurs existantes de promotion_groups (année → ses 2 semestres)
    if not db.execute('SELECT 1 FROM semester_groups LIMIT 1').fetchone():
        for r in db.execute('SELECT * FROM promotion_groups').fetchall():
            for sem in (f"S{2 * r['year_group'] - 1}", f"S{2 * r['year_group']}"):
                db.execute('''
                    INSERT OR IGNORE INTO semester_groups
                        (semester_code, formation_type, td_groups, tp_groups, tp8_groups, pt_groups)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (sem, r['formation_type'], r['td_groups'], r['tp_groups'],
                      r['tp8_groups'], r['pt_groups']))
    # Nombre de groupes d'apprentis (ALT) dans un TP mutualisé (TP communs) :
    # les apprentis occupent les DERNIERS groupes de la promo. Défini sur la
    # ligne « Promo » (formation_type=2) de chaque semestre.
    if 'alt_groups' not in [r[1] for r in db.execute("PRAGMA table_info(semester_groups)").fetchall()]:
        db.execute("ALTER TABLE semester_groups ADD COLUMN alt_groups INTEGER NOT NULL DEFAULT 1")
    # Compléter les lignes manquantes (semestres × faces) avec les défauts
    for _sem in ('S1', 'S2', 'S3', 'S4', 'S5', 'S6'):
        for _ft in (0, 1, 2):
            db.execute('''
                INSERT OR IGNORE INTO semester_groups
                    (semester_code, formation_type, td_groups, tp_groups, tp8_groups, pt_groups)
                VALUES (?, ?, 1, 1, 1, 1)
            ''', (_sem, _ft))
    # Flag « semestre mutualisé » (FTP + ALT ensemble) sur la table semesters
    if 'mutualized' not in [r[1] for r in db.execute("PRAGMA table_info(semesters)").fetchall()]:
        db.execute("ALTER TABLE semesters ADD COLUMN mutualized INTEGER NOT NULL DEFAULT 0")
    # Semestre mutualisé : TP FTP / ALT distincts (1, défaut — lignes séparées
    # pour la répartition calendaire) ou TP communs à la promo (0, ligne TP MUT)
    if 'tp_separate' not in [r[1] for r in db.execute("PRAGMA table_info(semesters)").fetchall()]:
        db.execute("ALTER TABLE semesters ADD COLUMN tp_separate INTEGER NOT NULL DEFAULT 1")

    # Placement des TP par groupe : chaque groupe (TP12_1, TP12_2, …) a son propre
    # planning hebdomadaire dans la répartition calendaire. On ajoute group_index à
    # weekly_hours (1 = données existantes) et on élargit la clé d'unicité à
    # (session, semaine, groupe). SQLite ne sait pas modifier une contrainte UNIQUE
    # en place → reconstruction de la table.
    if 'group_index' not in [r[1] for r in db.execute("PRAGMA table_info(weekly_hours)").fetchall()]:
        db.execute('DROP TABLE IF EXISTS weekly_hours_new')  # au cas où un essai précédent a échoué
        db.execute('''
            CREATE TABLE weekly_hours_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                course_session_id INTEGER NOT NULL,
                week_number INTEGER NOT NULL,
                semester_week INTEGER,
                hours REAL DEFAULT 0,
                group_index INTEGER NOT NULL DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (course_session_id) REFERENCES course_sessions(id) ON DELETE CASCADE,
                UNIQUE(course_session_id, week_number, group_index)
            )
        ''')
        # Recopie (groupe 1), en ignorant d'éventuelles lignes orphelines
        db.execute('''
            INSERT INTO weekly_hours_new
                (id, course_session_id, week_number, semester_week, hours, group_index, created_at)
            SELECT id, course_session_id, week_number, semester_week, hours, 1, created_at
            FROM weekly_hours
            WHERE course_session_id IN (SELECT id FROM course_sessions)
        ''')
        db.execute('DROP TABLE weekly_hours')
        db.execute('ALTER TABLE weekly_hours_new RENAME TO weekly_hours')
        db.execute('CREATE INDEX IF NOT EXISTS idx_weekly_hours_session ON weekly_hours(course_session_id)')

    # Heures HETD effectuées dans un autre département (onglet Mon Compte).
    # Saisies par l'enseignant lui-même et visibles par lui seul, sauf lignes
    # marquées publiques (is_public=1) : alors visibles AUSSI par l'admin
    # (Bilan global), jamais par les autres enseignants.
    db.execute('''
        CREATE TABLE IF NOT EXISTS teacher_external_hours (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            teacher_id INTEGER NOT NULL,
            label      TEXT NOT NULL,
            hetd       REAL NOT NULL DEFAULT 0,
            is_public  INTEGER NOT NULL DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (teacher_id) REFERENCES teachers(id) ON DELETE CASCADE
        )
    ''')
    if 'is_public' not in [r[1] for r in db.execute("PRAGMA table_info(teacher_external_hours)").fetchall()]:
        db.execute("ALTER TABLE teacher_external_hours ADD COLUMN is_public INTEGER NOT NULL DEFAULT 0")

    # ======================= GESTION DES NOTES (admin) =======================
    # Promotion de notes (1 fichier Apogée importé = 1 promo, ex: BUT1 FI S2)
    db.execute('''
        CREATE TABLE IF NOT EXISTS grade_promos (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            code        TEXT NOT NULL UNIQUE,   -- ex 'T3GIMS2FI1'
            label       TEXT,                   -- intitulé lu dans le fichier (E4)
            semester    TEXT,                   -- 'S2' / 'S4' / 'S6'
            formation   TEXT,                   -- 'FI' / 'APP'
            template     TEXT,                  -- chemin relatif du .xlsm modèle (export)
            template_orig TEXT,                 -- nom d'origine du fichier importé
            updated_at  TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    # Éléments pédagogiques (ELP) : SAÉ, ressources, portfolio — colonnes du fichier
    db.execute('''
        CREATE TABLE IF NOT EXISTS grade_elps (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            promo_id   INTEGER NOT NULL,
            ordre      INTEGER,
            code       TEXT,
            title      TEXT,
            kind       TEXT,        -- 'SAE' / 'RES' / 'PORT'
            note_col   INTEGER,     -- index colonne Note dans le modèle
            bareme_col INTEGER,     -- index colonne Barème dans le modèle
            FOREIGN KEY (promo_id) REFERENCES grade_promos(id) ON DELETE CASCADE
        )
    ''')
    # Étudiants (roster importé du fichier standard)
    db.execute('''
        CREATE TABLE IF NOT EXISTS grade_students (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            promo_id   INTEGER NOT NULL,
            ordre      INTEGER,
            src_row    INTEGER,     -- ligne d'origine dans le modèle (export)
            numero     TEXT,
            nom        TEXT,
            prenom     TEXT,
            naissance  TEXT,
            FOREIGN KEY (promo_id) REFERENCES grade_promos(id) ON DELETE CASCADE
        )
    ''')
    # Notes saisies (étudiant × ELP)
    db.execute('''
        CREATE TABLE IF NOT EXISTS grade_marks (
            student_id INTEGER NOT NULL,
            elp_id     INTEGER NOT NULL,
            note       REAL,
            PRIMARY KEY (student_id, elp_id),
            FOREIGN KEY (student_id) REFERENCES grade_students(id) ON DELETE CASCADE,
            FOREIGN KEY (elp_id)     REFERENCES grade_elps(id) ON DELETE CASCADE
        )
    ''')
    # Compétences (5 max BUT GIM) calculées à partir des ELP par coefficients
    db.execute('''
        CREATE TABLE IF NOT EXISTS grade_competences (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            promo_id  INTEGER NOT NULL,
            ordre     INTEGER,
            name      TEXT,
            FOREIGN KEY (promo_id) REFERENCES grade_promos(id) ON DELETE CASCADE
        )
    ''')
    # Coefficients compétence × ELP (renormalisation auto si note manquante)
    db.execute('''
        CREATE TABLE IF NOT EXISTS grade_coefficients (
            competence_id INTEGER NOT NULL,
            elp_id        INTEGER NOT NULL,
            coeff         REAL DEFAULT 0,
            PRIMARY KEY (competence_id, elp_id),
            FOREIGN KEY (competence_id) REFERENCES grade_competences(id) ON DELETE CASCADE,
            FOREIGN KEY (elp_id)        REFERENCES grade_elps(id) ON DELETE CASCADE
        )
    ''')
    db.commit()

def get_year_config():
    """Lit la configuration de l'année courante (semaines début/fin).
    Inclut une plage optionnelle PAR année de progression (1/2/3) ; si non
    définie pour une année, on retombe sur la plage globale."""
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT key, value FROM app_settings")
    cfg = {}
    for r in cursor.fetchall():
        try:
            cfg[r['key']] = int(r['value'])
        except (TypeError, ValueError):
            pass  # valeurs non entières (ex. hetd_coeffs) ignorées ici
    gstart = cfg.get('academic_start_week', 36)
    gend   = cfg.get('academic_end_week',   26)
    gmax   = get_academic_max_week()
    ranges = {}
    for yg in (1, 2, 3):
        ranges[yg] = {
            'start': cfg.get(f'academic_start_week_y{yg}', gstart),
            'end':   cfg.get(f'academic_end_week_y{yg}',   gend),
            'max':   cfg.get(f'academic_max_week_y{yg}',   gmax),
        }
    return {
        'academic_start_week': gstart,
        'academic_end_week':   gend,
        'academic_ranges':     ranges,
        'semester_odd_start_week':  cfg.get('semester_odd_start_week',  36),
        'semester_odd_end_week':    cfg.get('semester_odd_end_week',    4),
        'semester_even_start_week': cfg.get('semester_even_start_week', 5),
        'semester_even_end_week':   cfg.get('semester_even_end_week',   26),
        'academic_max_week':   gmax,
    }

def year_range_for_yg(cfg, yg):
    """(start, end, max) de la plage de l'année de progression yg (repli global)."""
    r = (cfg.get('academic_ranges') or {}).get(yg) or (cfg.get('academic_ranges') or {}).get(str(yg))
    if r:
        return r['start'], r['end'], r.get('max', cfg['academic_max_week'])
    return cfg['academic_start_week'], cfg['academic_end_week'], cfg['academic_max_week']

def get_valid_school_weeks(start_week=36, end_week=26, max_week=52):
    """Ensemble des semaines valides pour une plage scolaire (peut chevaucher S52/S53→S01).
    max_week = dernière semaine de la 1re année civile (52 ou 53 selon le calendrier ISO)."""
    if start_week <= end_week:
        return set(range(start_week, end_week + 1))
    return set(range(start_week, max_week + 1)) | set(range(1, end_week + 1))

def semester_default_weeks(semester_code, cfg):
    """Semaines (début, fin) par défaut selon la parité du semestre (impair S1/S3/S5,
    pair S2/S4/S6), lues depuis la config Calendrier. Renvoie (None, None) si inconnu."""
    digits = ''.join(ch for ch in (semester_code or '') if ch.isdigit())
    if not digits:
        return None, None
    n = int(digits)
    if n % 2 == 1:
        return cfg['semester_odd_start_week'], cfg['semester_odd_end_week']
    return cfg['semester_even_start_week'], cfg['semester_even_end_week']

def resolve_course_weeks(course, cfg):
    """Semaines (début, fin) effectives d'un cours : calcul dynamique selon le semestre
    si default_weeks est activé, sinon les valeurs stockées. `course` doit exposer
    'default_weeks', 'semester_code', 'start_week', 'end_week'."""
    if course.get('default_weeks'):
        s, e = semester_default_weeks(course.get('semester_code'), cfg)
        if s and e:
            return s, e
    return course.get('start_week'), course.get('end_week')

def _init_fresh_db(path):
    """Crée une DB vierge depuis le schéma SQL"""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    db = sqlite3.connect(path)
    with open(SCHEMA_PATH, 'r') as f:
        db.executescript(f.read())
    _apply_migrations(db)
    db.close()

def _migrate_all_years():
    """Applique les migrations de schéma à TOUTES les bases année (pas seulement
    l'année courante) : on peut cibler n'importe quelle année via ?year=."""
    for year in list_years():
        try:
            db = sqlite3.connect(db_path_for_year(year))
            db.row_factory = sqlite3.Row
            _apply_migrations(db)
            db.close()
        except sqlite3.Error as e:
            print(f"  [warn] migration de l'année {year} échouée : {e}")

# Helper function to initialize database
def init_db():
    """Initialise la DB pour l'année courante, migre depuis edt.db si besoin"""
    _migrate_db_layout()   # range les bases existantes dans databases/<année>/
    settings = get_settings()

    if not settings.get('current_year'):
        year = auto_detect_year()
        settings['current_year'] = year
        save_settings(settings)

    year = settings['current_year']
    db_path = db_path_for_year(year)

    if not os.path.exists(db_path):
        if os.path.exists(DATABASE):
            os.makedirs(os.path.dirname(db_path), exist_ok=True)
            shutil.copy2(DATABASE, db_path)   # migration depuis edt.db legacy
        else:
            _init_fresh_db(db_path)

    db = get_db()
    _apply_migrations(db)
    # Ne pas fermer si la connexion est gérée par le contexte Flask (g._db)
    try:
        if '_db' not in g or g._db is not db:
            db.close()
    except RuntimeError:
        db.close()  # Hors contexte Flask → fermer manuellement

    # Backfill : crée les années universitaires manquantes des promotions existantes.
    try:
        created = ensure_years_for_all_promotions()
        if created:
            print(f"  Années universitaires créées (backfill promos) : {', '.join(created)}")
    except Exception as e:
        print(f"  [warn] backfill des années des promotions échoué : {e}")

    # Migration du schéma sur TOUTES les bases année (pas seulement l'année
    # courante) : on peut cibler n'importe quelle année via ?year= (calendrier
    # cohorte). Sans cela, une vieille base peut manquer une table récente.
    _migrate_all_years()

# Helper function to convert sqlite3.Row to dict
def row_to_dict(row):
    """Convert sqlite3.Row to dictionary"""
    if row is None:
        return None
    return dict(row)

def rows_to_list(rows):
    """Convert list of sqlite3.Row to list of dictionaries"""
    return [dict(row) for row in rows]

# Error handler
def error_response(message, status_code=400):
    """Return error response"""
    return jsonify({'error': message}), status_code

# ======================= AUTHENTIFICATION (routes + garde) =======================

import re as _re
# Reconnaît /api/courses/<id>/content (autorisé aux intervenants en écriture)
_COURSE_CONTENT_RE = _re.compile(r'^/api/courses/\d+/content$')
# Reconnaît /api/constraints/me[...] (l'enseignant gère SES propres contraintes)
_CONSTRAINTS_SELF_RE = _re.compile(r'^/api/constraints/me(/file)?$')
# Reconnaît la saisie de notes par sous-matière (onglet Saisie Notes) : ouverte
# aux enseignants avec promo_access, l'autorisation fine (colonne éditable,
# référent) est vérifiée dans les handlers.
_SAISIE_WRITE_RE = _re.compile(r'^/api/promotions/\d+/saisie/S[1-6]/(notes|weights)$')

def _is_course_content_path(path):
    return bool(_COURSE_CONTENT_RE.match(path))

def _is_constraints_self_path(path):
    return bool(_CONSTRAINTS_SELF_RE.match(path))

def _is_saisie_write_path(path):
    return bool(_SAISIE_WRITE_RE.match(path))

def teacher_intervenes_in_course(course_id, teacher_name):
    """Vrai si l'enseignant (par nom) intervient dans une session de la matière."""
    if not teacher_name:
        return False
    db = get_db()
    row = db.execute('''
        SELECT 1 FROM course_sessions cs
        JOIN teachers t ON cs.teacher_id = t.id
        WHERE cs.course_id = ? AND LOWER(t.name) = LOWER(?)
        LIMIT 1
    ''', (course_id, teacher_name)).fetchone()
    return row is not None

def _is_https():
    """Vrai si la requête arrive en HTTPS (gère le proxy via X-Forwarded-Proto)."""
    return request.is_secure or request.headers.get('X-Forwarded-Proto', '') == 'https'

@app.before_request
def _force_https():
    """Redirige http -> https quand EDT_FORCE_HTTPS=1 (à activer en production)."""
    if os.environ.get('EDT_FORCE_HTTPS', '0') != '1':
        return
    if not _is_https():
        return redirect(request.url.replace('http://', 'https://', 1), code=301)

@app.before_request
def _daily_backup_hook():
    """Sauvegarde quotidienne déclenchée par le trafic (sans cron)."""
    try:
        _maybe_daily_backup()
    except Exception:
        pass

# Endpoints pouvant cibler une année universitaire précise via ?year=AAAA-AAAA :
# calendrier de cohorte (édition sur 3 années) et onglet Service enseignant
# (Bilan Global / Répartition Calendaire / Contraintes ont leur propre sélecteur
# d'année, sans impacter l'année active utilisée par les autres onglets).
_YEAR_OVERRIDE_PATHS = {
    '/api/special-calendar', '/api/config', '/api/week-comments',
    '/api/category-codes', '/api/semester-groups',
}
_YEAR_OVERRIDE_PREFIXES = (
    '/api/repartition',          # tableaux annuel + journalier
    '/api/checks/repartition',
    '/api/weekly-hours',         # édition des heures depuis la répartition
    '/api/course-sessions',
    '/api/teachers',
    '/api/courses',              # onglet Matières + modal d'ordonnancement
    '/api/semesters',
    '/api/rooms',                # listes du modal matière
    '/api/matieres-programme',
    '/api/matiere-referents',
    '/api/course-orderings',
    '/api/hetd-coeffs',
    '/api/constraints',
    '/api/my-account',           # onglet Mon Compte (contact + heures hors GIM)
    '/api/external-hours',       # heures hors GIM publiques (admin, Bilan Global)
    '/api/export/repartition',
    '/api/import/repartition',
)

@app.before_request
def _apply_year_override():
    p = request.path
    if p in _YEAR_OVERRIDE_PATHS or p.startswith(_YEAR_OVERRIDE_PREFIXES):
        y = request.args.get('year')
        if y and _is_year(y) and os.path.exists(db_path_for_year(y)):
            g._year_override = y

@app.before_request
def _require_auth():
    """Protège l'API : connexion obligatoire ; visiteur = lecture seule."""
    if request.method == 'OPTIONS':
        return
    path = request.path
    if not path.startswith('/api/'):
        return  # fichiers statiques (SPA + page de connexion) : libres
    if path in _PUBLIC_API:
        return
    role = session.get('role')
    if not role:
        return error_response('Authentification requise', 401)
    if role != 'admin' and request.method not in ('GET', 'HEAD'):
        # Exception : l'enseignant intervenant peut éditer le contenu de SA matière.
        # L'autorisation fine (intervenant ou non) est vérifiée dans le handler.
        if role == 'teacher' and (_is_course_content_path(path) or _is_constraints_self_path(path)
                                  or path.startswith('/api/my-account')
                                  or (_is_saisie_write_path(path) and session.get('promo_access'))):
            return
        # Le choix de l'année universitaire est personnel à la session : autorisé à tous
        if path == '/api/years/session':
            return
        # Toute autre écriture (y compris la liste des enseignants) est réservée à l'admin
        return error_response('Accès en lecture seule', 403)

_WRITE_METHODS = {'POST', 'PUT', 'DELETE', 'PATCH'}

@app.after_request
def _security_headers(resp):
    """En-têtes de sécurité HTTP appliqués à toutes les réponses."""
    resp.headers.setdefault('X-Content-Type-Options', 'nosniff')
    resp.headers.setdefault('X-Frame-Options', 'SAMEORIGIN')          # anti-clickjacking
    resp.headers.setdefault('Referrer-Policy', 'same-origin')
    # HSTS : force HTTPS pendant 6 mois (n'a d'effet que servi en HTTPS)
    if _is_https():
        resp.headers.setdefault('Strict-Transport-Security', 'max-age=15552000; includeSubDomains')
    return resp

@app.after_request
def _audit_writes(resp):
    """Journalise toute modification de données (POST/PUT/DELETE/PATCH sur l'API).
    Les connexions sont journalisées séparément dans la route /api/login."""
    try:
        if (request.path.startswith('/api/') and request.method in _WRITE_METHODS
                and request.path not in ('/api/login', '/api/logout')):
            _audit('WRITE', ip=_client_ip(), user=session.get('user') or '-',
                   role=session.get('role') or '-', method=request.method,
                   path=request.path, status=resp.status_code)
    except Exception:
        pass
    return resp

# ---- Limitation des tentatives de connexion (anti-bruteforce) ----
# Après un échec, toute nouvelle tentative depuis la même IP est refusée
# pendant un délai minimum de 5 s, qui double à chaque échec (5, 10, 20… max 5 min).
_LOGIN_LOCK = threading.Lock()
_login_state = {}            # ip -> {'fails': int, 'until': float}
_LOGIN_BASE_DELAY = 5        # secondes (minimum imposé)
_LOGIN_MAX_DELAY = 300       # plafond : 5 min

def _client_ip():
    """IP réelle du client (gère le proxy de PythonAnywhere via X-Forwarded-For)."""
    xff = request.headers.get('X-Forwarded-For', '')
    if xff:
        return xff.split(',')[0].strip()
    return request.remote_addr or 'unknown'

def _login_retry_after(ip):
    """Secondes restantes avant qu'une nouvelle tentative soit autorisée (0 = OK)."""
    st = _login_state.get(ip)
    if not st:
        return 0
    return max(0, int(round(st['until'] - time.time())))

def _register_login_failure(ip):
    """Enregistre un échec et arme le délai d'attente (back-off exponentiel)."""
    with _LOGIN_LOCK:
        st = _login_state.get(ip) or {'fails': 0, 'until': 0}
        st['fails'] += 1
        delay = min(_LOGIN_BASE_DELAY * (2 ** (st['fails'] - 1)), _LOGIN_MAX_DELAY)
        st['until'] = time.time() + delay
        _login_state[ip] = st

def _reset_login_failures(ip):
    with _LOGIN_LOCK:
        _login_state.pop(ip, None)

@app.route('/api/login', methods=['POST'])
def login():
    ip = _client_ip()
    wait = _login_retry_after(ip)
    if wait > 0:
        _audit('LOGIN_BLOCKED', ip=ip, wait=wait)
        return error_response(
            f'Trop de tentatives de connexion. Réessayez dans {wait} seconde(s).', 429)

    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    # Compte fixe (Admin), identifiant insensible à la casse
    key = next((k for k in AUTH_USERS if k.lower() == username.lower()), None)
    if key:
        user = AUTH_USERS[key]
        # Comparaison en temps constant (anti timing-attack)
        if not hmac.compare_digest(str(user['password']), str(password)):
            _register_login_failure(ip)
            _audit('LOGIN_FAIL', ip=ip, user=username, role='admin')
            return error_response('Identifiant ou mot de passe incorrect', 401)
        _reset_login_failures(ip)
        session.permanent = True
        session['user'] = key
        session['role'] = user['role']
        session.pop('teacher_name', None)
        _audit('LOGIN_OK', ip=ip, user=key, role='admin')
        return jsonify({'username': key, 'role': user['role']})
    # Sinon : connexion enseignant par nom de famille (insensible à la casse).
    # Sans mot de passe : accès EDT habituel. Avec mot de passe (défini par
    # l'admin dans la fiche enseignant) : la session débloque en plus l'onglet
    # Promotions en lecture seule (promo_access). Un mot de passe fourni mais
    # incorrect refuse TOUTE connexion (pas de session dégradée silencieuse).
    if username:
        db = get_db()
        row = db.execute('SELECT name, status, password_hash FROM teachers WHERE LOWER(name) = LOWER(?)', (username,)).fetchone()
        if row:
            promo_access = False
            if password:
                if not row['password_hash'] or not check_password_hash(row['password_hash'], password):
                    _register_login_failure(ip)
                    _audit('LOGIN_FAIL', ip=ip, user=username, role='teacher')
                    return error_response('Identifiant ou mot de passe incorrect', 401)
                promo_access = True
            _reset_login_failures(ip)
            status = row['status'] or 'Titulaire'
            session.permanent = True
            session['user'] = row['name']
            session['role'] = 'teacher'
            session['teacher_name'] = row['name']
            session['teacher_status'] = status
            session['promo_access'] = promo_access
            _audit('LOGIN_OK', ip=ip, user=row['name'], role='teacher', promo=int(promo_access))
            return jsonify({'username': row['name'], 'role': 'teacher',
                            'teacher': row['name'], 'status': status,
                            'promo_access': promo_access})
    _register_login_failure(ip)
    _audit('LOGIN_FAIL', ip=ip, user=username, role='-')
    return error_response('Identifiant ou mot de passe incorrect', 401)

@app.route('/api/logout', methods=['POST'])
def logout():
    if session.get('user'):
        _audit('LOGOUT', ip=_client_ip(), user=session.get('user'), role=session.get('role'))
    session.clear()
    return jsonify({'message': 'Déconnecté'})

# ===== Initialisation du mot de passe enseignant (auto-service, pré-connexion) =====
# L'enseignant choisit son mot de passe au login, uniquement si l'admin l'a
# autorisé (password_allowed=1) et tant qu'aucun mot de passe n'est défini.

_PWD_POLICY_MSG = ('Le mot de passe doit contenir au moins 8 caractères, '
                   'dont une majuscule et un caractère spécial')

def _password_policy_ok(pwd):
    return (len(pwd) >= 8
            and re.search(r'[A-Z]', pwd) is not None
            and re.search(r'[^A-Za-z0-9]', pwd) is not None)

@app.route('/api/password-init/status', methods=['GET'])
def password_init_status():
    """Indique si le bouton « Initialiser mon mot de passe » doit s'afficher
    pour cet identifiant. Ne révèle rien d'autre (inconnu = non autorisé)."""
    username = (request.args.get('username') or '').strip()
    can_init = False
    if username and username.lower() != 'admin':
        row = get_db().execute(
            'SELECT password_allowed, password_hash FROM teachers WHERE LOWER(name) = LOWER(?)',
            (username,)).fetchone()
        if row and row['password_allowed'] and not row['password_hash']:
            can_init = True
    return jsonify({'can_init': can_init})

@app.route('/api/password-init', methods=['POST'])
def password_init():
    """Création du mot de passe par l'enseignant lui-même (saisi deux fois).
    Refusé si non autorisé par l'admin ou si un mot de passe existe déjà."""
    ip = _client_ip()
    wait = _login_retry_after(ip)
    if wait > 0:
        _audit('PWD_INIT_BLOCKED', ip=ip, wait=wait)
        return error_response(
            f'Trop de tentatives. Réessayez dans {wait} seconde(s).', 429)
    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    p1 = data.get('password') or ''
    p2 = data.get('password2') or ''
    db = get_db()
    row = db.execute(
        'SELECT id, name, password_allowed, password_hash FROM teachers WHERE LOWER(name) = LOWER(?)',
        (username,)).fetchone() if username else None
    if not row or not row['password_allowed'] or row['password_hash']:
        _register_login_failure(ip)
        _audit('PWD_INIT_REFUSED', ip=ip, user=username)
        return error_response("Initialisation du mot de passe non autorisée pour cet identifiant", 403)
    if p1 != p2:
        return error_response('Les deux mots de passe ne correspondent pas')
    if not _password_policy_ok(p1):
        return error_response(_PWD_POLICY_MSG)
    db.execute('UPDATE teachers SET password_hash = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
               (generate_password_hash(p1), row['id']))
    db.commit()
    _reset_login_failures(ip)
    _audit('TEACHER_PWD_INIT', ip=ip, teacher=row['name'])
    return jsonify({'message': 'Mot de passe enregistré'}), 200

@app.route('/api/me', methods=['GET'])
def me():
    if session.get('role'):
        return jsonify({'username': session.get('user'), 'role': session.get('role'),
                        'teacher': session.get('teacher_name'),
                        'status': session.get('teacher_status'),
                        'promo_access': bool(session.get('promo_access'))})
    return error_response('Non authentifié', 401)

@app.route('/api/audit-log', methods=['GET'])
def audit_log_view():
    """Renvoie les dernières lignes du journal d'audit (admin uniquement)."""
    if session.get('role') != 'admin':
        return error_response('Accès réservé à l\'administrateur', 403)
    try:
        n = max(1, min(int(request.args.get('lines', 200)), 2000))
    except (TypeError, ValueError):
        n = 200
    if not os.path.isfile(AUDIT_LOG_PATH):
        return jsonify({'lines': []})
    with open(AUDIT_LOG_PATH, encoding='utf-8') as f:
        lines = f.readlines()
    return jsonify({'lines': [l.rstrip('\n') for l in lines[-n:]][::-1]})

@app.route('/api/backups', methods=['GET'])
def list_backups():
    """Liste les sauvegardes disponibles, toutes années (admin uniquement)."""
    if session.get('role') != 'admin':
        return error_response('Accès réservé à l\'administrateur', 403)
    items = []
    for year in list_years():
        bdir = _backup_dir_for_year(year)
        if not os.path.isdir(bdir):
            continue
        for f in os.listdir(bdir):
            if not f.endswith('.db'):
                continue
            p = os.path.join(bdir, f)
            items.append({
                'year': year,
                'file': f,
                'datetime': datetime.fromtimestamp(os.path.getmtime(p)).strftime('%Y-%m-%d %H:%M:%S'),
                'size_kb': round(os.path.getsize(p) / 1024),
            })
    items.sort(key=lambda x: x['datetime'], reverse=True)
    return jsonify({'backups': items})

@app.route('/api/backups', methods=['POST'])
def create_backup_now():
    """Crée immédiatement une sauvegarde de l'année courante (admin)."""
    if session.get('role') != 'admin':
        return error_response('Accès réservé à l\'administrateur', 403)
    year = (request.get_json() or {}).get('year') or get_current_year()
    if not year or not os.path.isfile(db_path_for_year(year)):
        return error_response('Année introuvable', 404)
    path = _create_backup(year, tag='manuel')
    _audit('BACKUP_MANUEL', ip=_client_ip(), user=session.get('user'), year=year,
           file=os.path.basename(path) if path else '-')
    return jsonify({'created': os.path.basename(path) if path else None}), 201

@app.route('/api/backups/restore', methods=['POST'])
def restore_backup():
    """Restaure une sauvegarde dans la base de son année (admin).
    Une copie de sécurité de l'état courant est créée avant écrasement."""
    if session.get('role') != 'admin':
        return error_response('Accès réservé à l\'administrateur', 403)
    data = request.get_json() or {}
    year = (data.get('year') or '').strip()
    fname = (data.get('file') or '').strip()
    if not _is_year(year) or not fname:
        return error_response('Paramètres invalides', 400)
    bdir = _backup_dir_for_year(year)
    path = os.path.join(bdir, fname)
    # Anti path-traversal : le fichier doit bien être dans le dossier de sauvegardes
    if (os.path.dirname(os.path.abspath(path)) != os.path.abspath(bdir)
            or not os.path.isfile(path)):
        return error_response('Sauvegarde introuvable', 404)
    target = db_path_for_year(year)
    if not os.path.isfile(target):
        return error_response('Base cible introuvable', 404)
    # Filet de sécurité : sauvegarde de l'état actuel avant restauration
    _create_backup(year, tag='avant_restauration')
    src = sqlite3.connect(path)
    dst = sqlite3.connect(target)
    try:
        with dst:
            src.backup(dst)
    finally:
        src.close()
        dst.close()
    _audit('RESTORE', ip=_client_ip(), user=session.get('user'), year=year, file=fname)
    return jsonify({'restored': fname, 'year': year})

@app.route('/api/backups', methods=['DELETE'])
def delete_backup():
    """Supprime une sauvegarde (admin)."""
    if session.get('role') != 'admin':
        return error_response('Accès réservé à l\'administrateur', 403)
    data = request.get_json() or {}
    year = (data.get('year') or '').strip()
    fname = (data.get('file') or '').strip()
    if not _is_year(year) or not fname:
        return error_response('Paramètres invalides', 400)
    bdir = _backup_dir_for_year(year)
    path = os.path.join(bdir, fname)
    # Anti path-traversal : le fichier doit bien être dans le dossier de sauvegardes
    if (os.path.dirname(os.path.abspath(path)) != os.path.abspath(bdir)
            or not os.path.isfile(path)):
        return error_response('Sauvegarde introuvable', 404)
    os.remove(path)
    _audit('BACKUP_DELETE', ip=_client_ip(), user=session.get('user'), year=year, file=fname)
    return jsonify({'deleted': fname, 'year': year})

# ======================= CONTRAINTES ENSEIGNANTS =======================
_ALLOWED_CONSTRAINT_EXT = {'.pdf', '.doc', '.docx', '.odt', '.rtf', '.txt',
                           '.xls', '.xlsx', '.ods', '.csv', '.png', '.jpg', '.jpeg'}
_MAX_CONSTRAINT_FILE = 10 * 1024 * 1024   # 10 Mo

def _current_teacher_id():
    """id de l'enseignant connecté (d'après son nom en session), ou None."""
    name = session.get('teacher_name')
    if not name:
        return None
    row = get_db().execute('SELECT id FROM teachers WHERE LOWER(name)=LOWER(?)', (name,)).fetchone()
    return row['id'] if row else None

def _constraints_year_dir():
    d = os.path.join(_CONSTRAINTS_DIR, get_current_year() or 'default')
    os.makedirs(d, exist_ok=True)
    return d

def _constraint_row(teacher_id):
    return get_db().execute(
        'SELECT teacher_id, content, file_path, file_original, updated_at '
        'FROM teacher_constraints WHERE teacher_id=?', (teacher_id,)).fetchone()

def _constraint_payload(row, teacher_name=None):
    return {
        'teacher_id': row['teacher_id'] if row else None,
        'teacher_name': teacher_name,
        'content': (row['content'] if row else '') or '',
        'file_original': (row['file_original'] if row else '') or '',
        'has_file': bool(row and row['file_path']),
        'updated_at': (row['updated_at'] if row else '') or '',
    }

def _send_constraint_file(row):
    """Envoie le fichier de contrainte en pièce jointe (anti path-traversal)."""
    p = os.path.join(_CONSTRAINTS_DIR, row['file_path'])
    base = os.path.abspath(_CONSTRAINTS_DIR)
    if os.path.commonpath([os.path.abspath(p), base]) != base or not os.path.isfile(p):
        return error_response('Fichier introuvable', 404)
    return send_file(p, as_attachment=True,
                     download_name=row['file_original'] or os.path.basename(p))

@app.route('/api/constraints/me', methods=['GET'])
def get_my_constraints():
    tid = _current_teacher_id()
    if not tid:
        return error_response('Réservé aux enseignants', 403)
    return jsonify(_constraint_payload(_constraint_row(tid), session.get('teacher_name')))

@app.route('/api/constraints/me', methods=['PUT'])
def save_my_constraints():
    tid = _current_teacher_id()
    if not tid:
        return error_response('Réservé aux enseignants', 403)
    content = (request.get_json() or {}).get('content')
    content = content.strip() if isinstance(content, str) else None
    db = get_db()
    db.execute('''INSERT INTO teacher_constraints(teacher_id, content, updated_at)
                  VALUES(?,?,CURRENT_TIMESTAMP)
                  ON CONFLICT(teacher_id) DO UPDATE SET
                    content=excluded.content, updated_at=CURRENT_TIMESTAMP''',
               (tid, content or None))
    db.commit()
    _audit('CONSTRAINTS_SAVE', ip=_client_ip(), user=session.get('user'))
    return jsonify({'content': content or ''})

@app.route('/api/constraints/me/file', methods=['POST'])
def upload_my_constraint_file():
    tid = _current_teacher_id()
    if not tid:
        return error_response('Réservé aux enseignants', 403)
    f = request.files.get('file')
    if not f or not f.filename:
        return error_response('Aucun fichier reçu', 400)
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in _ALLOWED_CONSTRAINT_EXT:
        return error_response('Type de fichier non autorisé', 400)
    if request.content_length and request.content_length > _MAX_CONSTRAINT_FILE:
        return error_response('Fichier trop volumineux (max 10 Mo)', 400)

    safe = secure_filename(f.filename) or ('fichier' + ext)
    dest = os.path.join(_constraints_year_dir(), f"{tid}_{safe}")
    db = get_db()
    old = _constraint_row(tid)
    if old and old['file_path']:
        oldp = os.path.join(_CONSTRAINTS_DIR, old['file_path'])
        if os.path.isfile(oldp) and os.path.abspath(oldp) != os.path.abspath(dest):
            try: os.remove(oldp)
            except OSError: pass
    f.save(dest)
    if os.path.getsize(dest) > _MAX_CONSTRAINT_FILE:
        os.remove(dest)
        return error_response('Fichier trop volumineux (max 10 Mo)', 400)
    rel = os.path.relpath(dest, _CONSTRAINTS_DIR)
    db.execute('''INSERT INTO teacher_constraints(teacher_id, file_path, file_original, updated_at)
                  VALUES(?,?,?,CURRENT_TIMESTAMP)
                  ON CONFLICT(teacher_id) DO UPDATE SET
                    file_path=excluded.file_path, file_original=excluded.file_original,
                    updated_at=CURRENT_TIMESTAMP''',
               (tid, rel, f.filename))
    db.commit()
    _audit('CONSTRAINTS_FILE', ip=_client_ip(), user=session.get('user'), file=f.filename)
    return jsonify({'file_original': f.filename, 'has_file': True})

@app.route('/api/constraints/me/file', methods=['GET'])
def download_my_constraint_file():
    tid = _current_teacher_id()
    if not tid:
        return error_response('Réservé aux enseignants', 403)
    row = _constraint_row(tid)
    if not row or not row['file_path']:
        return error_response('Aucun fichier', 404)
    return _send_constraint_file(row)

@app.route('/api/constraints/me/file', methods=['DELETE'])
def delete_my_constraint_file():
    tid = _current_teacher_id()
    if not tid:
        return error_response('Réservé aux enseignants', 403)
    db = get_db()
    row = _constraint_row(tid)
    if row and row['file_path']:
        p = os.path.join(_CONSTRAINTS_DIR, row['file_path'])
        if os.path.isfile(p):
            try: os.remove(p)
            except OSError: pass
    db.execute('''UPDATE teacher_constraints SET file_path=NULL, file_original=NULL,
                  updated_at=CURRENT_TIMESTAMP WHERE teacher_id=?''', (tid,))
    db.commit()
    return jsonify({'has_file': False})

@app.route('/api/constraints', methods=['GET'])
def list_constraints():
    """Liste les contraintes renseignées (admin uniquement)."""
    if session.get('role') != 'admin':
        return error_response('Accès réservé à l\'administrateur', 403)
    rows = get_db().execute('''
        SELECT t.id AS teacher_id, t.name AS teacher_name,
               tc.content, tc.file_path, tc.file_original, tc.updated_at
        FROM teacher_constraints tc
        JOIN teachers t ON t.id = tc.teacher_id
        WHERE (tc.content IS NOT NULL AND tc.content <> '') OR tc.file_path IS NOT NULL
        ORDER BY t.name''').fetchall()
    return jsonify([{
        'teacher_id': r['teacher_id'], 'teacher_name': r['teacher_name'],
        'content': r['content'] or '', 'file_original': r['file_original'] or '',
        'has_file': bool(r['file_path']), 'updated_at': r['updated_at'] or '',
    } for r in rows])

@app.route('/api/constraints/<int:teacher_id>/file', methods=['GET'])
def download_constraint_file(teacher_id):
    """Téléchargement du fichier d'un enseignant (admin uniquement)."""
    if session.get('role') != 'admin':
        return error_response('Accès réservé à l\'administrateur', 403)
    row = _constraint_row(teacher_id)
    if not row or not row['file_path']:
        return error_response('Aucun fichier', 404)
    return _send_constraint_file(row)

# ======================= GESTION DES NOTES (admin) =======================
# Onglet admin : import d'un fichier Apogée par promo (roster + ELP), saisie des
# notes, coefficients de compétences éditables (renormalisation si note absente),
# export du PV au format Apogée (.xlsm rempli).

_NOTES_DIR = os.environ.get('EDT_NOTES_DIR') or os.path.join(_BASE_DIR, 'uploads', 'notes')
_ALLOWED_GRADE_EXT = {'.xlsm', '.xlsx'}
_MAX_GRADE_FILE = 15 * 1024 * 1024   # 15 Mo

# Noms standard des compétences BUT GIM par semestre (source : fichier de calcul)
_COMP_NAMES_5 = ['Maintenir', 'Améliorer', 'Installer', 'Manager', 'Sécuriser']
_COMP_NAMES_3 = ['Améliorer', 'Installer', 'Manager']
_COMP_BY_SEM = {
    'S1': _COMP_NAMES_5, 'S2': _COMP_NAMES_5, 'S3': _COMP_NAMES_5,
    'S4': _COMP_NAMES_5, 'S5': _COMP_NAMES_3, 'S6': _COMP_NAMES_3,
}

# Référence des coefficients par semestre, extraite une fois pour toutes du fichier
# de calcul (data/coefficients.json). Sert de défaut ; les éventuelles modifications
# sont stockées par année dans app_settings['grade_coeff_reference'].
_REF_COEFF_FILE = os.path.join(_BASE_DIR, 'data', 'coefficients.json')

def _ref_coeffs_defaults():
    try:
        with open(_REF_COEFF_FILE, encoding='utf-8') as f:
            return json.load(f)
    except (OSError, ValueError):
        return {}

def _load_ref_coeffs(db):
    """Coefficients de référence : surcharge stockée en base si présente, sinon fichier."""
    row = db.execute("SELECT value FROM app_settings WHERE key='grade_coeff_reference'").fetchone()
    if row and row['value']:
        try:
            return json.loads(row['value'])
        except ValueError:
            pass
    return _ref_coeffs_defaults()

def _seed_promo_competences(db, promo_id, semester):
    """Crée les compétences + coefficients d'une promo depuis la référence du
    semestre (mapping composant→ELP par index : groupe IS = SAÉ+STAGE, PORT, RES)."""
    ref = _load_ref_coeffs(db).get(semester)
    elps = db.execute('SELECT id, kind FROM grade_elps WHERE promo_id=? ORDER BY ordre',
                      (promo_id,)).fetchall()
    sae = [e['id'] for e in elps if e['kind'] == 'SAE']
    port = [e['id'] for e in elps if e['kind'] == 'PORT']
    res = [e['id'] for e in elps if e['kind'] == 'RES']
    code2elp = {}
    if ref:
        ref_is = [c for c in ref['components'] if c['kind'] in ('SAE', 'STAGE')]
        ref_port = [c for c in ref['components'] if c['kind'] == 'PORT']
        ref_res = [c for c in ref['components'] if c['kind'] == 'RES']
        for grp, target in ((ref_is, sae), (ref_port, port), (ref_res, res)):
            for i, c in enumerate(grp):
                if i < len(target):
                    code2elp[c['code']] = target[i]
        comps = ref['competences']
    else:
        comps = [{'name': n, 'coeffs': {}} for n in _COMP_BY_SEM.get(semester, _COMP_NAMES_5)]
    for i, comp in enumerate(comps, start=1):
        cid = db.execute('INSERT INTO grade_competences(promo_id, ordre, name) VALUES(?,?,?)',
                         (promo_id, i, comp['name'])).lastrowid
        for code, coeff in (comp.get('coeffs') or {}).items():
            eid = code2elp.get(code)
            if eid and coeff:
                db.execute('INSERT INTO grade_coefficients(competence_id, elp_id, coeff) VALUES(?,?,?)',
                           (cid, eid, coeff))

def _require_admin():
    """Renvoie une réponse 403 si l'utilisateur courant n'est pas admin, sinon None."""
    if session.get('role') != 'admin':
        return error_response('Accès réservé à l\'administrateur', 403)
    return None

def _require_promo_read():
    """Lecture Promotions : admin, ou enseignant connecté avec mot de passe
    (session promo_access) — en GET/HEAD uniquement. Les écritures restent admin."""
    role = session.get('role')
    if role == 'admin':
        return None
    if role == 'teacher' and session.get('promo_access') and request.method in ('GET', 'HEAD'):
        return None
    return error_response('Accès réservé', 403)

def _notes_year_dir():
    d = os.path.join(_NOTES_DIR, get_current_year() or 'default')
    os.makedirs(d, exist_ok=True)
    return d

def _grade_promo_row(promo_id):
    return get_db().execute(
        'SELECT * FROM grade_promos WHERE id=?', (promo_id,)).fetchone()

def _promo_meta(filename, label):
    """Déduit (code, semestre, formation) du nom de fichier et de l'intitulé."""
    base = os.path.splitext(os.path.basename(filename))[0]
    text = (base + ' ' + (label or '')).upper()
    m = re.search(r'S\s*([1-6])', text)
    semester = ('S' + m.group(1)) if m else ''
    formation = 'APP' if 'APP' in text else ('FI' if 'FI' in text else '')
    return base, semester, formation

def _parse_apogee(path):
    """Lit un fichier Apogée : (label, [elps], [students])."""
    import openpyxl
    is_xlsm = path.lower().endswith('.xlsm')
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=is_xlsm)
    ws = wb.active
    label = ws.cell(4, 5).value or ws.cell(1, 10).value or ''
    label = str(label).strip()
    elps, ordre = [], 0
    for c in range(5, ws.max_column + 1):
        typ = ws.cell(13, c).value
        code = ws.cell(8, c).value
        if typ == 'N' and code:
            title = ws.cell(14, c).value or code
            bareme_col = c + 1 if ws.cell(13, c + 1).value == 'B' else None
            scode = str(code).upper()
            title_l = str(title).lower()
            if 'portfolio' in title_l:
                kind = 'PORT'
            elif 'IS' in scode:
                kind = 'SAE'
            elif 'IR' in scode:
                kind = 'RES'
            else:
                kind = 'AUTRE'
            ordre += 1
            elps.append({'ordre': ordre, 'code': str(code).strip(),
                         'title': str(title).strip(), 'kind': kind,
                         'note_col': c, 'bareme_col': bareme_col})
    students, r, o = [], 18, 0
    def _txt(v):
        return ('' if v is None else str(v)).strip()
    while r <= 1000:
        num = ws.cell(r, 1).value
        if num in (None, ''):
            break
        o += 1
        students.append({'ordre': o, 'src_row': r, 'numero': _txt(num),
                         'nom': _txt(ws.cell(r, 2).value),
                         'prenom': _txt(ws.cell(r, 3).value),
                         'naissance': _txt(ws.cell(r, 4).value)})
        r += 1
    return label, elps, students

def _competence_averages(db, promo_id):
    """Moyenne par compétence et par étudiant, avec renormalisation : seuls les
    coefficients dont la note existe entrent au numérateur ET au dénominateur."""
    elp_ids = [r['id'] for r in db.execute(
        'SELECT id FROM grade_elps WHERE promo_id=?', (promo_id,))]
    comps = db.execute(
        'SELECT id FROM grade_competences WHERE promo_id=? ORDER BY ordre, id',
        (promo_id,)).fetchall()
    coeffs = {}
    for r in db.execute('''SELECT c.competence_id, c.elp_id, c.coeff
                           FROM grade_coefficients c
                           JOIN grade_competences gc ON gc.id = c.competence_id
                           WHERE gc.promo_id=?''', (promo_id,)):
        coeffs[(r['competence_id'], r['elp_id'])] = r['coeff'] or 0
    marks = {}
    for r in db.execute('''SELECT m.student_id, m.elp_id, m.note FROM grade_marks m
                           JOIN grade_students s ON s.id = m.student_id
                           WHERE s.promo_id=? AND m.note IS NOT NULL''', (promo_id,)):
        marks[(r['student_id'], r['elp_id'])] = r['note']
    students = [r['id'] for r in db.execute(
        'SELECT id FROM grade_students WHERE promo_id=?', (promo_id,))]
    out = {}
    for sid in students:
        row = {}
        for c in comps:
            cid = c['id']
            num = den = 0.0
            for eid in elp_ids:
                coeff = coeffs.get((cid, eid), 0) or 0
                if coeff and (sid, eid) in marks:
                    num += coeff * marks[(sid, eid)]
                    den += coeff
            row[str(cid)] = round(num / den, 3) if den > 0 else None
        out[str(sid)] = row
    return out

def _promo_payload(db, promo_id):
    promo = _grade_promo_row(promo_id)
    if not promo:
        return None
    elps = [dict(r) for r in db.execute(
        'SELECT id, ordre, code, title, kind, note_col, bareme_col '
        'FROM grade_elps WHERE promo_id=? ORDER BY ordre, id', (promo_id,))]
    students = [dict(r) for r in db.execute(
        'SELECT id, ordre, numero, nom, prenom, naissance '
        'FROM grade_students WHERE promo_id=? ORDER BY ordre, id', (promo_id,))]
    comps = [dict(r) for r in db.execute(
        'SELECT id, ordre, name FROM grade_competences WHERE promo_id=? ORDER BY ordre, id',
        (promo_id,))]
    marks = {}
    for r in db.execute('''SELECT m.student_id, m.elp_id, m.note FROM grade_marks m
                           JOIN grade_students s ON s.id = m.student_id
                           WHERE s.promo_id=?''', (promo_id,)):
        marks[f"{r['student_id']}_{r['elp_id']}"] = r['note']
    coeffs = {}
    for r in db.execute('''SELECT c.competence_id, c.elp_id, c.coeff
                           FROM grade_coefficients c
                           JOIN grade_competences gc ON gc.id = c.competence_id
                           WHERE gc.promo_id=?''', (promo_id,)):
        coeffs[f"{r['competence_id']}_{r['elp_id']}"] = r['coeff']
    return {
        'promo': {'id': promo['id'], 'code': promo['code'], 'label': promo['label'],
                  'semester': promo['semester'], 'formation': promo['formation'],
                  'template_orig': promo['template_orig'], 'updated_at': promo['updated_at']},
        'elps': elps, 'students': students, 'competences': comps,
        'marks': marks, 'coefficients': coeffs,
        'averages': _competence_averages(db, promo_id),
    }

@app.route('/api/grades/promos', methods=['GET'])
def list_grade_promos():
    err = _require_admin()
    if err:
        return err
    rows = get_db().execute('''SELECT p.id, p.code, p.label, p.semester, p.formation,
                                      p.template_orig, p.updated_at,
                                      (SELECT COUNT(*) FROM grade_students s WHERE s.promo_id=p.id) AS nb_students,
                                      (SELECT COUNT(*) FROM grade_elps e WHERE e.promo_id=p.id) AS nb_elps
                               FROM grade_promos p ORDER BY p.semester, p.formation, p.code''').fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/grades/promos/<int:promo_id>', methods=['GET'])
def get_grade_promo(promo_id):
    err = _require_admin()
    if err:
        return err
    payload = _promo_payload(get_db(), promo_id)
    if not payload:
        return error_response('Promo introuvable', 404)
    return jsonify(payload)

@app.route('/api/grades/promos/import', methods=['POST'])
def import_grade_promo():
    err = _require_admin()
    if err:
        return err
    f = request.files.get('file')
    if not f or not f.filename:
        return error_response('Aucun fichier reçu', 400)
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in _ALLOWED_GRADE_EXT:
        return error_response('Type de fichier non autorisé (.xlsm/.xlsx)', 400)
    if request.content_length and request.content_length > _MAX_GRADE_FILE:
        return error_response('Fichier trop volumineux (max 15 Mo)', 400)

    # Sauvegarde temporaire pour lecture
    safe = secure_filename(f.filename) or ('promo' + ext)
    dest = os.path.join(_notes_year_dir(), safe)
    f.save(dest)
    if os.path.getsize(dest) > _MAX_GRADE_FILE:
        os.remove(dest)
        return error_response('Fichier trop volumineux (max 15 Mo)', 400)
    try:
        label, elps, students = _parse_apogee(dest)
    except Exception as e:
        try: os.remove(dest)
        except OSError: pass
        return error_response(f'Lecture impossible : {e}', 400)
    if not elps or not students:
        try: os.remove(dest)
        except OSError: pass
        return error_response('Fichier non reconnu (aucun ELP / étudiant détecté)', 400)

    code, semester, formation = _promo_meta(f.filename, label)
    rel = os.path.relpath(dest, _NOTES_DIR)
    db = get_db()
    # Remplace une promo de même code (réimport) en conservant l'id si possible
    existing = db.execute('SELECT id FROM grade_promos WHERE code=?', (code,)).fetchone()
    if existing:
        pid = existing['id']
        db.execute('DELETE FROM grade_elps WHERE promo_id=?', (pid,))
        db.execute('DELETE FROM grade_students WHERE promo_id=?', (pid,))
        db.execute('DELETE FROM grade_competences WHERE promo_id=?', (pid,))
        db.execute('''UPDATE grade_promos SET label=?, semester=?, formation=?,
                      template=?, template_orig=?, updated_at=CURRENT_TIMESTAMP WHERE id=?''',
                   (label, semester, formation, rel, f.filename, pid))
    else:
        cur = db.execute('''INSERT INTO grade_promos(code, label, semester, formation,
                            template, template_orig, updated_at)
                            VALUES(?,?,?,?,?,?,CURRENT_TIMESTAMP)''',
                         (code, label, semester, formation, rel, f.filename))
        pid = cur.lastrowid
    for e in elps:
        db.execute('''INSERT INTO grade_elps(promo_id, ordre, code, title, kind, note_col, bareme_col)
                      VALUES(?,?,?,?,?,?,?)''',
                   (pid, e['ordre'], e['code'], e['title'], e['kind'], e['note_col'], e['bareme_col']))
    for s in students:
        db.execute('''INSERT INTO grade_students(promo_id, ordre, src_row, numero, nom, prenom, naissance)
                      VALUES(?,?,?,?,?,?,?)''',
                   (pid, s['ordre'], s['src_row'], s['numero'], s['nom'], s['prenom'], s['naissance']))
    _seed_promo_competences(db, pid, semester)
    db.commit()
    _audit('GRADES_IMPORT', ip=_client_ip(), user=session.get('user'), file=f.filename)
    return jsonify(_promo_payload(db, pid))

@app.route('/api/grades/promos/<int:promo_id>', methods=['DELETE'])
def delete_grade_promo(promo_id):
    err = _require_admin()
    if err:
        return err
    db = get_db()
    promo = _grade_promo_row(promo_id)
    if not promo:
        return error_response('Promo introuvable', 404)
    if promo['template']:
        p = os.path.join(_NOTES_DIR, promo['template'])
        if os.path.isfile(p):
            try: os.remove(p)
            except OSError: pass
    db.execute('DELETE FROM grade_promos WHERE id=?', (promo_id,))
    db.commit()
    return jsonify({'deleted': True})

@app.route('/api/grades/promos/<int:promo_id>/marks', methods=['PUT'])
def save_grade_marks(promo_id):
    err = _require_admin()
    if err:
        return err
    if not _grade_promo_row(promo_id):
        return error_response('Promo introuvable', 404)
    data = request.get_json() or {}
    marks = data.get('marks') or []
    db = get_db()
    valid_students = {r['id'] for r in db.execute(
        'SELECT id FROM grade_students WHERE promo_id=?', (promo_id,))}
    valid_elps = {r['id'] for r in db.execute(
        'SELECT id FROM grade_elps WHERE promo_id=?', (promo_id,))}
    for m in marks:
        try:
            sid = int(m.get('student_id')); eid = int(m.get('elp_id'))
        except (TypeError, ValueError):
            continue
        if sid not in valid_students or eid not in valid_elps:
            continue
        note = m.get('note')
        if note in (None, ''):
            db.execute('DELETE FROM grade_marks WHERE student_id=? AND elp_id=?', (sid, eid))
        else:
            try:
                note = float(note)
            except (TypeError, ValueError):
                continue
            db.execute('''INSERT INTO grade_marks(student_id, elp_id, note) VALUES(?,?,?)
                          ON CONFLICT(student_id, elp_id) DO UPDATE SET note=excluded.note''',
                       (sid, eid, note))
    db.commit()
    return jsonify({'averages': _competence_averages(db, promo_id)})

@app.route('/api/grades/promos/<int:promo_id>/competences', methods=['PUT'])
def save_grade_competences(promo_id):
    """Remplace l'ensemble compétences + coefficients de la promo."""
    err = _require_admin()
    if err:
        return err
    if not _grade_promo_row(promo_id):
        return error_response('Promo introuvable', 404)
    data = request.get_json() or {}
    comps = data.get('competences') or []
    db = get_db()
    valid_elps = {r['id'] for r in db.execute(
        'SELECT id FROM grade_elps WHERE promo_id=?', (promo_id,))}
    db.execute('DELETE FROM grade_competences WHERE promo_id=?', (promo_id,))
    for i, c in enumerate(comps, start=1):
        name = (c.get('name') or '').strip() or f'Compétence {i}'
        cur = db.execute('INSERT INTO grade_competences(promo_id, ordre, name) VALUES(?,?,?)',
                         (promo_id, i, name))
        cid = cur.lastrowid
        for eid_str, coeff in (c.get('coeffs') or {}).items():
            try:
                eid = int(eid_str); cval = float(coeff)
            except (TypeError, ValueError):
                continue
            if eid in valid_elps and cval:
                db.execute('INSERT INTO grade_coefficients(competence_id, elp_id, coeff) VALUES(?,?,?)',
                           (cid, eid, cval))
    db.commit()
    return jsonify(_promo_payload(db, promo_id))

@app.route('/api/grades/coeff-reference', methods=['GET'])
def get_coeff_reference():
    """Coefficients de référence par semestre (édités dans Matières → Coefficients)."""
    err = _require_admin()
    if err:
        return err
    return jsonify(_load_ref_coeffs(get_db()))

@app.route('/api/grades/coeff-reference', methods=['PUT'])
def save_coeff_reference():
    err = _require_admin()
    if err:
        return err
    data = request.get_json()
    if not isinstance(data, dict):
        return error_response('Format invalide', 400)
    db = get_db()
    db.execute('''INSERT INTO app_settings(key, value) VALUES('grade_coeff_reference', ?)
                  ON CONFLICT(key) DO UPDATE SET value=excluded.value''',
               (json.dumps(data, ensure_ascii=False),))
    db.commit()
    return jsonify(_load_ref_coeffs(db))

@app.route('/api/grades/coeff-reference/reset', methods=['POST'])
def reset_coeff_reference():
    """Réinitialise les coefficients aux valeurs d'origine du fichier de calcul."""
    err = _require_admin()
    if err:
        return err
    db = get_db()
    db.execute("DELETE FROM app_settings WHERE key='grade_coeff_reference'")
    db.commit()
    return jsonify(_ref_coeffs_defaults())

@app.route('/api/grades/promos/<int:promo_id>/coefficients/reseed', methods=['POST'])
def reseed_grade_coefficients(promo_id):
    """Recharge les compétences/coefficients de la promo depuis la référence du semestre."""
    err = _require_admin()
    if err:
        return err
    promo = _grade_promo_row(promo_id)
    if not promo:
        return error_response('Promo introuvable', 404)
    db = get_db()
    db.execute('DELETE FROM grade_competences WHERE promo_id=?', (promo_id,))
    _seed_promo_competences(db, promo_id, promo['semester'])
    db.commit()
    return jsonify(_promo_payload(db, promo_id))

@app.route('/api/grades/promos/<int:promo_id>/export', methods=['GET'])
def export_grade_promo(promo_id):
    err = _require_admin()
    if err:
        return err
    import openpyxl, io
    db = get_db()
    promo = _grade_promo_row(promo_id)
    if not promo:
        return error_response('Promo introuvable', 404)
    if not promo['template']:
        return error_response('Aucun modèle disponible pour cette promo', 404)
    tpath = os.path.join(_NOTES_DIR, promo['template'])
    if not os.path.isfile(tpath):
        return error_response('Modèle introuvable sur le serveur', 404)
    is_xlsm = tpath.lower().endswith('.xlsm')
    wb = openpyxl.load_workbook(tpath, keep_vba=is_xlsm)
    ws = wb.active
    elps = db.execute('SELECT id, note_col, bareme_col FROM grade_elps WHERE promo_id=?',
                      (promo_id,)).fetchall()
    students = db.execute('SELECT id, src_row FROM grade_students WHERE promo_id=?',
                          (promo_id,)).fetchall()
    marks = {}
    for r in db.execute('''SELECT m.student_id, m.elp_id, m.note FROM grade_marks m
                           JOIN grade_students s ON s.id = m.student_id
                           WHERE s.promo_id=? AND m.note IS NOT NULL''', (promo_id,)):
        marks[(r['student_id'], r['elp_id'])] = r['note']
    for s in students:
        row = s['src_row']
        if not row:
            continue
        for e in elps:
            note = marks.get((s['id'], e['id']))
            if note is not None and e['note_col']:
                ws.cell(row=row, column=e['note_col']).value = note
                if e['bareme_col'] and ws.cell(row=row, column=e['bareme_col']).value in (None, ''):
                    ws.cell(row=row, column=e['bareme_col']).value = 20
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    ext = '.xlsm' if is_xlsm else '.xlsx'
    fname = (promo['template_orig'] or (promo['code'] + ext))
    mime = ('application/vnd.ms-excel.sheet.macroEnabled.12' if is_xlsm
            else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    _audit('GRADES_EXPORT', ip=_client_ip(), user=session.get('user'), promo=promo['code'])
    return send_file(bio, as_attachment=True, download_name=fname, mimetype=mime)

# ======================= GESTION DES PROMOTIONS (admin) =======================
# Une promotion = cohorte d'étudiants sur 3 ans (ex 25-28_ALT), stockée dans une
# base séparée (databases/promotions.db) pour un suivi pluriannuel des résultats.

def _promotion_name(start_year):
    """Nom d'une cohorte sur 3 ans (sans suffixe formation) : ex 2025 -> '25-28'."""
    return f"{start_year % 100:02d}-{(start_year + 3) % 100:02d}"

_SUBCOHORTS = ('FTP', 'ALT')   # sous-cohortes d'une promotion (cohorte)

def _parse_student_list(path):
    """Lit une liste d'étudiants : auto-détection format Apogée (.xlsm/.xlsx,
    roster ligne 18+) ou tableur simple (en-têtes ligne 1). Retourne
    [{numero, nom, prenom, naissance}]."""
    import openpyxl
    import unicodedata
    is_xlsm = path.lower().endswith('.xlsm')
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=is_xlsm)
    ws = wb.active

    def txt(v):
        return ('' if v is None else str(v)).strip()

    # PV de jury : n° étudiant en col B sous la ligne des codes ELP (T3IS../T3IR..),
    # NOM/Prénom en cols C/D. Pas de date de naissance dans ce modèle.
    if _find_pv_code_row(ws):
        _, rows = _pv_marks_ws(ws)
        return [{'numero': r['numero'], 'nom': r['nom'], 'prenom': r['prenom'],
                 'naissance': ''} for r in rows]

    # Détection Apogée : marqueur 'apoL_a01_code' en A6 ou 'Numéro' en A17
    a6 = txt(ws.cell(6, 1).value).lower()
    h17 = txt(ws.cell(17, 1).value).lower()
    if a6 == 'apol_a01_code' or h17 in ('numéro', 'numero'):
        students, r = [], 18
        while r <= 2000:
            num = ws.cell(r, 1).value
            if num in (None, ''):
                break
            students.append({'numero': txt(num), 'nom': txt(ws.cell(r, 2).value),
                             'prenom': txt(ws.cell(r, 3).value), 'naissance': txt(ws.cell(r, 4).value)})
            r += 1
        return students

    # Format simple : en-têtes ligne 1
    def norm(s):
        s = unicodedata.normalize('NFD', txt(s)).encode('ascii', 'ignore').decode().lower().strip()
        return s
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(1, c).value)
        if not h:
            continue
        if h in ('nom', 'name', 'nom de famille', 'last name'):
            headers.setdefault('nom', c)
        elif h in ('prenom', 'first name'):
            headers.setdefault('prenom', c)
        elif h in ('numero', 'n', 'no', 'num', 'code', 'code apogee', 'apogee', 'matricule', 'n etudiant'):
            headers.setdefault('numero', c)
        elif 'naiss' in h or h in ('ne le', 'nee le', 'dob', 'date de naissance'):
            headers.setdefault('naissance', c)
    students = []
    if not headers:
        return students
    r = 2
    while r <= 5000:
        if all(ws.cell(r, c).value in (None, '') for c in range(1, ws.max_column + 1)):
            break
        rec = {'numero': '', 'nom': '', 'prenom': '', 'naissance': ''}
        for key, col in headers.items():
            rec[key] = txt(ws.cell(r, col).value)
        if rec['nom'] or rec['prenom'] or rec['numero']:
            students.append(rec)
        r += 1
    return students

def _dedup_key(s):
    """Clé d'unicité d'un étudiant : numéro si présent, sinon nom+prénom+naissance."""
    num = (s.get('numero') or '').strip().lower()
    if num:
        return ('num', num)
    return ('id', (s.get('nom') or '').strip().lower(),
            (s.get('prenom') or '').strip().lower(),
            (s.get('naissance') or '').strip().lower())

def _promotion_payload(db, pid):
    promo = db.execute('SELECT * FROM promotions WHERE id=?', (pid,)).fetchone()
    if not promo:
        return None
    students = [dict(r) for r in db.execute(
        '''SELECT id, numero, nom, prenom, naissance, statut, abandon_semestre, formation
           FROM promotion_students WHERE promotion_id=?
           ORDER BY nom COLLATE NOCASE, prenom COLLATE NOCASE''', (pid,))]
    counts = {st: 0 for st in _STUDENT_STATUSES}
    # Compteurs par sous-cohorte : {'FTP': {statut:n,...,total}, 'ALT': {...}}
    sub_counts = {f: {st: 0 for st in _STUDENT_STATUSES} for f in _SUBCOHORTS}
    for s in students:
        counts[s['statut']] = counts.get(s['statut'], 0) + 1
        f = s.get('formation') if s.get('formation') in _SUBCOHORTS else 'FTP'
        sub_counts[f][s['statut']] = sub_counts[f].get(s['statut'], 0) + 1
    for f in _SUBCOHORTS:
        sub_counts[f]['total'] = sum(sub_counts[f][st] for st in _STUDENT_STATUSES)
    return {'promotion': dict(promo), 'students': students, 'counts': counts,
            'sub_counts': sub_counts, 'subcohorts': list(_SUBCOHORTS),
            'total': len(students), 'statuses': _STUDENT_STATUSES, 'semesters': _PROMO_SEMESTERS}

@app.route('/api/promotions', methods=['GET'])
def list_promotions():
    err = _require_promo_read()
    if err:
        return err
    rows = get_promotions_db().execute('''
        SELECT p.*,
               (SELECT COUNT(*) FROM promotion_students s WHERE s.promotion_id=p.id) AS total,
               (SELECT COUNT(*) FROM promotion_students s WHERE s.promotion_id=p.id AND s.statut='Actif') AS actifs
        FROM promotions p ORDER BY p.start_year DESC, p.name''').fetchall()
    # Nom du programme affecté (base programmes séparée → résolution en mémoire)
    prog_names = {}
    try:
        for r in get_programmes_db().execute('SELECT id, name FROM programmes').fetchall():
            prog_names[r['id']] = r['name']
    except sqlite3.Error:
        pass
    out = []
    for r in rows:
        d = dict(r)
        d['programme_name'] = prog_names.get(d.get('programme_id'))
        out.append(d)
    return jsonify(out)

@app.route('/api/promotions-lite', methods=['GET'])
def list_promotions_lite():
    """Liste légère des promotions (id, nom, formation, programme) — lecture seule,
    accessible à tout utilisateur connecté (sélecteur de l'onglet Matières)."""
    pdb = get_promotions_db()
    rows = pdb.execute('''SELECT id, name, formation, start_year, end_year, programme_id
                          FROM promotions ORDER BY start_year DESC, name''').fetchall()
    prog_names = {}
    try:
        for r in get_programmes_db().execute('SELECT id, name FROM programmes').fetchall():
            prog_names[r['id']] = r['name']
    except sqlite3.Error:
        pass
    out = []
    for r in rows:
        d = dict(r)
        d['programme_name'] = prog_names.get(d.get('programme_id'))
        out.append(d)
    return jsonify(out)

@app.route('/api/promotions/<int:pid>/matieres', methods=['GET'])
def promotion_matieres(pid):
    """Matières (composants) du programme affecté à la promotion, par semestre.
    Lecture seule (tous utilisateurs connectés) — alimente l'onglet Matières."""
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    has_prog = _promo_programme_id(pdb, pid) is not None
    data = _promo_coeffs(pdb, pid) if has_prog else {}
    out = {}
    for sem, d in (data or {}).items():
        comps = [{k: c.get(k) for k in ('code', 'label', 'short_label', 'apogee_code', 'kind', 'volumes')}
                 for c in d.get('components', [])]
        out[sem] = {'components': comps}
    return jsonify({'has_programme': has_prog, 'semesters': out})

@app.route('/api/matieres-programme', methods=['GET'])
def matieres_programme():
    """Pour l'ANNÉE ACTIVE : matières du programme par semestre, en résolvant
    automatiquement la cohorte de chaque semestre (S1/S2→1re année, S3/S4→2e, S5/S6→3e).
    FTP et ALT d'une même année partageant le même programme, on prend l'une ou l'autre.
    Lecture seule (tous utilisateurs connectés) — alimente l'onglet Matières."""
    try:
        start = int(str(get_current_year() or '').split('-')[0])
    except (ValueError, IndexError):
        start = None
    pdb = get_promotions_db()
    prog_names = {}
    try:
        for r in get_programmes_db().execute('SELECT id, name FROM programmes').fetchall():
            prog_names[r['id']] = r['name']
    except sqlite3.Error:
        pass
    out = {}
    any_prog = False
    for n in range(1, 7):
        sem = 'S%d' % n
        year_group = (n + 1) // 2          # S1/S2→1, S3/S4→2, S5/S6→3
        comps, promo_name, prog_name = [], None, None
        if start is not None:
            cohort_year = start - (year_group - 1)
            row = pdb.execute(
                '''SELECT id, name, programme_id FROM promotions WHERE start_year=?
                   ORDER BY (programme_id IS NULL), formation LIMIT 1''',
                (cohort_year,)).fetchone()
            if row:
                promo_name = row['name']
                if row['programme_id']:
                    d = (_promo_coeffs(pdb, row['id']) or {}).get(sem) or {}
                    comps = [{k: c.get(k) for k in ('code', 'label', 'short_label',
                                                    'apogee_code', 'kind', 'volumes')}
                             for c in d.get('components', [])]
                    prog_name = prog_names.get(row['programme_id'])
                    if comps:
                        any_prog = True
        out[sem] = {'components': comps, 'promo_name': promo_name, 'programme_name': prog_name}
    return jsonify({'has_any': any_prog, 'semesters': out})

@app.route('/api/promotions', methods=['POST'])
def create_promotion():
    err = _require_admin()
    if err:
        return err
    data = request.get_json() or {}
    try:
        start = int(data.get('start_year'))
    except (TypeError, ValueError):
        return error_response('Année de début invalide', 400)
    if start < 2000 or start > 2100:
        return error_response('Année de début hors limites', 400)
    name = _promotion_name(start)
    db = get_promotions_db()
    if db.execute('SELECT 1 FROM promotions WHERE name=?', (name,)).fetchone():
        return error_response(f'La promotion « {name} » existe déjà', 409)
    # Une cohorte contient les 2 sous-cohortes (FTP+ALT) ; formation='MUT' au niveau promo.
    cur = db.execute('''INSERT INTO promotions(name, formation, start_year, end_year, updated_at)
                        VALUES(?,'MUT',?,?,CURRENT_TIMESTAMP)''', (name, start, start + 3))
    db.commit()
    new_id = cur.lastrowid
    # Crée/complète les 3 années universitaires de la promo (start→+1→+2→+3),
    # chacune copiée de l'année précédente existante (matières, sessions,
    # répartition, calendrier). Sans effet si les années existent déjà.
    created_years = []
    try:
        created_years = _ensure_years_for_promo(start)
    except Exception as e:
        app.logger.warning('Création des années de la promo %s échouée : %s', new_id, e)
    _audit('PROMO_CREATE', ip=_client_ip(), user=session.get('user'), promo=name,
           years=','.join(created_years))
    payload = _promotion_payload(db, new_id)
    payload['created_years'] = created_years
    return jsonify(payload)

@app.route('/api/promotions/<int:pid>', methods=['GET'])
def get_promotion(pid):
    err = _require_promo_read()
    if err:
        return err
    payload = _promotion_payload(get_promotions_db(), pid)
    if not payload:
        return error_response('Promotion introuvable', 404)
    return jsonify(payload)

@app.route('/api/promotions/<int:pid>', methods=['DELETE'])
def delete_promotion(pid):
    err = _require_admin()
    if err:
        return err
    db = get_promotions_db()
    if not db.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    db.execute('DELETE FROM promotions WHERE id=?', (pid,))
    db.commit()
    # Les années universitaires ne sont PAS supprimées (partagées entre cohortes).
    return jsonify({'deleted': True})

@app.route('/api/promotions/<int:pid>/students', methods=['POST'])
def add_promotion_student(pid):
    err = _require_admin()
    if err:
        return err
    db = get_promotions_db()
    if not db.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    data = request.get_json() or {}
    nom = (data.get('nom') or '').strip()
    prenom = (data.get('prenom') or '').strip()
    numero = (data.get('numero') or '').strip()
    if not (nom or prenom or numero):
        return error_response('Renseignez au moins un nom ou un numéro', 400)
    statut = (data.get('statut') or 'Actif').strip()
    if statut not in _STUDENT_STATUSES:
        statut = 'Actif'
    sem = (data.get('abandon_semestre') or '').strip()
    if statut != 'Abandon' or sem not in _PROMO_SEMESTERS:
        sem = None
    formation = (data.get('formation') or '').strip().upper()
    if formation not in _SUBCOHORTS:
        formation = 'FTP'
    db.execute('''INSERT INTO promotion_students(promotion_id, numero, nom, prenom, naissance, statut, abandon_semestre, formation)
                  VALUES(?,?,?,?,?,?,?,?)''',
               (pid, numero, nom, prenom, (data.get('naissance') or '').strip(), statut, sem, formation))
    db.commit()
    return jsonify(_promotion_payload(db, pid))

@app.route('/api/promotions/<int:pid>/students/<int:sid>', methods=['PUT'])
def update_promotion_student(pid, sid):
    err = _require_admin()
    if err:
        return err
    db = get_promotions_db()
    row = db.execute('SELECT 1 FROM promotion_students WHERE id=? AND promotion_id=?', (sid, pid)).fetchone()
    if not row:
        return error_response('Étudiant introuvable', 404)
    data = request.get_json() or {}
    fields, params = [], []
    for key in ('numero', 'nom', 'prenom', 'naissance'):
        if key in data:
            fields.append(f'{key}=?'); params.append((data.get(key) or '').strip())
    # Profil d'entrée : BAC / cursus / recrutement (valeur autorisée ou vide)
    for key, allowed in _STUDENT_PROFILE.items():
        if key in data:
            val = (data.get(key) or '').strip()
            fields.append(f'{key}=?'); params.append(val if val in allowed else None)
    new_statut = None
    if 'statut' in data:
        st = (data.get('statut') or 'Actif').strip()
        if st in _STUDENT_STATUSES:
            new_statut = st
            fields.append('statut=?'); params.append(st)
    # Semestre d'abandon : effacé si le statut n'est plus « Abandon »
    if new_statut is not None and new_statut != 'Abandon':
        fields.append('abandon_semestre=?'); params.append(None)
    elif 'abandon_semestre' in data:
        sem = (data.get('abandon_semestre') or '').strip()
        fields.append('abandon_semestre=?'); params.append(sem if sem in _PROMO_SEMESTERS else None)
    if fields:
        params += [sid, pid]
        db.execute(f'UPDATE promotion_students SET {", ".join(fields)} WHERE id=? AND promotion_id=?', params)
        db.commit()
    return jsonify(_promotion_payload(db, pid))

@app.route('/api/promotions/<int:pid>/students/<int:sid>', methods=['DELETE'])
def delete_promotion_student(pid, sid):
    err = _require_admin()
    if err:
        return err
    db = get_promotions_db()
    db.execute('DELETE FROM promotion_students WHERE id=? AND promotion_id=?', (sid, pid))
    db.commit()
    return jsonify(_promotion_payload(db, pid))

@app.route('/api/promotions/<int:pid>/students/import', methods=['POST'])
def import_promotion_students(pid):
    err = _require_admin()
    if err:
        return err
    db = get_promotions_db()
    if not db.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    f = request.files.get('file')
    if not f or not f.filename:
        return error_response('Aucun fichier reçu', 400)
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in _ALLOWED_GRADE_EXT:
        return error_response('Type de fichier non autorisé (.xlsm/.xlsx)', 400)
    if request.content_length and request.content_length > _MAX_GRADE_FILE:
        return error_response('Fichier trop volumineux (max 15 Mo)', 400)
    import tempfile
    tmp = os.path.join(tempfile.gettempdir(), secure_filename(f.filename) or ('import' + ext))
    f.save(tmp)
    try:
        students = _parse_student_list(tmp)
    except Exception as e:
        return error_response(f'Lecture impossible : {e}', 400)
    finally:
        try: os.remove(tmp)
        except OSError: pass
    if not students:
        return error_response('Aucun étudiant détecté dans le fichier', 400)
    # Sous-cohorte cible (FTP/ALT) de cet import
    formation = (request.form.get('formation') or '').strip().upper()
    if formation not in _SUBCOHORTS:
        formation = 'FTP'
    # Dédoublonnage : on n'importe pas un étudiant déjà présent dans la promotion
    existing = {_dedup_key(dict(r)) for r in db.execute(
        'SELECT numero, nom, prenom, naissance FROM promotion_students WHERE promotion_id=?', (pid,))}
    imported = skipped = 0
    for s in students:
        key = _dedup_key(s)
        if key in existing:
            skipped += 1
            continue
        existing.add(key)
        db.execute('''INSERT INTO promotion_students(promotion_id, numero, nom, prenom, naissance, statut, formation)
                      VALUES(?,?,?,?,?,'Actif',?)''',
                   (pid, s['numero'], s['nom'], s['prenom'], s['naissance'], formation))
        imported += 1
    db.commit()
    _audit('PROMO_IMPORT', ip=_client_ip(), user=session.get('user'),
           imported=imported, skipped=skipped)
    payload = _promotion_payload(db, pid)
    payload['import_report'] = {'imported': imported, 'skipped': skipped, 'total_fichier': len(students)}
    return jsonify(payload)

# ---- Effectif par année d'étude (1..3) : report auto du jury + ajustements manuels ----

def _year_effectif_payload(pdb, pid, year):
    """Effectif d'une promotion pour une année d'étude (1..3) avec compteurs par
    sous-cohorte. Chaque étudiant est marqué « entrant » (entry_year==année) et, le cas
    échéant, ajusté à la main ('add'/'remove')."""
    promo = pdb.execute('SELECT * FROM promotions WHERE id=?', (pid,)).fetchone()
    if not promo:
        return None
    roster = _year_rosters(pdb, pid).get(year, set())
    overrides = {r['student_id']: r['action'] for r in pdb.execute(
        'SELECT student_id, action FROM promotion_year_override WHERE promotion_id=? AND year=?',
        (pid, year))}
    fm = _year_formation_map(pdb, pid, year)          # sous-cohorte propre à cette année
    students = []
    for r in pdb.execute('''SELECT id, numero, nom, prenom, naissance, statut, abandon_semestre,
                                   formation, entry_year, bac, cursus, recrutement
                            FROM promotion_students
                            WHERE promotion_id=? ORDER BY nom COLLATE NOCASE, prenom COLLATE NOCASE''',
                         (pid,)):
        if r['id'] in roster:
            d = dict(r)
            d['formation'] = fm.get(r['id'], d.get('formation') or 'FTP')
            d['entrant'] = (r['entry_year'] or 1) == year
            d['manual'] = overrides.get(r['id'])
            students.append(d)
    sub_counts = {f: {st: 0 for st in _STUDENT_STATUSES} for f in _SUBCOHORTS}
    for s in students:
        f = s.get('formation') if s.get('formation') in _SUBCOHORTS else 'FTP'
        sub_counts[f][s['statut']] = sub_counts[f].get(s['statut'], 0) + 1
    for f in _SUBCOHORTS:
        sub_counts[f]['total'] = sum(sub_counts[f][st] for st in _STUDENT_STATUSES)
    return {'promotion': dict(promo), 'year': year, 'students': students,
            'sub_counts': sub_counts, 'subcohorts': list(_SUBCOHORTS),
            'total': len(students), 'statuses': _STUDENT_STATUSES,
            'semesters': _PROMO_SEMESTERS, 'years': [1, 2, 3],
            'profile_options': {k: list(v) for k, v in _STUDENT_PROFILE.items()}}

@app.route('/api/promotions/<int:pid>/effectif/<int:year>', methods=['GET'])
def get_year_effectif(pid, year):
    err = _require_promo_read()
    if err:
        return err
    if year not in (1, 2, 3):
        return error_response('Année invalide', 400)
    payload = _year_effectif_payload(get_promotions_db(), pid, year)
    if not payload:
        return error_response('Promotion introuvable', 404)
    return jsonify(payload)

@app.route('/api/promotions/<int:pid>/tuteurs/<int:year>', methods=['GET'])
def get_year_tuteurs(pid, year):
    """Tuteurs (universitaire + entreprise) de l'effectif d'une année.
    ALT : une saisie faite une année reste valable les années suivantes tant
    qu'elle n'est pas remplacée (inherited_from = année d'origine). FTP : la
    saisie vaut uniquement pour l'année du stage (pas de report)."""
    err = _require_admin()
    if err:
        return err
    if year not in (1, 2, 3):
        return error_response('Année invalide', 400)
    pdb = get_promotions_db()
    payload = _year_effectif_payload(pdb, pid, year)
    if not payload:
        return error_response('Promotion introuvable', 404)
    by_student = {}
    for r in pdb.execute('''SELECT student_id, year, tuteur_univ, tuteur_entreprise, entreprise
                            FROM student_tutors WHERE promotion_id = ? AND year <= ?''',
                         (pid, year)):
        by_student.setdefault(r['student_id'], {})[r['year']] = r
    students = []
    for s in payload['students']:
        t = by_student.get(s['id'], {})
        entry, inherited = None, None
        if year in t:
            entry = t[year]
        elif (s.get('formation') or 'FTP') == 'ALT' and t:
            prev = max(t.keys())
            entry, inherited = t[prev], prev
        students.append({
            'id': s['id'], 'numero': s['numero'], 'nom': s['nom'], 'prenom': s['prenom'],
            'formation': s.get('formation') or 'FTP', 'statut': s['statut'],
            'tuteur_univ': (entry['tuteur_univ'] if entry else '') or '',
            'tuteur_entreprise': (entry['tuteur_entreprise'] if entry else '') or '',
            'entreprise': (entry['entreprise'] if entry else '') or '',
            'inherited_from': inherited,
        })
    return jsonify({'year': year, 'students': students,
                    'subcohorts': payload['subcohorts']})

@app.route('/api/promotions/<int:pid>/tuteurs/<int:year>/<int:sid>', methods=['PUT'])
def set_year_tuteurs(pid, year, sid):
    """Enregistre les tuteurs d'un étudiant pour une année d'étude. Tous les
    champs vides → suppression de la ligne de l'année (un ALT retombe alors sur
    la valeur héritée de l'année précédente, s'il y en a une)."""
    err = _require_admin()
    if err:
        return err
    if year not in (1, 2, 3):
        return error_response('Année invalide', 400)
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotion_students WHERE id = ? AND promotion_id = ?',
                       (sid, pid)).fetchone():
        return error_response('Étudiant introuvable', 404)
    data = request.get_json() or {}
    tu = (data.get('tuteur_univ') or '').strip()
    te = (data.get('tuteur_entreprise') or '').strip()
    en = (data.get('entreprise') or '').strip()
    if not (tu or te or en):
        pdb.execute('DELETE FROM student_tutors WHERE promotion_id = ? AND student_id = ? AND year = ?',
                    (pid, sid, year))
    else:
        pdb.execute('''INSERT INTO student_tutors
                           (promotion_id, student_id, year, tuteur_univ, tuteur_entreprise, entreprise)
                       VALUES (?, ?, ?, ?, ?, ?)
                       ON CONFLICT (student_id, year) DO UPDATE SET
                           tuteur_univ = excluded.tuteur_univ,
                           tuteur_entreprise = excluded.tuteur_entreprise,
                           entreprise = excluded.entreprise,
                           updated_at = CURRENT_TIMESTAMP''',
                    (pid, sid, year, tu, te, en))
    pdb.commit()
    return jsonify({'message': 'Tuteurs enregistrés'})

@app.route('/api/promotions/<int:pid>/effectif/<int:year>/students', methods=['POST'])
def add_year_student(pid, year):
    """Ajoute un étudiant à l'effectif d'une année : nouvel étudiant (créé avec
    entry_year=année) ou réintégration d'un étudiant existant (override 'add')."""
    err = _require_admin()
    if err:
        return err
    if year not in (1, 2, 3):
        return error_response('Année invalide', 400)
    db = get_promotions_db()
    if not db.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    data = request.get_json() or {}
    sid = data.get('student_id')
    if sid:                                  # réintégration d'un étudiant existant
        if not db.execute('SELECT 1 FROM promotion_students WHERE id=? AND promotion_id=?', (sid, pid)).fetchone():
            return error_response('Étudiant introuvable', 404)
        db.execute('''INSERT OR REPLACE INTO promotion_year_override(promotion_id, year, student_id, action)
                      VALUES(?,?,?,'add')''', (pid, year, sid))
    else:                                    # nouvel étudiant
        nom = (data.get('nom') or '').strip()
        prenom = (data.get('prenom') or '').strip()
        numero = (data.get('numero') or '').strip()
        if not (nom or prenom or numero):
            return error_response('Renseignez au moins un nom ou un numéro', 400)
        statut = (data.get('statut') or 'Actif').strip()
        if statut not in _STUDENT_STATUSES:
            statut = 'Actif'
        formation = (data.get('formation') or '').strip().upper()
        if formation not in _SUBCOHORTS:
            formation = 'FTP'
        db.execute('''INSERT INTO promotion_students(promotion_id, numero, nom, prenom, naissance,
                                                     statut, formation, entry_year)
                      VALUES(?,?,?,?,?,?,?,?)''',
                   (pid, numero, nom, prenom, (data.get('naissance') or '').strip(), statut, formation, year))
    db.commit()
    return jsonify(_year_effectif_payload(db, pid, year))

@app.route('/api/promotions/<int:pid>/effectif/<int:year>/students/<int:sid>', methods=['DELETE'])
def remove_year_student(pid, year, sid):
    """Retire un étudiant de l'effectif d'une année (ajustement 'remove'). N'efface ni
    l'étudiant ni ses notes ; il reste présent dans les autres années."""
    err = _require_admin()
    if err:
        return err
    if year not in (1, 2, 3):
        return error_response('Année invalide', 400)
    db = get_promotions_db()
    if not db.execute('SELECT 1 FROM promotion_students WHERE id=? AND promotion_id=?', (sid, pid)).fetchone():
        return error_response('Étudiant introuvable', 404)
    db.execute('''INSERT OR REPLACE INTO promotion_year_override(promotion_id, year, student_id, action)
                  VALUES(?,?,?,'remove')''', (pid, year, sid))
    db.commit()
    return jsonify(_year_effectif_payload(db, pid, year))

@app.route('/api/promotions/<int:pid>/effectif/<int:year>/students/<int:sid>/cohorte', methods=['POST'])
def set_year_cohorte(pid, year, sid):
    """Change la sous-cohorte (FTP/ALT) d'un étudiant À PARTIR de l'année `year` (les
    années antérieures sont conservées ; report aux années suivantes tant qu'un autre
    changement n'est pas posé)."""
    err = _require_admin()
    if err:
        return err
    if year not in (1, 2, 3):
        return error_response('Année invalide', 400)
    db = get_promotions_db()
    if not db.execute('SELECT 1 FROM promotion_students WHERE id=? AND promotion_id=?', (sid, pid)).fetchone():
        return error_response('Étudiant introuvable', 404)
    formation = ((request.get_json() or {}).get('formation') or '').strip().upper()
    if formation not in _SUBCOHORTS:
        return error_response('Sous-cohorte invalide', 400)
    # Formation héritée (année précédente / base) : si identique, pas d'override (nettoyage).
    inherited = _year_formation_map(db, pid, year - 1).get(sid, 'FTP')
    if formation == inherited:
        db.execute('DELETE FROM promotion_year_formation WHERE promotion_id=? AND year=? AND student_id=?',
                   (pid, year, sid))
    else:
        db.execute('''INSERT OR REPLACE INTO promotion_year_formation(promotion_id, year, student_id, formation)
                      VALUES(?,?,?,?)''', (pid, year, sid, formation))
    db.commit()
    return jsonify(_year_effectif_payload(db, pid, year))

@app.route('/api/promotions/<int:pid>/effectif/<int:year>/report', methods=['POST'])
def report_year_effectif(pid, year):
    """Réinitialise l'effectif de l'année sur le report automatique (jury) en supprimant
    les ajustements manuels de cette année."""
    err = _require_admin()
    if err:
        return err
    if year not in (1, 2, 3):
        return error_response('Année invalide', 400)
    db = get_promotions_db()
    if not db.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    db.execute('DELETE FROM promotion_year_override WHERE promotion_id=? AND year=?', (pid, year))
    db.commit()
    return jsonify(_year_effectif_payload(db, pid, year))

@app.route('/api/promotions/<int:pid>/effectif/<int:year>/import', methods=['POST'])
def import_year_students(pid, year):
    """Importe une liste d'étudiants dans l'effectif d'une année (entry_year=année).
    Dédoublonnage sur toute la promotion. Importer en année 1 alimente automatiquement
    les années 2 et 3 (report auto, aucune décision de jury encore)."""
    err = _require_admin()
    if err:
        return err
    if year not in (1, 2, 3):
        return error_response('Année invalide', 400)
    db = get_promotions_db()
    if not db.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    f = request.files.get('file')
    if not f or not f.filename:
        return error_response('Aucun fichier reçu', 400)
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in _ALLOWED_GRADE_EXT:
        return error_response('Type de fichier non autorisé (.xlsm/.xlsx)', 400)
    if request.content_length and request.content_length > _MAX_GRADE_FILE:
        return error_response('Fichier trop volumineux (max 15 Mo)', 400)
    formation = (request.form.get('formation') or '').strip().upper()
    if formation not in _SUBCOHORTS:
        formation = 'FTP'
    import tempfile
    tmp = os.path.join(tempfile.gettempdir(), secure_filename(f.filename) or ('import' + ext))
    f.save(tmp)
    try:
        students = _parse_student_list(tmp)
    except Exception as e:
        return error_response(f'Lecture impossible : {e}', 400)
    finally:
        try: os.remove(tmp)
        except OSError: pass
    if not students:
        return error_response('Aucun étudiant détecté dans le fichier', 400)
    existing = {_dedup_key(dict(r)) for r in db.execute(
        'SELECT numero, nom, prenom, naissance FROM promotion_students WHERE promotion_id=?', (pid,))}
    imported = skipped = 0
    for s in students:
        key = _dedup_key(s)
        if key in existing:
            skipped += 1
            continue
        existing.add(key)
        db.execute('''INSERT INTO promotion_students(promotion_id, numero, nom, prenom, naissance,
                                                     statut, formation, entry_year)
                      VALUES(?,?,?,?,?,'Actif',?,?)''',
                   (pid, s['numero'], s['nom'], s['prenom'], s['naissance'], formation, year))
        imported += 1
    db.commit()
    _audit('PROMO_IMPORT', ip=_client_ip(), user=session.get('user'),
           imported=imported, skipped=skipped, year=year)
    payload = _year_effectif_payload(db, pid, year)
    payload['import_report'] = {'imported': imported, 'skipped': skipped, 'total_fichier': len(students)}
    return jsonify(payload)

# ---- Redoublants (RED) : bascule vers la promo cible (année +1) ou Abandon ----

def _red_students(pdb, pid, comp=None):
    """{student_id(int): année} des redoublants (décision de jury RED) d'une promotion."""
    comp = comp or _jury_compute(pdb, pid)
    return {int(sid): y for (y, sid), dec in comp['decisions'].items() if dec == 'RED'}

def _red_target_promo(pdb, pid):
    """Promotion cible d'un redoublant = celle dont la 1re année est un an après
    (start_year + 1). Retourne (id ou None, nom affiché)."""
    promo = pdb.execute('SELECT start_year FROM promotions WHERE id=?', (pid,)).fetchone()
    if not promo:
        return None, None
    nxt = promo['start_year'] + 1
    row = pdb.execute('SELECT id, name FROM promotions WHERE start_year=?', (nxt,)).fetchone()
    return (row['id'] if row else None), (row['name'] if row else _promotion_name(nxt))

def _ensure_target_promo(db, pid):
    """Id de la promo cible d'un redoublant (start_year+1), CRÉÉE si absente (nom,
    3 années universitaires, même programme que la source). Cas des RED en BUT1 dont
    la cohorte suivante n'existe pas encore."""
    src = db.execute('SELECT start_year, programme_id FROM promotions WHERE id=?', (pid,)).fetchone()
    if not src:
        return None
    start = src['start_year'] + 1
    row = db.execute('SELECT id FROM promotions WHERE start_year=?', (start,)).fetchone()
    if row:
        return row['id']
    name = _promotion_name(start)
    cur = db.execute('''INSERT INTO promotions(name, formation, start_year, end_year, programme_id, updated_at)
                        VALUES(?,'MUT',?,?,?,CURRENT_TIMESTAMP)''',
                     (name, start, start + 3, src['programme_id']))
    new_id = cur.lastrowid
    db.commit()
    try:
        _ensure_years_for_promo(start)
    except Exception as e:
        app.logger.warning('Création des années de la promo cible %s échouée : %s', new_id, e)
    _audit('PROMO_CREATE', ip=_client_ip(), user=session.get('user'), promo=name, reason='RED_target')
    return new_id

def _red_payload(pdb, pid):
    promo = pdb.execute('SELECT * FROM promotions WHERE id=?', (pid,)).fetchone()
    if not promo:
        return None
    comp = _jury_compute(pdb, pid)
    _reconcile_red_transfers(pdb, pid, comp)     # nettoie les bascules RED caduques
    red = _red_students(pdb, pid, comp)
    target_id, target_name = _red_target_promo(pdb, pid)
    transfers = {r['student_id']: r['decision'] for r in pdb.execute(
        'SELECT student_id, decision FROM promotion_red_transfer WHERE promotion_id=?', (pid,))}
    info = {r['id']: r for r in pdb.execute(
        '''SELECT id, numero, nom, prenom, naissance, formation FROM promotion_students
           WHERE promotion_id=?''', (pid,))}
    rows = []
    for sid, year in red.items():
        s = info.get(sid)
        if not s:
            continue
        rows.append({'student_id': sid, 'numero': s['numero'], 'nom': s['nom'],
                     'prenom': s['prenom'], 'formation': s['formation'] or 'FTP',
                     'year': year, 'decision': transfers.get(sid)})
    rows.sort(key=lambda r: ((r['formation'] or ''), (r['nom'] or '').lower(), (r['prenom'] or '').lower()))
    return {'target_promo_id': target_id, 'target_promo_name': target_name,
            'target_exists': target_id is not None, 'students': rows}

@app.route('/api/promotions/<int:pid>/red', methods=['GET'])
def get_red_students(pid):
    err = _require_promo_read()
    if err:
        return err
    payload = _red_payload(get_promotions_db(), pid)
    if not payload:
        return error_response('Promotion introuvable', 404)
    return jsonify(payload)

def _revert_red_transfer(db, pid, sid, prev, restore_statut='RED'):
    """Annule une décision RED : supprime l'étudiant créé dans la promo cible et
    rétablit le statut source (`restore_statut`). Utilisé au changement de décision
    (reste 'RED') et à la réconciliation quand l'étudiant n'est plus RED ('Actif')."""
    if not prev:
        return
    if prev['target_student_id']:
        db.execute('DELETE FROM promotion_students WHERE id=?', (prev['target_student_id'],))
    db.execute('UPDATE promotion_students SET statut=?, abandon_semestre=NULL WHERE id=?',
               (restore_statut, sid))
    db.execute('DELETE FROM promotion_red_transfer WHERE promotion_id=? AND student_id=?', (pid, sid))

def _reconcile_red_transfers(db, pid, comp=None):
    """Annule les bascules RED devenues caduques : un étudiant avec un transfert RED
    mais dont la décision n'est PLUS RED (passée à ADM/ADMJ/AJAC/AJ) est retiré de la
    promo cible (étudiant créé supprimé) et repasse en statut 'Actif'. Idempotent."""
    red = _red_students(db, pid, comp)
    n = 0
    for t in db.execute('SELECT student_id, decision, target_student_id FROM promotion_red_transfer '
                        'WHERE promotion_id=?', (pid,)).fetchall():
        if t['student_id'] not in red:
            _revert_red_transfer(db, pid, t['student_id'], t, restore_statut='Actif')
            n += 1
    if n:
        db.commit()
    return n

@app.route('/api/promotions/<int:pid>/red/<int:sid>/decide', methods=['POST'])
def decide_red_student(pid, sid):
    """Décision pour un redoublant : 'FTP'/'ALT' (réinscription dans la promo cible à
    l'année redoublée) ou 'ABANDON' (statut Abandon). Vide/'NONE' efface la décision."""
    err = _require_admin()
    if err:
        return err
    db = get_promotions_db()
    s = db.execute('SELECT * FROM promotion_students WHERE id=? AND promotion_id=?', (sid, pid)).fetchone()
    if not s:
        return error_response('Étudiant introuvable', 404)
    red = _red_students(db, pid)
    if sid not in red:
        return error_response("Cet étudiant n'est pas redoublant (RED)", 400)
    year = red[sid]
    decision = ((request.get_json() or {}).get('decision') or '').strip().upper()
    prev = db.execute('SELECT decision, target_student_id FROM promotion_red_transfer '
                      'WHERE promotion_id=? AND student_id=?', (pid, sid)).fetchone()
    _revert_red_transfer(db, pid, sid, prev)
    if decision in ('', 'NONE'):
        db.commit()
        return jsonify(_red_payload(db, pid))
    if decision not in ('FTP', 'ALT', 'ABANDON'):
        return error_response('Décision invalide', 400)
    target_student_id = None
    if decision == 'ABANDON':
        db.execute("UPDATE promotion_students SET statut='Abandon', abandon_semestre=? WHERE id=?",
                   (f'S{year * 2}', sid))
    else:
        target_id = _ensure_target_promo(db, pid)   # crée la promo cible si absente
        if not target_id:
            return error_response('Promotion cible introuvable', 400)
        cur = db.execute('''INSERT INTO promotion_students(promotion_id, numero, nom, prenom, naissance,
                                                          statut, formation, entry_year)
                            VALUES(?,?,?,?,?,'Actif',?,?)''',
                         (target_id, s['numero'], s['nom'], s['prenom'], s['naissance'], decision, year))
        target_student_id = cur.lastrowid
        db.execute("UPDATE promotion_students SET statut='RED' WHERE id=?", (sid,))
    db.execute('''INSERT OR REPLACE INTO promotion_red_transfer(promotion_id, student_id, decision, target_student_id)
                  VALUES(?,?,?,?)''', (pid, sid, decision, target_student_id))
    db.commit()
    _audit('RED_DECIDE', ip=_client_ip(), user=session.get('user'),
           promo=pid, student=sid, decision=decision, year=year)
    return jsonify(_red_payload(db, pid))

# ---- Notes d'une promotion pour un semestre (matières/compétences = référence) ----

# Bonus / pénalité d'assiduité — barème IUT de Toulon (règlement intérieur des études en BUT).
def _penalty_points(hours):
    """Malus d'assiduité à retrancher de la moyenne de CHAQUE UE du semestre.
    Barème : 0 jusqu'à 8 h d'absences injustifiées (tolérance), −0,05 pt/h de la 9e à
    la 18e h, puis −0,1 pt/h à partir de la 19e h. Retourne un nombre de points ≥ 0."""
    try:
        h = float(hours)
    except (TypeError, ValueError):
        return 0.0
    if h <= 8:
        return 0.0
    p = 0.05 * (min(h, 18.0) - 8.0)
    if h > 18:
        p += 0.1 * (h - 18.0)
    return round(p, 2)

def _bonus_points(note):
    """Bonification sport/art : 0,05 pt par point de note (/20) au-dessus de 10,
    plafonnée à 0,5 pt. S'ajoute à la moyenne d'UE de l'année."""
    try:
        n = float(note)
    except (TypeError, ValueError):
        return 0.0
    return round(min(0.5, max(0.0, 0.05 * (n - 10.0))), 2)

def _code_by_kind(components, kind):
    for c in (components or []):
        if c.get('kind') == kind:
            return c.get('code')
    return None

def _semester_competence_averages(pdb, pid, semester, competences, components=None):
    """Moyennes de compétences {student_id: {ci: avg}} pour un semestre donné.
    Le malus d'assiduité (heures saisies sur la matière de type PEN) est retranché
    de la moyenne de chaque UE, conformément au règlement (IUT de Toulon)."""
    marks = {}
    for r in pdb.execute('''SELECT student_id, matiere_code, note FROM student_marks
                            WHERE promotion_id=? AND semester=? AND note IS NOT NULL''', (pid, semester)):
        marks[f"{r['student_id']}_{r['matiere_code']}"] = r['note']
    students = [r['id'] for r in pdb.execute(
        'SELECT id FROM promotion_students WHERE promotion_id=?', (pid,))]
    pen_code = _code_by_kind(components, 'PEN')
    averages = {}
    for sid in students:
        malus = _penalty_points(marks.get(f"{sid}_{pen_code}")) if pen_code else 0.0
        row = {}
        for ci, comp in enumerate(competences):
            num = den = 0.0
            for code, coeff in (comp.get('coeffs') or {}).items():
                coeff = coeff or 0
                key = f"{sid}_{code}"
                if coeff and key in marks:
                    num += coeff * marks[key]
                    den += coeff
            row[str(ci)] = round(max(0.0, num / den - malus), 2) if den > 0 else None
        averages[str(sid)] = row
    return averages

def _year_competence_averages(pdb, pid, s_odd, s_even, ref_all):
    """Moyenne annuelle par compétence = moyenne des 2 semestres (malus d'assiduité inclus),
    augmentée de la bonification sport/art (meilleure note des 2 semestres, max 0,5).
    Retourne (averages, noms)."""
    d_odd, d_even = (ref_all.get(s_odd) or {}), (ref_all.get(s_even) or {})
    comps_odd = d_odd.get('competences', [])
    names = [c.get('name', '') for c in comps_odd]
    avg_odd = _semester_competence_averages(pdb, pid, s_odd, comps_odd, d_odd.get('components', []))
    comps_even = d_even.get('competences', [])
    avg_even = _semester_competence_averages(pdb, pid, s_even, comps_even, d_even.get('components', []))
    # Note de bonus (sport/art) : on retient la meilleure note saisie sur les 2 semestres
    bonus_note = {}
    for sem, comps_d in ((s_odd, d_odd), (s_even, d_even)):
        bcode = _code_by_kind(comps_d.get('components', []), 'BONUS')
        if not bcode:
            continue
        for r in pdb.execute('''SELECT student_id, note FROM student_marks
                                WHERE promotion_id=? AND semester=? AND matiere_code=? AND note IS NOT NULL''',
                             (pid, sem, bcode)):
            sid = str(r['student_id'])
            bonus_note[sid] = max(bonus_note.get(sid, 0.0), r['note'])
    out = {}
    for r in pdb.execute('SELECT id FROM promotion_students WHERE promotion_id=?', (pid,)):
        sid = str(r['id'])
        bonus = _bonus_points(bonus_note.get(sid)) if sid in bonus_note else 0.0
        row = {}
        for ci in range(len(names)):
            vals = [v for v in ((avg_odd.get(sid) or {}).get(str(ci)),
                                (avg_even.get(sid) or {}).get(str(ci))) if v is not None]
            row[str(ci)] = round(min(20.0, sum(vals) / len(vals) + bonus), 2) if vals else None
        out[sid] = row
    return out, names

# ---- Coefficients propres à une promotion (copie indépendante de la référence) ----

def _store_promo_coeffs(pdb, pid, data):
    pdb.execute('''INSERT INTO promotion_coefficients(promotion_id, data, updated_at)
                   VALUES(?,?,CURRENT_TIMESTAMP)
                   ON CONFLICT(promotion_id) DO UPDATE
                     SET data=excluded.data, updated_at=CURRENT_TIMESTAMP''',
                (pid, json.dumps(data, ensure_ascii=False)))

def _ensure_bonus_pen(coeffs):
    """Garantit qu'un composant BONUS et un composant PEN existent pour chaque semestre
    (ajout non destructif depuis les valeurs par défaut). Retourne True si modifié."""
    defaults = _ref_coeffs_defaults()
    changed = False
    for sem, d in (coeffs or {}).items():
        comps = d.get('components')
        if comps is None:
            continue
        have = {c.get('kind') for c in comps}
        for kind in ('BONUS', 'PEN'):
            if kind not in have:
                src = next((c for c in defaults.get(sem, {}).get('components', [])
                            if c.get('kind') == kind), None)
                if src:
                    comps.append(dict(src))
                    changed = True
    return changed

# Types de matières retenus : SAE, RES, PEN, BONUS. Stage et Portfolio = SAÉ.
_KIND_REMAP = {'STAGE': 'SAE', 'PORT': 'SAE'}

def _normalize_kinds(coeffs):
    """Convertit les types Stage/Portfolio en SAÉ. Retourne True si modifié."""
    changed = False
    for d in (coeffs or {}).values():
        for c in d.get('components', []):
            if c.get('kind') in _KIND_REMAP:
                c['kind'] = _KIND_REMAP[c['kind']]
                changed = True
    return changed

_VOLUME_KEYS = ('cm', 'td', 'tp', 'pt')   # CM/TD/TP (ressources) + TD/TP/PT (SAE)

def _ensure_volumes(coeffs):
    """Garantit un objet 'volumes' {cm, td, tp, pt} (heures, défaut None) sur chaque
    composant (matière). Migre l'ancien champ scalaire 'hours' (→ td) si présent.
    Mute et retourne le dict passé."""
    for d in (coeffs or {}).values():
        for c in d.get('components', []):
            vol = c.get('volumes')
            if not isinstance(vol, dict):
                vol = {}
            # Migration éventuelle de l'ancien champ 'hours' scalaire
            if c.get('hours') not in (None, '') and not any(vol.get(k) for k in _VOLUME_KEYS):
                try:
                    vol['td'] = float(c['hours'])
                except (TypeError, ValueError):
                    pass
            for k in _VOLUME_KEYS:
                vol.setdefault(k, None)
            c['volumes'] = {k: vol.get(k) for k in _VOLUME_KEYS}
            c.pop('hours', None)
    return coeffs

# ---- Programmes (Programme National réutilisable) ----

def _store_programme_data(grdb, prog_id, data):
    grdb.execute('''INSERT INTO programme_data(programme_id, data, updated_at)
                    VALUES(?,?,CURRENT_TIMESTAMP)
                    ON CONFLICT(programme_id) DO UPDATE
                      SET data=excluded.data, updated_at=CURRENT_TIMESTAMP''',
                 (prog_id, json.dumps(data, ensure_ascii=False)))

def _programme_data(grdb, prog_id, seed=True):
    """Contenu d'un programme (matières/compétences/coeffs/heures par semestre).
    Stocké si présent, sinon initialisé depuis la référence puis persisté.
    Complète BONUS/PEN, normalise les types et garantit le champ heures."""
    row = grdb.execute('SELECT data FROM programme_data WHERE programme_id=?',
                       (prog_id,)).fetchone()
    if row and row['data']:
        try:
            data = json.loads(row['data'])
            changed = _ensure_bonus_pen(data)
            changed = _normalize_kinds(data) or changed
            _ensure_volumes(data)
            if changed:
                _store_programme_data(grdb, prog_id, data)
                grdb.commit()
            return data
        except ValueError:
            pass
    ref = _ensure_volumes(_ref_coeffs_defaults() or {})
    _ensure_bonus_pen(ref)
    _normalize_kinds(ref)
    if seed:
        _store_programme_data(grdb, prog_id, ref)
        grdb.commit()
    return ref

def _promo_programme_id(pdb, pid):
    """programme_id affecté à une promotion (None si aucun / colonne absente)."""
    try:
        row = pdb.execute('SELECT programme_id FROM promotions WHERE id=?', (pid,)).fetchone()
    except sqlite3.OperationalError:
        return None
    return row['programme_id'] if row else None

def _promo_coeffs(pdb, pid, seed=True):
    """Coefficients d'une promotion. Si la promo est liée à un PROGRAMME, c'est lui qui
    fait foi (source unique). Sinon (legacy) : copie stockée dans promotion_coefficients,
    sinon initialisée depuis la référence globale. BONUS/PEN complétés, types Stage/
    Portfolio normalisés en SAÉ, de façon non destructive."""
    prog_id = _promo_programme_id(pdb, pid)
    if prog_id:
        try:
            grdb = get_programmes_db()
            if grdb.execute('SELECT 1 FROM programmes WHERE id=?', (prog_id,)).fetchone():
                return _programme_data(grdb, prog_id)
        except sqlite3.Error:
            pass
    row = pdb.execute('SELECT data FROM promotion_coefficients WHERE promotion_id=?',
                      (pid,)).fetchone()
    if row and row['data']:
        try:
            coeffs = json.loads(row['data'])
            changed = _ensure_bonus_pen(coeffs)
            changed = _normalize_kinds(coeffs) or changed
            if changed:
                _store_promo_coeffs(pdb, pid, coeffs)
                pdb.commit()
            return coeffs
        except ValueError:
            pass
    ref = _load_ref_coeffs(get_db()) or {}
    _ensure_bonus_pen(ref)
    _normalize_kinds(ref)
    if seed:
        _store_promo_coeffs(pdb, pid, ref)
        pdb.commit()
    return ref

@app.route('/api/promotions/<int:pid>/coefficients', methods=['GET'])
def get_promotion_coefficients(pid):
    err = _require_admin()
    if err:
        return err
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    return jsonify(_promo_coeffs(pdb, pid))

@app.route('/api/promotions/<int:pid>/coefficients', methods=['PUT'])
def save_promotion_coefficients(pid):
    err = _require_admin()
    if err:
        return err
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    data = request.get_json()
    if not isinstance(data, dict):
        return error_response('Format invalide', 400)
    # Si la promo est liée à un programme, l'édition porte sur le programme (source unique).
    prog_id = _promo_programme_id(pdb, pid)
    if prog_id:
        grdb = get_programmes_db()
        if grdb.execute('SELECT 1 FROM programmes WHERE id=?', (prog_id,)).fetchone():
            _ensure_volumes(data)
            _store_programme_data(grdb, prog_id, data)
            grdb.commit()
            return jsonify(_promo_coeffs(pdb, pid))
    _store_promo_coeffs(pdb, pid, data)
    pdb.commit()
    return jsonify(_promo_coeffs(pdb, pid))

@app.route('/api/promotions/<int:pid>/coefficients/copy', methods=['POST'])
def copy_promotion_coefficients(pid):
    """Copie les semestres choisis (matières + coefficients) de la promotion source
    vers une ou plusieurs promotions cibles : les semestres concernés y sont remplacés."""
    err = _require_admin()
    if err:
        return err
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion source introuvable', 404)
    data = request.get_json() or {}
    semesters = [s for s in (data.get('semesters') or []) if s in _PROMO_SEMESTERS]
    targets = []
    for t in (data.get('targets') or []):
        try:
            tid = int(t)
        except (TypeError, ValueError):
            continue
        if tid != pid and tid not in targets:
            targets.append(tid)
    if not semesters or not targets:
        return error_response('Sélectionnez au moins un semestre et une promotion cible', 400)
    src = _promo_coeffs(pdb, pid)
    copied = done = 0
    for tid in targets:
        if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (tid,)).fetchone():
            continue
        tgt = _promo_coeffs(pdb, tid)
        for sem in semesters:
            if sem in src:
                tgt[sem] = json.loads(json.dumps(src[sem], ensure_ascii=False))
                copied += 1
        _store_promo_coeffs(pdb, tid, tgt)
        done += 1
    pdb.commit()
    return jsonify({'copied': copied, 'targets': done, 'semesters': semesters})

@app.route('/api/promotions/<int:pid>/coefficients/reset', methods=['POST'])
def reset_promotion_coefficients(pid):
    """Réinitialise les coefficients de la promotion aux valeurs d'origine du fichier."""
    err = _require_admin()
    if err:
        return err
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    ref = _ref_coeffs_defaults()
    _store_promo_coeffs(pdb, pid, ref)
    pdb.commit()
    return jsonify(ref)

@app.route('/api/promotions/<int:pid>/coefficients/save-as-default', methods=['POST'])
def save_promo_coeffs_as_default(pid):
    """Écrase le fichier d'origine (data/coefficients.json) avec les coefficients de la
    promotion : ils deviennent le modèle par défaut des nouvelles promotions et la cible
    du bouton « Réinitialiser ». L'ancien fichier est sauvegardé (horodaté)."""
    err = _require_admin()
    if err:
        return err
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    coeffs = _promo_coeffs(pdb, pid)
    backup = None
    try:
        os.makedirs(os.path.dirname(_REF_COEFF_FILE), exist_ok=True)
        if os.path.isfile(_REF_COEFF_FILE):
            ts = datetime.now().strftime('%Y%m%d-%H%M%S')
            backup = os.path.join(os.path.dirname(_REF_COEFF_FILE),
                                  f'coefficients.backup-{ts}.json')
            shutil.copy2(_REF_COEFF_FILE, backup)
        with open(_REF_COEFF_FILE, 'w', encoding='utf-8') as f:
            json.dump(coeffs, f, ensure_ascii=False, indent=2)
            f.write('\n')
    except OSError as e:
        return error_response(f'Écriture du fichier d\'origine impossible : {e}', 500)
    # On retire l'éventuelle surcharge en base pour que le fichier fasse foi
    db = get_db()
    db.execute("DELETE FROM app_settings WHERE key='grade_coeff_reference'")
    db.commit()
    _audit('COEFF_DEFAULT_OVERWRITE', ip=_client_ip(), user=session.get('user'), promo=pid)
    return jsonify({'ok': True, 'backup': os.path.basename(backup) if backup else None})

# ======================= PROGRAMMES (Programme National) =======================

def _programme_row(grdb, prog_id):
    return grdb.execute('SELECT * FROM programmes WHERE id=?', (prog_id,)).fetchone()

def _programme_promo_count(prog_id):
    """Nombre de promotions liées à ce programme (base promotions séparée)."""
    try:
        r = get_promotions_db().execute(
            'SELECT COUNT(*) AS n FROM promotions WHERE programme_id=?', (prog_id,)).fetchone()
        return r['n'] if r else 0
    except sqlite3.Error:
        return 0

@app.route('/api/programmes', methods=['GET'])
def list_programmes():
    # Lecture ouverte aux enseignants (consultation du programme) ; l'écriture
    # reste réservée à l'admin (before_request + _require_admin sur PUT/POST).
    grdb = get_programmes_db()
    rows = grdb.execute('SELECT * FROM programmes ORDER BY name').fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d['promotions'] = _programme_promo_count(r['id'])
        out.append(d)
    return jsonify(out)

@app.route('/api/programmes', methods=['POST'])
def create_programme():
    err = _require_admin()
    if err:
        return err
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return error_response('Nom du programme requis', 400)
    grdb = get_programmes_db()
    if grdb.execute('SELECT 1 FROM programmes WHERE name=?', (name,)).fetchone():
        return error_response(f'Le programme « {name} » existe déjà', 409)
    label = (data.get('label') or '').strip() or None
    cur = grdb.execute('INSERT INTO programmes(name, label, updated_at) VALUES(?,?,CURRENT_TIMESTAMP)',
                       (name, label))
    new_id = cur.lastrowid
    # Contenu initial : duplication d'un programme existant (from_id) sinon référence.
    base = None
    from_id = data.get('from_id')
    if from_id:
        try:
            src = _programme_data(grdb, int(from_id), seed=False)
            base = json.loads(json.dumps(src, ensure_ascii=False))
        except (TypeError, ValueError):
            base = None
    if base is None:
        base = _ensure_volumes(_ref_coeffs_defaults() or {})
    _store_programme_data(grdb, new_id, base)
    grdb.commit()
    return jsonify(dict(_programme_row(grdb, new_id)))

@app.route('/api/programmes/<int:prog_id>', methods=['PUT'])
def update_programme(prog_id):
    err = _require_admin()
    if err:
        return err
    grdb = get_programmes_db()
    if not _programme_row(grdb, prog_id):
        return error_response('Programme introuvable', 404)
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return error_response('Nom du programme requis', 400)
    if grdb.execute('SELECT 1 FROM programmes WHERE name=? AND id<>?', (name, prog_id)).fetchone():
        return error_response(f'Le programme « {name} » existe déjà', 409)
    label = (data.get('label') or '').strip() or None
    grdb.execute('UPDATE programmes SET name=?, label=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
                 (name, label, prog_id))
    grdb.commit()
    return jsonify(dict(_programme_row(grdb, prog_id)))

@app.route('/api/programmes/export', methods=['GET'])
def export_programmes():
    """Exporte tous les programmes (nom + données complètes, incl. heures) en JSON
    téléchargeable. Sert à transférer entre installations (local → serveur)."""
    err = _require_admin()
    if err:
        return err
    grdb = get_programmes_db()
    out = []
    for r in grdb.execute('SELECT id, name, label FROM programmes ORDER BY name').fetchall():
        out.append({
            'name': r['name'],
            'label': r['label'],
            'data': _programme_data(grdb, r['id'], seed=False),
        })
    payload = json.dumps({'programmes': out}, ensure_ascii=False, indent=2)
    resp = app.response_class(payload, mimetype='application/json')
    resp.headers['Content-Disposition'] = 'attachment; filename=programmes_export.json'
    return resp

@app.route('/api/programmes/import', methods=['POST'])
def import_programmes():
    """Importe des programmes exportés (upsert par NOM) : crée le programme s'il
    n'existe pas, puis remplace ses données (matières, coefficients, heures).
    Ne touche pas aux promotions (liens conservés via l'id existant)."""
    err = _require_admin()
    if err:
        return err
    payload = request.get_json(silent=True) or {}
    progs = payload.get('programmes')
    if not isinstance(progs, list) or not progs:
        return error_response('JSON invalide : clé « programmes » (liste) attendue', 400)
    grdb = get_programmes_db()
    report = {'created': 0, 'updated': 0}
    for p in progs:
        name = (p.get('name') or '').strip()
        data = p.get('data')
        if not name or not isinstance(data, dict):
            continue
        row = grdb.execute('SELECT id FROM programmes WHERE name=?', (name,)).fetchone()
        if row:
            pid = row['id']
            report['updated'] += 1
        else:
            label = (p.get('label') or '').strip() or None
            cur = grdb.execute('INSERT INTO programmes(name, label, updated_at) VALUES(?,?,CURRENT_TIMESTAMP)',
                               (name, label))
            pid = cur.lastrowid
            report['created'] += 1
        _store_programme_data(grdb, pid, _ensure_volumes(data))
        grdb.execute('UPDATE programmes SET updated_at=CURRENT_TIMESTAMP WHERE id=?', (pid,))
    grdb.commit()
    _audit('PROGRAMMES_IMPORT', ip=_client_ip(), user=session.get('user'), **report)
    return jsonify({'ok': True, 'report': report})

@app.route('/api/programmes/<int:prog_id>', methods=['DELETE'])
def delete_programme(prog_id):
    err = _require_admin()
    if err:
        return err
    grdb = get_programmes_db()
    if not _programme_row(grdb, prog_id):
        return error_response('Programme introuvable', 404)
    n = _programme_promo_count(prog_id)
    if n:
        return error_response(f'Programme utilisé par {n} promotion(s) — détachez-les d\'abord', 409)
    grdb.execute('DELETE FROM programmes WHERE id=?', (prog_id,))
    grdb.commit()
    return jsonify({'deleted': True})

@app.route('/api/programmes/<int:prog_id>/data', methods=['GET'])
def get_programme_data_route(prog_id):
    # Lecture ouverte aux enseignants (consultation) ; écriture admin uniquement.
    grdb = get_programmes_db()
    if not _programme_row(grdb, prog_id):
        return error_response('Programme introuvable', 404)
    return jsonify(_programme_data(grdb, prog_id))

@app.route('/api/programmes/<int:prog_id>/data', methods=['PUT'])
def save_programme_data_route(prog_id):
    err = _require_admin()
    if err:
        return err
    grdb = get_programmes_db()
    if not _programme_row(grdb, prog_id):
        return error_response('Programme introuvable', 404)
    data = request.get_json()
    if not isinstance(data, dict):
        return error_response('Format invalide', 400)
    _ensure_volumes(data)
    _store_programme_data(grdb, prog_id, data)
    grdb.commit()
    return jsonify(_programme_data(grdb, prog_id))

@app.route('/api/programmes/<int:prog_id>/data/reset', methods=['POST'])
def reset_programme_data_route(prog_id):
    err = _require_admin()
    if err:
        return err
    grdb = get_programmes_db()
    if not _programme_row(grdb, prog_id):
        return error_response('Programme introuvable', 404)
    ref = _ensure_volumes(_ref_coeffs_defaults() or {})
    _store_programme_data(grdb, prog_id, ref)
    grdb.commit()
    return jsonify(_programme_data(grdb, prog_id))

@app.route('/api/promotions/<int:pid>/programme', methods=['PUT'])
def assign_promotion_programme(pid):
    err = _require_admin()
    if err:
        return err
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    data = request.get_json() or {}
    raw = data.get('programme_id')
    prog_id = None
    if raw not in (None, '', 0, '0'):
        try:
            prog_id = int(raw)
        except (TypeError, ValueError):
            return error_response('programme_id invalide', 400)
        if not get_programmes_db().execute('SELECT 1 FROM programmes WHERE id=?', (prog_id,)).fetchone():
            return error_response('Programme introuvable', 404)
    pdb.execute('UPDATE promotions SET programme_id=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
                (prog_id, pid))
    pdb.commit()
    return jsonify({'ok': True, 'programme_id': prog_id})

_VALID_MIN_COMP = 3   # une année est validée si ≥ 3 compétences validées (règlement IUT Toulon)

def _year_validation(pdb, pid, ref_all, target_year):
    """Statut de validation de l'année `target_year` par étudiant, selon la règle :
    une année est validée si au moins 3 compétences sont validées (moyenne annuelle
    du regroupement ≥ 10). Cascade : valider une compétence à un niveau supérieur
    valide automatiquement les niveaux inférieurs (compétence repérée par son nom).
    Retourne {sid: {'validated', 'total', 'status'}} avec status ok/ko/partial/none."""
    # Moyennes annuelles par année et par compétence (nom) : {year: {nom: {sid: avg}}}
    by_year = {}
    for y in (1, 2, 3):
        avgs, names = _year_competence_averages(pdb, pid, f'S{2*y-1}', f'S{2*y}', ref_all)
        m = {}
        for ci, nm in enumerate(names):
            if nm:
                m[nm] = {sid: row.get(str(ci)) for sid, row in avgs.items()}
        by_year[y] = m
    target_names = list(by_year.get(target_year, {}).keys())
    total = len(target_names)
    out = {}
    for r in pdb.execute('SELECT id FROM promotion_students WHERE promotion_id=?', (pid,)):
        sid = str(r['id'])
        validated = evaluated = 0
        for nm in target_names:
            # Meilleure moyenne annuelle de cette compétence au niveau cible ou supérieur
            best = None
            for y in range(target_year, 4):
                avg = (by_year.get(y, {}).get(nm, {}) or {}).get(sid)
                if avg is not None:
                    best = avg if best is None else max(best, avg)
            if best is not None:
                evaluated += 1
                if best >= 10:
                    validated += 1
        if evaluated == 0:
            status = 'none'
        elif validated >= _VALID_MIN_COMP:
            status = 'ok'
        elif validated + (total - evaluated) < _VALID_MIN_COMP:
            status = 'ko'           # impossible d'atteindre 3 même en validant le reste
        else:
            status = 'partial'      # des compétences restent à évaluer
        out[sid] = {'validated': validated, 'total': total, 'status': status}
    return out

# ---- Jury : codes UE (ADM/AJ/ADMJ/CMP) et décision de passage (ADM/AJAC/AJ) ----
# Nomenclature BUT : UE{semestre}{compétence} (ex UE32 = UE2 au S3), UE{x}N{y} =
# moyenne annuelle de l'UE x au niveau y, GIM{y} = moyenne générale de l'année y.

def _jury_ue_numbers(ref_all):
    """Numérotation canonique des UE : ordre d'apparition des compétences S1→S6
    (Maintenir=1, Améliorer=2, Installer=3, Manager=4, Sécuriser=5).
    Retourne {nom: numéro}."""
    nums = {}
    for s in _PROMO_SEMESTERS:
        for c in (ref_all.get(s) or {}).get('competences', []):
            nm = (c.get('name') or '').strip()
            if nm and nm not in nums:
                nums[nm] = len(nums) + 1
    return nums

def _jury_compute(pdb, pid):
    """Cœur du calcul jury pour TOUTES les années (1..3) d'une promotion. Retourne les
    structures réutilisables : moyennes semestrielles/annuelles par UE, codes UE
    (ADM/ADMJ/CMP/AJ) et décision d'année par étudiant.
    Décision d'année :
      • ADM   — toutes les UE validées sur les moyennes ;
      • ADMJ  — toutes les UE validées mais grâce à ≥ 1 ADMJ posé par le jury sur une UE ;
      • AJAC  — ≥ 3 UE validées (années 1 et 2 ; les UE terminales doivent être validées) ;
      • AJ    — sinon ;
      • RED   — override manuel du jury sur un candidat AJ (redoublant), stocké en ue_num=0."""
    ref_all = _promo_coeffs(pdb, pid) or {}
    nums = _jury_ue_numbers(ref_all)
    sem_avgs, year_avgs, ue_by_year = {}, {}, {}
    for y in (1, 2, 3):
        so, se = f'S{2 * y - 1}', f'S{2 * y}'
        for s in (so, se):
            d = ref_all.get(s) or {}
            comps = d.get('competences', [])
            cnums = [nums.get((c.get('name') or '').strip()) for c in comps]
            raw = _semester_competence_averages(pdb, pid, s, comps, d.get('components', []))
            sem_avgs[s] = {sid: {cnums[int(ci)]: v for ci, v in row.items() if cnums[int(ci)]}
                           for sid, row in raw.items()}
        yavg, ynames = _year_competence_averages(pdb, pid, so, se, ref_all)
        ynums = [nums.get((n or '').strip()) for n in ynames]
        ue_by_year[y] = [n for n in ynums if n]
        year_avgs[y] = {sid: {ynums[int(ci)]: v for ci, v in row.items() if ynums[int(ci)]}
                        for sid, row in yavg.items()}
    ue_overrides, red_overrides = {}, set()
    for r in pdb.execute('SELECT year, student_id, ue_num, decision FROM jury_decisions '
                         'WHERE promotion_id=?', (pid,)):
        if r['ue_num'] == 0:                       # décision d'année (RED)
            if r['decision'] == 'RED':
                red_overrides.add((r['year'], str(r['student_id'])))
        else:
            ue_overrides[(r['year'], str(r['student_id']), r['ue_num'])] = r['decision']

    def validated_above(sid, ue, year):
        # UE validée (ADM ou ADMJ) à un niveau strictement supérieur → CMP en dessous
        for y2 in range(year + 1, 4):
            a2 = (year_avgs[y2].get(sid) or {}).get(ue)
            if (a2 is not None and a2 >= 10) or ue_overrides.get((y2, sid, ue)) == 'ADMJ':
                return True
        return False

    all_sids = [str(r['id']) for r in pdb.execute(
        'SELECT id FROM promotion_students WHERE promotion_id=?', (pid,))]
    ue_codes, decisions = {}, {}
    for y in (1, 2, 3):
        ues = ue_by_year.get(y) or []
        # UE « terminales » : absentes de l'année suivante (UE1/UE5 en 2e année). Elles ne
        # pourront plus être rattrapées → doivent être validées pour autoriser un AJAC.
        terminal = [u for u in ues if y < 3 and u not in (ue_by_year.get(y + 1) or [])]
        for sid in all_sids:
            codes = {}
            for ue in ues:
                annual = (year_avgs[y].get(sid) or {}).get(ue)
                admj = ue_overrides.get((y, sid, ue)) == 'ADMJ'
                code = None
                if annual is not None:
                    if annual >= 10:
                        code = 'ADM'
                    elif admj:
                        code = 'ADMJ'
                    elif validated_above(sid, ue, y):
                        code = 'CMP'
                    else:
                        code = 'AJ'
                codes[ue] = code
            ue_codes[(y, sid)] = codes
            if all(c is None for c in codes.values()):
                decisions[(y, sid)] = None
                continue
            valid = {u for u, c in codes.items() if c in ('ADM', 'ADMJ', 'CMP')}
            if len(valid) == len(ues):
                base = 'ADMJ' if any(codes[u] == 'ADMJ' for u in ues) else 'ADM'
            elif y < 3 and len(valid) >= _VALID_MIN_COMP and all(u in valid for u in terminal):
                base = 'AJAC'
            else:
                base = 'AJ'
            if base == 'AJ' and (y, sid) in red_overrides:
                base = 'RED'
            decisions[(y, sid)] = base
    return {'nums': nums, 'sem_avgs': sem_avgs, 'year_avgs': year_avgs,
            'ue_by_year': ue_by_year, 'ue_codes': ue_codes, 'decisions': decisions}

def _jury_failed_before(comp, year):
    """Étudiants (sid str) ajournés — décision AJ ou RED — à une année STRICTEMENT
    antérieure à `year`. Ils sont retirés des grilles Jury et Notes des années/semestres
    suivants (ils ne progressent pas). AJAC (autorisé à continuer) n'est PAS concerné."""
    return {sid for (y, sid), dec in comp['decisions'].items()
            if y < year and dec in ('AJ', 'RED')}

_STUDENT_LEFT_STATUSES = ('Abandon',)

def _year_formation_map(pdb, pid, year):
    """Sous-cohorte (FTP/ALT) de chaque étudiant POUR une année d'étude donnée : le
    dernier changement de cohorte d'une année <= `year`, sinon la formation de base.
    Un changement en année N vaut donc pour N et les années suivantes (report auto),
    sans modifier les années antérieures. Retourne {student_id: 'FTP'|'ALT'}."""
    base = {r['id']: (r['formation'] or 'FTP')
            for r in pdb.execute('SELECT id, formation FROM promotion_students WHERE promotion_id=?', (pid,))}
    over = {}
    for r in pdb.execute('''SELECT year, student_id, formation FROM promotion_year_formation
                            WHERE promotion_id=? AND year<=? ORDER BY year''', (pid, year)):
        if r['formation'] in _SUBCOHORTS:
            over[r['student_id']] = r['formation']   # ORDER BY year → le plus récent l'emporte
    return {sid: over.get(sid, base.get(sid, 'FTP')) for sid in base}

def _sem_year(sem_code):
    """Année d'étude (1..3) d'un code semestre 'S1'..'S6' ; 1 par défaut."""
    try:
        return (int(str(sem_code)[1:]) + 1) // 2
    except (ValueError, IndexError):
        return 1

def _year_rosters(pdb, pid, comp=None):
    """Effectif (ids d'étudiants) de chaque année 1..3 d'une promotion — « report auto ».
    Année 1 = étudiants entrés en 1re année. Chaque année suivante reprend les étudiants
    de l'année précédente qui PASSENT (décision jury ≠ AJ/RED) et ne sont pas partis
    (Abandon/Sortie), plus les redoublants entrant directement cette année-là (entry_year).
    Des ajustements manuels (promotion_year_override) retirent ('remove') ou réintègrent
    ('add') un étudiant sur une année. Retourne {1:set, 2:set, 3:set}."""
    meta = {}
    for r in pdb.execute('''SELECT id, statut, abandon_semestre, entry_year
                            FROM promotion_students WHERE promotion_id=?''', (pid,)):
        meta[r['id']] = {
            'statut': r['statut'],
            # Année d'abandon connue seulement si un semestre est renseigné ; sinon None
            # (l'étudiant reste visible partout tant que le semestre n'est pas précisé).
            'abandon_year': _sem_year(r['abandon_semestre']) if r['abandon_semestre'] else None,
            'entry': r['entry_year'] or 1}
    removes = {1: set(), 2: set(), 3: set()}
    adds = {1: set(), 2: set(), 3: set()}
    for r in pdb.execute('SELECT year, student_id, action FROM promotion_year_override '
                         'WHERE promotion_id=?', (pid,)):
        if r['year'] in removes:
            (removes if r['action'] == 'remove' else adds)[r['year']].add(r['student_id'])
    comp = comp or _jury_compute(pdb, pid)
    failed_at = {1: set(), 2: set(), 3: set()}
    for (y, sid), dec in comp['decisions'].items():
        if y in failed_at and dec in ('AJ', 'RED'):
            failed_at[y].add(int(sid))

    def left_before(sid, year):
        m = meta.get(sid)
        return bool(m and m['statut'] in _STUDENT_LEFT_STATUSES
                    and m['abandon_year'] is not None and m['abandon_year'] < year)

    rosters = {}
    for y in (1, 2, 3):
        entrants = {sid for sid, m in meta.items() if m['entry'] == y}
        if y == 1:
            base = entrants
        else:
            survivors = {sid for sid in rosters[y - 1]
                         if sid not in failed_at[y - 1] and not left_before(sid, y)}
            base = survivors | entrants
        rosters[y] = (base - removes[y]) | (adds[y] & set(meta.keys()))
    return rosters

def _jury_payload(pdb, pid, year, formation=None):
    """Tableau de jury de l'année `year` (1..3) : par étudiant et par UE, les moyennes
    semestrielles (UExy), la moyenne annuelle (UExNy), le code UE et la décision d'année.
    Les étudiants ajournés (AJ/RED) à une année antérieure sont retirés de la liste."""
    comp = _jury_compute(pdb, pid)
    nums, sem_avgs = comp['nums'], comp['sem_avgs']
    year_avgs, ue_by_year = comp['year_avgs'], comp['ue_by_year']
    ue_codes, decisions = comp['ue_codes'], comp['decisions']
    names_by_num = {v: k for k, v in nums.items()}
    students = [dict(r) for r in pdb.execute(
        '''SELECT id, numero, nom, prenom, statut, formation FROM promotion_students
           WHERE promotion_id=? ORDER BY nom COLLATE NOCASE, prenom COLLATE NOCASE''', (pid,))]
    roster = _year_rosters(pdb, pid, comp).get(year, set())
    fm = _year_formation_map(pdb, pid, year)          # sous-cohorte propre à cette année
    kept = []
    for s in students:
        if s['id'] not in roster:
            continue
        s['formation'] = fm.get(s['id'], s.get('formation') or 'FTP')
        if formation in _SUBCOHORTS and s['formation'] != formation:
            continue
        kept.append(s)
    students = kept
    s_odd, s_even = f'S{2 * year - 1}', f'S{2 * year}'
    ues = ue_by_year.get(year) or []
    terminal = [u for u in ues if year < 3 and u not in (ue_by_year.get(year + 1) or [])]
    rows = {}
    for st in students:
        sid = str(st['id'])
        codes = ue_codes.get((year, sid), {})
        per_ue = {}
        for ue in ues:
            per_ue[str(ue)] = {'odd': (sem_avgs.get(s_odd, {}).get(sid) or {}).get(ue),
                               'even': (sem_avgs.get(s_even, {}).get(sid) or {}).get(ue),
                               'annual': (year_avgs[year].get(sid) or {}).get(ue),
                               'code': codes.get(ue), 'admj': codes.get(ue) == 'ADMJ'}
        vals = [v['annual'] for v in per_ue.values() if v['annual'] is not None]
        gim = round(sum(vals) / len(vals), 2) if vals else None
        dec = decisions.get((year, sid))
        # red_eligible : candidat ajourné → le jury peut choisir AJ ou RED (menu déroulant)
        rows[sid] = {'ues': per_ue, 'gim': gim, 'decision': dec,
                     'red_eligible': dec in ('AJ', 'RED')}
    return {'year': year, 'semesters': [s_odd, s_even],
            'ues': [{'num': u, 'name': names_by_num.get(u, ''), 'terminal': u in terminal}
                    for u in ues],
            'students': students, 'rows': rows}

@app.route('/api/promotions/<int:pid>/jury/<int:year>', methods=['GET'])
def get_promotion_jury(pid, year):
    err = _require_promo_read()
    if err:
        return err
    if year not in (1, 2, 3):
        return error_response('Année invalide', 400)
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    formation = (request.args.get('formation') or '').strip().upper() or None
    return jsonify(_jury_payload(pdb, pid, year, formation))

@app.route('/api/promotions/<int:pid>/jury/<int:year>', methods=['PUT'])
def save_promotion_jury(pid, year):
    """Enregistre les décisions de jury :
      • {decisions: [{student_id, ue, admj}]}    — ADMJ posé/retiré sur une UE ;
      • {year_decisions: [{student_id, red}]}    — RED posé/retiré sur un candidat AJ
        (décision d'année, stockée en ue_num=0)."""
    err = _require_admin()
    if err:
        return err
    if year not in (1, 2, 3):
        return error_response('Année invalide', 400)
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    valid_students = {r['id'] for r in pdb.execute(
        'SELECT id FROM promotion_students WHERE promotion_id=?', (pid,))}
    body = request.get_json() or {}
    changed = 0
    for d in body.get('decisions') or []:
        try:
            sid, ue = int(d.get('student_id')), int(d.get('ue'))
        except (TypeError, ValueError):
            continue
        if sid not in valid_students or not (1 <= ue <= 9):
            continue
        if d.get('admj'):
            pdb.execute('''INSERT OR REPLACE INTO jury_decisions(promotion_id, year, student_id, ue_num, decision)
                           VALUES(?,?,?,?,'ADMJ')''', (pid, year, sid, ue))
        else:
            pdb.execute('''DELETE FROM jury_decisions WHERE promotion_id=? AND year=?
                           AND student_id=? AND ue_num=?''', (pid, year, sid, ue))
        changed += 1
    for d in body.get('year_decisions') or []:
        try:
            sid = int(d.get('student_id'))
        except (TypeError, ValueError):
            continue
        if sid not in valid_students:
            continue
        if d.get('red'):
            pdb.execute('''INSERT OR REPLACE INTO jury_decisions(promotion_id, year, student_id, ue_num, decision)
                           VALUES(?,?,?,0,'RED')''', (pid, year, sid))
        else:
            pdb.execute('''DELETE FROM jury_decisions WHERE promotion_id=? AND year=?
                           AND student_id=? AND ue_num=0''', (pid, year, sid))
        changed += 1
    pdb.commit()
    _reconcile_red_transfers(pdb, pid)   # décision passée de RED à AJAC/ADMJ/… → retire de la promo cible
    _audit('JURY_SAVE', ip=_client_ip(), user=session.get('user'),
           promo=pid, year=year, decisions=changed)
    formation = (request.args.get('formation') or '').strip().upper() or None
    return jsonify(_jury_payload(pdb, pid, year, formation))

def _coeff_codes(competences):
    """Codes des matières ayant un coefficient non nul dans au moins une compétence."""
    codes = set()
    for cc in competences or []:
        for code_k, val in (cc.get('coeffs') or {}).items():
            try:
                if float(val) != 0:
                    codes.add(code_k)
            except (TypeError, ValueError):
                pass
    return codes

def _visible_note_components(components, competences):
    """Matières affichées dans la grille de notes : celles ayant un coefficient dans le
    programme, plus les saisies spéciales BONUS/PEN. Écarte les matières sans coefficient
    (Aide à la réussite, Portfolio aux semestres impairs, etc.)."""
    codes = _coeff_codes(competences)
    return [m for m in (components or [])
            if m.get('kind') in ('BONUS', 'PEN') or m.get('code') in codes]

def _semester_roster(pdb, pid, semester, formation=None, comp=None):
    """Étudiants d'une grille de notes d'un semestre : effectif de l'année
    correspondante (report auto + ajustements), sous-cohorte résolue pour
    l'année, filtre optionnel FTP/ALT. Renvoie (students, comp jury)."""
    year = (int(semester[1:]) + 1) // 2   # 1,1,2,2,3,3
    students = [dict(r) for r in pdb.execute(
        '''SELECT id, numero, nom, prenom, statut, formation FROM promotion_students
           WHERE promotion_id=? ORDER BY nom COLLATE NOCASE, prenom COLLATE NOCASE''', (pid,))]
    if comp is None:
        comp = _jury_compute(pdb, pid)
    roster = _year_rosters(pdb, pid, comp).get(year, set())
    fm = _year_formation_map(pdb, pid, year)
    kept = []
    for s in students:
        if s['id'] not in roster:
            continue
        s['formation'] = fm.get(s['id'], s.get('formation') or 'FTP')
        if formation in _SUBCOHORTS and s['formation'] != formation:
            continue
        kept.append(s)
    return kept, comp

def _promo_notes_payload(pdb, pid, semester, formation=None):
    """Construit le tableau de notes : étudiants (effectif) × matières (référence),
    notes saisies, moyennes de compétences, rappels (semestre antérieur / année précédente)
    et moyenne annuelle courante. Si `formation` (FTP/ALT) est fourni, la grille ne montre
    que cette sous-cohorte."""
    ref_all = _promo_coeffs(pdb, pid) or {}
    ref = ref_all.get(semester) or {'components': [], 'competences': []}
    competences = ref.get('competences', [])
    # Matières sans coefficient dans le programme (Aide à la réussite, Portfolio impair…)
    # → masquées de la grille de notes.
    components = _visible_note_components(ref.get('components', []), competences)
    marks = {}
    for r in pdb.execute('''SELECT student_id, matiere_code, note, mention FROM student_marks
                            WHERE promotion_id=? AND semester=? AND note IS NOT NULL''', (pid, semester)):
        marks[f"{r['student_id']}_{r['matiere_code']}"] = r['mention'] if r['mention'] else r['note']
    averages = _semester_competence_averages(pdb, pid, semester, competences, components)
    sem_num = int(semester[1:])
    year = (sem_num + 1) // 2   # 1,1,2,2,3,3
    # Décisions de jury par année (ADM/ADMJ/AJAC/AJ/RED) — servent de « statut » dans la grille
    # + effectif de l'année (report auto + ajustements) et sous-cohorte propre à l'année
    students, comp = _semester_roster(pdb, pid, semester, formation)
    dec = comp['decisions']
    # Rappels : toutes les années précédentes (moyennes annuelles) puis le semestre antérieur.
    # kind ('year'/'semester') + year/sem servent au front pour grouper et coder les colonnes (UExy/UExNy).
    previous = []
    for py in range(1, year):
        po, pe = f'S{py * 2 - 1}', f'S{py * 2}'
        yavg, ynames = _year_competence_averages(pdb, pid, po, pe, ref_all)
        if ynames:
            previous.append({'short': f'{po}-{pe}', 'label': f'Année {py} ({po}–{pe})',
                             'kind': 'year', 'year': py,
                             'competences': ynames, 'averages': yavg,
                             'validation': _year_validation(pdb, pid, ref_all, py),
                             'decision': {str(s['id']): dec.get((py, str(s['id']))) for s in students}})
    if sem_num % 2 == 0:
        psem = f'S{sem_num - 1}'
        pdata = ref_all.get(psem) or {}
        pcomps = pdata.get('competences', [])
        if pcomps:
            previous.append({'short': psem, 'label': f'Semestre {psem}',
                             'kind': 'semester', 'sem': sem_num - 1,
                             'competences': [c.get('name', '') for c in pcomps],
                             'averages': _semester_competence_averages(
                                 pdb, pid, psem, pcomps, pdata.get('components', []))})
    # Moyenne annuelle courante = moyenne des 2 semestres de l'année (S1+S2, S3+S4, S5+S6)
    s_odd, s_even = f'S{year * 2 - 1}', f'S{year * 2}'
    year_averages, year_competences = _year_competence_averages(pdb, pid, s_odd, s_even, ref_all)
    return {'semester': semester, 'year': year, 'components': components,
            'competences': [{'name': c.get('name', ''), 'coeffs': c.get('coeffs', {})} for c in competences],
            # Statut de la saisie enseignant par matière (Saisie Notes) :
            # 'provisoire' / 'definitif' — indicateur affiché dans Bulletins
            'saisie_status': _saisie_status(pdb, pid, semester, formation),
            'students': students, 'marks': marks, 'averages': averages, 'previous': previous,
            'year_semesters': [s_odd, s_even], 'year_competences': year_competences,
            'year_averages': year_averages,
            'ue_numbers': _jury_ue_numbers(ref_all),
            'year_validation': _year_validation(pdb, pid, ref_all, year),
            'year_decision': {str(s['id']): dec.get((year, str(s['id']))) for s in students}}

@app.route('/api/promotions/<int:pid>/notes/<semester>', methods=['GET'])
def get_promotion_notes(pid, semester):
    err = _require_promo_read()
    if err:
        return err
    if semester not in _PROMO_SEMESTERS:
        return error_response('Semestre invalide', 400)
    pdb = get_promotions_db()
    promo = pdb.execute('SELECT id FROM promotions WHERE id=?', (pid,)).fetchone()
    if not promo:
        return error_response('Promotion introuvable', 404)
    formation = (request.args.get('formation') or '').strip().upper() or None
    return jsonify(_promo_notes_payload(pdb, pid, semester, formation))

@app.route('/api/promotions/<int:pid>/notes/<semester>', methods=['PUT'])
def save_promotion_notes(pid, semester):
    err = _require_admin()
    if err:
        return err
    if semester not in _PROMO_SEMESTERS:
        return error_response('Semestre invalide', 400)
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    valid_students = {r['id'] for r in pdb.execute(
        'SELECT id FROM promotion_students WHERE promotion_id=?', (pid,))}
    for m in (request.get_json() or {}).get('marks') or []:
        try:
            sid = int(m.get('student_id'))
        except (TypeError, ValueError):
            continue
        code = (m.get('matiere_code') or '').strip()
        if sid not in valid_students or not code:
            continue
        note = m.get('note')
        if note in (None, ''):
            pdb.execute('''DELETE FROM student_marks WHERE promotion_id=? AND semester=?
                           AND student_id=? AND matiere_code=?''', (pid, semester, sid, code))
            continue
        if isinstance(note, str) and note.strip().upper() == 'ABI':
            note_num, mention = 0.0, 'ABI'
        else:
            try:
                note_num, mention = float(note), None
            except (TypeError, ValueError):
                continue
        pdb.execute('''INSERT INTO student_marks(promotion_id, semester, student_id, matiere_code, note, mention)
                       VALUES(?,?,?,?,?,?)
                       ON CONFLICT(promotion_id, semester, student_id, matiere_code)
                       DO UPDATE SET note=excluded.note, mention=excluded.mention''',
                    (pid, semester, sid, code, note_num, mention))
    pdb.commit()
    _reconcile_red_transfers(pdb, pid)   # une note peut faire passer un RED en ADM/AJAC → nettoyage
    formation = (request.args.get('formation') or '').strip().upper() or None
    return jsonify(_promo_notes_payload(pdb, pid, semester, formation))

@app.route('/api/promotions/<int:pid>/notes/<semester>', methods=['DELETE'])
def delete_promotion_notes(pid, semester):
    """Supprime des notes saisies. Par défaut toutes les notes de la promo/semestre ;
    restreint à une sous-cohorte via ?formation=FTP|ALT, et/ou à un seul étudiant via
    ?student_id=N. Renvoie la grille rafraîchie (filtrée par la sous-cohorte courante)."""
    err = _require_admin()
    if err:
        return err
    if semester not in _PROMO_SEMESTERS:
        return error_response('Semestre invalide', 400)
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    formation = (request.args.get('formation') or '').strip().upper() or None
    sql = 'DELETE FROM student_marks WHERE promotion_id=? AND semester=?'
    params = [pid, semester]
    sid = request.args.get('student_id')
    if sid:
        try:
            sid_int = int(sid)
        except (TypeError, ValueError):
            return error_response('Étudiant invalide', 400)
        sql += ' AND student_id=?'
        params.append(sid_int)
    elif formation in _SUBCOHORTS:
        # « Tout le semestre » : borné à la sous-cohorte affichée (non destructif pour l'autre)
        sql += (' AND student_id IN (SELECT id FROM promotion_students'
                ' WHERE promotion_id=? AND formation=?)')
        params += [pid, formation]
    deleted = pdb.execute(sql, params).rowcount
    pdb.commit()
    _audit('NOTES_DELETE', ip=_client_ip(), user=session.get('user'),
           promo=pid, semester=semester, deleted=deleted,
           scope=('student:' + sid) if sid else (formation or 'all'))
    payload = _promo_notes_payload(pdb, pid, semester, formation)
    payload['delete_report'] = {'deleted': deleted}
    return jsonify(payload)

# ===== SAISIE NOTES (onglet Saisie Notes : notes par SOUS-MATIÈRE, enseignants) =====
# Chaque enseignant (connecté avec mot de passe → promo_access) saisit les notes
# des sous-matières qu'il a en charge. La note matière (student_marks, onglet
# Bulletins) = moyenne pondérée des sous-notes ; le RÉFÉRENT de la matière
# ([[referents]] de l'année) définit la pondération (0 = sous-matière exclue,
# pas de note à saisir). Statut provisoire/définitif par colonne (sous-matière
# × face). La note matière reportée reste modifiable par l'admin.

def _promo_semester_year(pdb, pid, semester):
    """Année universitaire (AAAA-AAAA) correspondant au semestre d'une promotion."""
    p = pdb.execute('SELECT start_year FROM promotions WHERE id=?', (pid,)).fetchone()
    if not p:
        return None
    y = p['start_year'] + (int(semester[1:]) + 1) // 2 - 1
    return f'{y}-{y + 1}'

def _saisie_groups(ydb, semester, formation):
    """Sous-matières (courses) d'un semestre de la base année, groupées par
    matière (_mat_base_key). Pour chaque sous-matière : enseignants actifs de
    la face demandée, `on_face` (a des séances sur la face) et `any_session`
    (a des séances tout court) — une sous-matière sans séance sur la face ne
    compte pas dans la moyenne de cette face (sauf si elle n'a aucune séance :
    saisie admin possible sur les deux faces)."""
    fts = (0, 2) if formation == 'FTP' else (1, 2)
    subs = {}
    for r in ydb.execute('''SELECT c.id, c.code, c.name FROM courses c
                            JOIN semesters s ON s.id = c.semester_id
                            WHERE s.code = ?''', (semester,)):
        subs[r['id']] = {'code': r['code'], 'name': r['name'],
                         'teachers': set(), 'on_face': False, 'any_session': False}
    for r in ydb.execute('''SELECT cs.course_id, cs.formation_type, t.name
                            FROM course_sessions cs JOIN teachers t ON t.id = cs.teacher_id
                            WHERE cs.nb_sessions > 0 OR cs.total_hours > 0'''):
        c = subs.get(r['course_id'])
        if not c:
            continue
        c['any_session'] = True
        if r['formation_type'] in fts:
            c['on_face'] = True
            c['teachers'].add(r['name'])
    groups = {}
    for c in subs.values():
        groups.setdefault(_mat_base_key(c['code']), []).append(c)
    return groups

def _saisie_counted(c, weight):
    """Une sous-matière compte dans la moyenne de la face si elle y a des
    séances (ou n'en a nulle part) et que sa pondération est > 0."""
    return (c['on_face'] or not c['any_session']) and weight > 0

def _saisie_payload(pdb, pid, semester, formation, teacher_name=None):
    """Grille de saisie : étudiants (sous-cohorte) × sous-matières groupées par
    matière. `teacher_name` (enseignant connecté) restreint aux matières où il
    intervient : ses colonnes sont éditables, les autres visibles en lecture
    seule (notes + pondérations). Admin (teacher_name=None) : tout éditable."""
    year_label = _promo_semester_year(pdb, pid, semester)
    path = db_path_for_year(year_label) if year_label else None
    base = {'semester': semester, 'formation': formation, 'year': year_label}
    if not (path and os.path.isfile(path)):
        return {**base, 'available': False, 'groups': [], 'students': [],
                'marks': {}, 'official': {}}
    ydb = _open_connection(path)
    try:
        groups_raw = _saisie_groups(ydb, semester, formation)
        refs = _matiere_referents(ydb)
    finally:
        ydb.close()
    ref = (_promo_coeffs(pdb, pid) or {}).get(semester) or {}
    components = _visible_note_components(ref.get('components', []), ref.get('competences', []))
    comp_by_code = {c.get('code'): c for c in components if c.get('code')}
    students, _ = _semester_roster(pdb, pid, semester, formation)
    meta = {r['code']: r for r in pdb.execute(
        '''SELECT code, weight, definitive FROM submatiere_meta
           WHERE promotion_id=? AND semester=? AND formation=?''', (pid, semester, formation))}
    tkey = (teacher_name or '').strip().lower()
    out_groups = []
    for key in sorted(groups_raw):
        compo = comp_by_code.get(key)
        if not compo:
            continue   # matière hors référence (Bonus, Aide à la réussite…) : pas de note
        r = (refs.get(key) or {}).get(formation) or {}
        referent = r.get('name')
        is_ref = bool(tkey) and bool(referent) and referent.strip().lower() == tkey
        mine, subs = False, []
        for c in sorted(groups_raw[key], key=lambda x: x['code']):
            m = meta.get(c['code'])
            w = m['weight'] if (m and m['weight'] is not None) else 1.0
            teaches = bool(tkey) and any(t.strip().lower() == tkey for t in c['teachers'])
            mine = mine or teaches
            subs.append({'code': c['code'], 'name': c['name'],
                         'teachers': sorted(c['teachers']),
                         'weight': w, 'counted': _saisie_counted(c, w),
                         # face_ok : la sous-matière existe sur la face (sert au
                         # recalcul live côté client quand la pondération change)
                         'face_ok': c['on_face'] or not c['any_session'],
                         'definitive': bool(m and m['definitive']),
                         'editable': (teacher_name is None or teaches) and _saisie_counted(c, w)})
        if teacher_name is not None and not (mine or is_ref):
            continue   # l'enseignant ne voit que les matières dont il a la charge
        out_groups.append({'key': key,
                           'label': compo.get('short_label') or compo.get('label') or key,
                           'referent': referent,
                           'is_referent': teacher_name is None or is_ref,
                           'subs': subs})
    codes = {s['code'] for grp in out_groups for s in grp['subs']}
    sids = {s['id'] for s in students}
    marks = {}
    for r in pdb.execute('''SELECT student_id, code, note, mention FROM submatiere_marks
                            WHERE promotion_id=? AND semester=? AND note IS NOT NULL''',
                         (pid, semester)):
        if r['student_id'] in sids and r['code'] in codes:
            marks[f"{r['student_id']}_{r['code']}"] = r['mention'] if r['mention'] else r['note']
    keys = {grp['key'] for grp in out_groups}
    official = {}   # note matière actuelle (Bulletins) — affichée à côté de la moyenne calculée
    for r in pdb.execute('''SELECT student_id, matiere_code, note, mention FROM student_marks
                            WHERE promotion_id=? AND semester=? AND note IS NOT NULL''',
                         (pid, semester)):
        if r['student_id'] in sids and r['matiere_code'] in keys:
            official[f"{r['student_id']}_{r['matiere_code']}"] = r['mention'] if r['mention'] else r['note']
    return {**base, 'available': True, 'groups': out_groups, 'students': students,
            'marks': marks, 'official': official}

def _recompute_matiere_marks(pdb, pid, semester, keys):
    """Reporte la moyenne pondérée des sous-notes dans la note matière
    (student_marks, utilisée par Bulletins/Jury) pour les matières `keys` —
    seulement quand TOUTES les sous-matières comptées (pondération > 0 sur la
    face de l'étudiant) ont une note. La note reportée reste modifiable par
    l'admin (écrasée au prochain enregistrement de saisie de la matière)."""
    year_label = _promo_semester_year(pdb, pid, semester)
    path = db_path_for_year(year_label) if year_label else None
    if not keys or not (path and os.path.isfile(path)):
        return
    ydb = _open_connection(path)
    try:
        by_face = {f: _saisie_groups(ydb, semester, f) for f in _SUBCOHORTS}
    finally:
        ydb.close()
    students, _ = _semester_roster(pdb, pid, semester)
    weights = {(r['formation'], r['code']): r['weight'] for r in pdb.execute(
        '''SELECT formation, code, weight FROM submatiere_meta
           WHERE promotion_id=? AND semester=?''', (pid, semester))}
    notes = {(r['student_id'], r['code']): r['note'] for r in pdb.execute(
        '''SELECT student_id, code, note FROM submatiere_marks
           WHERE promotion_id=? AND semester=? AND note IS NOT NULL''', (pid, semester))}
    for s in students:
        f = s['formation'] if s['formation'] in _SUBCOHORTS else 'FTP'
        for key in keys:
            wsubs = []
            for c in by_face[f].get(key) or []:
                w = weights.get((f, c['code']))
                w = 1.0 if w is None else w
                if _saisie_counted(c, w):
                    wsubs.append((c['code'], w))
            if not wsubs:
                continue
            vals = [(notes.get((s['id'], code)), w) for code, w in wsubs]
            if any(n is None for n, _ in vals):
                continue   # saisie incomplète : la note matière n'est pas touchée
            avg = round(sum(n * w for n, w in vals) / sum(w for _, w in vals), 2)
            pdb.execute('''INSERT INTO student_marks(promotion_id, semester, student_id, matiere_code, note, mention)
                           VALUES(?,?,?,?,?,NULL)
                           ON CONFLICT(promotion_id, semester, student_id, matiere_code)
                           DO UPDATE SET note=excluded.note, mention=NULL''',
                        (pid, semester, s['id'], key, avg))
    pdb.commit()

def _saisie_status(pdb, pid, semester, formation):
    """{matiere_code: 'definitif'|'provisoire'} pour les matières dont des
    sous-notes ont été saisies (indicateur affiché dans Bulletins)."""
    if formation not in _SUBCOHORTS:
        return {}
    year_label = _promo_semester_year(pdb, pid, semester)
    path = db_path_for_year(year_label) if year_label else None
    if not (path and os.path.isfile(path)):
        return {}
    ydb = _open_connection(path)
    try:
        groups = _saisie_groups(ydb, semester, formation)
    finally:
        ydb.close()
    meta = {r['code']: r for r in pdb.execute(
        '''SELECT code, weight, definitive FROM submatiere_meta
           WHERE promotion_id=? AND semester=? AND formation=?''', (pid, semester, formation))}
    with_marks = {r['code'] for r in pdb.execute(
        '''SELECT DISTINCT m.code FROM submatiere_marks m
           JOIN promotion_students ps ON ps.id = m.student_id
           WHERE m.promotion_id=? AND m.semester=? AND ps.formation=?''',
        (pid, semester, formation))}
    out = {}
    for key, subs in groups.items():
        counted = []
        for c in subs:
            m = meta.get(c['code'])
            w = m['weight'] if (m and m['weight'] is not None) else 1.0
            if _saisie_counted(c, w):
                counted.append(c['code'])
        if not counted or not any(code in with_marks for code in counted):
            continue   # aucune saisie enseignant sur cette matière
        alldef = all(meta.get(code) and meta[code]['definitive'] for code in counted)
        out[key] = 'definitif' if alldef else 'provisoire'
    return out

def _saisie_teacher_ctx():
    """(teacher_name, err) : None pour l'admin (tous droits) ; nom de
    l'enseignant si connecté avec mot de passe (promo_access) ; 403 sinon."""
    role = session.get('role')
    if role == 'admin':
        return None, None
    if role == 'teacher' and session.get('promo_access'):
        return session.get('teacher_name') or '', None
    return None, error_response('Accès réservé', 403)

@app.route('/api/promotions/<int:pid>/saisie/<semester>', methods=['GET'])
def get_promotion_saisie(pid, semester):
    teacher, err = _saisie_teacher_ctx()
    if err:
        return err
    if semester not in _PROMO_SEMESTERS:
        return error_response('Semestre invalide', 400)
    formation = (request.args.get('formation') or 'FTP').strip().upper()
    if formation not in _SUBCOHORTS:
        return error_response('Sous-cohorte invalide', 400)
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    return jsonify(_saisie_payload(pdb, pid, semester, formation, teacher))

@app.route('/api/promotions/<int:pid>/saisie/<semester>/notes', methods=['PUT'])
def save_promotion_saisie(pid, semester):
    """Enregistre des colonnes de saisie {formation, columns: [{code, notes:
    {student_id: note|''|ABI}, definitive}]}. Un enseignant ne peut écrire que
    les colonnes marquées éditables pour lui (il enseigne la sous-matière sur
    la face et la pondération est > 0) ; l'admin peut tout écrire."""
    teacher, err = _saisie_teacher_ctx()
    if err:
        return err
    if semester not in _PROMO_SEMESTERS:
        return error_response('Semestre invalide', 400)
    data = request.get_json() or {}
    formation = (data.get('formation') or '').strip().upper()
    if formation not in _SUBCOHORTS:
        return error_response('Sous-cohorte invalide', 400)
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    payload = _saisie_payload(pdb, pid, semester, formation, teacher)
    editable = {s['code']: grp['key'] for grp in payload['groups']
                for s in grp['subs'] if s['editable']}
    valid_sids = {s['id'] for s in payload['students']}
    keys = set()
    for col in (data.get('columns') or []):
        code = (col.get('code') or '').strip()
        if code not in editable:
            return error_response(f'Sous-matière {code} non modifiable', 403)
        keys.add(editable[code])
        for sid_s, val in (col.get('notes') or {}).items():
            try:
                sid = int(sid_s)
            except (TypeError, ValueError):
                continue
            if sid not in valid_sids:
                continue
            if val in (None, ''):
                pdb.execute('''DELETE FROM submatiere_marks WHERE promotion_id=?
                               AND semester=? AND student_id=? AND code=?''',
                            (pid, semester, sid, code))
                continue
            if isinstance(val, str) and val.strip().upper() == 'ABI':
                note, mention = 0.0, 'ABI'
            else:
                try:
                    note, mention = float(val), None
                except (TypeError, ValueError):
                    continue
            pdb.execute('''INSERT INTO submatiere_marks(promotion_id, semester, student_id, code, note, mention)
                           VALUES(?,?,?,?,?,?)
                           ON CONFLICT(promotion_id, semester, student_id, code)
                           DO UPDATE SET note=excluded.note, mention=excluded.mention''',
                        (pid, semester, sid, code, note, mention))
        if 'definitive' in col:
            pdb.execute('''INSERT INTO submatiere_meta(promotion_id, semester, formation, code, weight, definitive)
                           VALUES(?,?,?,?,NULL,?)
                           ON CONFLICT(promotion_id, semester, formation, code)
                           DO UPDATE SET definitive=excluded.definitive''',
                        (pid, semester, formation, code, 1 if col.get('definitive') else 0))
    pdb.commit()
    _recompute_matiere_marks(pdb, pid, semester, keys)
    _reconcile_red_transfers(pdb, pid)
    _audit('SAISIE_NOTES', ip=_client_ip(), user=session.get('user'),
           promo=pid, semester=semester, formation=formation,
           cols=len(data.get('columns') or []))
    return jsonify(_saisie_payload(pdb, pid, semester, formation, teacher))

@app.route('/api/promotions/<int:pid>/saisie/<semester>/weights', methods=['PUT'])
def save_promotion_saisie_weights(pid, semester):
    """Pondérations d'une matière {formation, matiere_key, weights: {code: w}}.
    Réservé à l'admin et au RÉFÉRENT de la matière sur cette face (w >= 0 ;
    0 exclut la sous-matière de la moyenne et de la saisie)."""
    teacher, err = _saisie_teacher_ctx()
    if err:
        return err
    if semester not in _PROMO_SEMESTERS:
        return error_response('Semestre invalide', 400)
    data = request.get_json() or {}
    formation = (data.get('formation') or '').strip().upper()
    if formation not in _SUBCOHORTS:
        return error_response('Sous-cohorte invalide', 400)
    key = (data.get('matiere_key') or '').strip()
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    payload = _saisie_payload(pdb, pid, semester, formation, teacher)
    grp = next((x for x in payload['groups'] if x['key'] == key), None)
    if not grp:
        return error_response('Matière introuvable', 404)
    if not grp['is_referent']:
        return error_response('Réservé au référent de la matière', 403)
    codes = {s['code'] for s in grp['subs']}
    for code, w in (data.get('weights') or {}).items():
        if code not in codes:
            continue
        try:
            w = float(w)
        except (TypeError, ValueError):
            return error_response(f'Pondération invalide pour {code}')
        if w < 0:
            return error_response('Une pondération ne peut pas être négative')
        pdb.execute('''INSERT INTO submatiere_meta(promotion_id, semester, formation, code, weight, definitive)
                       VALUES(?,?,?,?,?,0)
                       ON CONFLICT(promotion_id, semester, formation, code)
                       DO UPDATE SET weight=excluded.weight''',
                    (pid, semester, formation, code, w))
    pdb.commit()
    _recompute_matiere_marks(pdb, pid, semester, {key})
    _audit('SAISIE_WEIGHTS', ip=_client_ip(), user=session.get('user'),
           promo=pid, semester=semester, formation=formation, matiere=key)
    return jsonify(_saisie_payload(pdb, pid, semester, formation, teacher))

def _cell_txt(v):
    return ('' if v is None else str(v)).strip()

def _note_value(v):
    """Note d'une cellule : nombre tel quel, 'ABI' = mention (comptée 0), sinon None (absent)."""
    if isinstance(v, (int, float)):
        return float(v)
    if _cell_txt(v).upper() == 'ABI':
        return 'ABI'
    return None

def _kind_from_code(code, title=''):
    cu = str(code).upper()
    if 'portfolio' in str(title).lower():
        return 'PORT'
    if 'IS' in cu:
        return 'SAE'
    if 'IR' in cu:
        return 'RES'
    return 'AUTRE'

def _apogee_marks_ws(ws):
    """Layout Apogée : ELP en ligne 8 (code) / 13 (type N) / 14 (titre), notes lignes 18+, n° col A."""
    elps = []
    for c in range(5, ws.max_column + 1):
        if ws.cell(13, c).value == 'N' and ws.cell(8, c).value:
            elps.append({'note_col': c, 'code': _cell_txt(ws.cell(8, c).value),
                         'kind': _kind_from_code(ws.cell(8, c).value, ws.cell(14, c).value)})
    rows, r = [], 18
    while r <= 2000:
        num = ws.cell(r, 1).value
        if num in (None, ''):
            break
        notes = {}
        for e in elps:
            nv = _note_value(ws.cell(r, e['note_col']).value)
            if nv is not None:
                notes[e['note_col']] = nv
        rows.append({'numero': _cell_txt(num), 'notes': notes})
        r += 1
    return elps, rows

def _find_pv_code_row(ws, max_scan=30):
    """Ligne des codes ELP (T3IS.../T3IR...) d'un PV de jury. Repérée dynamiquement
    car sa position varie selon le modèle (ligne 6 sur un PV « matière », ligne 13
    sur un PV de jury général). Retourne l'indice de la ligne qui contient le plus
    de codes matière, ou None si aucune."""
    best_row, best_n = None, 0
    for r in range(1, min(max_scan, ws.max_row) + 1):
        n = sum(1 for c in range(1, ws.max_column + 1)
                if re.match(r'^T3I[SR]\d', _cell_txt(ws.cell(r, c).value), re.I))
        if n > best_n:
            best_row, best_n = r, n
    return best_row

def _pv_num(v):
    """Numéro d'étudiant en texte, sans décimale parasite (22503981.0 → '22503981')."""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return _cell_txt(v)

def _pv_name_cols(ws, hdr_row):
    """Colonnes NOM / Prénom repérées sur la ligne d'en-tête (accents ignorés).
    Repli sur C/D, disposition standard des PV de jury."""
    import unicodedata
    nom_col = prenom_col = None
    for c in range(1, ws.max_column + 1):
        t = _cell_txt(ws.cell(hdr_row, c).value)
        t = unicodedata.normalize('NFD', t).encode('ascii', 'ignore').decode().lower()
        if 'prenom' in t and prenom_col is None:
            prenom_col = c
        elif 'nom' in t and nom_col is None:
            nom_col = c
    return nom_col or 3, prenom_col or 4

def _pv_marks_ws(ws):
    """Layout PV de jury : codes matière T3IS../T3IR.. sur une ligne d'en-tête (repérée
    dynamiquement), notes en dessous, n° étudiant en col B, NOM/Prénom en cols C/D.
    Retourne (elps, rows) — rows : {numero, nom, prenom, notes:{note_col:val}}."""
    code_row = _find_pv_code_row(ws)
    if not code_row:
        return [], []
    elps = []
    for c in range(1, ws.max_column + 1):
        code = _cell_txt(ws.cell(code_row, c).value)
        if re.match(r'^T3I[SR]\d', code, re.I):
            elps.append({'note_col': c, 'code': code.upper(),
                         'kind': _kind_from_code(code, ws.cell(code_row - 1, c).value)})
    nom_col, prenom_col = _pv_name_cols(ws, code_row - 1)
    rows = []
    for r in range(code_row + 1, ws.max_row + 1):
        num = _pv_num(ws.cell(r, 2).value)
        if not num or not num[0].isdigit():
            continue
        notes = {}
        for e in elps:
            nv = _note_value(ws.cell(r, e['note_col']).value)
            if nv is not None:
                notes[e['note_col']] = nv
        rows.append({'numero': num,
                     'nom': _cell_txt(ws.cell(r, nom_col).value),
                     'prenom': _cell_txt(ws.cell(r, prenom_col).value),
                     'notes': notes})
    return elps, rows

def _parse_notes_file(path):
    """Lit un fichier de notes (Apogée ou PV de jury), auto-détection.
    Retourne ([elps {note_col, kind}], [rows {numero, notes:{note_col:val}}])."""
    import openpyxl
    is_xlsm = path.lower().endswith('.xlsm')
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=is_xlsm)
    ws = wb.active
    # PV de jury : présence d'une ligne de codes matière T3IS../T3IR.. (position variable)
    if _find_pv_code_row(ws):
        return _pv_marks_ws(ws)
    return _apogee_marks_ws(ws)

@app.route('/api/promotions/<int:pid>/notes/<semester>/import', methods=['POST'])
def import_promotion_notes(pid, semester):
    """Importe les notes depuis un fichier Apogée OU un PV de jury (auto-détection).
    Les colonnes de notes sont mappées aux matières de la référence (par index SAÉ/Stage/PORT/RES) ;
    les notes sont rattachées aux étudiants de l'effectif via le n° Apogée."""
    err = _require_admin()
    if err:
        return err
    if semester not in _PROMO_SEMESTERS:
        return error_response('Semestre invalide', 400)
    pdb = get_promotions_db()
    if not pdb.execute('SELECT 1 FROM promotions WHERE id=?', (pid,)).fetchone():
        return error_response('Promotion introuvable', 404)
    f = request.files.get('file')
    if not f or not f.filename:
        return error_response('Aucun fichier reçu', 400)
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in _ALLOWED_GRADE_EXT:
        return error_response('Type de fichier non autorisé (.xlsm/.xlsx)', 400)
    if request.content_length and request.content_length > _MAX_GRADE_FILE:
        return error_response('Fichier trop volumineux (max 15 Mo)', 400)
    import tempfile
    tmp = os.path.join(tempfile.gettempdir(), secure_filename(f.filename) or ('notes' + ext))
    f.save(tmp)
    try:
        elps, rows = _parse_notes_file(tmp)
    except Exception as e:
        return error_response(f'Lecture impossible : {e}', 400)
    finally:
        try: os.remove(tmp)
        except OSError: pass
    if not elps:
        return error_response('Aucune colonne de note détectée (fichier Apogée attendu)', 400)

    # Mapping colonne note (fichier) -> code matière (référence du semestre).
    # Priorité à la correspondance EXACTE par code Apogée (T3IR202, …) ; repli
    # sur l'appariement par index/type (SAÉ/Stage, Portfolio, Ressource).
    ref = (_promo_coeffs(pdb, pid) or {}).get(semester) or {}
    components = ref.get('components', [])
    apo2disp = {c['apogee_code']: c['code'] for c in components if c.get('apogee_code')}
    col2code = {}
    for e in elps:
        disp = apo2disp.get(e.get('code'))
        if disp:
            col2code[e['note_col']] = disp
    if not col2code:
        is_codes = [c['code'] for c in components if c.get('kind') in ('SAE', 'STAGE')]
        port_codes = [c['code'] for c in components if c.get('kind') == 'PORT']
        res_codes = [c['code'] for c in components if c.get('kind') == 'RES']
        for grp_kind, codes in (('SAE', is_codes), ('PORT', port_codes), ('RES', res_codes)):
            i = 0
            for e in elps:
                if e['kind'] == grp_kind:
                    if i < len(codes):
                        col2code[e['note_col']] = codes[i]
                    i += 1
    if not col2code:
        return error_response('Aucune correspondance matière (vérifiez la référence des coefficients)', 400)

    # Index des étudiants de l'effectif par numéro
    students_by_num = {}
    for s in pdb.execute('SELECT id, numero FROM promotion_students WHERE promotion_id=?', (pid,)):
        num = (s['numero'] or '').strip().lower()
        if num:
            students_by_num[num] = s['id']
    notes_set = 0
    matched = unmatched = 0
    for row in rows:
        sid = students_by_num.get(row['numero'].lower())
        if not sid:
            unmatched += 1
            continue
        matched += 1
        for note_col, val in row['notes'].items():
            code = col2code.get(note_col)
            if not code:
                continue
            note_num, mention = (0.0, 'ABI') if isinstance(val, str) else (val, None)
            pdb.execute('''INSERT INTO student_marks(promotion_id, semester, student_id, matiere_code, note, mention)
                           VALUES(?,?,?,?,?,?)
                           ON CONFLICT(promotion_id, semester, student_id, matiere_code)
                           DO UPDATE SET note=excluded.note, mention=excluded.mention''',
                        (pid, semester, sid, code, note_num, mention))
            notes_set += 1
    pdb.commit()
    _audit('NOTES_IMPORT', ip=_client_ip(), user=session.get('user'),
           promo=pid, semester=semester, notes=notes_set)
    formation = (request.form.get('formation') or '').strip().upper() or None
    payload = _promo_notes_payload(pdb, pid, semester, formation)
    payload['import_report'] = {'notes': notes_set, 'etudiants_rapproches': matched,
                                'etudiants_non_trouves': unmatched, 'matieres_mappees': len(col2code)}
    return jsonify(payload)

# Couleur d'en-tête par type de matière (cohérent avec _coeffKindColor côté frontend)
_NOTES_KIND_FILL = {'SAE': 'DBEAFE', 'RES': 'DCFCE7', 'PORT': 'FEF3C7',
                    'STAGE': 'FEE2E2'}

@app.route('/api/promotions/<int:pid>/notes/<semester>/export', methods=['GET'])
def export_promotion_notes(pid, semester):
    """Exporte les notes d'une promo/semestre en .xlsx (disposition type Apogée :
    Numéro / Nom / Prénom / Naissance + une colonne par matière)."""
    err = _require_admin()
    if err:
        return err
    if semester not in _PROMO_SEMESTERS:
        return error_response('Semestre invalide', 400)
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    pdb = get_promotions_db()
    promo = pdb.execute('SELECT id, name, formation FROM promotions WHERE id=?', (pid,)).fetchone()
    if not promo:
        return error_response('Promotion introuvable', 404)

    ref = (_promo_coeffs(pdb, pid) or {}).get(semester) or {}
    # Matières sans coefficient dans le programme (Aide à la réussite, Portfolio impair…)
    # → masquées, cohérent avec la grille de saisie.
    components = _visible_note_components(ref.get('components', []), ref.get('competences', []))
    students = pdb.execute(
        '''SELECT id, numero, nom, prenom, naissance FROM promotion_students
           WHERE promotion_id=? ORDER BY nom COLLATE NOCASE, prenom COLLATE NOCASE''', (pid,)).fetchall()
    marks = {}
    for r in pdb.execute('''SELECT student_id, matiere_code, note, mention FROM student_marks
                            WHERE promotion_id=? AND semester=? AND note IS NOT NULL''', (pid, semester)):
        marks[(r['student_id'], r['matiere_code'])] = r['mention'] if r['mention'] else r['note']

    formation = (promo['formation'] or '').upper()
    year = get_current_year() or ''
    year_start = year[:4] if len(year) >= 4 and year[:4].isdigit() else ''
    year_slash = year.replace('-', '/')

    # Disposition reproduite à l'identique de la feuille `imp_exp` du fichier
    # d'import Apogée (lignes apoL_*/ELP/Code/Type Rés./libellés, étudiants à
    # partir de la ligne 18, paires Note/Barème) — sans les macros. La secrétaire
    # n'a plus qu'à copier-coller le bloc dans le vrai fichier Apogée.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'imp_exp'
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold = Font(bold=True)

    n = len(components)
    last_bareme_col = 4 + 2 * n   # barème de la dernière matière

    # --- Métadonnées (lignes 1-4) ---
    ws.cell(1, 5, 'Apogée')
    ws.cell(3, 1, ' Fichier :')
    ws.cell(3, 2, f"t3gim{semester}{formation}{year}.txt".lower())
    ws.cell(3, 5, 'Version de diplôme :')
    ws.cell(3, 7, 'T3GIM'); ws.cell(3, 8, 840)
    ws.cell(3, 9, 'BUT TER GIM ISP')
    ws.cell(4, 5, f"{semester} {formation} {year}")
    ws.cell(4, 12, ' Année :'); ws.cell(4, 13, year_slash)

    # --- Étiquettes de structure en colonne A ---
    ws.cell(6, 1, 'apoL_a01_code'); ws.cell(6, 2, 'apoL_a02_nom')
    ws.cell(6, 3, 'apoL_a03_prenom'); ws.cell(6, 4, 'apoL_a04_naissance')
    ws.cell(7, 1, 'Type Objet'); ws.cell(8, 1, 'Code'); ws.cell(9, 1, 'Version')
    ws.cell(10, 1, 'Année'); ws.cell(11, 1, 'Session')
    ws.cell(12, 1, 'Admission/Admissibilité'); ws.cell(13, 1, 'Type Rés.')
    ws.cell(15, 4, 'Session'); ws.cell(16, 1, 'Etudiant'); ws.cell(16, 4, 'Admissibilité')
    for col, txt in enumerate(['Numéro', 'Nom', 'Prénom', 'Naissance'], start=1):
        ws.cell(17, col, txt).font = bold

    # --- Colonnes des matières (paires Note/Barème) ---
    for i, m in enumerate(components):
        nc = 5 + 2 * i          # colonne Note
        bc = nc + 1             # colonne Barème
        # Code et nom exacts Apogée (le code d'affichage R/SAE n'est qu'interne)
        apo = m.get('apogee_code') or m.get('code', '')
        aname = m.get('apogee_name') or m.get('label', '')
        kind_fill = PatternFill('solid', fgColor=_NOTES_KIND_FILL.get(m.get('kind'), 'F3F4F6'))
        ws.cell(6, nc, f"apoL_c{2 * i + 1:04d}"); ws.cell(6, bc, f"apoL_c{2 * i + 2:04d}")
        ws.cell(7, nc, 'ELP')
        ws.cell(7, bc, 'APO_COL_VAL_FIN' if bc == last_bareme_col else 'ELP')
        ws.cell(8, nc, apo); ws.cell(8, bc, apo)
        if year_start:
            ws.cell(10, nc, int(year_start)); ws.cell(10, bc, int(year_start))
        ws.cell(11, nc, 0); ws.cell(11, bc, 0)
        ws.cell(12, nc, 1); ws.cell(12, bc, 1)
        ws.cell(13, nc, 'N'); ws.cell(13, bc, 'B')
        ws.cell(15, nc, 0); ws.cell(15, bc, 0)
        lab = ws.cell(14, nc, f"{apo} - {aname}")
        lab.font = Font(bold=True, size=9); lab.alignment = center; lab.fill = kind_fill
        ws.cell(17, nc, 'Note').font = bold
        ws.cell(17, bc, 'Barème').font = bold

    # --- Étudiants (ligne 18+) : notes dans les colonnes Note, barème = 20 partout ---
    row = 18
    for s in students:
        ws.cell(row, 1, s['numero']); ws.cell(row, 2, s['nom'])
        ws.cell(row, 3, s['prenom']); ws.cell(row, 4, s['naissance'])
        for i, m in enumerate(components):
            nc = 5 + 2 * i
            val = marks.get((s['id'], m.get('code')))
            ws.cell(row, nc, val if val is not None else None).alignment = center
            ws.cell(row, nc + 1, 20).alignment = center   # barème toujours 20, prérempli
        row += 1

    for col, wd in ((1, 12), (2, 18), (3, 14), (4, 12)):
        ws.column_dimensions[get_column_letter(col)].width = wd
    for i in range(2 * n):
        ws.column_dimensions[get_column_letter(5 + i)].width = 8
    ws.row_dimensions[14].height = 42
    ws.freeze_panes = ws.cell(row=18, column=5)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"T3GIM_{semester}_{formation}_{year}.xlsx"
    _audit('NOTES_EXPORT', ip=_client_ip(), user=session.get('user'),
           promo=pid, semester=semester)
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ======================= STATIC FILES =======================

_STATIC_DIR = os.path.join(_BASE_DIR, 'static')

@app.route('/')
def serve_root():
    """Serve index.html at root"""
    index_path = os.path.join(_STATIC_DIR, 'index.html')
    if os.path.exists(index_path):
        return send_file(index_path)
    return jsonify({'message': 'Welcome to IUT EDT Management System'}), 200

@app.route('/<path:path>')
def serve_static(path):
    """Serve static files"""
    return send_from_directory(_STATIC_DIR, path)

# Dossier des documents téléchargeables (PDF du Programme national, etc.)
_DOCS_DIR = os.path.join(_BASE_DIR, 'docs')

def _find_pn_pdf():
    """Chemin du PDF du Programme national s'il existe.
    Priorité à la variable d'env EDT_PN_PDF, sinon le 1er .pdf du dossier docs/."""
    env = os.environ.get('EDT_PN_PDF')
    if env and os.path.isfile(env):
        return env
    if os.path.isdir(_DOCS_DIR):
        pdfs = sorted(f for f in os.listdir(_DOCS_DIR) if f.lower().endswith('.pdf'))
        if pdfs:
            return os.path.join(_DOCS_DIR, pdfs[0])
    return None

@app.route('/api/pn-pdf/info', methods=['GET'])
def pn_pdf_info():
    """Indique si un PDF du Programme national est disponible au téléchargement."""
    p = _find_pn_pdf()
    if not p:
        return jsonify({'available': False})
    return jsonify({'available': True, 'filename': 'PN_GIM_ISP.pdf'})

@app.route('/api/pn-pdf', methods=['GET'])
def pn_pdf_download():
    """Télécharge le PDF du Programme national déposé dans le dossier docs/."""
    p = _find_pn_pdf()
    if not p:
        return error_response('Aucun PDF du Programme national disponible', 404)
    return send_file(p, mimetype='application/pdf', as_attachment=True,
                     download_name='PN_GIM_ISP.pdf')

# ======================= TEACHERS CRUD =======================

def _teacher_public(d):
    """Version « publique » d'une fiche enseignant : le hash de mot de passe ne
    sort jamais de la base, seuls les booléens has_password (mot de passe créé)
    et password_allowed (création autorisée par l'admin) sont exposés à l'API."""
    d['has_password'] = bool(d.pop('password_hash', None))
    d['password_allowed'] = bool(d.get('password_allowed'))
    return d

@app.route('/api/teachers', methods=['GET'])
def get_teachers():
    """Get all teachers"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT * FROM teachers ORDER BY name')
        teachers = [_teacher_public(t) for t in rows_to_list(cursor.fetchall())]
        return jsonify(teachers), 200
    except Exception as e:
        return error_response(f'Error fetching teachers: {str(e)}', 500)

@app.route('/api/teachers', methods=['POST'])
def create_teacher():
    """Create a new teacher"""
    try:
        data = request.get_json()
        if not data or not data.get('name'):
            return error_response('Teacher name is required')
        
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            INSERT INTO teachers (name, email, phone, structure, corps_code, status, max_hours_day, priority)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['name'],
            data.get('email'),
            data.get('phone'),
            data.get('structure'),
            data.get('corps_code'),
            data.get('status') or 'Titulaire',
            data.get('max_hours_day', 6),
            data.get('priority', 1)
        ))
        db.commit()
        teacher_id = cursor.lastrowid
        cursor.execute('SELECT * FROM teachers WHERE id = ?', (teacher_id,))
        teacher = _teacher_public(row_to_dict(cursor.fetchone()))
        return jsonify(teacher), 201
    except sqlite3.IntegrityError:
        return error_response('Teacher name already exists')
    except Exception as e:
        return error_response(f'Error creating teacher: {str(e)}', 500)

@app.route('/api/teachers/<int:teacher_id>', methods=['GET'])
def get_teacher(teacher_id):
    """Get a specific teacher"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT * FROM teachers WHERE id = ?', (teacher_id,))
        teacher = cursor.fetchone()
        
        if not teacher:
            return error_response('Teacher not found', 404)

        return jsonify(_teacher_public(row_to_dict(teacher))), 200
    except Exception as e:
        return error_response(f'Error fetching teacher: {str(e)}', 500)

@app.route('/api/teachers/<int:teacher_id>', methods=['PUT'])
def update_teacher(teacher_id):
    """Update a teacher"""
    try:
        data = request.get_json()
        db = get_db()
        cursor = db.cursor()
        
        # Check if teacher exists
        cursor.execute('SELECT * FROM teachers WHERE id = ?', (teacher_id,))
        if not cursor.fetchone():
            return error_response('Teacher not found', 404)
        
        # Update fields
        update_fields = []
        values = []
        for field in ['name', 'email', 'phone', 'structure', 'corps_code', 'status', 'max_hours_day', 'priority']:
            if field in data:
                update_fields.append(f'{field} = ?')
                values.append(data[field])
        
        if not update_fields:
            return error_response('No fields to update')
        
        values.append(teacher_id)
        query = f'UPDATE teachers SET {", ".join(update_fields)}, updated_at = CURRENT_TIMESTAMP WHERE id = ?'
        cursor.execute(query, values)
        db.commit()
        
        cursor.execute('SELECT * FROM teachers WHERE id = ?', (teacher_id,))
        teacher = _teacher_public(row_to_dict(cursor.fetchone()))
        return jsonify(teacher), 200
    except sqlite3.IntegrityError:
        return error_response('Teacher name already exists')
    except Exception as e:
        return error_response(f'Error updating teacher: {str(e)}', 500)

@app.route('/api/teachers/<int:teacher_id>', methods=['DELETE'])
def delete_teacher(teacher_id):
    """Delete a teacher"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Check if teacher exists
        cursor.execute('SELECT * FROM teachers WHERE id = ?', (teacher_id,))
        if not cursor.fetchone():
            return error_response('Teacher not found', 404)
        
        cursor.execute('DELETE FROM teachers WHERE id = ?', (teacher_id,))
        db.commit()
        return jsonify({'message': 'Teacher deleted'}), 200
    except Exception as e:
        return error_response(f'Error deleting teacher: {str(e)}', 500)

@app.route('/api/teachers/<int:teacher_id>/password-access', methods=['PUT'])
def set_teacher_password_access(teacher_id):
    """Gère l'accès Promotions d'un enseignant (admin seul). Body {action} :
      • allow  → autorise l'enseignant à créer son mot de passe au login ;
      • reset  → supprime le mot de passe actuel (le bouton d'initialisation
                 réapparaît au login, l'autorisation est conservée) ;
      • revoke → supprime le mot de passe ET retire l'autorisation.
    Le mot de passe lui-même n'est jamais choisi ni connu de l'admin."""
    err = _require_admin()
    if err:
        return err
    db = get_db()
    row = db.execute('SELECT name FROM teachers WHERE id = ?', (teacher_id,)).fetchone()
    if not row:
        return error_response('Teacher not found', 404)
    action = (request.get_json() or {}).get('action') or ''
    if action == 'allow':
        db.execute('UPDATE teachers SET password_allowed = 1, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
                   (teacher_id,))
        event = 'TEACHER_PWD_ALLOWED'
    elif action == 'reset':
        db.execute('''UPDATE teachers SET password_hash = NULL, password_allowed = 1,
                      updated_at = CURRENT_TIMESTAMP WHERE id = ?''', (teacher_id,))
        event = 'TEACHER_PWD_RESET'
    elif action == 'revoke':
        db.execute('''UPDATE teachers SET password_hash = NULL, password_allowed = 0,
                      updated_at = CURRENT_TIMESTAMP WHERE id = ?''', (teacher_id,))
        event = 'TEACHER_PWD_REVOKED'
    else:
        return error_response('Action invalide (allow, reset ou revoke)')
    db.commit()
    _audit(event, ip=_client_ip(), teacher=row['name'])
    t = db.execute('SELECT password_allowed, password_hash FROM teachers WHERE id = ?',
                   (teacher_id,)).fetchone()
    return jsonify({'password_allowed': bool(t['password_allowed']),
                    'has_password': bool(t['password_hash'])}), 200

# ======================= MON COMPTE (enseignant connecté) =======================
# Fiche contact + heures HETD effectuées hors GIM. Ces heures sont personnelles :
# elles ne sont renvoyées que par ces endpoints, à l'enseignant lui-même — sauf
# les lignes marquées publiques, exposées à l'admin via /api/external-hours/public.

def _my_teacher_row(db):
    """Fiche teachers (base de l'année consultée) de l'enseignant connecté, ou None."""
    name = session.get('teacher_name')
    if session.get('role') != 'teacher' or not name:
        return None
    return db.execute('SELECT * FROM teachers WHERE LOWER(name) = LOWER(?)', (name,)).fetchone()

def _my_account_payload(db, t):
    ext = rows_to_list(db.execute(
        'SELECT id, label, hetd, is_public FROM teacher_external_hours WHERE teacher_id = ? ORDER BY id',
        (t['id'],)).fetchall())
    return {
        'teacher': {'id': t['id'], 'name': t['name'], 'email': t['email'], 'phone': t['phone'],
                    'structure': t['structure'], 'corps_code': t['corps_code'], 'status': t['status']},
        'external_hours': ext,
        'total_external_hetd': round(sum(float(e['hetd'] or 0) for e in ext), 2),
    }

@app.route('/api/my-account', methods=['GET'])
def my_account_get():
    db = get_db()
    t = _my_teacher_row(db)
    if not t:
        return error_response('Réservé aux enseignants connectés', 403)
    return jsonify(_my_account_payload(db, t)), 200

@app.route('/api/my-account', methods=['PUT'])
def my_account_update():
    """L'enseignant met à jour ses coordonnées : email, téléphone, employeur.
    (Nom, corps et statut restent gérés par l'admin dans la fiche enseignant.)"""
    db = get_db()
    t = _my_teacher_row(db)
    if not t:
        return error_response('Réservé aux enseignants connectés', 403)
    data = request.get_json() or {}
    fields, values = [], []
    for field in ['email', 'phone', 'structure']:
        if field in data:
            v = (data.get(field) or '').strip() or None
            fields.append(f'{field} = ?')
            values.append(v)
    if not fields:
        return error_response('Aucun champ à mettre à jour')
    values.append(t['id'])
    db.execute(f'UPDATE teachers SET {", ".join(fields)}, updated_at = CURRENT_TIMESTAMP WHERE id = ?', values)
    db.commit()
    t = db.execute('SELECT * FROM teachers WHERE id = ?', (t['id'],)).fetchone()
    return jsonify(_my_account_payload(db, t)), 200

@app.route('/api/my-account/external-hours', methods=['POST'])
def my_account_add_external():
    """Ajoute une ligne d'heures hors GIM. Body : {label, hetd}."""
    db = get_db()
    t = _my_teacher_row(db)
    if not t:
        return error_response('Réservé aux enseignants connectés', 403)
    data = request.get_json() or {}
    label = (data.get('label') or '').strip()
    if not label:
        return error_response('Intitulé requis (ex. département, établissement)')
    try:
        hetd = float(data.get('hetd'))
    except (TypeError, ValueError):
        return error_response('Nombre d\'HETD invalide')
    if not (0 < hetd <= 1000):
        return error_response('Nombre d\'HETD invalide (entre 0 et 1000)')
    db.execute('INSERT INTO teacher_external_hours (teacher_id, label, hetd, is_public) VALUES (?, ?, ?, ?)',
               (t['id'], label, hetd, 1 if data.get('is_public') else 0))
    db.commit()
    return jsonify(_my_account_payload(db, t)), 201

@app.route('/api/my-account/external-hours/<int:ext_id>', methods=['PUT'])
def my_account_update_external(ext_id):
    """Modifie une ligne d'heures hors GIM (uniquement les siennes).
    Body : {is_public} et/ou {label, hetd}."""
    db = get_db()
    t = _my_teacher_row(db)
    if not t:
        return error_response('Réservé aux enseignants connectés', 403)
    row = db.execute('SELECT * FROM teacher_external_hours WHERE id = ? AND teacher_id = ?',
                     (ext_id, t['id'])).fetchone()
    if not row:
        return error_response('Ligne introuvable', 404)
    data = request.get_json() or {}
    label = (data.get('label') or '').strip() if 'label' in data else row['label']
    if not label:
        return error_response('Intitulé requis')
    hetd = row['hetd']
    if 'hetd' in data:
        try:
            hetd = float(data.get('hetd'))
        except (TypeError, ValueError):
            return error_response('Nombre d\'HETD invalide')
        if not (0 < hetd <= 1000):
            return error_response('Nombre d\'HETD invalide (entre 0 et 1000)')
    is_public = (1 if data.get('is_public') else 0) if 'is_public' in data else row['is_public']
    db.execute('UPDATE teacher_external_hours SET label = ?, hetd = ?, is_public = ? WHERE id = ?',
               (label, hetd, is_public, ext_id))
    db.commit()
    return jsonify(_my_account_payload(db, t)), 200

@app.route('/api/my-account/external-hours/<int:ext_id>', methods=['DELETE'])
def my_account_delete_external(ext_id):
    """Supprime une ligne d'heures hors GIM (uniquement les siennes)."""
    db = get_db()
    t = _my_teacher_row(db)
    if not t:
        return error_response('Réservé aux enseignants connectés', 403)
    cur = db.execute('DELETE FROM teacher_external_hours WHERE id = ? AND teacher_id = ?',
                     (ext_id, t['id']))
    db.commit()
    if cur.rowcount == 0:
        return error_response('Ligne introuvable', 404)
    return jsonify(_my_account_payload(db, t)), 200

@app.route('/api/external-hours/public', methods=['GET'])
def public_external_hours():
    """Lignes d'heures hors GIM déclarées PUBLIQUES par leur enseignant.
    Réservé à l'admin (Bilan global) — jamais exposé aux autres enseignants."""
    if session.get('role') != 'admin':
        return error_response('Accès réservé à l\'administrateur', 403)
    rows = get_db().execute('''
        SELECT teacher_id, label, hetd FROM teacher_external_hours
        WHERE is_public = 1 ORDER BY teacher_id, id
    ''').fetchall()
    return jsonify(rows_to_list(rows)), 200

@app.route('/api/teachers/<int:teacher_id>/availability', methods=['GET'])
def get_teacher_availability(teacher_id):
    """Get teacher availability"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Check if teacher exists
        cursor.execute('SELECT * FROM teachers WHERE id = ?', (teacher_id,))
        if not cursor.fetchone():
            return error_response('Teacher not found', 404)
        
        cursor.execute('SELECT * FROM teacher_availability WHERE teacher_id = ? ORDER BY day_of_week, start_time', (teacher_id,))
        availability = rows_to_list(cursor.fetchall())
        return jsonify(availability), 200
    except Exception as e:
        return error_response(f'Error fetching availability: {str(e)}', 500)

@app.route('/api/teachers/<int:teacher_id>/availability', methods=['POST'])
def create_teacher_availability(teacher_id):
    """Add teacher availability"""
    try:
        data = request.get_json()
        db = get_db()
        cursor = db.cursor()
        
        # Check if teacher exists
        cursor.execute('SELECT * FROM teachers WHERE id = ?', (teacher_id,))
        if not cursor.fetchone():
            return error_response('Teacher not found', 404)
        
        cursor.execute('''
            INSERT INTO teacher_availability (teacher_id, day_of_week, start_time, end_time, available, priority)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            teacher_id,
            data['day_of_week'],
            data['start_time'],
            data['end_time'],
            data.get('available', 1),
            data.get('priority', 1)
        ))
        db.commit()
        availability_id = cursor.lastrowid
        cursor.execute('SELECT * FROM teacher_availability WHERE id = ?', (availability_id,))
        availability = row_to_dict(cursor.fetchone())
        return jsonify(availability), 201
    except Exception as e:
        return error_response(f'Error creating availability: {str(e)}', 500)

# ======================= ROOMS CRUD =======================

@app.route('/api/rooms', methods=['GET'])
def get_rooms():
    """Get all rooms"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT * FROM rooms ORDER BY name')
        rooms = rows_to_list(cursor.fetchall())
        return jsonify(rooms), 200
    except Exception as e:
        return error_response(f'Error fetching rooms: {str(e)}', 500)

@app.route('/api/rooms', methods=['POST'])
def create_room():
    """Create a new room"""
    try:
        data = request.get_json()
        if not data or not data.get('name'):
            return error_response('Room name is required')
        
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            INSERT INTO rooms (name, capacity, room_type, location)
            VALUES (?, ?, ?, ?)
        ''', (
            data['name'],
            data.get('capacity'),
            data.get('room_type', 'standard'),
            data.get('location')
        ))
        db.commit()
        room_id = cursor.lastrowid
        cursor.execute('SELECT * FROM rooms WHERE id = ?', (room_id,))
        room = row_to_dict(cursor.fetchone())
        return jsonify(room), 201
    except sqlite3.IntegrityError:
        return error_response('Room name already exists')
    except Exception as e:
        return error_response(f'Error creating room: {str(e)}', 500)

@app.route('/api/rooms/<int:room_id>', methods=['GET'])
def get_room(room_id):
    """Get a specific room"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT * FROM rooms WHERE id = ?', (room_id,))
        room = cursor.fetchone()
        
        if not room:
            return error_response('Room not found', 404)
        
        return jsonify(row_to_dict(room)), 200
    except Exception as e:
        return error_response(f'Error fetching room: {str(e)}', 500)

@app.route('/api/rooms/<int:room_id>', methods=['PUT'])
def update_room(room_id):
    """Update a room"""
    try:
        data = request.get_json()
        db = get_db()
        cursor = db.cursor()
        
        # Check if room exists
        cursor.execute('SELECT * FROM rooms WHERE id = ?', (room_id,))
        if not cursor.fetchone():
            return error_response('Room not found', 404)
        
        # Update fields
        update_fields = []
        values = []
        for field in ['name', 'capacity', 'room_type', 'location']:
            if field in data:
                update_fields.append(f'{field} = ?')
                values.append(data[field])
        
        if not update_fields:
            return error_response('No fields to update')
        
        values.append(room_id)
        query = f'UPDATE rooms SET {", ".join(update_fields)}, updated_at = CURRENT_TIMESTAMP WHERE id = ?'
        cursor.execute(query, values)
        db.commit()
        
        cursor.execute('SELECT * FROM rooms WHERE id = ?', (room_id,))
        room = row_to_dict(cursor.fetchone())
        return jsonify(room), 200
    except sqlite3.IntegrityError:
        return error_response('Room name already exists')
    except Exception as e:
        return error_response(f'Error updating room: {str(e)}', 500)

@app.route('/api/rooms/<int:room_id>', methods=['DELETE'])
def delete_room(room_id):
    """Delete a room"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Check if room exists
        cursor.execute('SELECT * FROM rooms WHERE id = ?', (room_id,))
        if not cursor.fetchone():
            return error_response('Room not found', 404)
        
        cursor.execute('DELETE FROM rooms WHERE id = ?', (room_id,))
        db.commit()
        return jsonify({'message': 'Room deleted'}), 200
    except Exception as e:
        return error_response(f'Error deleting room: {str(e)}', 500)

# ======================= SEMESTERS CRUD =======================

@app.route('/api/semesters', methods=['GET'])
def get_semesters():
    """Get all semesters"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT * FROM semesters ORDER BY code')
        semesters = rows_to_list(cursor.fetchall())
        return jsonify(semesters), 200
    except Exception as e:
        return error_response(f'Error fetching semesters: {str(e)}', 500)

@app.route('/api/semesters', methods=['POST'])
def create_semester():
    """Create a new semester"""
    try:
        data = request.get_json()
        if not data or not data.get('code') or not data.get('year_group'):
            return error_response('Semester code and year_group are required')
        
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            INSERT INTO semesters (code, year_group, name, start_week, end_week)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            data['code'],
            data['year_group'],
            data.get('name'),
            data.get('start_week'),
            data.get('end_week')
        ))
        db.commit()
        semester_id = cursor.lastrowid
        cursor.execute('SELECT * FROM semesters WHERE id = ?', (semester_id,))
        semester = row_to_dict(cursor.fetchone())
        return jsonify(semester), 201
    except sqlite3.IntegrityError:
        return error_response('Semester code already exists')
    except Exception as e:
        return error_response(f'Error creating semester: {str(e)}', 500)

@app.route('/api/semesters/<int:semester_id>', methods=['GET'])
def get_semester(semester_id):
    """Get a specific semester"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT * FROM semesters WHERE id = ?', (semester_id,))
        semester = cursor.fetchone()
        
        if not semester:
            return error_response('Semester not found', 404)
        
        return jsonify(row_to_dict(semester)), 200
    except Exception as e:
        return error_response(f'Error fetching semester: {str(e)}', 500)

@app.route('/api/semesters/<int:semester_id>', methods=['PUT'])
def update_semester(semester_id):
    """Update a semester"""
    try:
        data = request.get_json()
        db = get_db()
        cursor = db.cursor()
        
        # Check if semester exists
        cursor.execute('SELECT * FROM semesters WHERE id = ?', (semester_id,))
        if not cursor.fetchone():
            return error_response('Semester not found', 404)
        
        # Update fields
        update_fields = []
        values = []
        for field in ['code', 'year_group', 'name', 'start_week', 'end_week']:
            if field in data:
                update_fields.append(f'{field} = ?')
                values.append(data[field])
        
        if not update_fields:
            return error_response('No fields to update')
        
        values.append(semester_id)
        query = f'UPDATE semesters SET {", ".join(update_fields)}, updated_at = CURRENT_TIMESTAMP WHERE id = ?'
        cursor.execute(query, values)
        db.commit()
        
        cursor.execute('SELECT * FROM semesters WHERE id = ?', (semester_id,))
        semester = row_to_dict(cursor.fetchone())
        return jsonify(semester), 200
    except sqlite3.IntegrityError:
        return error_response('Semester code already exists')
    except Exception as e:
        return error_response(f'Error updating semester: {str(e)}', 500)

@app.route('/api/semesters/<int:semester_id>', methods=['DELETE'])
def delete_semester(semester_id):
    """Delete a semester"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Check if semester exists
        cursor.execute('SELECT * FROM semesters WHERE id = ?', (semester_id,))
        if not cursor.fetchone():
            return error_response('Semester not found', 404)
        
        cursor.execute('DELETE FROM semesters WHERE id = ?', (semester_id,))
        db.commit()
        return jsonify({'message': 'Semester deleted'}), 200
    except Exception as e:
        return error_response(f'Error deleting semester: {str(e)}', 500)

@app.route('/api/semesters/<int:sem_id>/special-weeks', methods=['GET'])
def get_special_weeks(sem_id):
    """Return special weeks for a semester: {vacation_ftp: [38,43,...], company_alt: [36,37,...]}"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            'SELECT week_number, week_type FROM semester_special_weeks WHERE semester_id = ? ORDER BY week_type, week_number',
            (sem_id,)
        )
        result = {'vacation_ftp': [], 'company_alt': []}
        for r in cursor.fetchall():
            if r['week_type'] in result:
                result[r['week_type']].append(r['week_number'])
        return jsonify(result), 200
    except Exception as e:
        return error_response(str(e), 500)

@app.route('/api/semesters/<int:sem_id>/special-weeks', methods=['PUT'])
def set_special_weeks(sem_id):
    """Replace all special weeks for a semester.
    Body: {vacation_ftp: [38, 43], company_alt: [36, 37]}
    """
    try:
        data = request.get_json() or {}
        db = get_db()
        cursor = db.cursor()
        cursor.execute('DELETE FROM semester_special_weeks WHERE semester_id = ?', (sem_id,))
        for wtype in ['vacation_ftp', 'company_alt']:
            for wnum in (data.get(wtype) or []):
                cursor.execute(
                    'INSERT OR IGNORE INTO semester_special_weeks (semester_id, week_number, week_type) VALUES (?, ?, ?)',
                    (sem_id, int(wnum), wtype)
                )
        db.commit()
        return jsonify({'message': 'Special weeks updated'}), 200
    except Exception as e:
        return error_response(str(e), 500)

# ======================= SPECIAL CALENDAR =======================

_SPECIAL_TYPES = ['vacation_ftp_y1', 'vacation_ftp_y2', 'vacation_ftp_y3',
                  'stage_ftp_y1', 'stage_ftp_y2', 'stage_ftp_y3',
                  'company_alt_y1', 'company_alt_y2', 'company_alt_y3']

@app.route('/api/special-calendar', methods=['GET'])
def get_special_calendar():
    """Retourne le calendrier spécial global : {vacation_ftp: [], stage_ftp_y2: [], ...}"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT week_number, week_type FROM special_calendar ORDER BY week_type, week_number')
        result = {t: [] for t in _SPECIAL_TYPES}
        for r in cursor.fetchall():
            if r['week_type'] in result:
                result[r['week_type']].append(r['week_number'])
        return jsonify(result), 200
    except Exception as e:
        return error_response(str(e), 500)

@app.route('/api/special-calendar', methods=['PUT'])
def set_special_calendar():
    """Remplace les semaines de chaque catégorie fournie.
    Body: {vacation_ftp: [38,43], stage_ftp_y2: [...], ...}
    """
    try:
        data = request.get_json() or {}
        db = get_db()
        cursor = db.cursor()
        for wtype in _SPECIAL_TYPES:
            if wtype in data:
                cursor.execute('DELETE FROM special_calendar WHERE week_type = ?', (wtype,))
                for wnum in (data.get(wtype) or []):
                    cursor.execute(
                        'INSERT OR IGNORE INTO special_calendar (week_number, week_type) VALUES (?, ?)',
                        (int(wnum), wtype)
                    )
        db.commit()
        return jsonify({'message': 'Calendrier mis à jour'}), 200
    except Exception as e:
        return error_response(str(e), 500)

@app.route('/api/week-comments', methods=['GET'])
def get_week_comments():
    """Retourne les commentaires par semaine :
    {week_number: {comment: str, formations: [..]}}"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT week_number, comment, formations, years FROM week_comments')
        return jsonify({str(r['week_number']): {
            'comment': r['comment'],
            'formations': [f for f in (r['formations'] or 'FTP,ALT').split(',') if f],
            'years': [int(y) for y in (r['years'] or '1,2,3').split(',') if y],
        } for r in cursor.fetchall()}), 200
    except Exception as e:
        return error_response(str(e), 500)

@app.route('/api/week-comments', methods=['PUT'])
def set_week_comments():
    """Remplace l'ensemble des commentaires.
    Body: {week_number: {comment: str, formations: [..]}, ...}
    (accepte aussi l'ancien format {week_number: "texte"}). Vide = suppression."""
    try:
        data = request.get_json() or {}
        db = get_db()
        cursor = db.cursor()
        cursor.execute('DELETE FROM week_comments')
        for wnum, val in data.items():
            if isinstance(val, dict):
                text = (val.get('comment') or '').strip()
                forms = [f for f in (val.get('formations') or []) if f in ('FTP', 'ALT')]
                yrs = [str(y) for y in (val.get('years') or []) if str(y) in ('1', '2', '3')]
            else:
                text = (val or '').strip()
                forms, yrs = [], []
            formations = ','.join(forms) if forms else 'FTP,ALT'
            years = ','.join(yrs) if yrs else '1,2,3'
            if text:
                cursor.execute(
                    'INSERT OR REPLACE INTO week_comments (week_number, comment, formations, years) VALUES (?, ?, ?, ?)',
                    (int(wnum), text, formations, years)
                )
        db.commit()
        return jsonify({'message': 'Commentaires mis à jour'}), 200
    except Exception as e:
        return error_response(str(e), 500)

@app.route('/api/category-codes', methods=['GET'])
def get_category_codes():
    """Retourne le code texte par catégorie : {week_type: code}"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT week_type, code FROM special_category_codes')
        return jsonify({r['week_type']: r['code'] for r in cursor.fetchall()}), 200
    except Exception as e:
        return error_response(str(e), 500)

@app.route('/api/category-codes', methods=['PUT'])
def set_category_codes():
    """Remplace les codes des catégories. Body: {week_type: code, ...}
    Une valeur vide supprime le code."""
    try:
        data = request.get_json() or {}
        db = get_db()
        cursor = db.cursor()
        cursor.execute('DELETE FROM special_category_codes')
        for wtype, code in data.items():
            if wtype not in _SPECIAL_TYPES:
                continue
            text = (code or '').strip()
            if text:
                cursor.execute(
                    'INSERT OR REPLACE INTO special_category_codes (week_type, code) VALUES (?, ?)',
                    (wtype, text)
                )
        db.commit()
        return jsonify({'message': 'Codes mis à jour'}), 200
    except Exception as e:
        return error_response(str(e), 500)

# ======================= COURSES CRUD =======================

def _mat_base_key(code):
    """Clé de regroupement matière (= matBaseKey côté front) : enlève une lettre
    finale (R1.03a→R1.03) et normalise le dernier groupe numérique sur 2 chiffres
    (SAE1.1→SAE1.01)."""
    c = re.sub(r'[a-z]$', '', code or '')
    return re.sub(r'\.(\d+)$', lambda m: '.' + str(int(m.group(1))).zfill(2), c)

def _matiere_referents(db):
    """Référent par MATIÈRE (groupe de sous-matières partageant la même clé) et
    par face (FTP/ALT). Intervenants d'une face = enseignants des sessions
    actives de toutes les sous-matières (FTP : formation_type 0 et 2 ; ALT : 1
    et 2). Choix admin (matiere_referents) prioritaire tant que l'enseignant
    intervient encore ; sinon un titulaire UNIQUE parmi les intervenants →
    référent automatique ; sinon à définir."""
    faces = {}   # (matiere_key, 'FTP'/'ALT') -> {teacher_id: {id, name, titulaire}}
    for r in db.execute('''
            SELECT c.code, cs.formation_type, t.id AS tid, t.name, t.status
            FROM course_sessions cs
            JOIN teachers t ON cs.teacher_id = t.id
            JOIN courses c ON c.id = cs.course_id
            WHERE cs.nb_sessions > 0 OR cs.total_hours > 0'''):
        key = _mat_base_key(r['code'])
        for face in (('FTP',) if r['formation_type'] == 0 else
                     ('ALT',) if r['formation_type'] == 1 else
                     ('FTP', 'ALT') if r['formation_type'] == 2 else ()):
            faces.setdefault((key, face), {})[r['tid']] = {
                'id': r['tid'], 'name': r['name'],
                'titulaire': (r['status'] or 'Titulaire') == 'Titulaire'}
    manual = {(r['matiere_key'], r['formation']): r['teacher_id'] for r in
              db.execute('SELECT matiere_key, formation, teacher_id FROM matiere_referents')}
    out = {}
    for (key, face), teachers in faces.items():
        tits = [t for t in teachers.values() if t['titulaire']]
        chosen = teachers.get(manual.get((key, face)))
        if chosen:
            ref = {'teacher_id': chosen['id'], 'name': chosen['name'], 'auto': False}
        elif len(tits) == 1:
            ref = {'teacher_id': tits[0]['id'], 'name': tits[0]['name'], 'auto': True}
        else:
            ref = {'teacher_id': None, 'name': None, 'auto': False}
        ref['options'] = sorted(
            ({'id': t['id'], 'name': t['name'], 'titulaire': t['titulaire']}
             for t in teachers.values()),
            key=lambda x: x['name'])
        out.setdefault(key, {})[face] = ref
    return out

@app.route('/api/matiere-referents', methods=['PUT'])
def set_matiere_referent():
    """Choix admin du référent d'une face (FTP/ALT) d'une matière — prioritaire
    sur la présélection automatique. Body {matiere_key, formation, teacher_id} ;
    teacher_id vide → efface le choix (retour à la règle automatique)."""
    err = _require_admin()
    if err:
        return err
    data = request.get_json() or {}
    key = (data.get('matiere_key') or '').strip()
    formation = (data.get('formation') or '').upper()
    if not key:
        return error_response('matiere_key requis')
    if formation not in ('FTP', 'ALT'):
        return error_response("formation doit valoir 'FTP' ou 'ALT'")
    db = get_db()
    tid = data.get('teacher_id')
    if tid in ('', None):
        db.execute('DELETE FROM matiere_referents WHERE matiere_key = ? AND formation = ?',
                   (key, formation))
    else:
        try:
            tid = int(tid)
        except (TypeError, ValueError):
            return error_response('teacher_id invalide')
        fts = (0, 2) if formation == 'FTP' else (1, 2)
        # L'enseignant doit intervenir sur cette face d'une sous-matière du groupe.
        rows = db.execute('''SELECT DISTINCT c.code FROM course_sessions cs
                             JOIN courses c ON c.id = cs.course_id
                             WHERE cs.teacher_id = ? AND cs.formation_type IN (?, ?)
                               AND (cs.nb_sessions > 0 OR cs.total_hours > 0)''',
                          (tid, *fts)).fetchall()
        if not any(_mat_base_key(r['code']) == key for r in rows):
            return error_response(f"Cet enseignant n'intervient pas sur la face {formation} de cette matière")
        db.execute('''INSERT INTO matiere_referents (matiere_key, formation, teacher_id)
                      VALUES (?, ?, ?)
                      ON CONFLICT (matiere_key, formation) DO UPDATE SET
                          teacher_id = excluded.teacher_id''',
                   (key, formation, tid))
    db.commit()
    return jsonify({'message': 'Référent enregistré'})

@app.route('/api/courses', methods=['GET'])
def get_courses():
    """Get all courses with hours and rooms per teaching type"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            SELECT c.id, c.code, c.name, c.semester_id, c.course_type,
                   c.start_week, c.end_week, c.default_weeks, c.mutualized, c.tp_type,
                   c.content, c.content_pn, c.created_at, c.updated_at,
                   s.code as semester_code,
                   GROUP_CONCAT(DISTINCT t.name) as teacher_names,
                   COALESCE(SUM(CASE WHEN cs.teaching_type='CM' AND cs.formation_type IN (0,2)
                                     THEN cs.total_hours ELSE 0 END), 0) as cm_hours,
                   MAX(CASE WHEN cs.teaching_type='CM' AND cs.formation_type IN (0,2)
                            THEN cs.slot_duration END) as cm_slot_duration,
                   MAX(CASE WHEN cs.teaching_type='CM' AND cs.formation_type IN (0,2)
                            THEN cs.room_name END) as cm_room,
                   MAX(CASE WHEN cs.teaching_type='CM' AND cs.formation_type IN (0,2)
                            THEN t.name END) as cm_teacher,
                   COALESCE(SUM(CASE WHEN cs.teaching_type='TD' AND cs.formation_type IN (0,2)
                                     THEN cs.total_hours ELSE 0 END), 0) as td_hours,
                   MAX(CASE WHEN cs.teaching_type='TD' AND cs.formation_type IN (0,2)
                            THEN cs.slot_duration END) as td_slot_duration,
                   MAX(CASE WHEN cs.teaching_type='TD' AND cs.formation_type IN (0,2)
                            THEN cs.room_name END) as td_room,
                   MAX(CASE WHEN cs.teaching_type='TD' AND cs.formation_type IN (0,2)
                            THEN t.name END) as td_teacher,
                   COALESCE(SUM(CASE WHEN cs.teaching_type='TP' AND cs.formation_type IN (0,2)
                                     THEN cs.total_hours ELSE 0 END), 0) as tp_hours,
                   MAX(CASE WHEN cs.teaching_type='TP' AND cs.formation_type IN (0,2)
                            THEN cs.slot_duration END) as tp_slot_duration,
                   MAX(CASE WHEN cs.teaching_type='TP' AND cs.formation_type IN (0,2)
                            THEN cs.room_name END) as tp_room,
                   MAX(CASE WHEN cs.teaching_type='TP' AND cs.formation_type IN (0,2)
                            THEN t.name END) as tp_teacher,
                   COALESCE(SUM(CASE WHEN cs.teaching_type='PT' AND cs.formation_type=0
                                     THEN cs.total_hours ELSE 0 END), 0) as pt_hours,
                   MAX(CASE WHEN cs.teaching_type='PT' AND cs.formation_type=0
                            THEN cs.slot_duration END) as pt_slot_duration,
                   MAX(CASE WHEN cs.teaching_type='PT' AND cs.formation_type=0
                            THEN cs.room_name END) as pt_room,
                   MAX(CASE WHEN cs.teaching_type='PT' AND cs.formation_type=0
                            THEN t.name END) as pt_teacher,
                   COALESCE(SUM(CASE WHEN cs.teaching_type='CM' AND cs.formation_type=1
                                     THEN cs.total_hours ELSE 0 END), 0) as alt_cm_hours,
                   MAX(CASE WHEN cs.teaching_type='CM' AND cs.formation_type=1
                            THEN cs.slot_duration END) as alt_cm_slot_duration,
                   MAX(CASE WHEN cs.teaching_type='CM' AND cs.formation_type=1
                            THEN cs.room_name END) as alt_cm_room,
                   MAX(CASE WHEN cs.teaching_type='CM' AND cs.formation_type=1
                            THEN t.name END) as alt_cm_teacher,
                   COALESCE(SUM(CASE WHEN cs.teaching_type='TD' AND cs.formation_type=1
                                     THEN cs.total_hours ELSE 0 END), 0) as alt_td_hours,
                   MAX(CASE WHEN cs.teaching_type='TD' AND cs.formation_type=1
                            THEN cs.slot_duration END) as alt_td_slot_duration,
                   MAX(CASE WHEN cs.teaching_type='TD' AND cs.formation_type=1
                            THEN cs.room_name END) as alt_td_room,
                   MAX(CASE WHEN cs.teaching_type='TD' AND cs.formation_type=1
                            THEN t.name END) as alt_td_teacher,
                   COALESCE(SUM(CASE WHEN cs.teaching_type='TP' AND cs.formation_type=1
                                     THEN cs.total_hours ELSE 0 END), 0) as alt_tp_hours,
                   MAX(CASE WHEN cs.teaching_type='TP' AND cs.formation_type=1
                            THEN cs.slot_duration END) as alt_tp_slot_duration,
                   MAX(CASE WHEN cs.teaching_type='TP' AND cs.formation_type=1
                            THEN cs.room_name END) as alt_tp_room,
                   MAX(CASE WHEN cs.teaching_type='TP' AND cs.formation_type=1
                            THEN t.name END) as alt_tp_teacher,
                   COALESCE(SUM(CASE WHEN cs.teaching_type='PT' AND cs.formation_type=1
                                     THEN cs.total_hours ELSE 0 END), 0) as alt_pt_hours,
                   MAX(CASE WHEN cs.teaching_type='PT' AND cs.formation_type=1
                            THEN cs.slot_duration END) as alt_pt_slot_duration,
                   MAX(CASE WHEN cs.teaching_type='PT' AND cs.formation_type=1
                            THEN cs.room_name END) as alt_pt_room,
                   MAX(CASE WHEN cs.teaching_type='PT' AND cs.formation_type=1
                            THEN t.name END) as alt_pt_teacher,
                   COUNT(DISTINCT cs.id) as session_count
            FROM courses c
            JOIN semesters s ON c.semester_id = s.id
            LEFT JOIN course_sessions cs ON cs.course_id = c.id
            LEFT JOIN teachers t ON cs.teacher_id = t.id
            GROUP BY c.id, c.code, c.name, c.semester_id, c.course_type,
                     c.start_week, c.end_week, c.default_weeks, c.mutualized, c.tp_type,
                     c.content, c.content_pn, c.created_at, c.updated_at, s.code
            ORDER BY s.code, c.code
        ''')
        courses = rows_to_list(cursor.fetchall())
        # Résolution dynamique des semaines pour les cours en "dates par défaut"
        cfg = get_year_config()
        refs = _matiere_referents(db)
        for c in courses:
            sw, ew = resolve_course_weeks(c, cfg)
            c['start_week'], c['end_week'] = sw, ew
            # Référents au niveau matière : identiques pour toutes les
            # sous-matières d'un même groupe (clé _mat_base_key).
            c['referents'] = refs.get(_mat_base_key(c['code']), {})
        return jsonify(courses), 200
    except Exception as e:
        return error_response(f'Error fetching courses: {str(e)}', 500)

@app.route('/api/courses', methods=['POST'])
def create_course():
    """Create a new course with optional teaching hours per type"""
    try:
        data = request.get_json()
        if not data or not data.get('code') or not data.get('name') or not data.get('semester_id'):
            return error_response('code, name and semester_id are required')

        db = get_db()
        cursor = db.cursor()

        cursor.execute('SELECT * FROM semesters WHERE id = ?', (data['semester_id'],))
        if not cursor.fetchone():
            return error_response('Semester not found', 404)

        default_weeks = 1 if data.get('default_weeks') else 0
        # En mode "dates par défaut", les semaines sont dynamiques → ne rien figer
        start_week = None if default_weeks else data.get('start_week')
        end_week   = None if default_weeks else data.get('end_week')
        cfg = get_year_config()
        valid_weeks = get_valid_school_weeks(cfg['academic_start_week'], cfg['academic_end_week'], cfg['academic_max_week'])
        sw_label = f"S{cfg['academic_start_week']:02d}–S{cfg['academic_end_week']:02d}"
        if start_week is not None and int(start_week) not in valid_weeks:
            return error_response(f'Semaine de début hors plage année scolaire ({sw_label})', 400)
        if end_week is not None and int(end_week) not in valid_weeks:
            return error_response(f'Semaine de fin hors plage année scolaire ({sw_label})', 400)

        cursor.execute('''
            INSERT INTO courses (code, name, semester_id, course_type, start_week, end_week, default_weeks, mutualized, tp_type)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['code'],
            data['name'],
            data['semester_id'],
            data.get('course_type', 'Ressource'),
            start_week,
            end_week,
            default_weeks,
            1 if data.get('mutualized') else 0,
            'TP8' if str(data.get('tp_type', '')).upper() == 'TP8' else 'TP12',
        ))
        db.commit()
        course_id = cursor.lastrowid

        # Les sessions (FTP/ALT) sont créées via PUT /teaching-hours appelé juste après

        cursor.execute('''
            SELECT c.*, s.code as semester_code
            FROM courses c JOIN semesters s ON c.semester_id = s.id
            WHERE c.id = ?
        ''', (course_id,))
        course = row_to_dict(cursor.fetchone())
        return jsonify(course), 201
    except sqlite3.IntegrityError:
        return error_response('Un cours avec ce code existe déjà pour ce semestre')
    except Exception as e:
        return error_response(f'Error creating course: {str(e)}', 500)

@app.route('/api/courses/<int:course_id>', methods=['GET'])
def get_course(course_id):
    """Get a specific course"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            SELECT c.*, s.code as semester_code 
            FROM courses c 
            JOIN semesters s ON c.semester_id = s.id 
            WHERE c.id = ?
        ''', (course_id,))
        course = cursor.fetchone()
        
        if not course:
            return error_response('Course not found', 404)
        
        return jsonify(row_to_dict(course)), 200
    except Exception as e:
        return error_response(f'Error fetching course: {str(e)}', 500)

@app.route('/api/courses/<int:course_id>/content', methods=['PUT'])
def update_course_content(course_id):
    """Met à jour la description / contenu pédagogique d'une matière.
    - `content` (contenu Enseignant) : éditable par l'admin ou un enseignant intervenant.
    - `content_pn` (Programme national) : éditable par l'admin uniquement."""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT content, content_pn FROM courses WHERE id = ?', (course_id,))
        row = cursor.fetchone()
        if not row:
            return error_response('Course not found', 404)

        role = session.get('role')
        if role != 'admin':
            teacher_name = session.get('teacher_name')
            if not teacher_intervenes_in_course(course_id, teacher_name):
                return error_response("Vous n'intervenez pas dans cette matière", 403)

        data = request.get_json() or {}

        def _norm(v):
            return v.strip() or None if isinstance(v, str) else None

        content = row['content']
        if 'content' in data:
            content = _norm(data.get('content'))

        content_pn = row['content_pn']
        if 'content_pn' in data:
            if role != 'admin':
                return error_response("Seul l'administrateur peut modifier le Programme national", 403)
            content_pn = _norm(data.get('content_pn'))

        cursor.execute(
            'UPDATE courses SET content = ?, content_pn = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
            (content, content_pn, course_id)
        )
        db.commit()
        return jsonify({'id': course_id, 'content': content or '', 'content_pn': content_pn or ''}), 200
    except Exception as e:
        return error_response(f'Error updating course content: {str(e)}', 500)

@app.route('/api/courses/<int:course_id>', methods=['PUT'])
def update_course(course_id):
    """Update a course"""
    try:
        data = request.get_json()
        db = get_db()
        cursor = db.cursor()
        
        # Check if course exists
        cursor.execute('SELECT * FROM courses WHERE id = ?', (course_id,))
        if not cursor.fetchone():
            return error_response('Course not found', 404)
        
        # En mode "dates par défaut", les semaines sont dynamiques → ne rien figer
        if 'default_weeks' in data and data.get('default_weeks'):
            data['start_week'] = None
            data['end_week']   = None

        # Validation semaines scolaires
        cfg = get_year_config()
        valid_weeks = get_valid_school_weeks(cfg['academic_start_week'], cfg['academic_end_week'], cfg['academic_max_week'])
        sw_label = f"S{cfg['academic_start_week']:02d}–S{cfg['academic_end_week']:02d}"
        for wk_field in ['start_week', 'end_week']:
            if wk_field in data and data[wk_field] is not None:
                if int(data[wk_field]) not in valid_weeks:
                    label = 'début' if wk_field == 'start_week' else 'fin'
                    return error_response(f'Semaine de {label} hors plage année scolaire ({sw_label})', 400)

        # Normaliser le type de TP (TP12 par défaut)
        if 'tp_type' in data:
            data['tp_type'] = 'TP8' if str(data.get('tp_type') or '').upper() == 'TP8' else 'TP12'

        # Update fields
        update_fields = []
        values = []
        for field in ['code', 'name', 'semester_id', 'course_type', 'start_week', 'end_week', 'default_weeks', 'mutualized', 'tp_type']:
            if field in data:
                update_fields.append(f'{field} = ?')
                values.append(data[field])
        
        if not update_fields:
            return error_response('No fields to update')
        
        values.append(course_id)
        query = f'UPDATE courses SET {", ".join(update_fields)}, updated_at = CURRENT_TIMESTAMP WHERE id = ?'
        cursor.execute(query, values)
        db.commit()
        
        cursor.execute('''
            SELECT c.*, s.code as semester_code 
            FROM courses c 
            JOIN semesters s ON c.semester_id = s.id 
            WHERE c.id = ?
        ''', (course_id,))
        course = row_to_dict(cursor.fetchone())
        return jsonify(course), 200
    except Exception as e:
        return error_response(f'Error updating course: {str(e)}', 500)

@app.route('/api/courses/<int:course_id>', methods=['DELETE'])
def delete_course(course_id):
    """Delete a course (cascade deletes its sessions)"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT * FROM courses WHERE id = ?', (course_id,))
        if not cursor.fetchone():
            return error_response('Course not found', 404)
        cursor.execute('DELETE FROM courses WHERE id = ?', (course_id,))
        db.commit()
        return jsonify({'message': 'Course deleted'}), 200
    except Exception as e:
        return error_response(f'Error deleting course: {str(e)}', 500)

@app.route('/api/courses/<int:course_id>/teaching-hours', methods=['GET'])
def get_course_teaching_hours(course_id):
    """Get hours, slot_duration and room per teaching type.
    Query param: formation_type (0=FTP default, 1=ALT)
    """
    try:
        formation_type = int(request.args.get('formation_type', 0))
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT id FROM courses WHERE id = ?', (course_id,))
        if not cursor.fetchone():
            return error_response('Course not found', 404)
        cursor.execute('''
            SELECT teaching_type, total_hours, slot_duration, room_name, teacher_id,
                   COALESCE(sessions_per_week_max, 1) as sessions_per_week_max
            FROM course_sessions
            WHERE course_id = ? AND formation_type = ?
        ''', (course_id, formation_type))
        rows = cursor.fetchall()
        result = {t: {'hours': 0, 'slot_duration': 1.5, 'room': '', 'teacher_id': None, 'sessions_per_week_max': 1} for t in ['CM', 'TD', 'TP', 'PT']}
        for r in rows:
            t = r['teaching_type']
            if t in result:
                result[t] = {
                    'hours': r['total_hours'] or 0,
                    'slot_duration': r['slot_duration'] or 1.5,
                    'room': r['room_name'] or '',
                    'teacher_id': r['teacher_id'],
                    'sessions_per_week_max': r['sessions_per_week_max'] or 1,
                }
        return jsonify(result), 200
    except Exception as e:
        return error_response(str(e), 500)

@app.route('/api/courses/<int:course_id>/teaching-hours', methods=['PUT'])
def update_course_teaching_hours(course_id):
    """Upsert hours, slot_duration and room per teaching type.
    Query param: formation_type (0=FTP default, 1=ALT)
    Body: {"CM": {"hours": 10, "slot_duration": 1.5, "room": "Amphi A"}, ...}
    """
    try:
        formation_type = int(request.args.get('formation_type', 0))
        data = request.get_json()
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT id FROM courses WHERE id = ?', (course_id,))
        if not cursor.fetchone():
            return error_response('Course not found', 404)

        for ttype in ['CM', 'TD', 'TP', 'PT']:
            info = (data or {}).get(ttype, {})
            hours = float(info.get('hours') or 0)
            slot_dur = float(info.get('slot_duration') or 1.5)
            room = (info.get('room') or '').strip() or None
            teacher_id = info.get('teacher_id') or None
            spw = max(1, int(info.get('sessions_per_week_max') or 1))
            nb = round(hours / slot_dur) if slot_dur > 0 and hours > 0 else 0

            if hours > 0:
                cursor.execute('''
                    INSERT INTO course_sessions
                        (course_id, formation_type, teaching_type,
                         total_hours, slot_duration, nb_sessions, room_name, teacher_id,
                         sessions_per_week_max)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(course_id, formation_type, teaching_type) DO UPDATE SET
                        total_hours = excluded.total_hours,
                        slot_duration = excluded.slot_duration,
                        nb_sessions = excluded.nb_sessions,
                        room_name = excluded.room_name,
                        teacher_id = excluded.teacher_id,
                        sessions_per_week_max = excluded.sessions_per_week_max,
                        updated_at = CURRENT_TIMESTAMP
                ''', (course_id, formation_type, ttype, hours, slot_dur, nb, room, teacher_id, spw))
            else:
                # Nettoyer les weekly_hours avant de supprimer la session
                cursor.execute('''
                    DELETE FROM weekly_hours WHERE course_session_id IN (
                        SELECT id FROM course_sessions
                        WHERE course_id = ? AND teaching_type = ? AND formation_type = ?
                    )
                ''', (course_id, ttype, formation_type))
                cursor.execute('''
                    DELETE FROM course_sessions
                    WHERE course_id = ? AND teaching_type = ? AND formation_type = ?
                ''', (course_id, ttype, formation_type))

        db.commit()
        return jsonify({'message': 'Teaching hours updated'}), 200
    except Exception as e:
        return error_response(str(e), 500)

@app.route('/api/courses/<int:course_id>/weekly-distribution', methods=['GET'])
def get_weekly_distribution(course_id):
    """Get weekly hours distribution by teaching type.
    Query param: formation_type (0=FTP default, 1=ALT)
    Returns: {CM: {36: 1.5, 37: 1.5}, TD: {...}, TP: {}, PT: {}}
    """
    try:
        formation_type = int(request.args.get('formation_type', 0))
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT id FROM courses WHERE id = ?', (course_id,))
        if not cursor.fetchone():
            return error_response('Course not found', 404)
        cursor.execute('''
            SELECT cs.teaching_type, wh.week_number, wh.hours
            FROM course_sessions cs
            JOIN weekly_hours wh ON wh.course_session_id = cs.id
            WHERE cs.course_id = ? AND cs.formation_type = ? AND wh.group_index = 1
            ORDER BY cs.teaching_type, wh.week_number
        ''', (course_id, formation_type))
        rows = cursor.fetchall()
        result = {t: {} for t in ['CM', 'TD', 'TP', 'PT']}
        for r in rows:
            t = r['teaching_type']
            if t in result:
                result[t][r['week_number']] = r['hours']
        return jsonify(result), 200
    except Exception as e:
        return error_response(str(e), 500)

@app.route('/api/courses/<int:course_id>/weekly-distribution', methods=['PUT'])
def update_weekly_distribution(course_id):
    """Save weekly hours per teaching type.
    Query param: formation_type (0=FTP default, 1=ALT)
    Body: {CM: {"36": 1.5, "37": 1.5}, TD: {"36": 2}, ...}
    """
    try:
        formation_type = int(request.args.get('formation_type', 0))
        data = request.get_json() or {}
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT id FROM courses WHERE id = ?', (course_id,))
        if not cursor.fetchone():
            return error_response('Course not found', 404)

        for ttype in ['CM', 'TD', 'TP', 'PT']:
            week_data = data.get(ttype, {})
            cursor.execute('''
                SELECT id FROM course_sessions
                WHERE course_id = ? AND teaching_type = ? AND formation_type = ?
            ''', (course_id, ttype, formation_type))
            session = cursor.fetchone()
            if not session:
                continue
            session_id = session['id']
            # Groupe 1 uniquement : préserve le placement par groupe des TP
            # (groupes 2..N édités dans la répartition calendaire).
            cursor.execute('DELETE FROM weekly_hours WHERE course_session_id = ? AND group_index = 1', (session_id,))
            for week_str, hours in week_data.items():
                h = float(hours or 0)
                if h > 0:
                    cursor.execute('''
                        INSERT INTO weekly_hours (course_session_id, week_number, hours, group_index)
                        VALUES (?, ?, ?, 1)
                    ''', (session_id, int(week_str), h))

        db.commit()
        return jsonify({'message': 'Weekly distribution updated'}), 200
    except Exception as e:
        return error_response(str(e), 500)

# ======================= COURSE SESSIONS CRUD =======================

@app.route('/api/course-sessions', methods=['GET'])
def get_course_sessions():
    """Get all course sessions"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            SELECT cs.*, c.code as course_code, c.name as course_name,
                   c.semester_id, c.tp_type, s.code as semester_code, s.year_group,
                   t.name as teacher_name
            FROM course_sessions cs
            JOIN courses c ON cs.course_id = c.id
            JOIN semesters s ON c.semester_id = s.id
            LEFT JOIN teachers t ON cs.teacher_id = t.id
            ORDER BY s.code, c.code, cs.formation_type,
                     CASE cs.teaching_type WHEN 'CM' THEN 0 WHEN 'TD' THEN 1 WHEN 'TP' THEN 2 WHEN 'PT' THEN 3 ELSE 4 END
        ''')
        sessions = rows_to_list(cursor.fetchall())
        return jsonify(sessions), 200
    except Exception as e:
        return error_response(f'Error fetching course sessions: {str(e)}', 500)

@app.route('/api/course-sessions', methods=['POST'])
def create_course_session():
    """Create a new course session"""
    try:
        data = request.get_json()
        if not data or not data.get('course_id') or not data.get('teaching_type') or data.get('formation_type') is None:
            return error_response('course_id, teaching_type, and formation_type are required')
        
        db = get_db()
        cursor = db.cursor()
        
        # Check if course exists
        cursor.execute('SELECT * FROM courses WHERE id = ?', (data['course_id'],))
        if not cursor.fetchone():
            return error_response('Course not found', 404)
        
        cursor.execute('''
            INSERT INTO course_sessions (course_id, teacher_id, formation_type, teaching_type, 
                                        nb_sessions, total_hours, slot_duration, room_name, promo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['course_id'],
            data.get('teacher_id'),
            data['formation_type'],
            data['teaching_type'],
            data.get('nb_sessions', 0),
            data.get('total_hours', 0),
            data.get('slot_duration', 1.5),
            data.get('room_name'),
            data.get('promo')
        ))
        db.commit()
        session_id = cursor.lastrowid
        cursor.execute('''
            SELECT cs.*, c.code as course_code, c.name as course_name, t.name as teacher_name
            FROM course_sessions cs
            JOIN courses c ON cs.course_id = c.id
            LEFT JOIN teachers t ON cs.teacher_id = t.id
            WHERE cs.id = ?
        ''', (session_id,))
        session = row_to_dict(cursor.fetchone())
        return jsonify(session), 201
    except Exception as e:
        return error_response(f'Error creating course session: {str(e)}', 500)

@app.route('/api/course-sessions/<int:session_id>', methods=['GET'])
def get_course_session(session_id):
    """Get a specific course session"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            SELECT cs.*, c.code as course_code, c.name as course_name, t.name as teacher_name
            FROM course_sessions cs
            JOIN courses c ON cs.course_id = c.id
            LEFT JOIN teachers t ON cs.teacher_id = t.id
            WHERE cs.id = ?
        ''', (session_id,))
        session = cursor.fetchone()
        
        if not session:
            return error_response('Course session not found', 404)
        
        return jsonify(row_to_dict(session)), 200
    except Exception as e:
        return error_response(f'Error fetching course session: {str(e)}', 500)

@app.route('/api/course-sessions/<int:session_id>', methods=['PUT'])
def update_course_session(session_id):
    """Update a course session"""
    try:
        data = request.get_json()
        db = get_db()
        cursor = db.cursor()
        
        # Check if session exists
        cursor.execute('SELECT * FROM course_sessions WHERE id = ?', (session_id,))
        if not cursor.fetchone():
            return error_response('Course session not found', 404)
        
        # Update fields
        update_fields = []
        values = []
        for field in ['course_id', 'teacher_id', 'formation_type', 'teaching_type', 'nb_sessions', 
                      'total_hours', 'slot_duration', 'room_name', 'promo']:
            if field in data:
                update_fields.append(f'{field} = ?')
                values.append(data[field])
        
        if not update_fields:
            return error_response('No fields to update')
        
        values.append(session_id)
        query = f'UPDATE course_sessions SET {", ".join(update_fields)}, updated_at = CURRENT_TIMESTAMP WHERE id = ?'
        cursor.execute(query, values)
        db.commit()
        
        cursor.execute('''
            SELECT cs.*, c.code as course_code, c.name as course_name, t.name as teacher_name
            FROM course_sessions cs
            JOIN courses c ON cs.course_id = c.id
            LEFT JOIN teachers t ON cs.teacher_id = t.id
            WHERE cs.id = ?
        ''', (session_id,))
        session = row_to_dict(cursor.fetchone())
        return jsonify(session), 200
    except Exception as e:
        return error_response(f'Error updating course session: {str(e)}', 500)

@app.route('/api/course-sessions/<int:session_id>', methods=['DELETE'])
def delete_course_session(session_id):
    """Delete a course session"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Check if session exists
        cursor.execute('SELECT * FROM course_sessions WHERE id = ?', (session_id,))
        if not cursor.fetchone():
            return error_response('Course session not found', 404)
        
        cursor.execute('DELETE FROM weekly_hours WHERE course_session_id = ?', (session_id,))
        cursor.execute('DELETE FROM course_sessions WHERE id = ?', (session_id,))
        db.commit()
        return jsonify({'message': 'Course session deleted'}), 200
    except Exception as e:
        return error_response(f'Error deleting course session: {str(e)}', 500)

@app.route('/api/course-sessions/by-course/<int:course_id>', methods=['GET'])
def get_course_sessions_by_course(course_id):
    """Get all sessions for a course"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Check if course exists
        cursor.execute('SELECT * FROM courses WHERE id = ?', (course_id,))
        if not cursor.fetchone():
            return error_response('Course not found', 404)
        
        cursor.execute('''
            SELECT cs.*, c.code as course_code, c.name as course_name, t.name as teacher_name
            FROM course_sessions cs
            JOIN courses c ON cs.course_id = c.id
            LEFT JOIN teachers t ON cs.teacher_id = t.id
            WHERE cs.course_id = ?
            ORDER BY CASE cs.teaching_type WHEN 'CM' THEN 0 WHEN 'TD' THEN 1 WHEN 'TP' THEN 2 WHEN 'PT' THEN 3 ELSE 4 END
        ''', (course_id,))
        sessions = rows_to_list(cursor.fetchall())
        return jsonify(sessions), 200
    except Exception as e:
        return error_response(f'Error fetching course sessions: {str(e)}', 500)

# ======================= WEEKLY HOURS =======================

@app.route('/api/weekly-hours/<int:session_id>', methods=['GET'])
def get_weekly_hours(session_id):
    """Get weekly hours for a session"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Check if session exists
        cursor.execute('SELECT * FROM course_sessions WHERE id = ?', (session_id,))
        if not cursor.fetchone():
            return error_response('Course session not found', 404)
        
        cursor.execute('''
            SELECT * FROM weekly_hours 
            WHERE course_session_id = ? 
            ORDER BY week_number
        ''', (session_id,))
        hours = rows_to_list(cursor.fetchall())
        return jsonify(hours), 200
    except Exception as e:
        return error_response(f'Error fetching weekly hours: {str(e)}', 500)

@app.route('/api/weekly-hours/<int:session_id>', methods=['POST'])
def create_weekly_hours(session_id):
    """Create weekly hours entry"""
    try:
        data = request.get_json()
        if not data or data.get('week_number') is None or data.get('hours') is None:
            return error_response('week_number and hours are required')
        
        db = get_db()
        cursor = db.cursor()
        
        # Check if session exists
        cursor.execute('SELECT * FROM course_sessions WHERE id = ?', (session_id,))
        if not cursor.fetchone():
            return error_response('Course session not found', 404)
        
        cursor.execute('''
            INSERT INTO weekly_hours (course_session_id, week_number, semester_week, hours)
            VALUES (?, ?, ?, ?)
        ''', (
            session_id,
            data['week_number'],
            data.get('semester_week'),
            data['hours']
        ))
        db.commit()
        hours_id = cursor.lastrowid
        cursor.execute('SELECT * FROM weekly_hours WHERE id = ?', (hours_id,))
        hours = row_to_dict(cursor.fetchone())
        return jsonify(hours), 201
    except sqlite3.IntegrityError:
        return error_response('Weekly hours entry already exists for this session and week')
    except Exception as e:
        return error_response(f'Error creating weekly hours: {str(e)}', 500)

@app.route('/api/weekly-hours/batch', methods=['PUT'])
def update_weekly_hours_batch():
    """Batch update weekly hours"""
    try:
        data = request.get_json()
        if not data or not isinstance(data, list):
            return error_response('Request body must be a list of entries')
        
        db = get_db()
        cursor = db.cursor()
        updated_count = 0
        
        for entry in data:
            if not entry.get('id'):
                continue
            
            cursor.execute('''
                UPDATE weekly_hours 
                SET hours = ?, semester_week = ?
                WHERE id = ?
            ''', (
                entry.get('hours', 0),
                entry.get('semester_week'),
                entry['id']
            ))
            updated_count += 1
        
        db.commit()
        return jsonify({'updated': updated_count}), 200
    except Exception as e:
        return error_response(f'Error updating weekly hours: {str(e)}', 500)

@app.route('/api/weekly-hours/<int:session_id>/<int:week_number>', methods=['PUT'])
def upsert_weekly_hours(session_id, week_number):
    """Upsert a single weekly_hours cell (hours > 0 → insert/replace, hours == 0 → delete)"""
    try:
        data = request.get_json() or {}
        hours = float(data.get('hours', 0))
        group_index = int(data.get('group_index', 1) or 1)
        db = get_db()
        cursor = db.cursor()
        if hours > 0:
            cursor.execute('''
                INSERT OR REPLACE INTO weekly_hours (course_session_id, week_number, hours, group_index)
                VALUES (?, ?, ?, ?)
            ''', (session_id, week_number, hours, group_index))
        else:
            cursor.execute(
                'DELETE FROM weekly_hours WHERE course_session_id = ? AND week_number = ? AND group_index = ?',
                (session_id, week_number, group_index)
            )
            hours = 0
        db.commit()
        return jsonify({'session_id': session_id, 'week_number': week_number,
                        'group_index': group_index, 'hours': hours}), 200
    except Exception as e:
        return error_response(f'Error upserting weekly hours: {str(e)}', 500)

# ======================= SERVICE CALCULATION =======================

def _tp_type_label(teaching_type, tp_type):
    """Libellé de type affiché : TP → TP12 / TP8 selon la taille de groupe de la matière."""
    if (teaching_type or '').upper() == 'TP':
        return 'TP8' if str(tp_type or '').upper().replace(' ', '') == 'TP8' else 'TP12'
    return teaching_type or ''

def calculate_hetd(teaching_type, total_hours):
    """Calculate HETD based on teaching type"""
    coefficients = {
        'CM': 1.5,
        'TD': 1.0,
        'TP': 2/3,
        'PT': 1.0
    }
    coefficient = coefficients.get(teaching_type, 1.0)
    return total_hours * coefficient

def _load_semester_groups(db):
    """Retourne ({(semestre, face): {td,tp,tp8,pt,alt}}, {sem mutualisés}, {sem à TP distincts})."""
    cur = db.execute('SELECT semester_code, formation_type, td_groups, tp_groups, tp8_groups, pt_groups, alt_groups FROM semester_groups')
    sg = {(r['semester_code'], r['formation_type']): {
        'TD': r['td_groups'], 'TP': r['tp_groups'],
        'TP8': r['tp8_groups'], 'PT': r['pt_groups'],
        'ALT': r['alt_groups'] if r['alt_groups'] is not None else 1,
    } for r in cur.fetchall()}
    mut, tpsep = set(), set()
    for r in db.execute('SELECT code, mutualized, tp_separate FROM semesters').fetchall():
        if r['mutualized']:
            mut.add(r['code'])
        if r['tp_separate'] is None or r['tp_separate']:
            tpsep.add(r['code'])
    return sg, mut, tpsep

def _group_multiplier(sg_map, mut_set, tpsep_set, semester_code, formation_type, teaching_type, tp_type=None):
    """Nb de groupes pour (semestre, face, type d'enseignement). CM = 1 groupe.
    Semestre mutualisé : groupes « Promo » (ft=2) pour tout, SAUF les TP quand
    ils restent distincts FTP/ALT (tp_separate) → groupes de la face.
    Pour les TP, `tp_type` de la matière (TP12/TP8) choisit la ligne."""
    tt = teaching_type.upper().replace(' ', '')
    if tt == 'CM':
        return 1
    ft = formation_type
    if semester_code in mut_set and not (tt.startswith('TP') and semester_code in tpsep_set):
        ft = 2
    groups = sg_map.get((semester_code, ft))
    if not groups:
        return 1
    if tt.startswith('TP'):
        tt = 'TP8' if (tp_type or '').upper().replace(' ', '') == 'TP8' else 'TP'
    return groups.get(tt, 1)

@app.route('/api/semester-groups', methods=['GET'])
def get_semester_groups():
    """Groupes par (semestre, face) + flags « semestre mutualisé ».
    Retour : {groups: [...], mutualized: {S1: 0/1, ...}}"""
    try:
        db = get_db()
        cur = db.execute('''
            SELECT semester_code, formation_type, td_groups, tp_groups, tp8_groups, pt_groups, alt_groups
            FROM semester_groups
            ORDER BY semester_code, formation_type
        ''')
        mut, tpsep = {}, {}
        for r in db.execute('SELECT code, mutualized, tp_separate FROM semesters').fetchall():
            mut[r['code']] = int(r['mutualized'] or 0)
            tpsep[r['code']] = 1 if (r['tp_separate'] is None or r['tp_separate']) else 0
        return jsonify({'groups': rows_to_list(cur.fetchall()),
                        'mutualized': mut, 'tp_separate': tpsep}), 200
    except Exception as e:
        return error_response(f'Error loading semester groups: {str(e)}', 500)

@app.route('/api/semester-groups', methods=['PUT'])
def update_semester_groups():
    """Mise à jour bulk des groupes.
    Body: [{semester_code, formation_type, td_groups, tp_groups, tp8_groups, pt_groups}, ...]"""
    try:
        data = request.get_json()
        if not isinstance(data, list):
            return error_response('Expected a list of semester group entries')
        db = get_db()
        for entry in data:
            sem = str(entry.get('semester_code') or '').upper()
            if sem not in ('S1', 'S2', 'S3', 'S4', 'S5', 'S6'):
                return error_response(f'Semestre invalide : {sem}')
            ft = int(entry.get('formation_type'))
            td = max(1, int(entry.get('td_groups', 1)))
            tp = max(1, int(entry.get('tp_groups', 1)))
            tp8 = max(1, int(entry.get('tp8_groups', entry.get('tp_groups', 1))))
            pt = max(1, int(entry.get('pt_groups', 1)))
            alt = max(1, int(entry.get('alt_groups', 1)))   # nb de groupes apprentis (dernier(s) groupe(s))
            db.execute('''
                INSERT INTO semester_groups (semester_code, formation_type, td_groups, tp_groups, tp8_groups, pt_groups, alt_groups)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(semester_code, formation_type) DO UPDATE SET
                    td_groups = excluded.td_groups,
                    tp_groups = excluded.tp_groups,
                    tp8_groups = excluded.tp8_groups,
                    pt_groups = excluded.pt_groups,
                    alt_groups = excluded.alt_groups
            ''', (sem, ft, td, tp, tp8, pt, alt))
        db.commit()
        return jsonify({'success': True}), 200
    except Exception as e:
        return error_response(f'Error updating semester groups: {str(e)}', 500)

def _merge_sessions_to_mut(db, course_id, teaching_types):
    """Fusionne les sessions FTP/ALT des types donnés en une session MUT (ft=2).
    En cas de doublon, la face FTP est conservée (ALT supprimée)."""
    def _delete_session(cs_id):
        db.execute('DELETE FROM weekly_hours WHERE course_session_id = ?', (cs_id,))
        db.execute('DELETE FROM course_sessions WHERE id = ?', (cs_id,))
    for tt in teaching_types:
        rows = {r['formation_type']: r for r in db.execute(
            'SELECT * FROM course_sessions WHERE course_id = ? AND teaching_type = ?',
            (course_id, tt)).fetchall()}
        if 2 in rows:
            for ft in (0, 1):
                if ft in rows:
                    _delete_session(rows[ft]['id'])
        elif 0 in rows:
            if 1 in rows:
                _delete_session(rows[1]['id'])
            db.execute('UPDATE course_sessions SET formation_type = 2 WHERE id = ?', (rows[0]['id'],))
        elif 1 in rows:
            db.execute('UPDATE course_sessions SET formation_type = 2 WHERE id = ?', (rows[1]['id'],))

def _split_sessions_from_mut(db, course_id, teaching_types):
    """Recopie les sessions MUT (ft=2) des types donnés sur les faces FTP et ALT."""
    def _delete_session(cs_id):
        db.execute('DELETE FROM weekly_hours WHERE course_session_id = ?', (cs_id,))
        db.execute('DELETE FROM course_sessions WHERE id = ?', (cs_id,))
    for tt in teaching_types:
        rows = {r['formation_type']: r for r in db.execute(
            'SELECT * FROM course_sessions WHERE course_id = ? AND teaching_type = ?',
            (course_id, tt)).fetchall()}
        if 2 not in rows:
            continue
        m = rows[2]
        if 0 in rows:
            _delete_session(m['id'])
        else:
            db.execute('UPDATE course_sessions SET formation_type = 0 WHERE id = ?', (m['id'],))
        if 1 not in rows:
            db.execute('''
                INSERT INTO course_sessions
                    (course_id, teacher_id, formation_type, teaching_type,
                     nb_sessions, total_hours, slot_duration, room_name, promo,
                     sessions_per_week_max)
                VALUES (?, ?, 1, ?, ?, ?, ?, ?, ?, ?)
            ''', (m['course_id'], m['teacher_id'], tt, m['nb_sessions'],
                  m['total_hours'], m['slot_duration'], m['room_name'],
                  m['promo'], m['sessions_per_week_max']))

@app.route('/api/semesters/<sem_code>/mutualized', methods=['PUT'])
def set_semester_mutualized(sem_code):
    """Bascule « semestre mutualisé » (FTP + ALT ensemble). Body : {mutualized: 0/1}.
    Les SAÉ ne sont PAS concernées : elles gardent leur réglage individuel
    (case « Mutualisé » de la fiche matière).
    • Passage en mutualisé : les ressources du semestre passent en MUT ;
      leurs CM/TD FTP/ALT sont fusionnés en sessions MUT (face FTP prioritaire),
      et les TP aussi si le semestre est en « TP communs » (tp_separate=0).
    • Retour en non mutualisé : les ressources repassent en non-MUT ; les CM/TD
      (et TP MUT éventuels) sont recopiés sur les deux faces FTP et ALT."""
    try:
        sem_code = (sem_code or '').upper()
        db = get_db()
        srow = db.execute('SELECT id, mutualized, tp_separate FROM semesters WHERE code = ?', (sem_code,)).fetchone()
        if not srow:
            return error_response('Semestre introuvable', 404)
        target = 1 if (request.get_json() or {}).get('mutualized') else 0
        sid = srow['id']
        tp_separate = 1 if (srow['tp_separate'] is None or srow['tp_separate']) else 0

        courses = db.execute(
            "SELECT id FROM courses WHERE semester_id = ? AND UPPER(COALESCE(course_type,'')) != 'SAE'",
            (sid,)).fetchall()
        for c in courses:
            if target:
                types = ('CM', 'TD') if tp_separate else ('CM', 'TD', 'TP')
                _merge_sessions_to_mut(db, c['id'], types)
            else:
                _split_sessions_from_mut(db, c['id'], ('CM', 'TD', 'TP'))
        db.execute('''UPDATE courses SET mutualized = ?, updated_at = CURRENT_TIMESTAMP
                      WHERE semester_id = ? AND UPPER(COALESCE(course_type,'')) != 'SAE' ''',
                   (target, sid))
        db.execute('UPDATE semesters SET mutualized = ? WHERE id = ?', (target, sid))
        db.commit()
        _audit('SEMESTER_MUTUALIZED', ip=_client_ip(), user=session.get('user'),
               semester=sem_code, mutualized=target, courses=len(courses))
        return jsonify({'semester': sem_code, 'mutualized': target, 'courses': len(courses)}), 200
    except Exception as e:
        return error_response(f'Error toggling mutualized semester: {str(e)}', 500)

@app.route('/api/semesters/<sem_code>/tp-separate', methods=['PUT'])
def set_semester_tp_separate(sem_code):
    """Semestre mutualisé : TP FTP / ALT distincts ou communs. Body : {tp_separate: 0/1}.
    Les SAÉ ne sont pas concernées (réglage individuel dans la fiche matière).
    • distincts (1) : les TP restent (ou redeviennent) des lignes FTP et ALT
      séparées — utile pour une répartition calendaire différente par face ;
    • communs (0) : les TP de chaque ressource sont fusionnés en une ligne TP MUT
      (face FTP conservée en cas de doublon)."""
    try:
        sem_code = (sem_code or '').upper()
        db = get_db()
        srow = db.execute('SELECT id, mutualized FROM semesters WHERE code = ?', (sem_code,)).fetchone()
        if not srow:
            return error_response('Semestre introuvable', 404)
        target = 1 if (request.get_json() or {}).get('tp_separate') else 0
        sid = srow['id']

        courses = db.execute(
            "SELECT id FROM courses WHERE semester_id = ? AND UPPER(COALESCE(course_type,'')) != 'SAE'",
            (sid,)).fetchall()
        # Restructuration des sessions TP uniquement si le semestre est mutualisé
        if srow['mutualized']:
            for c in courses:
                if target:
                    _split_sessions_from_mut(db, c['id'], ('TP',))
                else:
                    _merge_sessions_to_mut(db, c['id'], ('TP',))
        db.execute('UPDATE semesters SET tp_separate = ? WHERE id = ?', (target, sid))
        db.commit()
        _audit('SEMESTER_TP_SEPARATE', ip=_client_ip(), user=session.get('user'),
               semester=sem_code, tp_separate=target, courses=len(courses))
        return jsonify({'semester': sem_code, 'tp_separate': target,
                        'mutualized': int(srow['mutualized'] or 0), 'courses': len(courses)}), 200
    except Exception as e:
        return error_response(f'Error toggling TP separate: {str(e)}', 500)

@app.route('/api/service/teacher/<int:teacher_id>', methods=['GET'])
def get_teacher_service(teacher_id):
    """Get teacher service hours (HETD calculation)"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check if teacher exists
        cursor.execute('SELECT * FROM teachers WHERE id = ?', (teacher_id,))
        teacher = cursor.fetchone()
        if not teacher:
            return error_response('Teacher not found', 404)

        cursor.execute('''
            SELECT cs.id, c.code as course_code, cs.teaching_type, cs.total_hours,
                   cs.formation_type, s.code as semester_code, c.tp_type
            FROM course_sessions cs
            JOIN courses c ON cs.course_id = c.id
            JOIN semesters s ON c.semester_id = s.id
            WHERE cs.teacher_id = ?
        ''', (teacher_id,))
        sessions = cursor.fetchall()

        sg_map, mut_set, tpsep_set = _load_semester_groups(db)

        service_details = []
        total_hetd = 0

        for session in sessions:
            mult = _group_multiplier(sg_map, mut_set, tpsep_set, session['semester_code'], session['formation_type'], session['teaching_type'], session['tp_type'])
            effective_hours = (session['total_hours'] or 0) * mult
            hetd = calculate_hetd(session['teaching_type'], effective_hours)
            total_hetd += hetd
            service_details.append({
                'course_code': session['course_code'],
                'teaching_type': session['teaching_type'],
                'total_hours': session['total_hours'],
                'nb_groups': mult,
                'effective_hours': effective_hours,
                'hetd': hetd
            })

        return jsonify({
            'teacher_id': teacher_id,
            'teacher_name': teacher['name'],
            'service_details': service_details,
            'total_hetd': total_hetd
        }), 200
    except Exception as e:
        return error_response(f'Error calculating service: {str(e)}', 500)

@app.route('/api/service/all', methods=['GET'])
def get_all_service():
    """Get all teachers' service hours with breakdown by type.
    Heures multipliées par le nb de groupes (table semester_groups) selon (semestre, face),
    groupes « Promo » si le semestre est mutualisé."""
    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute('''
            SELECT t.id AS teacher_id, t.name AS teacher_name,
                   cs.teaching_type, cs.total_hours,
                   cs.formation_type, s.code AS semester_code, c.tp_type
            FROM teachers t
            JOIN course_sessions cs ON cs.teacher_id = t.id
            JOIN courses c ON cs.course_id = c.id
            JOIN semesters s ON c.semester_id = s.id
            ORDER BY t.name
        ''')
        rows = cursor.fetchall()
        sg_map, mut_set, tpsep_set = _load_semester_groups(db)

        # Agréger par enseignant en multipliant par le nb de groupes
        agg = {}
        for r in rows:
            tid = r['teacher_id']
            if tid not in agg:
                agg[tid] = {'teacher_id': tid, 'teacher_name': r['teacher_name'],
                            'cm_hours': 0, 'td_hours': 0, 'tp_hours': 0, 'pt_hours': 0}
            mult = _group_multiplier(sg_map, mut_set, tpsep_set, r['semester_code'], r['formation_type'], r['teaching_type'], r['tp_type'])
            h = (r['total_hours'] or 0) * mult
            tt = r['teaching_type'].upper().replace(' ', '')
            if tt == 'CM':
                agg[tid]['cm_hours'] += h
            elif tt == 'TD':
                agg[tid]['td_hours'] += h
            elif tt.startswith('TP'):
                agg[tid]['tp_hours'] += h
            elif tt == 'PT':
                agg[tid]['pt_hours'] += h

        services = []
        for s in agg.values():
            cm_h, td_h, tp_h, pt_h = s['cm_hours'], s['td_hours'], s['tp_hours'], s['pt_hours']
            hetd = cm_h * 1.5 + td_h * 1.0 + tp_h * (2.0/3.0) + pt_h * 1.0
            total_h = cm_h + td_h + tp_h + pt_h
            if total_h > 0:
                services.append({
                    'teacher_id': s['teacher_id'],
                    'teacher_name': s['teacher_name'],
                    'cm_hours': round(cm_h, 1),
                    'td_hours': round(td_h, 1),
                    'tp_hours': round(tp_h, 1),
                    'pt_hours': round(pt_h, 1),
                    'total_hours': round(total_h, 1),
                    'total_hetd': round(hetd, 2),
                })
        services.sort(key=lambda x: x['teacher_name'])
        return jsonify(services), 200
    except Exception as e:
        return error_response(f'Error calculating services: {str(e)}', 500)

# ======================= TIMETABLE GENERATION =======================

@app.route('/api/generate-timetable', methods=['POST'])
def generate_timetable():
    """Generate timetable with constraint-based scheduling.

    Constraints:
    - ALT: max 40h/week
    - FTP: target ~25-30h/week
    - Teacher: max 6h CM/TD per day
    - No room conflicts (same room, same time)
    - No teacher conflicts (same teacher, same time)
    - No class/formation conflicts (same formation, same time)
    - Prefer physical info room (123) over mobile; max 2 courses needing fixed info room
    """
    try:
        data = request.get_json() or {}
        weeks = data.get('weeks', [])
        day_start = data.get('day_start', '08:00')
        day_end = data.get('day_end', '18:00')

        db = get_db()
        cursor = db.cursor()

        # If no weeks specified, get all weeks that have data
        if not weeks:
            cursor.execute('SELECT DISTINCT week_number FROM weekly_hours ORDER BY week_number')
            weeks = [r['week_number'] for r in cursor.fetchall()]

        if not weeks:
            return error_response('No weeks to generate')

        # Clear existing timetable for these weeks
        placeholders = ','.join('?' * len(weeks))
        cursor.execute(f'DELETE FROM timetable_slots WHERE week_number IN ({placeholders})', weeks)

        # Build time grid: 30-min slots from day_start to day_end
        start_h, start_m = map(int, day_start.split(':'))
        end_h, end_m = map(int, day_end.split(':'))
        time_slots = []
        h, m = start_h, start_m
        while h * 60 + m < end_h * 60 + end_m:
            time_slots.append(f'{h:02d}:{m:02d}')
            m += 30
            if m >= 60:
                h += 1
                m = 0

        # Load all sessions with their weekly hours
        cursor.execute('''
            SELECT cs.id, cs.course_id, cs.teacher_id, cs.formation_type,
                   cs.teaching_type, cs.slot_duration, cs.room_name, cs.nb_sessions,
                   c.code as course_code, c.semester_id, s.year_group
            FROM course_sessions cs
            JOIN courses c ON cs.course_id = c.id
            JOIN semesters s ON c.semester_id = s.id
        ''')
        all_sessions = [dict(r) for r in cursor.fetchall()]

        # Load room IDs
        room_map = {}
        cursor.execute('SELECT id, name FROM rooms')
        for r in cursor.fetchall():
            room_map[r['name']] = r['id']

        # Pre-load all weekly_hours for requested weeks (avoid N+1 queries)
        wh_map = {}  # (course_session_id, week_number) -> hours
        cursor.execute(
            f'SELECT course_session_id, week_number, hours FROM weekly_hours WHERE week_number IN ({placeholders})',
            weeks)
        for r in cursor.fetchall():
            if r['hours'] and r['hours'] > 0:
                wh_map[(r['course_session_id'], r['week_number'])] = r['hours']

        generated_slots = 0
        conflicts = []

        for week in weeks:
            # Track occupancy: key = (day, time_slot_index) -> set of resources
            room_occupied = {}    # (day, time_idx, room_id) -> True
            teacher_occupied = {} # (day, time_idx, teacher_id) -> True
            formation_occupied = {} # (day, time_idx, year_group, formation) -> True
            teacher_day_cm_td = {} # (day, teacher_id) -> hours of CM/TD
            week_hours_ftp = 0
            week_hours_alt = 0

            # Get sessions that have hours this week
            week_sessions = []
            for sess in all_sessions:
                h = wh_map.get((sess['id'], week))
                if h:
                    sess_copy = dict(sess)
                    sess_copy['week_hours'] = h
                    week_sessions.append(sess_copy)

            # Sort: CM first (harder to place), then TD, then TP, then PT
            type_order = {'CM': 0, 'TD': 1, 'TP': 2, 'PT': 3}
            week_sessions.sort(key=lambda s: (type_order.get(s['teaching_type'], 4), -(s['week_hours'])))

            for sess in week_sessions:
                hours_to_place = sess['week_hours']
                duration = sess.get('slot_duration') or 1.5
                num_slots_needed = max(1, round(hours_to_place / duration))
                duration_in_30min = int(duration * 2)

                room_id = room_map.get(sess['room_name'])
                teacher_id = sess['teacher_id']
                formation = sess['formation_type']
                year_group = sess['year_group']
                ttype = sess['teaching_type']
                is_cm_td = ttype in ('CM', 'TD')

                placed = 0
                for slot_attempt in range(num_slots_needed):
                    best_day = None
                    best_time = None

                    for day in range(5):  # Mon-Fri
                        # Check teacher CM/TD limit (6h/day)
                        if is_cm_td and teacher_id:
                            current_cm_td = teacher_day_cm_td.get((day, teacher_id), 0)
                            if current_cm_td + duration > 6:
                                continue

                        for t_idx in range(len(time_slots) - duration_in_30min + 1):
                            # Check all 30-min sub-slots
                            conflict = False
                            for dt in range(duration_in_30min):
                                ti = t_idx + dt
                                if room_id and (day, ti, room_id) in room_occupied:
                                    conflict = True
                                    break
                                if teacher_id and (day, ti, teacher_id) in teacher_occupied:
                                    conflict = True
                                    break
                                if (day, ti, year_group, formation) in formation_occupied:
                                    conflict = True
                                    break
                            if not conflict:
                                best_day = day
                                best_time = t_idx
                                break
                        if best_day is not None:
                            break

                    if best_day is not None:
                        # Place the slot
                        start_t = time_slots[best_time]
                        end_idx = best_time + duration_in_30min
                        if end_idx < len(time_slots):
                            end_t = time_slots[end_idx]
                        else:
                            eh = int(start_t.split(':')[0]) + int(duration)
                            em = int(start_t.split(':')[1]) + int((duration % 1) * 60)
                            if em >= 60:
                                eh += 1
                                em -= 60
                            end_t = f'{eh:02d}:{em:02d}'

                        cursor.execute('''
                            INSERT INTO timetable_slots
                            (course_session_id, week_number, day_of_week, start_time, end_time,
                             room_id, teacher_id, formation_type)
                            VALUES (?,?,?,?,?,?,?,?)
                        ''', (sess['id'], week, best_day, start_t, end_t,
                              room_id, teacher_id, formation))
                        generated_slots += 1
                        placed += 1

                        # Mark occupied
                        for dt in range(duration_in_30min):
                            ti = best_time + dt
                            if room_id:
                                room_occupied[(best_day, ti, room_id)] = True
                            if teacher_id:
                                teacher_occupied[(best_day, ti, teacher_id)] = True
                            formation_occupied[(best_day, ti, year_group, formation)] = True

                        if is_cm_td and teacher_id:
                            teacher_day_cm_td[(best_day, teacher_id)] = \
                                teacher_day_cm_td.get((best_day, teacher_id), 0) + duration

                        if formation == 0:
                            week_hours_ftp += duration
                        elif formation == 1:
                            week_hours_alt += duration
                    else:
                        conflicts.append({
                            'week': week,
                            'session_id': sess['id'],
                            'course': sess['course_code'],
                            'type': ttype,
                            'reason': 'No available slot'
                        })

        db.commit()

        return jsonify({
            'generated_slots': generated_slots,
            'weeks_processed': len(weeks),
            'conflicts': conflicts[:50]  # Limit conflict list
        }), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return error_response(f'Error generating timetable: {str(e)}', 500)

@app.route('/api/available-weeks', methods=['GET'])
def get_available_weeks():
    """Get all weeks that have data"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT DISTINCT week_number FROM weekly_hours ORDER BY week_number')
        weeks = [r['week_number'] for r in cursor.fetchall()]
        return jsonify(weeks), 200
    except Exception as e:
        return error_response(f'Error: {str(e)}', 500)

@app.route('/api/timetable/week/<int:week_number>', methods=['GET'])
def get_timetable_week(week_number):
    """Get timetable for a specific week"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            SELECT ts.*, cs.teaching_type, cs.formation_type as form_type,
                   c.code as course_code, c.name as course_name,
                   s.code as semester_code, s.year_group,
                   t.name as teacher_name, r.name as room_name
            FROM timetable_slots ts
            JOIN course_sessions cs ON ts.course_session_id = cs.id
            JOIN courses c ON cs.course_id = c.id
            JOIN semesters s ON c.semester_id = s.id
            LEFT JOIN teachers t ON ts.teacher_id = t.id
            LEFT JOIN rooms r ON ts.room_id = r.id
            WHERE ts.week_number = ?
            ORDER BY ts.day_of_week, ts.start_time
        ''', (week_number,))
        slots = rows_to_list(cursor.fetchall())
        return jsonify({
            'week_number': week_number,
            'slots': slots
        }), 200
    except Exception as e:
        return error_response(f'Error fetching timetable: {str(e)}', 500)

@app.route('/api/timetable/teacher/<int:teacher_id>/week/<int:week_number>', methods=['GET'])
def get_teacher_timetable_week(teacher_id, week_number):
    """Get timetable for a teacher in a specific week"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Check if teacher exists
        cursor.execute('SELECT * FROM teachers WHERE id = ?', (teacher_id,))
        if not cursor.fetchone():
            return error_response('Teacher not found', 404)
        
        cursor.execute('''
            SELECT ts.*, c.code as course_code, c.name as course_name, 
                   r.name as room_name
            FROM timetable_slots ts
            JOIN course_sessions cs ON ts.course_session_id = cs.id
            JOIN courses c ON cs.course_id = c.id
            LEFT JOIN rooms r ON ts.room_id = r.id
            WHERE ts.teacher_id = ? AND ts.week_number = ?
            ORDER BY ts.day_of_week, ts.start_time
        ''', (teacher_id, week_number))
        slots = rows_to_list(cursor.fetchall())
        
        return jsonify({
            'teacher_id': teacher_id,
            'week_number': week_number,
            'slots': slots
        }), 200
    except Exception as e:
        return error_response(f'Error fetching teacher timetable: {str(e)}', 500)

@app.route('/api/timetable/week/<int:week_number>', methods=['DELETE'])
def clear_timetable_week(week_number):
    """Clear timetable for a specific week"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('DELETE FROM timetable_slots WHERE week_number = ?', (week_number,))
        deleted = cursor.rowcount
        db.commit()
        
        return jsonify({
            'week_number': week_number,
            'deleted': deleted
        }), 200
    except Exception as e:
        return error_response(f'Error clearing timetable: {str(e)}', 500)

# ======================= CALENDAR =======================

@app.route('/api/calendar', methods=['GET'])
def get_calendar():
    """Get calendar events"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT * FROM calendar_events ORDER BY week_number')
        events = rows_to_list(cursor.fetchall())
        return jsonify(events), 200
    except Exception as e:
        return error_response(f'Error fetching calendar: {str(e)}', 500)

@app.route('/api/calendar', methods=['POST'])
def create_calendar_event():
    """Create calendar event"""
    try:
        data = request.get_json()
        if not data or not data.get('event_type'):
            return error_response('event_type is required')
        
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            INSERT INTO calendar_events (week_number, date, event_type, description)
            VALUES (?, ?, ?, ?)
        ''', (
            data.get('week_number'),
            data.get('date'),
            data['event_type'],
            data.get('description')
        ))
        db.commit()
        event_id = cursor.lastrowid
        cursor.execute('SELECT * FROM calendar_events WHERE id = ?', (event_id,))
        event = row_to_dict(cursor.fetchone())
        return jsonify(event), 201
    except Exception as e:
        return error_response(f'Error creating calendar event: {str(e)}', 500)

# ======================= IMPORT =======================

@app.route('/api/import/excel', methods=['POST'])
def import_excel():
    """Import data from Excel file"""
    try:
        if 'file' not in request.files:
            return error_response('No file part')
        
        file = request.files['file']
        if file.filename == '':
            return error_response('No selected file')
        
        # This is a placeholder for Excel import
        # Actual implementation would require openpyxl or pandas library
        return jsonify({
            'message': 'Excel import not yet implemented',
            'note': 'Install openpyxl or pandas to enable Excel import'
        }), 501
    except Exception as e:
        return error_response(f'Error importing file: {str(e)}', 500)

# ======================= REPARTITION HEBDOMADAIRE =======================

@app.route('/api/config', methods=['GET'])
def get_config():
    return jsonify(get_year_config())

@app.route('/api/config', methods=['PUT'])
def update_config():
    data = request.get_json() or {}
    db = get_db()
    keys = ['academic_start_week', 'academic_end_week',
            'semester_odd_start_week', 'semester_odd_end_week',
            'semester_even_start_week', 'semester_even_end_week']
    # Plages par année de progression (1/2/3), optionnelles
    for yg in (1, 2, 3):
        keys += [f'academic_start_week_y{yg}', f'academic_end_week_y{yg}']
    for key in keys:
        if key in data:
            try:
                val = int(data[key])
            except (TypeError, ValueError):
                return error_response(f'{key} doit être un entier', 400)
            if val < 1 or val > 53:
                return error_response(f'{key} hors plage (1–53)', 400)
            db.execute('INSERT OR REPLACE INTO app_settings (key, value) VALUES (?, ?)', (key, str(val)))
    # Nombre de semaines de l'année (52 ou 53) par année de progression
    for yg in (1, 2, 3):
        key = f'academic_max_week_y{yg}'
        if key in data:
            try:
                val = int(data[key])
            except (TypeError, ValueError):
                return error_response(f'{key} doit être un entier', 400)
            if val not in (52, 53):
                return error_response(f'{key} doit valoir 52 ou 53', 400)
            db.execute('INSERT OR REPLACE INTO app_settings (key, value) VALUES (?, ?)', (key, str(val)))
    db.commit()
    return jsonify({'message': 'Configuration mise à jour'})

# Coefficients de conversion en HETD (Heures Équivalent TD) — réglage unique (onglet
# Paramètres), utilisé par le Bilan global ET la colonne Heures/HETD des Matières.
_HETD_DEFAULTS = {'cm': 1.5, 'td': 1.0, 'tp': 2.0 / 3.0, 'pt': 1.0}

def _get_hetd_coeffs(db):
    row = db.execute("SELECT value FROM app_settings WHERE key='hetd_coeffs'").fetchone()
    if row and row['value']:
        try:
            d = json.loads(row['value'])
            return {k: float(d.get(k, _HETD_DEFAULTS[k])) for k in _HETD_DEFAULTS}
        except (ValueError, TypeError):
            pass
    return dict(_HETD_DEFAULTS)

@app.route('/api/hetd-coeffs', methods=['GET'])
def get_hetd_coeffs():
    """Coefficients HETD (CM/TD/TP/PT) — lecture seule, tous utilisateurs connectés."""
    return jsonify(_get_hetd_coeffs(get_db()))

@app.route('/api/hetd-coeffs', methods=['PUT'])
def put_hetd_coeffs():
    err = _require_admin()
    if err:
        return err
    data = request.get_json() or {}
    coeffs = {}
    for k in _HETD_DEFAULTS:
        try:
            v = float(data.get(k))
        except (TypeError, ValueError):
            return error_response(f'Coefficient {k.upper()} invalide', 400)
        if v < 0 or v > 10:
            return error_response(f'Coefficient {k.upper()} hors plage (0–10)', 400)
        coeffs[k] = v
    db = get_db()
    db.execute("INSERT OR REPLACE INTO app_settings (key, value) VALUES ('hetd_coeffs', ?)",
               (json.dumps(coeffs),))
    db.commit()
    _audit('HETD_COEFFS_UPDATE', ip=_client_ip(), user=session.get('user'))
    return jsonify(coeffs)

@app.route('/api/repartition', methods=['GET'])
def get_repartition():
    """Tableau pivot : pour un semestre donné, heures par cours et par semaine"""
    semester_code = request.args.get('semester', '')
    try:
        cfg = get_year_config()
        start_week = cfg['academic_start_week']
        end_week_cfg = cfg['academic_end_week']
        max_week = cfg['academic_max_week']
        _key = lambda w: school_week_key(w, start_week, max_week)

        # Colonnes = plage complète de l'année académique (identique pour tous les semestres)
        weeks = sorted(get_valid_school_weeks(start_week, end_week_cfg, max_week), key=_key)

        db = get_db()
        cursor = db.cursor()

        # Détecter l'année de promotion à partir du code semestre
        year_group = None
        if semester_code:
            s_code = semester_code.split('+')[0]  # ex: "S1+S2" → "S1"
            try:
                sn = int(s_code[1])          # S1→1, S3→3, S5→5
                year_group = (sn + 1) // 2   # 1,2→1 ; 3,4→2 ; 5,6→3
            except Exception:
                pass

        # Plage des colonnes = plage propre à l'année de progression du semestre
        # (début/fin + nb de semaines 52/53), repli global si non définie.
        if year_group:
            start_week, end_week_cfg, max_week = year_range_for_yg(cfg, year_group)
            _key = lambda w: school_week_key(w, start_week, max_week)
            weeks = sorted(get_valid_school_weeks(start_week, end_week_cfg, max_week), key=_key)

        # Charger le calendrier spécial global
        cursor.execute('SELECT week_number, week_type FROM special_calendar ORDER BY week_type, week_number')
        sc = {t: [] for t in _SPECIAL_TYPES}
        for r in cursor.fetchall():
            if r['week_type'] in sc:
                sc[r['week_type']].append(r['week_number'])

        special_weeks = {
            'vacation_ftp': sc.get(f'vacation_ftp_y{year_group}', []) if year_group else [],
            'stage_ftp':    sc.get(f'stage_ftp_y{year_group}', []) if year_group else [],
            'company_alt':  sc.get(f'company_alt_y{year_group}', []) if year_group else [],
        }

        # Lignes de données avec heures hebdomadaires
        if semester_code:
            cursor.execute('''
                SELECT cs.id AS session_id,
                       c.code AS course_code, c.name AS course_name,
                       c.tp_type,
                       s.code AS semester_code,
                       cs.teaching_type, cs.formation_type,
                       cs.total_hours, cs.nb_sessions, cs.slot_duration,
                       cs.room_name,
                       t.name AS teacher_name,
                       wh.week_number, wh.semester_week, wh.hours, wh.group_index
                FROM course_sessions cs
                JOIN courses c ON cs.course_id = c.id
                JOIN semesters s ON c.semester_id = s.id
                LEFT JOIN teachers t ON cs.teacher_id = t.id
                LEFT JOIN weekly_hours wh ON wh.course_session_id = cs.id
                WHERE s.code = ? AND (cs.nb_sessions > 0 OR cs.total_hours > 0)
                ORDER BY c.code,
                         CASE cs.teaching_type WHEN 'CM' THEN 0 WHEN 'TD' THEN 1 WHEN 'TP' THEN 2 WHEN 'PT' THEN 3 ELSE 4 END,
                         cs.formation_type, wh.week_number
            ''', (semester_code,))
        else:
            cursor.execute('''
                SELECT cs.id AS session_id,
                       c.code AS course_code, c.name AS course_name,
                       c.tp_type,
                       s.code AS semester_code,
                       cs.teaching_type, cs.formation_type,
                       cs.total_hours, cs.nb_sessions, cs.slot_duration,
                       cs.room_name,
                       t.name AS teacher_name,
                       wh.week_number, wh.semester_week, wh.hours, wh.group_index
                FROM course_sessions cs
                JOIN courses c ON cs.course_id = c.id
                JOIN semesters s ON c.semester_id = s.id
                LEFT JOIN teachers t ON cs.teacher_id = t.id
                LEFT JOIN weekly_hours wh ON wh.course_session_id = cs.id
                WHERE cs.nb_sessions > 0 OR cs.total_hours > 0
                ORDER BY s.code, c.code,
                         CASE cs.teaching_type WHEN 'CM' THEN 0 WHEN 'TD' THEN 1 WHEN 'TP' THEN 2 WHEN 'PT' THEN 3 ELSE 4 END,
                         cs.formation_type, wh.week_number
            ''')
        raw = cursor.fetchall()

        # Nb de groupes par (semestre, face, type) : les TP se scindent en une ligne
        # par groupe dans la répartition calendaire.
        sg_map, mut_set, tpsep_set = _load_semester_groups(db)

        # Pivot : regrouper par session, en conservant les heures par groupe
        sessions = {}
        order = []
        for r in raw:
            sid = r['session_id']
            if sid not in sessions:
                tt = r['teaching_type'] or ''
                # Nb de groupes pour ce type (CM=1, TD, TP12/TP8, PT) selon semestre/face.
                # Sert au comptage des créneaux (répartition journalière) et, pour les TP,
                # à scinder la session en une ligne par groupe.
                try:
                    group_count = int(_group_multiplier(
                        sg_map, mut_set, tpsep_set, r['semester_code'],
                        r['formation_type'], tt, r['tp_type']) or 1)
                except Exception:
                    group_count = 1
                if group_count < 1:
                    group_count = 1
                n_groups = group_count if tt.upper().replace(' ', '').startswith('TP') else 1
                sessions[sid] = {
                    'session_id': sid,
                    'course_code': r['course_code'] or '',
                    'course_name': r['course_name'] or '',
                    'semester': r['semester_code'] or '',
                    'type': r['teaching_type'] or '',
                    'tp_type': r['tp_type'] or 'TP12',
                    'formation': {0: 'FTP', 1: 'ALT', 2: 'MUT', 3: 'OTHER'}.get(r['formation_type'], str(r['formation_type'] or '')),
                    'total_hours': r['total_hours'] or 0,
                    'nb_sessions': r['nb_sessions'] or 0,
                    'slot_duration': r['slot_duration'] or 1.5,
                    'teacher': r['teacher_name'] or '',
                    'room': r['room_name'] or '',
                    'n_groups': n_groups,
                    'group_count': group_count,
                    'by_group': {},   # {group_index: {week: hours}}
                }
                order.append(sid)
            if r['week_number'] is not None and (r['hours'] or 0) > 0:
                g = r['group_index'] or 1
                sessions[sid]['by_group'].setdefault(g, {})[r['week_number']] = r['hours'] or 0

        # Une ligne par session (CM/TD/PT) ou une ligne par groupe (TP à plusieurs groupes)
        _BASE = ('session_id', 'course_code', 'course_name', 'semester', 'type',
                 'tp_type', 'formation', 'total_hours', 'nb_sessions', 'slot_duration',
                 'teacher', 'room', 'group_count')
        out_rows = []
        for sid in order:
            s = sessions[sid]
            base = {k: s[k] for k in _BASE}
            n = s['n_groups']
            if n <= 1:
                # Fusionne tous les groupes éventuels sur une seule ligne
                merged = {}
                for gw in s['by_group'].values():
                    for w, h in gw.items():
                        merged[w] = merged.get(w, 0) + h
                out_rows.append({**base, 'group_index': 1, 'n_groups': 1, 'by_week': merged,
                                 'alt_group': False, 'group_label': ''})
            else:
                # TP mutualisé à TP communs : les apprentis (ALT) occupent les
                # DERNIERS groupes (nb = alt_groups de la promo).
                sem = s['semester']
                # Uniquement les matières réellement mutualisées (session MUT) ont des
                # groupes d'apprentis. Les SAE (face FTP ou ALT) gardent tous leurs
                # groupes dans LEUR cohorte — pas de dernier groupe « ALT ».
                is_common_mut = (sem in mut_set) and (sem not in tpsep_set) and (s['formation'] == 'MUT')
                k_alt = 0
                if is_common_mut:
                    k_alt = int((sg_map.get((sem, 2)) or {}).get('ALT', 1) or 0)
                    k_alt = max(0, min(k_alt, n))
                n_ftp = n - k_alt
                for g in range(1, n + 1):
                    is_alt = g > n_ftp
                    if is_alt:
                        rank = g - n_ftp
                        label = 'ALT' if k_alt == 1 else f'ALT {rank}'
                    else:
                        label = f'gr.{g}'
                    out_rows.append({**base, 'group_index': g, 'n_groups': n,
                                     'by_week': dict(s['by_group'].get(g, {})),
                                     'alt_group': is_alt, 'group_label': label})

        return jsonify({
            'weeks': weeks,
            'rows': out_rows,
            'special_weeks': special_weeks,
        }), 200
    except Exception as e:
        return error_response(f'Error fetching repartition: {str(e)}', 500)

@app.route('/api/checks/repartition', methods=['GET'])
def checks_repartition():
    """Contrôles de validation : charge enseignants par semaine + conflits salles"""
    try:
        db = get_db()
        cursor = db.cursor()

        # 1. Charge enseignants par semaine
        cursor.execute('''
            SELECT t.id AS teacher_id, t.name AS teacher_name,
                   wh.week_number,
                   SUM(wh.hours) AS total_hours,
                   GROUP_CONCAT(DISTINCT c.code || ' ' || cs.teaching_type) AS courses
            FROM weekly_hours wh
            JOIN course_sessions cs ON wh.course_session_id = cs.id
            JOIN courses c ON cs.course_id = c.id
            JOIN teachers t ON cs.teacher_id = t.id
            WHERE wh.hours > 0
            GROUP BY t.id, wh.week_number
            HAVING SUM(wh.hours) > 20
            ORDER BY t.name, wh.week_number
        ''')
        teacher_load = [dict(r) for r in cursor.fetchall()]

        # 2. Conflits salles : semaines où une salle est utilisée par plus de 2 matières
        #    - STD (salles standard, pool générique) : exclue du contrôle
        #    - 123 : label couvrant deux salles info → nombre de matières divisé par deux
        cursor.execute('''
            SELECT cs.room_name, wh.week_number,
                   CASE WHEN cs.room_name = '123'
                        THEN COUNT(DISTINCT c.id) / 2.0
                        ELSE COUNT(DISTINCT c.id) END AS nb_courses,
                   GROUP_CONCAT(DISTINCT c.code || ' ' || cs.teaching_type) AS courses
            FROM weekly_hours wh
            JOIN course_sessions cs ON wh.course_session_id = cs.id
            JOIN courses c ON cs.course_id = c.id
            WHERE wh.hours > 0 AND cs.room_name IS NOT NULL AND cs.room_name != ''
              AND cs.room_name != 'STD'
            GROUP BY cs.room_name, wh.week_number
            HAVING (CASE WHEN cs.room_name = '123'
                         THEN COUNT(DISTINCT c.id) / 2.0
                         ELSE COUNT(DISTINCT c.id) END) > 2
            ORDER BY cs.room_name, wh.week_number
        ''')
        room_conflicts = [dict(r) for r in cursor.fetchall()]

        return jsonify({
            'teacher_load': teacher_load,
            'room_conflicts': room_conflicts,
        }), 200
    except Exception as e:
        return error_response(f'Error running checks: {str(e)}', 500)

@app.route('/api/export/repartition', methods=['GET'])
def export_repartition_excel():
    """Export la répartition des 6 semestres en fichier Excel (un onglet par année)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    from copy import copy as _copy

    # Application de style mémoïsée : openpyxl re-hache chaque objet de style à
    # CHAQUE affectation (.font/.fill/.border/…), ce qui rend l'export très lent
    # (dizaines de milliers de cellules). On compose le StyleArray une seule fois
    # par combinaison puis on le réutilise → ~15× plus rapide, rendu identique.
    # Clé par id() des objets (réutilisés, épinglés dans la valeur pour garder
    # leur id stable) + le format numérique (chaîne).
    _style_memo = {}
    def _sc(cell, font=None, fill=None, border=None, alignment=None, number_format=None):
        key = (id(font), id(fill), id(border), id(alignment), number_format)
        entry = _style_memo.get(key)
        if entry is not None:
            cell._style = _copy(entry[0])
            return cell
        if font is not None: cell.font = font
        if fill is not None: cell.fill = fill
        if border is not None: cell.border = border
        if alignment is not None: cell.alignment = alignment
        if number_format is not None: cell.number_format = number_format
        # Copie figée : openpyxl mute le StyleArray en place si la cellule source est
        # re-stylée ensuite → sans copie, l'entrée du cache serait corrompue.
        _style_memo[key] = (_copy(cell._style), font, fill, border, alignment)
        return cell

    try:
        db = get_db()
        cursor = db.cursor()

        # Filtres d'affichage (repris de l'onglet Répartition) : semestres, formation, enseignant
        sel_sems = [s.strip() for s in (request.args.get('semesters') or '').split(',') if s.strip()]
        sel_sems = set(sel_sems) or None
        form_filter = (request.args.get('formation') or '').lower()   # '', 'ftp', 'alt'
        teacher_filter = (request.args.get('teacher') or '').strip() or None

        def session_included(s):
            if sel_sems is not None and s['semester'] not in sel_sems:
                return False
            if teacher_filter and (s['teacher'] or '') != teacher_filter:
                return False
            if form_filter == 'ftp' and s['formation'] not in ('FTP', 'MUT'):
                return False
            if form_filter == 'alt' and s['formation'] not in ('ALT', 'MUT'):
                return False
            return True

        cfg = get_year_config()
        start_week = cfg['academic_start_week']
        end_week_cfg = cfg['academic_end_week']
        max_week = cfg['academic_max_week']
        _key = lambda w: school_week_key(w, start_week, max_week)
        weeks = sorted(get_valid_school_weeks(start_week, end_week_cfg, max_week), key=_key)

        # Charger le calendrier spécial
        cursor.execute('SELECT week_number, week_type FROM special_calendar ORDER BY week_type, week_number')
        sc = {t: set() for t in _SPECIAL_TYPES}
        for r in cursor.fetchall():
            if r['week_type'] in sc:
                sc[r['week_type']].add(r['week_number'])

        # Charger les commentaires par semaine (affichés au-dessus du tableau)
        cursor.execute('SELECT week_number, comment, formations, years FROM week_comments')
        week_comments = {r['week_number']: {
            'comment': r['comment'],
            'years': [int(y) for y in (r['years'] or '1,2,3').split(',') if y],
        } for r in cursor.fetchall()}

        # Charger les codes des catégories (affichés en 1re ligne de chaque onglet)
        cursor.execute('SELECT week_type, code FROM special_category_codes')
        category_codes = {r['week_type']: r['code'] for r in cursor.fetchall()}

        # Charger toutes les sessions avec heures hebdomadaires
        cursor.execute('''
            SELECT cs.id AS session_id,
                   c.code AS course_code, c.name AS course_name,
                   c.tp_type,
                   s.code AS semester_code, s.year_group,
                   cs.teaching_type, cs.formation_type,
                   cs.total_hours, cs.nb_sessions, cs.slot_duration,
                   cs.room_name,
                   t.name AS teacher_name,
                   wh.week_number, wh.hours, wh.group_index
            FROM course_sessions cs
            JOIN courses c ON cs.course_id = c.id
            JOIN semesters s ON c.semester_id = s.id
            LEFT JOIN teachers t ON cs.teacher_id = t.id
            LEFT JOIN weekly_hours wh ON wh.course_session_id = cs.id
            WHERE cs.nb_sessions > 0 OR cs.total_hours > 0
            ORDER BY s.year_group, s.code, c.code,
                     CASE cs.teaching_type WHEN 'CM' THEN 0 WHEN 'TD' THEN 1 WHEN 'TP' THEN 2 WHEN 'PT' THEN 3 ELSE 4 END,
                     cs.formation_type, wh.week_number
        ''')
        raw = cursor.fetchall()

        # Nb de groupes (TP scindés en une ligne par groupe, comme à l'écran)
        sg_map, mut_set, tpsep_set = _load_semester_groups(db)

        # Pivoter par session, en conservant les heures par groupe
        sessions = {}
        order = []
        for r in raw:
            sid = r['session_id']
            if sid not in sessions:
                tt = r['teaching_type'] or ''
                n_groups = 1
                if tt.upper().replace(' ', '').startswith('TP'):
                    try:
                        n_groups = int(_group_multiplier(
                            sg_map, mut_set, tpsep_set, r['semester_code'],
                            r['formation_type'], tt, r['tp_type']) or 1)
                    except Exception:
                        n_groups = 1
                    if n_groups < 1:
                        n_groups = 1
                sessions[sid] = {
                    'session_id': sid,
                    'course_code': r['course_code'] or '',
                    'course_name': r['course_name'] or '',
                    'semester': r['semester_code'] or '',
                    'year_group': r['year_group'],
                    'type': r['teaching_type'] or '',
                    'type_disp': _tp_type_label(r['teaching_type'], r['tp_type']),
                    'formation': {0: 'FTP', 1: 'ALT', 2: 'MUT'}.get(r['formation_type'], '?'),
                    'formation_type': r['formation_type'],
                    'total_hours': r['total_hours'] or 0,
                    'teacher': r['teacher_name'] or '',
                    'room': r['room_name'] or '',
                    'n_groups': n_groups,
                    'by_group': {},   # {group_index: {week: hours}}
                }
                order.append(sid)
            if r['week_number'] is not None and (r['hours'] or 0) > 0:
                g = r['group_index'] or 1
                sessions[sid]['by_group'].setdefault(g, {})[r['week_number']] = r['hours']

        # Développe une ligne par groupe (TP à plusieurs groupes), sinon une ligne
        expanded = []
        for sid in order:
            s = sessions[sid]
            n = s['n_groups']
            common = {k: s[k] for k in ('session_id', 'course_code', 'course_name',
                                        'semester', 'year_group', 'type', 'formation',
                                        'formation_type', 'total_hours', 'teacher', 'room')}
            if n <= 1:
                merged = {}
                for gw in s['by_group'].values():
                    for w, h in gw.items():
                        merged[w] = merged.get(w, 0) + h
                expanded.append({**common, 'type_disp': s['type_disp'],
                                 'group_index': 1, 'n_groups': 1, 'by_week': merged})
            else:
                sem = s['semester']
                # Uniquement les matières réellement mutualisées (session MUT) ont des
                # groupes d'apprentis. Les SAE (face FTP ou ALT) gardent tous leurs
                # groupes dans LEUR cohorte — pas de dernier groupe « ALT ».
                is_common_mut = (sem in mut_set) and (sem not in tpsep_set) and (s['formation'] == 'MUT')
                k_alt = 0
                if is_common_mut:
                    k_alt = int((sg_map.get((sem, 2)) or {}).get('ALT', 1) or 0)
                    k_alt = max(0, min(k_alt, n))
                n_ftp = n - k_alt
                for gi in range(1, n + 1):
                    if gi > n_ftp:
                        rank = gi - n_ftp
                        lbl = f"{s['type_disp']} ALT" if k_alt == 1 else f"{s['type_disp']} ALT {rank}"
                    else:
                        lbl = f"{s['type_disp']} (gr.{gi})"
                    expanded.append({**common, 'type_disp': lbl,
                                     'group_index': gi, 'n_groups': n,
                                     'by_week': dict(s['by_group'].get(gi, {}))})

        # Styles
        thin = Side(style='thin', color='374151')
        thick = Side(style='medium', color='374151')
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        font_header = Font(bold=True, size=9)
        font_data = Font(size=8)
        font_mat = Font(bold=True, size=9, color='4F46E5')
        font_form_ftp = Font(bold=True, size=8, color='0369A1')
        font_form_alt = Font(bold=True, size=8, color='9A3412')
        font_form_mut = Font(bold=True, size=8, color='6B21A8')
        font_total_ok = Font(bold=True, size=8, color='4F46E5')
        font_total_bad = Font(bold=True, size=8, color='DC2626')
        fill_ftp = PatternFill('solid', fgColor='E0F2FE')
        fill_alt = PatternFill('solid', fgColor='FFF7ED')
        fill_mut = PatternFill('solid', fgColor='F3E8FF')
        fill_mat = PatternFill('solid', fgColor='EEF2FF')
        fill_header = PatternFill('solid', fgColor='F3F4F6')
        fill_total = PatternFill('solid', fgColor='EEF2FF')
        # Semaines spéciales (vacances / stage / entreprise) : couleur orange unique,
        # cohérente avec l'affichage à l'écran.
        fill_vacation = PatternFill('solid', fgColor='FDBA74')
        fill_stage    = PatternFill('solid', fgColor='FDBA74')
        fill_company  = PatternFill('solid', fgColor='FDBA74')
        align_center = Alignment(horizontal='center', vertical='center')
        # Objets hoistés (évite de recréer le même style à chaque itération de boucle,
        # ce qui casserait la mémoïsation par id()).
        align_right = Alignment(horizontal='right', vertical='center')
        font_bold8 = Font(bold=True, size=8)
        fill_overload = PatternFill('solid', fgColor='FCA5A5')
        # Annotations (ligne 1) — partagées par tous les onglets
        font_annot = Font(italic=True, size=7, color='666666')
        fill_annot = PatternFill('solid', fgColor='FFFFEE')
        align_annot = Alignment(horizontal='center', vertical='center', text_rotation=90)
        font_annot_lbl = Font(italic=True, size=8)

        wb = Workbook()
        wb.remove(wb.active)

        # Ordre d'affichage imposé des matières S5/S6 (entrelacé), identique à l'écran.
        # Les codes absents conservent leur ordre naturel (tri stable).
        s56_order = {code: i for i, code in enumerate([
            'R5.01', 'R6.01', 'R5.02', 'R5.03a', 'R5.03b', 'R6.02', 'R5.04', 'R6.03',
            'R5.05', 'R6.04', 'R5.06', 'R5.07', 'R5.08', 'R5.09', 'R5.10', 'R6.05',
            'R5.11', 'R6.06', 'R5.12', 'R6.07', 'R5.13', 'R5.14', 'R6.08',
            'SAE5.1', 'SAE5.2', 'SAE5.3', 'SAE6.1', 'SAE6.2',
        ])}
        # Repositionnements ponctuels « placer ce code juste après cet autre code ».
        course_after = {'R3.11': 'R3.08'}

        year_groups = [(1, 'S1+S2', ['S1', 'S2']), (2, 'S3+S4', ['S3', 'S4']), (3, 'S5+S6', ['S5', 'S6'])]

        for yg, sheet_name, sem_codes in year_groups:
            # Filtrer les sessions pour cette année + selon l'affichage (semestre/formation/enseignant)
            year_sessions = [s for s in expanded
                             if s['semester'] in sem_codes and session_included(s)]
            if not year_sessions:
                continue  # onglet non créé s'il n'y a rien à afficher

            ws = wb.create_sheet(title=sheet_name)

            # Semaines spéciales pour cette année
            sp_vacation = sc.get(f'vacation_ftp_y{yg}', set())
            sp_stage = sc.get(f'stage_ftp_y{yg}', set())
            sp_company = sc.get(f'company_alt_y{yg}', set())

            # Grouper par matière
            by_code = {}
            code_order = []
            for s in year_sessions:
                key = s['course_code'] + '|' + s['semester']
                if key not in by_code:
                    by_code[key] = {'code': s['course_code'], 'name': s['course_name'],
                                    'sem': s['semester'], 'ftp': [], 'alt': [], 'mut': []}
                    code_order.append(key)
                if s['formation'] == 'FTP': by_code[key]['ftp'].append(s)
                elif s['formation'] == 'ALT': by_code[key]['alt'].append(s)
                elif s['formation'] == 'MUT': by_code[key]['mut'].append(s)

            # Ordre imposé S5/S6 (entrelacé) ; autres années : ordre naturel conservé.
            code_order.sort(key=lambda k: s56_order.get(by_code[k]['code'], 10**6))
            # Repositionnements ponctuels (ex. R3.11 juste après R3.08)
            for move_code, after_code in course_after.items():
                mi = next((i for i, k in enumerate(code_order) if by_code[k]['code'] == move_code), -1)
                if mi < 0 or not any(by_code[k]['code'] == after_code for k in code_order):
                    continue
                item = code_order.pop(mi)
                ai = next((i for i, k in enumerate(code_order) if by_code[k]['code'] == after_code), -1)
                code_order.insert(ai + 1, item)

            # Colonnes : Formation | Matière | Type | Enseignant | Salle | Heures | Reste | S36 | S37 | ...
            # La colonne Formation est fusionnée verticalement par groupe (FTP/ALT/MUT).
            COL_FORM = 1
            COL_MAT = 2
            COL_TYPE = 3
            COL_TEACH = 4
            COL_ROOM = 5
            COL_TOTAL = 6
            COL_RESTE = 7
            COL_FIRST_WEEK = 8

            nw = len(weeks)
            # Colonne d'appoint (cachée) : formation par ligne, pour les SUMIF des totaux
            # (les cellules fusionnées de COL_FORM sont vides hors ancre, inutilisables en formule).
            COL_FORMKEY = COL_FIRST_WEEK + nw
            first_wk_letter = get_column_letter(COL_FIRST_WEEK)
            last_wk_letter = get_column_letter(COL_FIRST_WEEK + nw - 1)
            formkey_col = get_column_letter(COL_FORMKEY)
            form_range = f'${formkey_col}$6:${formkey_col}$2000'

            # --- Ligne 1 : Annotations (semaines spéciales, éditable) ---
            _sc(ws.cell(row=1, column=1, value='Annotations'), font=font_annot_lbl, fill=fill_annot, alignment=align_center)
            for col in range(2, COL_FIRST_WEEK):
                _sc(ws.cell(row=1, column=col), fill=fill_annot, border=border_all)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_FIRST_WEEK - 1)
            has_comment = False
            for i, w in enumerate(weeks):
                labels = []
                # Code de la catégorie si défini, sinon libellé complet
                if w in sp_vacation:
                    labels.append(category_codes.get(f'vacation_ftp_y{yg}') or 'Vacances')
                if w in sp_stage:
                    labels.append(category_codes.get(f'stage_ftp_y{yg}') or 'Stage FTP')
                if w in sp_company:
                    labels.append(category_codes.get(f'company_alt_y{yg}') or 'Entreprise ALT')
                wc = week_comments.get(w)
                if wc and (not wc['years'] or yg in wc['years']):
                    labels.append(wc['comment'])
                    has_comment = True
                _sc(ws.cell(row=1, column=COL_FIRST_WEEK + i, value=' / '.join(labels) if labels else ''),
                    font=font_annot, fill=fill_annot, alignment=align_annot, border=border_all)
            # Plus de hauteur si des commentaires (texte vertical) sont présents
            ws.row_dimensions[1].height = 140 if has_comment else 60

            # --- Ligne 2 : Totaux FTP (formules SUMIF) ---
            _sc(ws.cell(row=2, column=1, value='FTP'), font=font_form_ftp, fill=fill_ftp, alignment=align_center)
            for col in range(2, COL_FIRST_WEEK):
                _sc(ws.cell(row=2, column=col), fill=fill_ftp, border=border_all)
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=COL_FIRST_WEEK - 1)
            for i in range(nw):
                wk_col = get_column_letter(COL_FIRST_WEEK + i)
                wk_range = f'{wk_col}$6:{wk_col}$2000'
                formula = f'=SUMIF({form_range},"FTP",{wk_range})+SUMIF({form_range},"MUT",{wk_range})'
                _sc(ws.cell(row=2, column=COL_FIRST_WEEK + i, value=formula),
                    font=font_form_ftp, fill=fill_ftp, alignment=align_center, border=border_all, number_format='0.##;-0.##;0')

            # --- Ligne 3 : Totaux ALT (formules SUMIF) ---
            _sc(ws.cell(row=3, column=1, value='ALT'), font=font_form_alt, fill=fill_alt, alignment=align_center)
            for col in range(2, COL_FIRST_WEEK):
                _sc(ws.cell(row=3, column=col), fill=fill_alt, border=border_all)
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=COL_FIRST_WEEK - 1)
            for i in range(nw):
                wk_col = get_column_letter(COL_FIRST_WEEK + i)
                wk_range = f'{wk_col}$6:{wk_col}$2000'
                formula = f'=SUMIF({form_range},"ALT",{wk_range})+SUMIF({form_range},"MUT",{wk_range})'
                _sc(ws.cell(row=3, column=COL_FIRST_WEEK + i, value=formula),
                    font=font_form_alt, fill=fill_alt, alignment=align_center, border=border_all, number_format='0.##;-0.##;0')

            # --- Ligne 4 : En-têtes ---
            headers = ['Formation', 'Matière', 'Type', 'Enseignant', 'Salle', 'Heures', 'Reste']
            for i, h in enumerate(headers):
                _sc(ws.cell(row=4, column=i+1, value=h), font=font_header, fill=fill_header, border=border_all, alignment=align_center)
            for i, w in enumerate(weeks):
                _sc(ws.cell(row=4, column=COL_FIRST_WEEK + i, value=f'S{w:02d}'),
                    font=font_header, fill=fill_header, border=border_all, alignment=align_center)

            # --- Ligne 5+ : Données ---
            row = 5
            total_col_letter = get_column_letter(COL_TOTAL)
            for key in code_order:
                block = by_code[key]
                all_rows = block['mut'] + block['ftp'] + block['alt']
                if not all_rows:
                    continue

                # Ligne en-tête matière (le libellé occupe la colonne Matière)
                ws.cell(row=row, column=COL_MAT, value=f"{block['code']} — {block['name']} ({block['sem']})")
                for col in range(1, COL_FIRST_WEEK + nw):
                    if col == COL_MAT:
                        _sc(ws.cell(row=row, column=col), font=font_mat, fill=fill_mat, border=border_all)
                    else:
                        _sc(ws.cell(row=row, column=col), fill=fill_mat, border=border_all)
                row += 1

                # Ligne « Total promo » : total d'heures de la matière par promo (MUT compté
                # pour FTP et ALT)
                ftp_total = sum((s['total_hours'] or 0) for s in block['mut'] + block['ftp'])
                alt_total = sum((s['total_hours'] or 0) for s in block['mut'] + block['alt'])
                for col in range(1, COL_FIRST_WEEK + nw):
                    _sc(ws.cell(row=row, column=col), fill=fill_total, border=border_all)
                _sc(ws.cell(row=row, column=COL_FORM, value='Total promo'),
                    font=font_bold8, fill=fill_total, border=border_all, alignment=align_right)
                ws.merge_cells(start_row=row, start_column=COL_FORM, end_row=row, end_column=COL_MAT)
                _sc(ws.cell(row=row, column=COL_TYPE, value='FTP'),
                    font=font_form_ftp, fill=fill_ftp, alignment=align_center, border=border_all)
                _sc(ws.cell(row=row, column=COL_TEACH, value=ftp_total or None),
                    font=font_form_ftp, fill=fill_ftp, alignment=align_center, border=border_all, number_format='0.##')
                _sc(ws.cell(row=row, column=COL_ROOM, value='ALT'),
                    font=font_form_alt, fill=fill_alt, alignment=align_center, border=border_all)
                _sc(ws.cell(row=row, column=COL_TOTAL, value=alt_total or None),
                    font=font_form_alt, fill=fill_alt, alignment=align_center, border=border_all, number_format='0.##')
                row += 1

                # Une « bande » par formation : libellé fusionné verticalement dans COL_FORM
                for form, grp in (('MUT', block['mut']), ('FTP', block['ftp']), ('ALT', block['alt'])):
                    if not grp:
                        continue
                    group_start = row
                    form_font = font_form_mut if form == 'MUT' else font_form_ftp if form == 'FTP' else font_form_alt
                    form_fill = fill_mut if form == 'MUT' else fill_ftp if form == 'FTP' else fill_alt

                    for s in grp:
                        # Colonne Formation : style posé sur chaque cellule (avant fusion)
                        _sc(ws.cell(row=row, column=COL_FORM), fill=form_fill, border=border_all, alignment=align_center)
                        # Colonne d'appoint cachée : formation par ligne (pour les SUMIF)
                        _sc(ws.cell(row=row, column=COL_FORMKEY, value=form), font=font_data)
                        _sc(ws.cell(row=row, column=COL_TYPE, value=s.get('type_disp') or s['type']),
                            font=font_data, alignment=align_center, border=border_all)
                        _sc(ws.cell(row=row, column=COL_TEACH, value=s['teacher']), font=font_data, border=border_all)
                        _sc(ws.cell(row=row, column=COL_ROOM, value=s['room']), font=font_data, border=border_all)
                        # Heures prévues
                        _sc(ws.cell(row=row, column=COL_TOTAL, value=s['total_hours'] if s['total_hours'] > 0 else None),
                            font=font_data, alignment=align_center, border=border_all)
                        # Reste (formule Excel)
                        reste_formula = f'={total_col_letter}{row}-SUM({first_wk_letter}{row}:{last_wk_letter}{row})'
                        _sc(ws.cell(row=row, column=COL_RESTE, value=reste_formula),
                            font=font_total_ok, fill=fill_total, alignment=align_center, border=border_all, number_format='0.##;-0.##;0')

                        for i, w in enumerate(weeks):
                            h = s['by_week'].get(w, 0)
                            fillw = None
                            if form == 'MUT':
                                # CM/TD mutualisés : FTP (stage/vacances) + ALT (entreprise)
                                if w in sp_stage or w in sp_vacation or w in sp_company: fillw = fill_stage
                            elif form == 'FTP':
                                if w in sp_stage: fillw = fill_stage
                                elif w in sp_vacation: fillw = fill_vacation
                            elif form == 'ALT':
                                if w in sp_company: fillw = fill_company
                            _sc(ws.cell(row=row, column=COL_FIRST_WEEK + i, value=h if h else None),
                                font=font_data, fill=fillw, alignment=align_center, border=border_all, number_format='0.##;-0.##;0')

                        row += 1

                    # Libellé de la formation dans la cellule fusionnée (ancre = 1re ligne)
                    group_end = row - 1
                    _sc(ws.cell(row=group_start, column=COL_FORM, value=form),
                        font=form_font, fill=form_fill, alignment=align_center, border=border_all)
                    if group_end > group_start:
                        ws.merge_cells(start_row=group_start, start_column=COL_FORM,
                                       end_row=group_end, end_column=COL_FORM)

            # Mise en forme conditionnelle : Reste rouge si != 0
            reste_letter = get_column_letter(COL_RESTE)
            reste_range_cf = f'{reste_letter}6:{reste_letter}{max(row - 1, 6)}'
            ws.conditional_formatting.add(reste_range_cf, CellIsRule(
                operator='notEqual', formula=['0'],
                font=Font(bold=True, size=8, color='DC2626'),
                fill=PatternFill('solid', fgColor='FEE2E2')
            ))

            # Largeurs de colonnes
            ws.column_dimensions[get_column_letter(COL_FORM)].width = 5
            ws.column_dimensions[get_column_letter(COL_MAT)].width = 6
            ws.column_dimensions[get_column_letter(COL_TYPE)].width = 5
            ws.column_dimensions[get_column_letter(COL_TEACH)].width = 16
            ws.column_dimensions[get_column_letter(COL_ROOM)].width = 10
            ws.column_dimensions[get_column_letter(COL_TOTAL)].width = 6
            ws.column_dimensions[get_column_letter(COL_RESTE)].width = 6
            for i in range(nw):
                ws.column_dimensions[get_column_letter(COL_FIRST_WEEK + i)].width = 4.5
            # Colonne d'appoint formation : cachée (sert uniquement aux SUMIF)
            ws.column_dimensions[get_column_letter(COL_FORMKEY)].hidden = True

            # Figer les volets
            ws.freeze_panes = ws.cell(row=5, column=COL_FIRST_WEEK)

        # --- Feuille Contacts enseignants ---
        ws_contacts = wb.create_sheet(title='Contacts')
        contact_headers = ['Nom', 'Email', 'Téléphone', 'Structure', 'Corps', 'Statut']
        for i, h in enumerate(contact_headers, 1):
            c = ws_contacts.cell(row=1, column=i, value=h)
            c.font = font_header
            c.fill = fill_header
            c.border = border_all
            c.alignment = align_center

        cursor.execute('SELECT name, email, phone, structure, corps_code, status FROM teachers ORDER BY name')
        teachers_list = cursor.fetchall()
        # Ne garder que les enseignants présents dans l'affichage filtré
        incl_names = {(s['teacher'] or '').strip()
                      for s in expanded if session_included(s)}
        incl_names.discard('')
        teachers_list = [t for t in teachers_list if t['name'] in incl_names]
        for r_idx, t in enumerate(teachers_list, 2):
            for col_idx, field in enumerate(['name', 'email', 'phone', 'structure', 'corps_code', 'status'], 1):
                c = ws_contacts.cell(row=r_idx, column=col_idx, value=t[field])
                c.font = font_data
                c.border = border_all

        ws_contacts.column_dimensions['A'].width = 25
        ws_contacts.column_dimensions['B'].width = 30
        ws_contacts.column_dimensions['C'].width = 15
        ws_contacts.column_dimensions['D'].width = 20
        ws_contacts.column_dimensions['E'].width = 15
        ws_contacts.column_dimensions['F'].width = 12
        last_contact_row = len(teachers_list) + 1
        ws_contacts.auto_filter.ref = f'A1:F{last_contact_row}'
        ws_contacts.freeze_panes = 'A2'

        # --- Un onglet par enseignant : bilan + calendrier de ses heures ---
        import re

        def _safe_sheet_name(name, used):
            base = re.sub(r'[:\\/?*\[\]]', ' ', name).strip() or 'Enseignant'
            base = base[:31]
            title, n = base, 2
            while title.lower() in used:
                suffix = f' ({n})'
                title = base[:31 - len(suffix)] + suffix
                n += 1
            used.add(title.lower())
            return title

        # Grouper les sessions (filtrées selon l'affichage) par enseignant
        by_teacher = {}
        for s in expanded:
            if not session_included(s):
                continue
            tname = (s['teacher'] or '').strip()
            if tname:
                by_teacher.setdefault(tname, []).append(s)

        used_titles = {'contacts', 's1+s2', 's3+s4', 's5+s6'}
        type_rank = {'CM': 0, 'TD': 1, 'TP': 2, 'PT': 3}
        # (fill_overload, font_annot, fill_annot, align_annot : hoistés en tête)

        # Colonnes au format Répartition (sans la colonne Enseignant, redondante ici)
        TF_FORM, TF_MAT, TF_TYPE, TF_SALLE, TF_TOTAL, TF_RESTE = 1, 2, 3, 4, 5, 6
        TF_FIRST = 7
        nw = len(weeks)
        TF_FORMKEY = TF_FIRST + nw
        tf_total_letter = get_column_letter(TF_TOTAL)
        tf_first_letter = get_column_letter(TF_FIRST)
        tf_last_letter = get_column_letter(TF_FIRST + nw - 1)
        tf_formkey_letter = get_column_letter(TF_FORMKEY)
        tf_form_range = f'${tf_formkey_letter}$5:${tf_formkey_letter}$2000'

        for tname in sorted(by_teacher.keys(), key=lambda x: x.lower()):
            tsessions = by_teacher[tname]
            ws = wb.create_sheet(title=_safe_sheet_name(tname, used_titles))

            teacher_ygs = sorted({s['year_group'] for s in tsessions if s['year_group']})
            sp_vac = set()
            for g in teacher_ygs:
                sp_vac |= sc.get(f'vacation_ftp_y{g}', set())
            total_prev = sum(s['total_hours'] or 0 for s in tsessions)
            total_placed = sum(sum(s['by_week'].get(w, 0) for w in weeks) for s in tsessions)

            # --- Ligne 1 : nom + bilan (gauche) + annotations/codes par semaine (vertical) ---
            c = ws.cell(row=1, column=1,
                        value=f'{tname}  —  Prévu {total_prev:g} h  /  Placé {total_placed:g} h  /  '
                              f'Reste {total_prev - total_placed:g} h')
            c.font = Font(bold=True, size=10, color='4F46E5')
            c.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TF_RESTE)
            has_comment = False
            for i, w in enumerate(weeks):
                labels = []
                if w in sp_vac:
                    labels.append('Vacances')
                for g in teacher_ygs:
                    if w in sc.get(f'stage_ftp_y{g}', set()):
                        labels.append(category_codes.get(f'stage_ftp_y{g}') or 'Stage FTP')
                    if w in sc.get(f'company_alt_y{g}', set()):
                        labels.append(category_codes.get(f'company_alt_y{g}') or 'Entreprise ALT')
                seen = set()
                labels = [x for x in labels if not (x in seen or seen.add(x))]
                _sc(ws.cell(row=1, column=TF_FIRST + i, value=' / '.join(labels) if labels else ''),
                    font=font_annot, fill=fill_annot, alignment=align_annot, border=border_all)
                if labels:
                    has_comment = True
            ws.row_dimensions[1].height = 140 if has_comment else 40

            # --- Lignes 2 et 3 : totaux FTP / ALT (formules SUMIF) ---
            for trow, crit, lab_fill, lab_font in (
                (2, 'FTP', fill_ftp, font_form_ftp),
                (3, 'ALT', fill_alt, font_form_alt),
            ):
                _sc(ws.cell(row=trow, column=1, value=crit), font=lab_font, fill=lab_fill, alignment=align_center, border=border_all)
                for col in range(2, TF_FIRST):
                    _sc(ws.cell(row=trow, column=col), fill=lab_fill, border=border_all)
                ws.merge_cells(start_row=trow, start_column=1, end_row=trow, end_column=TF_FIRST - 1)
                for i in range(nw):
                    wk_col = get_column_letter(TF_FIRST + i)
                    wk_range = f'{wk_col}$5:{wk_col}$2000'
                    _sc(ws.cell(row=trow, column=TF_FIRST + i,
                                value=f'=SUMIF({tf_form_range},"{crit}",{wk_range})+SUMIF({tf_form_range},"MUT",{wk_range})'),
                        font=lab_font, fill=lab_fill, alignment=align_center, border=border_all, number_format='0.##;-0.##;0')

            # --- Ligne 4 : en-têtes ---
            for i, h in enumerate(['Form.', 'Matière', 'Type', 'Salle', 'Heures', 'Reste'], 1):
                _sc(ws.cell(row=4, column=i, value=h), font=font_header, fill=fill_header, border=border_all, alignment=align_center)
            for i, w in enumerate(weeks):
                _sc(ws.cell(row=4, column=TF_FIRST + i, value=f'S{w:02d}'),
                    font=font_header, fill=fill_header, border=border_all, alignment=align_center)

            # --- Lignes 5+ : données groupées par matière (colonne Formation fusionnée) ---
            by_code, code_order = {}, []
            for s in sorted(tsessions, key=lambda s: (s['semester'], s['course_code'], type_rank.get(s['type'], 9))):
                key = s['course_code'] + '|' + s['semester']
                if key not in by_code:
                    by_code[key] = {'code': s['course_code'], 'name': s['course_name'],
                                    'sem': s['semester'], 'mut': [], 'ftp': [], 'alt': []}
                    code_order.append(key)
                f = s['formation']
                (by_code[key]['mut'] if f == 'MUT' else by_code[key]['ftp'] if f == 'FTP' else by_code[key]['alt']).append(s)

            row = 5
            week_totals = {w: 0 for w in weeks}
            for key in code_order:
                block = by_code[key]
                ws.cell(row=row, column=TF_MAT, value=f"{block['code']} — {block['name']} ({block['sem']})")
                for col in range(1, TF_FIRST + nw):
                    if col == TF_MAT:
                        _sc(ws.cell(row=row, column=col), font=font_mat, fill=fill_mat, border=border_all)
                    else:
                        _sc(ws.cell(row=row, column=col), fill=fill_mat, border=border_all)
                row += 1

                for form, grp in (('MUT', block['mut']), ('FTP', block['ftp']), ('ALT', block['alt'])):
                    if not grp:
                        continue
                    group_start = row
                    form_font = font_form_mut if form == 'MUT' else font_form_ftp if form == 'FTP' else font_form_alt
                    form_fill = fill_mut if form == 'MUT' else fill_ftp if form == 'FTP' else fill_alt
                    for s in grp:
                        yg_s = s['year_group']
                        sp_v = sc.get(f'vacation_ftp_y{yg_s}', set()) if yg_s else set()
                        sp_st = sc.get(f'stage_ftp_y{yg_s}', set()) if yg_s else set()
                        sp_co = sc.get(f'company_alt_y{yg_s}', set())
                        _sc(ws.cell(row=row, column=TF_FORM), fill=form_fill, border=border_all, alignment=align_center)
                        _sc(ws.cell(row=row, column=TF_FORMKEY, value=form), font=font_data)
                        _sc(ws.cell(row=row, column=TF_TYPE, value=s.get('type_disp') or s['type']),
                            font=font_data, border=border_all, alignment=align_center)
                        _sc(ws.cell(row=row, column=TF_SALLE, value=s['room']),
                            font=font_data, border=border_all, alignment=align_center)
                        _sc(ws.cell(row=row, column=TF_TOTAL, value=s['total_hours'] or None),
                            font=font_data, border=border_all, alignment=align_center)
                        _sc(ws.cell(row=row, column=TF_RESTE,
                                    value=f'={tf_total_letter}{row}-SUM({tf_first_letter}{row}:{tf_last_letter}{row})'),
                            font=font_total_ok, fill=fill_total, border=border_all, alignment=align_center, number_format='0.##;-0.##;0')
                        for i, w in enumerate(weeks):
                            h = s['by_week'].get(w, 0)
                            week_totals[w] += h
                            fillw = None
                            if form == 'MUT':
                                if w in sp_st or w in sp_v or w in sp_co: fillw = fill_stage
                            elif form == 'FTP':
                                if w in sp_st: fillw = fill_stage
                                elif w in sp_v: fillw = fill_vacation
                            elif form == 'ALT':
                                if w in sp_co: fillw = fill_company
                            _sc(ws.cell(row=row, column=TF_FIRST + i, value=h if h else None),
                                font=font_data, fill=fillw, alignment=align_center, border=border_all, number_format='0.##;-0.##;0')
                        row += 1
                    group_end = row - 1
                    _sc(ws.cell(row=group_start, column=TF_FORM, value=form),
                        font=form_font, fill=form_fill, alignment=align_center, border=border_all)
                    if group_end > group_start:
                        ws.merge_cells(start_row=group_start, start_column=TF_FORM,
                                       end_row=group_end, end_column=TF_FORM)

            # Ligne charge totale de l'enseignant / semaine (toutes formations, MUT compté 1 fois)
            _sc(ws.cell(row=row, column=TF_MAT, value='Total / sem.'), font=font_bold8, fill=fill_header, border=border_all, alignment=align_center)
            _sc(ws.cell(row=row, column=TF_TOTAL, value=total_placed or None), font=font_bold8, fill=fill_header, border=border_all, alignment=align_center)
            for col in range(1, TF_FIRST):
                if col in (TF_MAT, TF_TOTAL):
                    continue
                _sc(ws.cell(row=row, column=col), fill=fill_header, border=border_all, alignment=align_center)
            for i, w in enumerate(weeks):
                tot = week_totals[w]
                _sc(ws.cell(row=row, column=TF_FIRST + i, value=tot if tot else None),
                    font=font_bold8, fill=(fill_overload if tot > 20 else fill_header),
                    alignment=align_center, border=border_all, number_format='0.##;-0.##;0')
            total_row = row

            # Mise en forme conditionnelle : Reste rouge si != 0
            reste_letter = get_column_letter(TF_RESTE)
            ws.conditional_formatting.add(
                f'{reste_letter}5:{reste_letter}{max(total_row - 1, 5)}',
                CellIsRule(operator='notEqual', formula=['0'],
                           font=Font(bold=True, size=8, color='DC2626'),
                           fill=PatternFill('solid', fgColor='FEE2E2')))

            # Largeurs + colonne d'appoint cachée + volets figés
            for col, wd in ((TF_FORM, 5), (TF_MAT, 16), (TF_TYPE, 5), (TF_SALLE, 10), (TF_TOTAL, 6), (TF_RESTE, 6)):
                ws.column_dimensions[get_column_letter(col)].width = wd
            for i in range(nw):
                ws.column_dimensions[get_column_letter(TF_FIRST + i)].width = 4.5
            ws.column_dimensions[get_column_letter(TF_FORMKEY)].hidden = True
            ws.freeze_panes = ws.cell(row=5, column=TF_FIRST)

        # Générer le fichier
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        year = get_current_year() or 'export'
        filename = f'repartition_{year}.xlsx'
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        return error_response(f'Error exporting: {str(e)}', 500)

@app.route('/api/import/repartition', methods=['POST'])
def import_repartition_excel():
    """Importe la répartition depuis un fichier Excel au format export"""
    from openpyxl import load_workbook

    if 'file' not in request.files:
        return error_response('Aucun fichier fourni', 400)

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return error_response('Le fichier doit être au format .xlsx', 400)

    try:
        wb = load_workbook(file, data_only=True)
        db = get_db()
        cursor = db.cursor()

        formation_map = {'FTP': 0, 'ALT': 1, 'MUT': 2}
        imported = 0
        errors = []

        year_sheet_names = {'S1+S2', 'S3+S4', 'S5+S6'}
        for ws in wb.worksheets:
            # N'importer que les onglets de répartition par année (ignorer Contacts + onglets enseignants)
            if ws.title not in year_sheet_names:
                continue

            # Trouver la ligne d'en-tête (celle qui contient "Matière") et localiser les
            # colonnes par leur libellé (l'ordre peut varier : Formation est désormais à gauche).
            header_row = None
            col_mat = col_type = col_form = None
            for r in range(1, 10):
                labels = {}
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is not None:
                        labels[str(v).strip()] = c
                if 'Matière' in labels:
                    header_row = r
                    col_mat = labels.get('Matière')
                    col_type = labels.get('Type')
                    col_form = labels.get('Formation')
                    break

            if header_row is None or not col_mat or not col_type or not col_form:
                errors.append(f"Onglet {ws.title}: en-tête introuvable")
                continue

            # Extraire les numéros de semaine depuis les en-têtes (S36, S37, ...)
            week_columns = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=header_row, column=col).value
                if val and str(val).startswith('S') and len(str(val)) == 3:
                    try:
                        week_num = int(str(val)[1:])
                        week_columns.append((col, week_num))
                    except ValueError:
                        pass

            if not week_columns:
                errors.append(f"Onglet {ws.title}: colonnes semaines introuvables")
                continue

            # Parcourir les lignes de données
            current_code = None
            current_sem = None
            last_form = None  # colonne Formation fusionnée : report de la dernière valeur lue

            for r in range(header_row + 1, ws.max_row + 1):
                mat_val = ws.cell(row=r, column=col_mat).value
                type_val = ws.cell(row=r, column=col_type).value
                form_cell = ws.cell(row=r, column=col_form).value
                # Cellule fusionnée : seule l'ancre porte la valeur → report sur les lignes suivantes
                if form_cell is not None and str(form_cell).strip():
                    last_form = str(form_cell).strip()
                form_val = last_form

                # Ligne en-tête matière : "R1.01 — Mécanique (S1)"
                if mat_val and not type_val:
                    mat_str = str(mat_val)
                    parts = mat_str.split(' — ', 1)
                    if len(parts) == 2:
                        current_code = parts[0].strip()
                        name_part = parts[1]
                        if '(' in name_part and name_part.endswith(')'):
                            sem_start = name_part.rfind('(')
                            current_sem = name_part[sem_start + 1:-1].strip()
                    continue

                if not type_val or not form_val:
                    continue
                if not current_code or not current_sem:
                    continue

                teaching_type = str(type_val).strip()
                formation_str = str(form_val).strip()
                formation_type = formation_map.get(formation_str)
                if formation_type is None:
                    continue

                # Trouver la session correspondante
                cursor.execute('''
                    SELECT cs.id FROM course_sessions cs
                    JOIN courses c ON cs.course_id = c.id
                    JOIN semesters s ON c.semester_id = s.id
                    WHERE c.code = ? AND s.code = ?
                      AND cs.teaching_type = ? AND cs.formation_type = ?
                ''', (current_code, current_sem, teaching_type, formation_type))

                row_result = cursor.fetchone()
                if not row_result:
                    errors.append(f"{current_code} {current_sem} {teaching_type} {formation_str}: session introuvable")
                    continue

                session_id = row_result['id']

                # Importer les heures par semaine
                for col, week_num in week_columns:
                    cell_val = ws.cell(row=r, column=col).value
                    try:
                        hours = float(cell_val) if cell_val else 0
                    except (ValueError, TypeError):
                        hours = 0

                    if hours > 0:
                        cursor.execute('''
                            INSERT OR REPLACE INTO weekly_hours (course_session_id, week_number, hours)
                            VALUES (?, ?, ?)
                        ''', (session_id, week_num, hours))
                        imported += 1
                    else:
                        cursor.execute(
                            'DELETE FROM weekly_hours WHERE course_session_id = ? AND week_number = ?',
                            (session_id, week_num)
                        )

        db.commit()
        return jsonify({
            'imported': imported,
            'errors': errors,
            'message': f'{imported} cellules importées' + (f', {len(errors)} erreurs' if errors else '')
        })
    except Exception as e:
        return error_response(f'Erreur import: {str(e)}', 500)

# ======================= ANNÉES UNIVERSITAIRES =======================

@app.route('/api/years', methods=['GET'])
def get_years():
    return jsonify({'years': list_years(), 'current': get_current_year()})

@app.route('/api/years', methods=['POST'])
def create_year():
    data = request.get_json() or {}
    new_year = (data.get('year') or '').strip()
    copy_from = (data.get('copy_from') or '').strip()

    if len(new_year) != 9 or new_year[4] != '-' or not new_year[:4].isdigit() or not new_year[5:].isdigit():
        return error_response('Format invalide — attendu AAAA-AAAA (ex : 2025-2026)', 400)

    new_path = db_path_for_year(new_year)
    if os.path.exists(new_path):
        return error_response(f"L'année {new_year} existe déjà", 400)

    if copy_from:
        src_path = db_path_for_year(copy_from)
        if not os.path.exists(src_path):
            return error_response(f"Année source {copy_from} introuvable", 404)
        os.makedirs(os.path.dirname(new_path), exist_ok=True)
        shutil.copy2(src_path, new_path)
        # S'assurer que la copie est à jour schéma
        db = sqlite3.connect(new_path)
        db.row_factory = sqlite3.Row
        _apply_migrations(db)
        db.close()
    else:
        _init_fresh_db(new_path)

    return jsonify({'year': new_year, 'years': list_years()}), 201

@app.route('/api/years/current', methods=['PUT'])
def set_current_year():
    """Change le défaut global de l'établissement (admin uniquement)."""
    err = _require_admin()
    if err:
        return err
    data = request.get_json() or {}
    year = (data.get('year') or '').strip()
    if not year:
        return error_response('year requis', 400)
    if not os.path.exists(db_path_for_year(year)):
        return error_response(f"Année {year} introuvable", 404)
    settings = get_settings()
    settings['current_year'] = year
    save_settings(settings)
    return jsonify({'current': year})

@app.route('/api/years/session', methods=['PUT'])
def set_session_year():
    """Choisit l'année universitaire pour CETTE session uniquement.
    Vue personnelle (enseignant ou admin) sans impact sur le défaut global."""
    if not session.get('role'):
        return error_response('Authentification requise', 401)
    data = request.get_json() or {}
    year = (data.get('year') or '').strip()
    if not year:
        # Réinitialise : revient au défaut global de l'établissement
        session.pop('year', None)
        return jsonify({'current': get_current_year()})
    if not os.path.exists(db_path_for_year(year)):
        return error_response(f"Année {year} introuvable", 404)
    session['year'] = year
    return jsonify({'current': year})

# ======================= ERROR HANDLERS =======================

@app.errorhandler(404)
def not_found(error):
    """Handle 404 errors"""
    return error_response('Resource not found', 404)

@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors"""
    return error_response('Internal server error', 500)

# ======================= OPTIMISATION CP-SAT =======================

@app.route('/api/course-orderings', methods=['GET'])
def get_course_orderings():
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            SELECT co.id, co.min_gap_weeks,
                   cp.id AS pred_id, cp.code AS pred_code, cp.name AS pred_name,
                   cs2.id AS succ_id, cs2.code AS succ_code, cs2.name AS succ_name
            FROM course_ordering co
            JOIN courses cp  ON co.course_id_pred = cp.id
            JOIN courses cs2 ON co.course_id_succ = cs2.id
            ORDER BY cp.code, cs2.code
        ''')
        rows = rows_to_list(cursor.fetchall())
        return jsonify(rows), 200
    except Exception as e:
        return error_response(str(e), 500)

@app.route('/api/course-orderings', methods=['POST'])
def create_course_ordering():
    try:
        data = request.get_json() or {}
        pred = data.get('course_id_pred')
        succ = data.get('course_id_succ')
        gap  = int(data.get('min_gap_weeks', 0))
        if not pred or not succ:
            return error_response('course_id_pred et course_id_succ requis')
        if pred == succ:
            return error_response('Une matière ne peut pas être contrainte par elle-même')
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            'INSERT INTO course_ordering (course_id_pred, course_id_succ, min_gap_weeks) VALUES (?, ?, ?)',
            (pred, succ, gap)
        )
        db.commit()
        oid = cursor.lastrowid
        cursor.execute('''
            SELECT co.id, co.min_gap_weeks,
                   cp.id AS pred_id, cp.code AS pred_code, cp.name AS pred_name,
                   cs2.id AS succ_id, cs2.code AS succ_code, cs2.name AS succ_name
            FROM course_ordering co
            JOIN courses cp  ON co.course_id_pred = cp.id
            JOIN courses cs2 ON co.course_id_succ = cs2.id
            WHERE co.id = ?
        ''', (oid,))
        row = row_to_dict(cursor.fetchone())
        return jsonify(row), 201
    except sqlite3.IntegrityError:
        return error_response('Cette contrainte existe déjà')
    except Exception as e:
        return error_response(str(e), 500)

@app.route('/api/course-orderings/<int:oid>', methods=['DELETE'])
def delete_course_ordering(oid):
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('DELETE FROM course_ordering WHERE id = ?', (oid,))
        db.commit()
        return jsonify({'message': 'Contrainte supprimée'}), 200
    except Exception as e:
        return error_response(str(e), 500)

@app.route('/api/optimize', methods=['POST'])
def optimize_repartition():
    """Optimise la répartition hebdomadaire via OR-Tools CP-SAT.
    Body: { "semester": "S1+S2" }  ou  { "semester": "S1" }
    Écrase weekly_hours pour les sessions concernées.
    """
    try:
        from ortools.sat.python import cp_model
    except ImportError:
        return error_response('OR-Tools non installé. Exécuter : pip install ortools', 500)

    data = request.get_json() or {}
    semester_param = data.get('semester', '')
    formation_param = data.get('formation', '')  # '', 'ftp', 'alt'

    sem_codes = semester_param.split('+') if '+' in semester_param else (
        [semester_param] if semester_param else []
    )

    # Filtrage par formation : ftp → ft in {0,2}, alt → ft in {1,2}, '' → tous
    if formation_param == 'ftp':
        ft_filter = {0, 2}
    elif formation_param == 'alt':
        ft_filter = {1, 2}
    else:
        ft_filter = {0, 1, 2}

    try:
        cfg = get_year_config()
        start_week   = cfg['academic_start_week']
        end_week_cfg = cfg['academic_end_week']
        max_week     = cfg['academic_max_week']
        # Plage (début/fin + nb semaines 52/53) propre à l'année de progression
        # optimisée — un appel = un semestre = une seule année (repli global).
        if sem_codes:
            try:
                yg0 = (int(sem_codes[0].lstrip('S')) + 1) // 2
                start_week, end_week_cfg, max_week = year_range_for_yg(cfg, yg0)
            except (ValueError, IndexError):
                pass
        _key = lambda w: school_week_key(w, start_week, max_week)
        all_valid_weeks = get_valid_school_weeks(start_week, end_week_cfg, max_week)

        db = get_db()
        cursor = db.cursor()

        # --- Sessions à optimiser ---
        q_sessions = '''
            SELECT cs.id, cs.nb_sessions, cs.slot_duration, cs.formation_type,
                   cs.teaching_type,
                   COALESCE(cs.sessions_per_week_max, 1) AS sessions_per_week_max,
                   cs.room_name,
                   c.id   AS course_id,
                   c.code AS course_code,
                   c.name AS course_name,
                   c.start_week AS course_start_week,
                   c.end_week   AS course_end_week,
                   c.default_weeks AS course_default_weeks,
                   s.code AS semester_code,
                   s.year_group AS year_group
            FROM course_sessions cs
            JOIN courses c  ON cs.course_id  = c.id
            JOIN semesters s ON c.semester_id = s.id
            WHERE {where} AND cs.nb_sessions > 0
            ORDER BY cs.id
        '''
        if sem_codes:
            ph = ','.join('?' * len(sem_codes))
            cursor.execute(q_sessions.format(where=f's.code IN ({ph})'), sem_codes)
        else:
            cursor.execute(q_sessions.format(where='1=1'))
        sessions = [s for s in cursor.fetchall() if s['formation_type'] in ft_filter]

        if not sessions:
            return jsonify({'message': 'Aucune session à optimiser', 'weeks_assigned': 0}), 200

        # --- Calendrier spécial ---
        cursor.execute('SELECT week_number, week_type FROM special_calendar')
        sc = {t: set() for t in _SPECIAL_TYPES}
        for r in cursor.fetchall():
            if r['week_type'] in sc:
                sc[r['week_type']].add(r['week_number'])

        # --- Salles par type ---
        cursor.execute('SELECT name, room_type FROM rooms')
        room_type_by_name = {r['name']: r['room_type'] for r in cursor.fetchall()}
        cursor.execute('SELECT room_type, COUNT(*) AS cnt FROM rooms GROUP BY room_type')
        room_counts = {r['room_type']: r['cnt'] for r in cursor.fetchall()}

        # --- Contraintes d'ordonnancement ---
        course_ids_in_scope = {s['course_id'] for s in sessions}
        cursor.execute('''
            SELECT id, course_id_pred, course_id_succ, min_gap_weeks
            FROM course_ordering
        ''')
        orderings = [r for r in cursor.fetchall()
                     if r['course_id_pred'] in course_ids_in_scope
                     and r['course_id_succ'] in course_ids_in_scope]


        # Index des cours → sessions
        course_to_sessions = {}
        for s in sessions:
            course_to_sessions.setdefault(s['course_id'], []).append(s['id'])

        # =========================================================
        # Construction du modèle CP-SAT
        # =========================================================
        model  = cp_model.CpModel()
        SCALE  = 10      # 1h → 10 unités (gère les 0.5h)
        ROOM_H = 30      # heures disponibles par salle par semaine
        MAX_H  = 40      # plafond souple heures/semaine/formation
        ABS_MAX_H = 50   # plafond absolu infranchissable
        TGT_FTP = 25     # cible heures/semaine FTP
        TGT_ALT = 35     # cible heures/semaine ALT

        sorted_weeks  = sorted(all_valid_weeks, key=_key)
        week_to_idx   = {w: i for i, w in enumerate(sorted_weeks)}

        _FT_LABELS = {0: 'FTP', 1: 'ALT', 2: 'MUT'}

        # === Fusion CM+TD : même cours/formation → mêmes variables x ===
        _CMTD = {'CM', 'TD'}
        _by_cf = {}   # (course_id, ft) → [sessions CM/TD]
        for s in sessions:
            if s['teaching_type'] in _CMTD:
                _by_cf.setdefault((s['course_id'], s['formation_type']), []).append(s)

        _merged_secondary = {}   # secondary_sid → primary_sid
        _merged_nb = {}          # primary_sid → max(nb) du groupe
        for group in _by_cf.values():
            if len(group) < 2:
                continue
            group.sort(key=lambda s: s['teaching_type'])   # CM avant TD
            primary = group[0]
            max_nb = max(s['nb_sessions'] for s in group)
            _merged_nb[primary['id']] = max_nb
            for sec in group[1:]:
                _merged_secondary[sec['id']] = primary['id']

        x             = {}   # x[sid][week] = IntVar/BoolVar(nb séances)
        session_weeks = {}   # sid → liste des semaines valides triées (numérotation parallèle)
        infeasible_sids = {}  # sid → reason string
        _contig_data  = {}   # sid → (n_valid, acts_list, min_span) pour pénalité souple

        for s in sessions:
            sid    = s['id']

            # Les sessions secondaires (TD fusionné avec CM) sont traitées après
            if sid in _merged_secondary:
                continue

            ft     = s['formation_type']   # 0=FTP 1=ALT 2=MUT
            sw, ew = resolve_course_weeks({
                'default_weeks': s['course_default_weeks'],
                'semester_code': s['semester_code'],
                'start_week':    s['course_start_week'],
                'end_week':      s['course_end_week'],
            }, cfg)
            nb     = _merged_nb.get(sid, s['nb_sessions'])  # nb fusionné si CM+TD
            max_pw = max(1, int(s['sessions_per_week_max'] or 1))

            valid = get_valid_school_weeks(sw, ew, max_week) & all_valid_weeks if (sw and ew) else set(all_valid_weeks)
            yg = s['year_group']
            excluded_types = []
            if ft == 0:   # FTP : vacances + stages selon année
                valid -= sc.get(f'vacation_ftp_y{yg}', set())
                excluded_types.append('vacances FTP')
                stg = sc.get(f'stage_ftp_y{yg}', set())
                if stg:
                    valid -= stg
                    excluded_types.append(f'stages FTP {yg}A')
            elif ft == 1:   # ALT : semaines entreprise selon année
                excl = sc.get(f'company_alt_y{yg}', set())
                valid -= excl
                if excl:
                    excluded_types.append(f'entreprise ALT {yg}A')
            elif ft == 2:   # MUT : compte pour FTP → exclure vacances FTP
                valid -= sc.get(f'vacation_ftp_y{yg}', set())
                excluded_types.append('vacances FTP')

            sw_list = sorted(valid, key=_key)
            session_weeks[sid] = sw_list
            x[sid] = {}

            sess_label = f"{s['course_code']} {s['teaching_type']} {_FT_LABELS[ft]}"

            if not sw_list:
                reason = f"Aucune semaine valide"
                if sw and ew:
                    reason += f" (plage S{sw}→S{ew}"
                    if excluded_types:
                        reason += f", excl. {', '.join(excluded_types)}"
                    reason += ")"
                infeasible_sids[sid] = {'label': sess_label, 'reason': reason}
                continue

            n_valid   = len(sw_list)
            min_span  = math.ceil(nb / max_pw)   # semaines minimales nécessaires

            if n_valid < min_span:
                infeasible_sids[sid] = {
                    'label': sess_label,
                    'reason': f"{n_valid} semaines valides, besoin de {min_span} min. "
                              f"({nb} séances, max {max_pw}/sem)"
                }
                continue

            # ---------------------------------------------------------------
            # Variables de session par semaine.
            # Contiguïté gérée en souple dans l'objectif (pénalité de trous).
            # ---------------------------------------------------------------
            acts_list = []
            for i, w in enumerate(sw_list):
                xv  = model.new_int_var(0, max_pw + 1, f'x_{sid}_{w}')
                act = model.new_bool_var(f'act_{sid}_{i}')
                model.add(xv >= 1).only_enforce_if(act)
                model.add(xv == 0).only_enforce_if(act.Not())
                x[sid][w] = xv
                acts_list.append(act)

            # Nombre total de séances = nb
            model.add(cp_model.LinearExpr.Sum(
                [x[sid][w] for w in sw_list]
            ) == nb)

            # Mémoriser pour la pénalité de contiguïté (objectif)
            _contig_data[sid] = (n_valid, acts_list, min_span)

        # === Alias des sessions secondaires (TD fusionné avec CM) ===
        for sec_sid, pri_sid in _merged_secondary.items():
            if pri_sid in infeasible_sids:
                infeasible_sids[sec_sid] = infeasible_sids[pri_sid]
                session_weeks[sec_sid] = session_weeks.get(pri_sid, [])
                x[sec_sid] = {}
            else:
                x[sec_sid] = x[pri_sid]
                session_weeks[sec_sid] = session_weeks[pri_sid]

        # --- Contrainte plafond absolu / semaine / formation ---
        # Le plafond souple (MAX_H=40h) est géré dans l'objectif ;
        # ici on impose le plafond absolu (ABS_MAX_H) pour éviter
        # des solutions aberrantes tout en laissant de la flexibilité
        # quand la contiguïté l'exige.
        cap_groups = []
        if ft_filter & {0, 2}:
            cap_groups.append(({0, 2}, 'ftp'))
        if ft_filter & {1, 2}:
            cap_groups.append(({1, 2}, 'alt'))
        for w in sorted_weeks:
            for ft_set, label in cap_groups:
                items = [
                    (x[s['id']][w], int(round(s['slot_duration'] * SCALE)))
                    for s in sessions
                    if s['formation_type'] in ft_set and w in x[s['id']]
                ]
                if items:
                    model.add(
                        cp_model.LinearExpr.WeightedSum(
                            [v for v, _ in items], [d for _, d in items]
                        ) <= ABS_MAX_H * SCALE
                    )

        # --- Contraintes de capacité salle ---
        room_load = {}
        for s in sessions:
            sid = s['id']
            if sid in infeasible_sids or not session_weeks[sid] or not s['room_name']:
                continue
            rt = room_type_by_name.get(s['room_name'])
            if not rt or rt not in room_counts:
                continue
            sd_scaled = int(round(s['slot_duration'] * SCALE))
            for w in session_weeks[sid]:
                room_load.setdefault((rt, w), []).append((x[sid][w], sd_scaled))

        for (rt, w), items in room_load.items():
            cap = int(room_counts[rt] * ROOM_H * SCALE)
            model.add(
                cp_model.LinearExpr.WeightedSum(
                    [v for v, _ in items], [d for _, d in items]
                ) <= cap
            )

        # --- Contraintes d'ordonnancement (A avant B avec écart min) ---
        # Variables booléennes ob[sid][w] = 1 ssi x[sid][w] >= 1
        ob = {}
        def get_ob(sid, w):
            if sid not in ob:
                ob[sid] = {}
            if w not in ob[sid]:
                xv = x[sid].get(w)
                if xv is None:
                    ob[sid][w] = None
                    return None
                bv = model.new_bool_var(f'ob_{sid}_{w}')
                model.add(xv >= 1).only_enforce_if(bv)
                model.add(xv == 0).only_enforce_if(bv.Not())
                ob[sid][w] = bv
            return ob[sid][w]

        for ord_row in orderings:
            gap       = ord_row['min_gap_weeks']
            pred_sids = course_to_sessions.get(ord_row['course_id_pred'], [])
            succ_sids = course_to_sessions.get(ord_row['course_id_succ'], [])
            for ps in pred_sids:
                for ss_ in succ_sids:
                    for w_p in session_weeks.get(ps, []):
                        for w_s in session_weeks.get(ss_, []):
                            # succ doit être au moins (gap+1) semaines après pred
                            if week_to_idx.get(w_s, 0) <= week_to_idx.get(w_p, 0) + gap:
                                bp = get_ob(ps,  w_p)
                                bs = get_ob(ss_, w_s)
                                if bp is not None and bs is not None:
                                    model.add(bp + bs <= 1)

        # =========================================================
        # Objectif multi-critères
        # =========================================================
        obj_terms = []

        # (0) Pénalité forte pour dépassement du plafond souple (MAX_H)
        #     Priorité : contiguïté (hard) > plafond 40h (soft fort) > cible (soft)
        W_OVERFLOW = 100
        for w in sorted_weeks:
            for ft_set, label in cap_groups:
                items_ovf = [
                    (x[s['id']][w], int(round(s['slot_duration'] * SCALE)))
                    for s in sessions
                    if s['formation_type'] in ft_set and w in x[s['id']]
                ]
                if items_ovf:
                    load_ovf = cp_model.LinearExpr.WeightedSum(
                        [v for v, _ in items_ovf], [d for _, d in items_ovf]
                    )
                    ovf = model.new_int_var(0, (ABS_MAX_H - MAX_H) * SCALE,
                                            f'ovf_{label}_{w}')
                    model.add(ovf >= load_ovf - MAX_H * SCALE)
                    obj_terms.append(W_OVERFLOW * ovf)

        # (1) Cibles de charge hebdomadaire (poids fort)
        W_BALANCE = 10
        obj_groups = []
        if ft_filter & {0, 2}:
            obj_groups.append(({0, 2}, TGT_FTP, 'ftp'))
        if ft_filter & {1, 2}:
            obj_groups.append(({1, 2}, TGT_ALT, 'alt'))
        for w in sorted_weeks:
            for ft_set, target, label in obj_groups:
                items = [
                    (x[s['id']][w], int(round(s['slot_duration'] * SCALE)))
                    for s in sessions
                    if s['formation_type'] in ft_set and w in x[s['id']]
                ]
                if not items:
                    continue
                load_expr = cp_model.LinearExpr.WeightedSum(
                    [v for v, _ in items], [d for _, d in items]
                )
                target_scaled = target * SCALE
                max_load = ABS_MAX_H * SCALE
                dev_pos = model.new_int_var(0, max_load, f'dp_{label}_{w}')
                dev_neg = model.new_int_var(0, max_load, f'dn_{label}_{w}')
                model.add(load_expr - target_scaled <= dev_pos)
                model.add(target_scaled - load_expr <= dev_neg)
                obj_terms += [W_BALANCE * dev_pos, W_BALANCE * dev_neg]

        # (2) Étalement uniforme par cours — utile uniquement si max_pw > 1
        W_SPREAD = 1
        for s in sessions:
            sid    = s['id']
            max_pw = max(1, int(s['sessions_per_week_max'] or 1))
            if not session_weeks.get(sid) or sid in infeasible_sids or sid in _merged_secondary or max_pw <= 1:
                continue
            nb = s['nb_sessions']
            mw = model.new_int_var(0, nb, f'mw_{sid}')
            for w in session_weeks[sid]:
                if w in x[sid]:
                    model.add(mw >= x[sid][w])
            obj_terms.append(W_SPREAD * mw)

        # (3) Contiguïté souple : pénaliser les trous dans le bloc actif
        #     gap = (dernière semaine active - première + 1) - nb semaines actives
        W_CONTIG = 50
        for sid, (n_valid, acts_list, min_span) in _contig_data.items():
            if n_valid <= 1:
                continue
            first_idx = model.new_int_var(0, n_valid - 1, f'first_{sid}')
            last_idx  = model.new_int_var(0, n_valid - 1, f'last_{sid}')
            for i, act in enumerate(acts_list):
                model.add(first_idx <= i).only_enforce_if(act)
                model.add(last_idx  >= i).only_enforce_if(act)
            num_active = cp_model.LinearExpr.Sum(acts_list)
            gap_count = model.new_int_var(0, n_valid, f'gap_{sid}')
            model.add(gap_count >= last_idx - first_idx + 1 - num_active)
            obj_terms.append(W_CONTIG * gap_count)

        if obj_terms:
            model.minimize(cp_model.LinearExpr.Sum(obj_terms))

        # --- Résolution ---
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 60.0
        solver.parameters.num_search_workers  = 4
        status = solver.solve(model)

        # --- Helper : diagnostic de charge par formation ---
        # Formations à afficher dans le diagnostic (selon le choix utilisateur)
        diag_labels = []
        if formation_param == 'ftp':
            diag_labels = [('FTP', {0, 2})]
        elif formation_param == 'alt':
            diag_labels = [('ALT', {1, 2})]
        else:
            diag_labels = [('FTP', {0, 2}), ('ALT', {1, 2})]

        def _build_diagnostics():
            problems = []

            # 1. Sessions pré-exclues
            if infeasible_sids:
                for v in infeasible_sids.values():
                    problems.append(f"⛔ {v['label']} : {v['reason']}")

            # 2. Charge totale par formation
            load = {}
            wks  = {}
            for label, ft_set in diag_labels:
                load[label] = 0.0
                wks[label] = set()
            for s in sessions:
                sid = s['id']
                if sid in infeasible_sids:
                    continue
                ft = s['formation_type']
                total_h = s['nb_sessions'] * s['slot_duration']
                for label, ft_set in diag_labels:
                    if ft in ft_set:
                        load[label] += total_h
                        wks[label].update(session_weeks.get(sid, []))

            for label, _ in diag_labels:
                nw = len(wks[label])
                if nw > 0:
                    avg = round(load[label] / nw, 1)
                    marker = '🔴' if load[label] > MAX_H * nw else '📊'
                    problems.append(
                        f"{marker} {label} : {round(load[label], 1)}h à répartir "
                        f"sur {nw} semaines (moy. {avg}h/sem, max {MAX_H}h/sem)"
                    )

            # 3. Charge forcée minimum par semaine
            week_forced = {}
            for s in sessions:
                sid = s['id']
                if sid in infeasible_sids or not session_weeks.get(sid):
                    continue
                ft = s['formation_type']
                sw_list = session_weeks[sid]
                n_valid = len(sw_list)
                nb = s['nb_sessions']
                forced_h = s['slot_duration'] * nb / n_valid if n_valid > 0 else 0
                for w in sw_list:
                    for label, ft_set in diag_labels:
                        if ft in ft_set:
                            week_forced[(label, w)] = week_forced.get((label, w), 0) + forced_h

            overloaded_weeks = sorted(
                [(fl, w, h) for (fl, w), h in week_forced.items() if h > MAX_H * 0.9],
                key=lambda x: -x[2]
            )[:8]
            for fl, w, h in overloaded_weeks:
                marker = '🔴' if h > MAX_H else '🟡'
                problems.append(
                    f"{marker} {fl} semaine {w} : charge estimée ~{round(h, 1)}h "
                    f"(max {MAX_H}h)"
                )

            # 4. Sessions les plus contraintes (marge < 30%)
            tight = []
            for s in sessions:
                sid = s['id']
                if sid in infeasible_sids:
                    continue
                sw_list = session_weeks.get(sid, [])
                if not sw_list:
                    continue
                nb = s['nb_sessions']
                max_pw = max(1, int(s['sessions_per_week_max'] or 1))
                min_span = math.ceil(nb / max_pw)
                n_valid = len(sw_list)
                slack = (n_valid - min_span) / max(min_span, 1)
                if slack < 0.3:
                    tight.append((s, n_valid, min_span, slack))
            tight.sort(key=lambda x: x[3])
            for s, nv, ms, sl in tight[:8]:
                ft_label = _FT_LABELS[s['formation_type']]
                problems.append(
                    f"⚠️ {s['course_code']} {s['teaching_type']} {ft_label} : "
                    f"{nv} semaines dispo pour {ms} nécessaires "
                    f"({s['nb_sessions']} séances, max {int(s['sessions_per_week_max'] or 1)}/sem) — "
                    f"marge {round(sl*100)}%"
                )

            # 5. Contraintes d'ordonnancement problématiques
            if orderings:
                for ord_row in orderings:
                    pred_sids_list = course_to_sessions.get(ord_row['course_id_pred'], [])
                    succ_sids_list = course_to_sessions.get(ord_row['course_id_succ'], [])
                    for ps in pred_sids_list:
                        for ss_ in succ_sids_list:
                            pw = session_weeks.get(ps, [])
                            sw_ = session_weeks.get(ss_, [])
                            if pw and sw_:
                                last_pred = max(week_to_idx.get(w, 0) for w in pw)
                                first_succ = min(week_to_idx.get(w, 0) for w in sw_)
                                gap = ord_row['min_gap_weeks']
                                if first_succ <= last_pred + gap:
                                    ps_info = next((s for s in sessions if s['id'] == ps), None)
                                    ss_info = next((s for s in sessions if s['id'] == ss_), None)
                                    if ps_info and ss_info:
                                        problems.append(
                                            f"🔗 Ordre : {ps_info['course_code']} doit finir "
                                            f"{gap} sem. avant {ss_info['course_code']} "
                                            f"— plages se chevauchent"
                                        )

            return problems

        if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            problems = _build_diagnostics()

            # --- Diagnostic par élimination : identifier la contrainte bloquante ---
            root_causes = []

            def _diag_model_base():
                """Modèle minimal : juste les variables + somme = nb."""
                dm = cp_model.CpModel()
                dx = {}
                for s in sessions:
                    sid = s['id']
                    if sid in infeasible_sids or sid in _merged_secondary:
                        continue
                    nb = _merged_nb.get(sid, s['nb_sessions'])
                    mpw = max(1, int(s['sessions_per_week_max'] or 1))
                    dx[sid] = {}
                    for w in session_weeks[sid]:
                        dx[sid][w] = dm.new_int_var(0, mpw, f'dx_{sid}_{w}')
                    dm.add(cp_model.LinearExpr.Sum(
                        [dx[sid][w] for w in session_weeks[sid]]
                    ) == nb)
                # Alias secondaires CM+TD
                for sec_sid, pri_sid in _merged_secondary.items():
                    if sec_sid not in infeasible_sids and pri_sid in dx:
                        dx[sec_sid] = dx[pri_sid]
                return dm, dx

            def _diag_solve(dm, timeout=10.0):
                ds = cp_model.CpSolver()
                ds.parameters.max_time_in_seconds = timeout
                ds.parameters.num_search_workers = 4
                return ds.solve(dm) in (cp_model.OPTIMAL, cp_model.FEASIBLE)

            try:
                # Test 1 : sans contiguïté, sans cap, sans salles
                dm0, _ = _diag_model_base()
                base_ok = _diag_solve(dm0, 5.0)
                if not base_ok:
                    root_causes.append(
                        '🔴 Certaines sessions ont des plages de semaines '
                        'trop étroites pour y caser toutes les séances '
                        '(même sans limite de 40h/sem, salles ou contiguïté).'
                    )
                else:
                    # Test 2 : sans contiguïté, avec cap 40h, sans salles
                    dm1, dx1 = _diag_model_base()
                    for w in sorted_weeks:
                        for ft_set in cap_groups:
                            items = [
                                (dx1[s['id']][w], int(round(s['slot_duration'] * SCALE)))
                                for s in sessions
                                if s['formation_type'] in ft_set[0]
                                   and s['id'] not in infeasible_sids
                                   and w in dx1.get(s['id'], {})
                            ]
                            if items:
                                dm1.add(cp_model.LinearExpr.WeightedSum(
                                    [v for v, _ in items], [d for _, d in items]
                                ) <= MAX_H * SCALE)
                    cap_ok = _diag_solve(dm1)
                    if not cap_ok:
                        root_causes.append(
                            '🔴 La CAPACITÉ de 40h/semaine est dépassée '
                            '(même sans contiguïté ni contrainte de salles). '
                            'Solutions : élargir les plages de semaines '
                            'ou réduire le volume horaire.'
                        )
                    else:
                        # Test 3 : sans contiguïté, avec cap + salles
                        # (le modèle principal a contiguïté + cap + salles)
                        # Si sans contiguïté + cap ça passe,
                        # c'est la contiguïté et/ou les salles qui bloquent.
                        # Tester contiguïté seule (sans salles) :
                        # on réutilise le modèle principal mais sans room_load
                        # → plus simple : on teste si base + cap + salles passe
                        dm2, dx2 = _diag_model_base()
                        for w in sorted_weeks:
                            for ft_set in cap_groups:
                                items = [
                                    (dx2[s['id']][w], int(round(s['slot_duration'] * SCALE)))
                                    for s in sessions
                                    if s['formation_type'] in ft_set[0]
                                       and s['id'] not in infeasible_sids
                                       and w in dx2.get(s['id'], {})
                                ]
                                if items:
                                    dm2.add(cp_model.LinearExpr.WeightedSum(
                                        [v for v, _ in items], [d for _, d in items]
                                    ) <= MAX_H * SCALE)
                        for (rt, w), items_r in room_load.items():
                            cap = int(room_counts[rt] * ROOM_H * SCALE)
                            r_items = []
                            for s in sessions:
                                sid = s['id']
                                if sid in infeasible_sids or not s['room_name']:
                                    continue
                                if room_type_by_name.get(s['room_name']) == rt and w in dx2.get(sid, {}):
                                    r_items.append((dx2[sid][w], int(round(s['slot_duration'] * SCALE))))
                            if r_items:
                                dm2.add(cp_model.LinearExpr.WeightedSum(
                                    [v for v, _ in r_items], [d for _, d in r_items]
                                ) <= cap)
                        cap_room_ok = _diag_solve(dm2)
                        if not cap_room_ok:
                            root_causes.append(
                                '🔴 La CAPACITÉ DES SALLES est insuffisante '
                                '(sans contiguïté). '
                                'Solutions : ajouter des salles, modifier les types, '
                                'ou élargir les plages.'
                            )
                        else:
                            # Tous les tests de diagnostic passent
                            # → le solveur n'a pas trouvé de solution dans le temps imparti
                            root_causes.append(
                                '🟡 Les contraintes de base sont satisfaisables mais le '
                                'solveur n\'a pas trouvé de solution dans le temps imparti. '
                                'Essayez de relancer ou d\'élargir les plages de semaines.'
                            )
            except Exception:
                pass  # diagnostic optionnel

            if root_causes:
                problems = root_causes + [''] + problems

            return jsonify({
                'error': f'Aucune solution trouvée ({solver.status_name(status)})',
                'problems': problems if problems else [
                    'Cause indéterminée — vérifiez les plages de semaines, '
                    'le max séances/sem et les contraintes d\'ordonnancement.'
                ],
            }), 422

        # --- Écriture dans weekly_hours ---
        db = get_db()
        cursor = db.cursor()
        weeks_assigned = 0

        for s in sessions:
            sid = s['id']
            if not session_weeks[sid]:
                continue
            slot_dur = s['slot_duration']
            cursor.execute('DELETE FROM weekly_hours WHERE course_session_id = ?', (sid,))
            for w in session_weeks[sid]:
                if w not in x[sid]:
                    continue
                n = solver.value(x[sid][w])
                if n > 0:
                    cursor.execute(
                        'INSERT INTO weekly_hours (course_session_id, week_number, hours) VALUES (?, ?, ?)',
                        (sid, w, round(n * slot_dur, 2))
                    )
                    weeks_assigned += 1

        db.commit()

        # Détecter les dépassements du plafond souple (MAX_H)
        overflow_weeks = []
        for w in sorted_weeks:
            for ft_set, label in cap_groups:
                total_h = sum(
                    solver.value(x[s['id']][w]) * s['slot_duration']
                    for s in sessions
                    if s['formation_type'] in ft_set and w in x.get(s['id'], {})
                )
                if total_h > MAX_H:
                    overflow_weeks.append(
                        f"S{w} {label.upper()}: {round(total_h, 1)}h (plafond souple {MAX_H}h)"
                    )

        n_infeasible = len(infeasible_sids)
        infeasible_list = [
            f"{v['label']} : {v['reason']}" for v in infeasible_sids.values()
        ] if infeasible_sids else None

        # Détecter les sessions avec trous de contiguïté
        non_contig = []
        for sid, (n_valid, acts_list, min_span) in _contig_data.items():
            if n_valid <= 1:
                continue
            active_indices = [i for i, act in enumerate(acts_list) if solver.value(act)]
            if len(active_indices) >= 2:
                span = active_indices[-1] - active_indices[0] + 1
                gaps = span - len(active_indices)
                if gaps > 0:
                    s_info = next((s for s in sessions if s['id'] == sid), None)
                    if s_info:
                        non_contig.append(
                            f"{s_info['course_code']} {s_info['teaching_type']} "
                            f"{_FT_LABELS[s_info['formation_type']]}: {gaps} trou(s)"
                        )

        warnings = []
        if n_infeasible:
            warnings.append(
                f'{n_infeasible} session(s) ignorée(s) : pas assez de semaines valides '
                f'(vérifier plages de semaines / calendrier spécial).'
            )
        if overflow_weeks:
            warnings.append(
                f'{len(overflow_weeks)} semaine(s) dépassent {MAX_H}h : '
                + ', '.join(overflow_weeks)
            )
        if non_contig:
            warnings.append(
                f'{len(non_contig)} session(s) non contiguë(s) : '
                + ', '.join(non_contig)
            )

        return jsonify({
            'status':               solver.status_name(status),
            'sessions_optimized':   len([s for s in sessions if s['id'] not in infeasible_sids
                                         and session_weeks.get(s['id'])]),
            'weeks_assigned':       weeks_assigned,
            'infeasible_sessions':  n_infeasible,
            'infeasible_details':   infeasible_list,
            'overflow_weeks':       overflow_weeks if overflow_weeks else None,
            'non_contiguous':       non_contig if non_contig else None,
            'warning':              ' | '.join(warnings) if warnings else None,
        }), 200

    except Exception as e:
        return error_response(f'Erreur optimisation : {str(e)}', 500)

# ======================= APPLICATION STARTUP =======================

# Initialize database on module load (works with any WSGI runner)
os.makedirs('static', exist_ok=True)
init_db()
_init_promotions_db()
_init_programmes_db()

if __name__ == '__main__':
    # Run Flask app
    # debug est désactivé par défaut ; activez-le en local avec EDT_DEBUG=1.
    app.run(
        host='0.0.0.0',
        port=int(os.environ.get('PORT', 5000)),
        debug=os.environ.get('EDT_DEBUG', '0') == '1'
    )
